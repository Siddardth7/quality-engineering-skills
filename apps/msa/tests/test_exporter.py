"""Tests for msa_app/exporter.py — Gage R&R report export (W08-3).

MSA-specific export config composed over quality_core.io.export. These tests
load the produced .xlsx back with openpyxl and assert structure/values, decode
the CSVs, assert user-derived strings are formula-injection-escaped, and check
PDFs for valid bytes. The shared export machinery itself is tested in
quality-core.
"""

from __future__ import annotations

import io
import re
import zlib
from typing import Any

import msa_app.exporter as exporter
import openpyxl
import pandas as pd
import pytest
from msa_app.exporter import (
    VERDICT_SENTENCES,
    GageStudyReport,
    _detail_rows,
    export_csv,
    export_excel,
    export_pdf,
    export_results_csv,
    verdict_sentence,
)
from msa_app.gage_rr_engine import METHOD, METHOD_NOTE, compute_gage_rr

# --- Fixtures ----------------------------------------------------------------

STUDY_ROWS = [
    {"part": "P01", "appraiser": "A", "trial": 1, "measurement": 10.05},
    {"part": "P01", "appraiser": "A", "trial": 2, "measurement": 10.02},
    {"part": "P02", "appraiser": "B", "trial": 1, "measurement": 9.98},
    {"part": "P02", "appraiser": "B", "trial": 2, "measurement": 10.01},
]

RESULTS = {
    "ev": 0.03,
    "av": 0.02,
    "grr": 0.036,
    "pv": 0.5,
    "tv": 0.501,
    "pev_study": 5.99,
    "pav_study": 3.99,
    "pgrr_study": 7.19,
    "ppv_study": 99.80,
    "pev_tolerance": 10.0,
    "pav_tolerance": 8.0,
    "pgrr_tolerance": 12.0,
    "ppv_tolerance": 150.0,
    "ndc": 6,
    "verdict": "Accept",
    "mean": 10.0,
    "n_parts": 2,
    "n_appraisers": 2,
    "n_trials": 2,
    "is_balanced": True,
    "method": "average_and_range",
    "method_note": "Average-and-Range method: the part x appraiser interaction is NOT estimated.",
}


def _report(**overrides: Any) -> GageStudyReport:
    raw_res = overrides.pop("results", {})
    results_override = raw_res if isinstance(raw_res, dict) else {}
    results = {**RESULTS, **results_override}
    kwargs = {"usl": 10.5, "lsl": 9.5, **overrides}
    return GageStudyReport(
        study=pd.DataFrame(STUDY_ROWS),
        results=results,
        **kwargs,  # type: ignore[arg-type]
    )


def _no_tolerance_report() -> GageStudyReport:
    return GageStudyReport(
        study=pd.DataFrame(STUDY_ROWS),
        results={
            **RESULTS,
            "pev_tolerance": None,
            "pav_tolerance": None,
            "pgrr_tolerance": None,
            "ppv_tolerance": None,
        },
        usl=None,
        lsl=None,
    )


def _kv_sheet_to_dict(ws) -> dict[str, object]:
    return {row[0].value: row[1].value for row in ws.iter_rows(min_row=1, max_col=2) if row[0].value}


# --- verdict_sentence ----------------------------------------------------------


@pytest.mark.parametrize("verdict", ["Accept", "Marginal", "Reject"])
def test_verdict_sentence_returns_configured_text(verdict):
    assert verdict_sentence(verdict) == VERDICT_SENTENCES[verdict]


def test_verdict_sentence_fallback_for_unknown_verdict():
    assert verdict_sentence("Bogus") == "Unrecognized verdict — review the study inputs."


# --- _detail_rows --------------------------------------------------------------


def test_detail_rows_tolerance_present():
    rows = dict(_detail_rows(_report()))
    assert rows["%GRR (Tolerance)"] == "12.00%"
    assert rows["USL"] == "10.500000"
    assert rows["LSL"] == "9.500000"
    assert rows["Verdict"] == "Accept"
    assert rows["Verdict Interpretation"] == VERDICT_SENTENCES["Accept"]


def test_detail_rows_tolerance_absent():
    rows = dict(_detail_rows(_no_tolerance_report()))
    assert rows["%GRR (Tolerance)"] == "N/A"
    assert rows["USL"] == "N/A"
    assert rows["LSL"] == "N/A"


# --- export_csv (study data) ----------------------------------------------------


def test_export_csv_header_and_row_count():
    data = export_csv(_report())
    frame = pd.read_csv(io.BytesIO(data))
    assert list(frame.columns) == ["part", "appraiser", "trial", "measurement"]
    assert len(frame) == len(STUDY_ROWS)


def test_export_csv_escapes_formula_injection_in_part():
    report = GageStudyReport(
        study=pd.DataFrame(
            [{"part": "=cmd", "appraiser": "A", "trial": 1, "measurement": 10.0}]
        ),
        results=RESULTS,
        usl=None,
        lsl=None,
    )
    data = export_csv(report)
    text = data.decode("utf-8")
    assert "'=cmd" in text


def test_export_csv_escapes_formula_injection_in_appraiser():
    report = GageStudyReport(
        study=pd.DataFrame(
            [{"part": "P01", "appraiser": "+SUM(A1)", "trial": 1, "measurement": 10.0}]
        ),
        results=RESULTS,
        usl=None,
        lsl=None,
    )
    data = export_csv(report)
    text = data.decode("utf-8")
    assert "'+SUM(A1)" in text


# --- export_results_csv (R1, flat metrics) --------------------------------------


def test_export_results_csv_contains_expected_columns_and_values():
    data = export_results_csv(_report())
    frame = pd.read_csv(io.BytesIO(data))
    assert list(frame.columns) == [
        "EV",
        "AV",
        "GRR",
        "PV",
        "TV",
        "%EV Study",
        "%AV Study",
        "%GRR Study",
        "%PV Study",
        "%EV Tolerance",
        "%AV Tolerance",
        "%GRR Tolerance",
        "%PV Tolerance",
        "ndc",
        "Verdict",
        "Verdict Interpretation",
        "Method",
        "Method Limitation",
    ]
    assert len(frame) == 1
    row = frame.iloc[0]
    assert row["ndc"] == 6
    assert row["Verdict"] == "Accept"
    assert row["%GRR Tolerance"] == "12.00%"


def test_export_results_csv_reports_na_when_tolerance_absent():
    data = export_results_csv(_no_tolerance_report())
    frame = pd.read_csv(io.BytesIO(data), na_filter=False)
    assert frame.iloc[0]["%GRR Tolerance"] == "N/A"


@pytest.mark.parametrize("verdict", ["Accept", "Marginal", "Reject"])
def test_export_results_csv_verdict_interpretation_per_verdict(verdict):
    data = export_results_csv(_report(results={"verdict": verdict}))
    frame = pd.read_csv(io.BytesIO(data))
    assert frame.iloc[0]["Verdict Interpretation"] == VERDICT_SENTENCES[verdict]


def test_export_results_csv_fallback_for_unknown_verdict():
    data = export_results_csv(_report(results={"verdict": "Bogus"}))
    frame = pd.read_csv(io.BytesIO(data))
    assert frame.iloc[0]["Verdict Interpretation"] == "Unrecognized verdict — review the study inputs."


# --- export_excel ----------------------------------------------------------------


def test_export_excel_structure_and_values():
    data = export_excel(_report())
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Summary", "Study Data"]

    summary = _kv_sheet_to_dict(wb["Summary"])
    assert summary["Verdict"] == "Accept"
    assert summary["%GRR (Study)"] == "7.19%"
    assert summary["%GRR (Tolerance)"] == "12.00%"
    assert summary["Tool Version"]

    data_ws = wb["Study Data"]
    assert [c.value for c in data_ws[1]] == ["part", "appraiser", "trial", "measurement"]
    assert data_ws.max_row == 1 + len(STUDY_ROWS)


def test_export_excel_reports_na_when_tolerance_and_limits_absent():
    wb = openpyxl.load_workbook(io.BytesIO(export_excel(_no_tolerance_report())))
    summary = _kv_sheet_to_dict(wb["Summary"])
    assert summary["%GRR (Tolerance)"] == "N/A"
    assert summary["USL"] == "N/A"
    assert summary["LSL"] == "N/A"


def test_export_excel_escapes_injection_in_study_data():
    report = GageStudyReport(
        study=pd.DataFrame(
            [{"part": "=cmd", "appraiser": "A", "trial": 1, "measurement": 10.0}]
        ),
        results=RESULTS,
        usl=None,
        lsl=None,
    )
    wb = openpyxl.load_workbook(io.BytesIO(export_excel(report)))
    data_ws = wb["Study Data"]
    assert str(data_ws.cell(row=2, column=1).value).startswith("'=")


# --- export_pdf --------------------------------------------------------------------


def test_export_pdf_is_valid_bytes():
    data = export_pdf(_report())
    assert data.startswith(b"%PDF")
    assert len(data) > 500


def test_export_pdf_handles_tolerance_absent():
    data = export_pdf(_no_tolerance_report())
    assert data.startswith(b"%PDF")


# --- Method declaration reaches the artifacts (#194 / audit A10) ----------------
# The finding's harm is a consumer receiving a %GRR indistinguishable from an
# ANOVA result. The consumer-facing surface is these four files, so the
# declaration has to survive the trip through each of them, unaltered.


def _production_report() -> GageStudyReport:
    """A report carrying the *real* engine output, including the full-length METHOD_NOTE.

    The module-level RESULTS fixture deliberately shortens ``method_note`` for
    readability; this builds the genuine 351-char production string so the
    exporters are exercised with what they will actually receive.
    """
    rows = [
        {
            "part": f"P{p}",
            "appraiser": a,
            "trial": t,
            "measurement": float(p) + 0.01 * t + bias,
        }
        for p in range(1, 7)
        for a, bias in [("A", 0.0), ("B", 0.02)]
        for t in (1, 2)
    ]
    study = pd.DataFrame(rows)
    return GageStudyReport(
        study=study,
        results=compute_gage_rr(study, tolerance=4.0),
        usl=12.0,
        lsl=8.0,
    )


def test_detail_rows_carry_the_method_declaration():
    rows = dict(_detail_rows(_report()))
    assert rows["Method"] == "average_and_range"
    assert rows["Method Limitation"] == RESULTS["method_note"]


def test_detail_rows_carry_the_method_declaration_without_tolerance():
    """The declaration is not tolerance-conditional in the export either."""
    rows = dict(_detail_rows(_no_tolerance_report()))
    assert rows["Method"] == "average_and_range"
    assert rows["Method Limitation"] == RESULTS["method_note"]


def test_export_results_csv_carries_the_method_declaration():
    frame = pd.read_csv(io.BytesIO(export_results_csv(_report())))
    row = frame.iloc[0]
    assert row["Method"] == "average_and_range"
    assert row["Method Limitation"] == RESULTS["method_note"]


def test_export_excel_summary_carries_the_method_declaration():
    wb = openpyxl.load_workbook(io.BytesIO(export_excel(_report())))
    summary = _kv_sheet_to_dict(wb["Summary"])
    assert summary["Method"] == "average_and_range"
    assert summary["Method Limitation"] == RESULTS["method_note"]


def test_results_csv_round_trips_the_full_length_production_method_note():
    """METHOD_NOTE contains embedded double quotes — it must survive CSV quoting intact."""
    frame = pd.read_csv(io.BytesIO(export_results_csv(_production_report())))
    assert frame.iloc[0]["Method"] == METHOD
    assert frame.iloc[0]["Method Limitation"] == METHOD_NOTE
    assert len(frame) == 1  # the quoted note did not spill into a second row


def test_export_excel_carries_the_full_length_production_method_note():
    wb = openpyxl.load_workbook(io.BytesIO(export_excel(_production_report())))
    summary = _kv_sheet_to_dict(wb["Summary"])
    assert summary["Method"] == METHOD
    assert summary["Method Limitation"] == METHOD_NOTE


def _pdf_text(data: bytes) -> str:
    """Recover the visible text of a PDF by inflating its content streams.

    fpdf2 Flate-compresses page content, so a substring search over the raw
    bytes finds nothing — the stream must be inflated first. This helper is
    deliberately insensitive to fpdf2's byte-level output, which varies by
    platform (macOS locally vs the Linux CI runner) even at a pinned version:

    - Streams are decoded with a ``decompressobj`` fed everything after the
      ``stream`` keyword. It stops at the zlib stream's own end, so it is immune
      to where the trailing ``endstream`` falls — unlike a ``stream...endstream``
      delimiter regex, which truncates if the compressed bytes happen to contain
      that marker and then loses all text to a failed decompress.
    - Every parenthesised string literal is collected, regardless of the show
      operator (``Tj`` vs a ``[...] TJ`` array). The literal parens inside
      METHOD_NOTE are escaped as ``\\(`` / ``\\)``, so the regex skips escaped
      characters rather than stopping at the first ``)``; the escapes are undone
      afterwards. In a page content stream every ``(...)`` literal is show-text.
    """
    chunk_re = re.compile(rb"\((?P<body>(?:[^\\()]|\\.)*)\)", re.S)
    out = []
    for m in re.finditer(rb"(?<!end)stream\r?\n", data):
        tail = data[m.end():]
        try:
            body = zlib.decompressobj().decompress(tail)
        except zlib.error:
            end = tail.find(b"endstream")
            body = tail[:end] if end != -1 else tail
        out.append(b"".join(mm.group("body") for mm in chunk_re.finditer(body)))
    joined = b"".join(out)
    joined = joined.replace(rb"\(", b"(").replace(rb"\)", b")")
    return joined.decode("latin-1")


def test_export_pdf_renders_with_the_full_length_production_method_note():
    """fpdf2/latin-1 guard: the real 351-char note must not raise or truncate the file.

    The baseline is the SAME report with only ``method_note`` emptied. An earlier
    version built it from the short module-level RESULTS fixture — a different
    report with different metrics and a different verdict — so the length delta
    did not isolate the note, and truncating or deleting the method rows passed
    the test anyway.
    """
    report = _production_report()
    baseline = export_pdf(
        GageStudyReport(
            study=report.study,
            results={**report.results, "method_note": ""},
            usl=report.usl,
            lsl=report.lsl,
        )
    )
    data = export_pdf(report)
    assert data.startswith(b"%PDF")
    # Same report, same everything but the note -> the delta IS the note.
    assert len(data) > len(baseline)

    # A length delta only proves SOMETHING was added, so it survives the note
    # being truncated. Assert the note's own text is in the page stream: every
    # character has to reach the PDF, not just enough of them to grow the file.
    text = _pdf_text(data)
    assert METHOD in text
    assert METHOD_NOTE in text
    assert METHOD_NOTE not in _pdf_text(baseline)


# =============================================================================
# Eight percentages reach the artifacts (#225, audit A10-c)
# =============================================================================
# _detail_rows went from two %GRR rows to eight; export_results_csv from two
# columns to eight. The PDF headline strip (D-7) deliberately did NOT change.

_STUDY_LABELS = ["%EV (Study)", "%AV (Study)", "%GRR (Study)", "%PV (Study)"]
_TOLERANCE_LABELS = [
    "%EV (Tolerance)",
    "%AV (Tolerance)",
    "%GRR (Tolerance)",
    "%PV (Tolerance)",
]


def test_detail_rows_carry_all_eight_percentages_in_aiag_form_order():
    """AIAG's form column reads EV, AV, GRR, PV (Figure III-B 16); study block first."""
    labels = [label for label, _ in _detail_rows(_report())]
    percentage_labels = [lb for lb in labels if lb.startswith("%")]
    assert percentage_labels == _STUDY_LABELS + _TOLERANCE_LABELS


def test_detail_rows_render_each_percentage_from_its_own_result_key():
    """Each row shows its own metric -- no row is wired to the wrong key.

    The RESULTS fixture gives all eight distinct values precisely so a copy-paste
    of the %GRR row into the %EV slot is visible here.
    """
    rows = dict(_detail_rows(_report()))
    assert rows["%EV (Study)"] == "5.99%"
    assert rows["%AV (Study)"] == "3.99%"
    assert rows["%GRR (Study)"] == "7.19%"
    assert rows["%PV (Study)"] == "99.80%"
    assert rows["%EV (Tolerance)"] == "10.00%"
    assert rows["%AV (Tolerance)"] == "8.00%"
    assert rows["%GRR (Tolerance)"] == "12.00%"
    assert rows["%PV (Tolerance)"] == "150.00%"


def test_detail_rows_tolerance_percentages_are_all_na_without_tolerance():
    """E-2 through the export layer: four N/A, and the study four still render."""
    rows = dict(_detail_rows(_no_tolerance_report()))
    for label in _TOLERANCE_LABELS:
        assert rows[label] == "N/A", label
    for label in _STUDY_LABELS:
        assert rows[label].endswith("%") and rows[label] != "N/A", label


def test_percentages_over_100_are_not_clamped_by_the_exporters():
    """E-3: %PV vs tolerance legitimately exceeds 100% and must survive verbatim."""
    rows = dict(_detail_rows(_report()))
    assert rows["%PV (Tolerance)"] == "150.00%"

    frame = pd.read_csv(io.BytesIO(export_results_csv(_report())))
    assert frame.iloc[0]["%PV Tolerance"] == "150.00%"

    summary = _kv_sheet_to_dict(
        openpyxl.load_workbook(io.BytesIO(export_excel(_report())))["Summary"]
    )
    assert summary["%PV (Tolerance)"] == "150.00%"

    assert "150.00%" in _pdf_text(export_pdf(_report()))


def test_percentages_over_100_survive_real_engine_output_end_to_end():
    """Same as above but with figures the engine actually produced, not a fixture."""
    report = _production_report()
    assert report.results["ppv_tolerance"] > 100.0

    frame = pd.read_csv(io.BytesIO(export_results_csv(report)))
    expected = f"{report.results['ppv_tolerance']:.2f}%"
    assert frame.iloc[0]["%PV Tolerance"] == expected
    assert float(expected.rstrip("%")) > 100.0


def test_infinite_study_percentages_render_through_the_exporters():
    """E-5: TV == 0 gives inf; _fmt_pct emits 'inf%' rather than raising."""
    infinite = {label: float("inf") for label in ("pev_study", "pav_study", "pgrr_study", "ppv_study")}
    rows = dict(_detail_rows(_report(results=infinite)))
    for label in _STUDY_LABELS:
        assert rows[label] == "inf%", label


def test_export_results_csv_carries_all_eight_percentage_values():
    frame = pd.read_csv(io.BytesIO(export_results_csv(_report())))
    row = frame.iloc[0]
    assert row["%EV Study"] == "5.99%"
    assert row["%AV Study"] == "3.99%"
    assert row["%GRR Study"] == "7.19%"
    assert row["%PV Study"] == "99.80%"
    assert row["%EV Tolerance"] == "10.00%"
    assert row["%AV Tolerance"] == "8.00%"
    assert row["%GRR Tolerance"] == "12.00%"
    assert row["%PV Tolerance"] == "150.00%"


def test_export_results_csv_all_tolerance_percentages_na_when_absent():
    frame = pd.read_csv(io.BytesIO(export_results_csv(_no_tolerance_report())), na_filter=False)
    row = frame.iloc[0]
    for column in ("%EV Tolerance", "%AV Tolerance", "%GRR Tolerance", "%PV Tolerance"):
        assert row[column] == "N/A", column


def test_export_excel_summary_carries_all_eight_percentages():
    summary = _kv_sheet_to_dict(
        openpyxl.load_workbook(io.BytesIO(export_excel(_report())))["Summary"]
    )
    for label in _STUDY_LABELS + _TOLERANCE_LABELS:
        assert label in summary, label
    assert summary["%EV (Study)"] == "5.99%"
    assert summary["%PV (Tolerance)"] == "150.00%"


def test_export_pdf_headline_strip_still_has_exactly_four_cells(monkeypatch):
    """D-7: the eight percentages reach the PDF via the detail table, not the strip.

    Feeding ten cells to pdf_summary_cells would wreck the fixed 4-metric layout.
    """
    captured = []
    original = exporter.pdf_summary_cells
    monkeypatch.setattr(
        exporter,
        "pdf_summary_cells",
        lambda pdf, cells: (captured.append(list(cells)), original(pdf, cells))[1],
    )
    export_pdf(_report())

    assert len(captured) == 1
    assert [label for label, _ in captured[0]] == [
        "%GRR (Study)",
        "%GRR (Tol)",
        "ndc",
        "Verdict",
    ]


def test_export_pdf_detail_table_carries_all_eight_percentage_labels():
    text = _pdf_text(export_pdf(_report()))
    for label in _STUDY_LABELS + _TOLERANCE_LABELS:
        assert label in text, label
