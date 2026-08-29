"""
quality_core/io/export_fmea.py
FMEA Excel exporter — one live ``=S*O*D`` RPN formula per row.

Sibling of the domain-agnostic ``quality_core.io.export`` primitives: that module owns
the cross-cutting machinery (formula-injection escaping, openpyxl table styling, the
``Formula`` opt-in marker) and stays free of domain knowledge; this module supplies the
FMEA *config* — column order, widths, and which single column is a live formula.

What ships in the workbook:

  - the RPN column is the **only** :class:`~quality_core.io.export.Formula` cell. It is
    written as ``=<S><r>*<O><r>*<D><r>``, referencing that same row's Severity,
    Occurrence, and Detection data cells, so an analyst editing an S/O/D rating in Excel
    sees RPN recalculate rather than reading a stale cached number.
  - the Action Priority column is a **structured lookup value**, not a formula: the
    ``"High"``/``"Medium"``/``"Low"`` string returned by
    :func:`quality_core.scoring.action_priority`. The AIAG-VDA 2019 AP table is not
    expressible as elementary spreadsheet arithmetic and is never re-derived here.
  - every other cell is plain data from :class:`~quality_core.schema.fmea.FMEARow` and
    goes through ``write_table_sheet``'s ordinary ``sanitize_cell`` escaping. Nothing in
    this module ever constructs a ``Formula`` from ingested text — only from the fixed
    formula template below — so a free-text field that happens to start with ``"="``
    stays inert.

Row order is the caller's: rows are exported in ``dataset.rows`` list order, **never
re-sorted** by ID, RPN, or AP. That is the pinned contract — the per-row formula
references stay correct under any order, so a "fix" that sorted them would only surprise
callers. Column letters are derived from ``FMEA_EXPORT_COLUMNS`` via
``get_column_letter``, never hand-typed, so reordering the column tuple cannot silently
desync a formula from its own header.

No new standards constant lives here — see ``io/ASSUMPTIONS_LOG.md`` (RULE-IO-003).
"""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from quality_core.io.export import Formula, write_table_sheet
from quality_core.schema.fmea import FMEADataset, FMEARow
from quality_core.scoring import action_priority

# ===========================================================================
# Sheet layout
# ===========================================================================

#: Exported columns, in sheet order (1-based column index = position + 1).
FMEA_EXPORT_COLUMNS: tuple[str, ...] = (
    "ID",
    "Process_Step",
    "Component",
    "Function",
    "Failure_Mode",
    "Effect",
    "Severity",
    "Cause",
    "Occurrence",
    "Current_Control",
    "Detection",
    "RPN",
    "Action_Priority",
)

#: Column widths keyed by column name — wide for free text, narrow for ratings.
#: Presentation only; ``write_table_sheet`` falls back to ``default_width`` for
#: anything missing, but every named column is given an explicit width here.
FMEA_COL_WIDTHS: dict[str, float] = {
    "ID": 6,
    "Process_Step": 24,
    "Component": 20,
    "Function": 26,
    "Failure_Mode": 30,
    "Effect": 28,
    "Severity": 8,
    "Cause": 26,
    "Occurrence": 10,
    "Current_Control": 28,
    "Detection": 10,
    "RPN": 8,
    "Action_Priority": 14,
}

#: Sheet row of the first data row: row 1 is the header written by
#: ``write_table_sheet``, whose body enumeration starts at 2.
_FIRST_DATA_ROW = 2


def _column_letter(name: str) -> str:
    """Return the Excel column letter for ``name``'s position in the export layout."""
    return get_column_letter(FMEA_EXPORT_COLUMNS.index(name) + 1)


_SEVERITY_COL = _column_letter("Severity")
_OCCURRENCE_COL = _column_letter("Occurrence")
_DETECTION_COL = _column_letter("Detection")


def _rpn_formula(sheet_row: int) -> Formula:
    """Return the live ``=S*O*D`` RPN formula referencing ``sheet_row``'s ratings.

    Exporter-authored: the only ``Formula`` this module ever builds. No number format
    is applied — RPN is a plain integer product, so the cell keeps Excel's ``General``.
    """
    return Formula(
        f"={_SEVERITY_COL}{sheet_row}*{_OCCURRENCE_COL}{sheet_row}*{_DETECTION_COL}{sheet_row}"
    )


def _row_record(row: FMEARow, sheet_row: int) -> dict[str, object]:
    """Map one ``FMEARow`` to its sheet record, keyed by ``FMEA_EXPORT_COLUMNS``."""
    return {
        "ID": row.ID,
        "Process_Step": row.Process_Step,
        "Component": row.Component,
        "Function": row.Function,
        "Failure_Mode": row.Failure_Mode,
        "Effect": row.Effect,
        "Severity": row.Severity,
        "Cause": row.Cause,
        "Occurrence": row.Occurrence,
        "Current_Control": row.Current_Control,
        "Detection": row.Detection,
        "RPN": _rpn_formula(sheet_row),
        "Action_Priority": action_priority(row.Severity, row.Occurrence, row.Detection),
    }


# ===========================================================================
# Public API
# ===========================================================================


def export_fmea_workbook(dataset: FMEADataset, *, title: str = "FMEA") -> bytes:
    """Export ``dataset`` to an .xlsx workbook with a live RPN formula per row.

    Returns the serialized single-sheet workbook as bytes, ready to hand to a download
    button or write to disk. ``title`` names the worksheet (default ``"FMEA"``); it is
    passed through to ``write_table_sheet`` unvalidated, matching every other caller.

    Rows are written in ``dataset.rows`` order. An empty dataset yields a valid workbook
    carrying the header row only.
    """
    records = [
        _row_record(row, index + _FIRST_DATA_ROW) for index, row in enumerate(dataset.rows)
    ]
    # Explicit ``columns=`` so an empty dataset still carries the full header: a
    # DataFrame built from an empty record list would otherwise have no columns at
    # all, and ``write_table_sheet`` would filter every column out.
    df = pd.DataFrame(records, columns=list(FMEA_EXPORT_COLUMNS))

    wb = Workbook()
    write_table_sheet(
        wb.active,
        df,
        title=title,
        columns=FMEA_EXPORT_COLUMNS,
        col_widths=FMEA_COL_WIDTHS,
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===========================================================================
# Benchmark dataset
# ===========================================================================

# Self-contained AIAG-style PFMEA rows. Deliberately NOT imported from
# ``quality_core.canvas`` — canvas sits structurally above ``io``, so depending on it
# here would invert the layering. The S/O/D ratings span distinct AIAG-VDA bands so the
# exported Action Priority column exercises High, Medium, and Low.

_BENCHMARK_ROWS: tuple[FMEARow, ...] = (
    FMEARow(
        ID=1,
        Process_Step="Inverter SMT Assembly",
        Component="Traction Inverter",
        Function="Gate driver desaturation protection",
        Failure_Mode="Gate driver desaturation during high-torque acceleration",
        Effect="Inverter shutdown, loss of vehicle propulsion",
        Severity=10,
        Cause="IGBT transient overvoltage spike",
        Occurrence=6,
        Current_Control="Periodic sampling of gate-drive waveform",
        Detection=7,
    ),
    FMEARow(
        ID=2,
        Process_Step="Brake Module Machining",
        Component="Brake-by-Wire Actuator",
        Function="Hydraulic pressure measurement",
        Failure_Mode="Pressure transducer calibration drift",
        Effect="Degraded pedal feel, warning indicator displayed",
        Severity=9,
        Cause="Piezoresistive element thermal fatigue",
        Occurrence=4,
        Current_Control="100% end-of-line calibrated pressure sweep",
        Detection=1,
    ),
    FMEARow(
        ID=3,
        Process_Step="Steering Sensor Assembly",
        Component="Steering Column Module",
        Function="Steering angle measurement",
        Failure_Mode="Steering angle sensor CAN frame drop",
        Effect="Transient loss of lane centering assist",
        Severity=8,
        Cause="Connector pin fretting corrosion",
        Occurrence=3,
        Current_Control="Operator visual check of connector seating",
        Detection=8,
    ),
    FMEARow(
        ID=4,
        Process_Step="Stator Winding",
        Component="Traction Motor",
        Function="Stator phase-to-phase insulation",
        Failure_Mode="Stator insulation dielectric breakdown",
        Effect="Motor ground fault, reduced torque output",
        Severity=7,
        Cause="Varnish void during trickle impregnation",
        Occurrence=8,
        Current_Control="In-line high-potential dielectric test",
        Detection=5,
    ),
    FMEARow(
        ID=5,
        Process_Step="Gearbox Final Assembly",
        Component="Reduction Gearbox",
        Function="Lubricant retention",
        Failure_Mode="Output shaft rotary oil seal extrusion",
        Effect="Oil seepage onto underbody, non-safety effect",
        Severity=5,
        Cause="Seal lip deformation under thermal cycling",
        Occurrence=4,
        Current_Control="Press-force depth monitoring",
        Detection=3,
    ),
    FMEARow(
        ID=6,
        Process_Step="Trim and Final Line",
        Component="Interior Trim Panel",
        Function="Panel gap appearance",
        Failure_Mode="Door trim panel gap above appearance limit",
        Effect="Cosmetic gap visible at customer delivery",
        Severity=3,
        Cause="Clip retention tolerance stack-up",
        Occurrence=2,
        Current_Control="Gauge check on every shift's first-off part",
        Detection=2,
    ),
)


def benchmark_fmea_dataset() -> FMEADataset:
    """Return a self-contained AIAG-style benchmark ``FMEADataset`` for exports/demos.

    Six automotive PFMEA rows whose ratings span the AIAG-VDA Severity/Occurrence/
    Detection bands, so the exported Action Priority column contains High, Medium, and
    Low. A fresh ``FMEADataset`` is built per call, so a caller mutating the returned
    dataset cannot affect the next one.
    """
    return FMEADataset(rows=list(_BENCHMARK_ROWS))
