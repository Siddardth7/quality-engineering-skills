"""
tests/test_ppap_export.py
Tests for quality_core/ppap/export.py — the PPAP submission-readiness live-formula .xlsx
exporter (#148).

Modelled on the sibling Control Plan suite (#145). The only live formulas in the whole
workbook are the four ``Completeness`` roll-ups (``B6`` ``=COUNTA``, ``B7``/``B8``/``B9``
``=COUNTIF(...,"<verdict>")``) and the single ``Capability Gate`` ``B4`` ``=IF/IF`` band
cascade — and ``B4`` only when the engine's own ``ProcessStudyResult.band is not None``.
Every other cell (the 18-row checklist, all metric labels, every literal value) is inert.

The two mandatory negative controls: a literal-value metric build must FAIL
``assert_cell_is_formula`` (proving the check is load-bearing, not tautological), and a
free-text field starting with ``"="`` must render inert while the Completeness roll-ups stay
live and still count that row (per-cell isolation). The accuracy scorecard reads each
formula's OWN referenced range/search-string back out of the *saved* sheet and recomputes
COUNTA/COUNTIF and the IF/IF cascade in Python against the IMPORTED threshold constants — it
never trusts ``data_only=True`` (openpyxl caches no value for a never-recalculated formula,
so that path reads ``None``).

The authority invariant (Section 5 dispositions Approved / Interim Approval / Rejected are
the customer's alone and never appear here) gets its own positive regression guard.
"""

from __future__ import annotations

import io
import re

import openpyxl
import pytest
from _xlsx_formula_audit import assert_cell_is_formula
from openpyxl.utils import get_column_letter
from quality_core.ppap import (
    ACCEPTANCE_THRESHOLD_CAPABLE,
    ACCEPTANCE_THRESHOLD_POTENTIALLY_CAPABLE,
    ACTION_ATTRIBUTE_DATA,
    ACTION_UNSTABLE,
    assess_initial_process_study,
    audit_ppap_package,
    benchmark_ppap_package,
    build_ppap_workbook,
    export_ppap_workbook,
)
from quality_core.ppap.export import PPAP_COL_WIDTHS, PPAP_EXPORT_COLUMNS
from quality_core.ppap.schema import PPAP_ELEMENT_IDS

# Column letters derived exactly as the exporter derives them, so a reorder of
# PPAP_EXPORT_COLUMNS moves the test's expectations with the implementation's.
_ELEMENT_ID_COL = get_column_letter(PPAP_EXPORT_COLUMNS.index("Element_ID") + 1)
_AUDIT_VERDICT_COL = get_column_letter(PPAP_EXPORT_COLUMNS.index("Audit_Verdict") + 1)
_DOCREF_COL = get_column_letter(PPAP_EXPORT_COLUMNS.index("Document_Reference") + 1)

_DEFAULT_TITLE = "PPAP Checklist"
_COMPLETENESS = "Completeness"
_CAPABILITY = "Capability Gate"

# The three COUNTIF-rolled verdicts, paired with their B-row on the Completeness sheet.
_VERDICT_ROWS: tuple[tuple[str, str], ...] = (
    ("B7", "SUBMITTED"),
    ("B8", "RETAINED_ON_FILE"),
    ("B9", "MISSING"),
)


def _saved(wb: openpyxl.Workbook) -> bytes:
    """Serialize a workbook to .xlsx bytes (tests never touch disk)."""
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _load(wb_bytes: bytes) -> openpyxl.Workbook:
    """Reload saved bytes with formulas intact (never data_only — no cache exists)."""
    return openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)


_RANGE_RE = re.compile(r"([A-Z]+)(\d+):([A-Z]+)(\d+)")
_COUNTIF_CRITERION_RE = re.compile(r'COUNTIF\([^,]+,"([^"]+)"\)')
# Two decimal thresholds embedded in the B4 IF/IF cascade, in file order (capable, then
# potentially-capable).
_THRESHOLD_RE = re.compile(r">=?(\d+\.\d+)")


def _referenced_row_span(formula: str) -> tuple[str, int, int]:
    """Parse ``col``, ``first_row``, ``last_row`` out of a saved COUNTA/COUNTIF formula.

    Reads the formula's OWN referenced range so the recomputation below counts exactly the
    cells Excel would, rather than a range the test re-derives independently.
    """
    m = _RANGE_RE.search(formula)
    assert m is not None, f"no A1 range found in formula {formula!r}"
    col, r1, r2 = m.group(1), int(m.group(2)), int(m.group(4))
    assert col == m.group(3), f"range spans two columns in {formula!r}"
    return col, r1, r2


def _counta(ws: openpyxl.worksheet.worksheet.Worksheet, col: str, r1: int, r2: int) -> int:
    """Python equivalent of COUNTA over ``col`` rows ``r1..r2`` (count non-empty)."""
    return sum(1 for r in range(r1, r2 + 1) if ws[f"{col}{r}"].value not in (None, ""))


def _countif(
    ws: openpyxl.worksheet.worksheet.Worksheet, col: str, r1: int, r2: int, needle: str
) -> int:
    """Python equivalent of COUNTIF(range, "needle") over ``col`` rows ``r1..r2``."""
    return sum(1 for r in range(r1, r2 + 1) if ws[f"{col}{r}"].value == needle)


def _eval_band_gate(index_value: float, capable: float, potentially: float) -> str:
    """Python transcription of the ``B4`` IF/IF cascade, over the imported thresholds."""
    if index_value > capable:
        return "ACCEPTABLE"
    if index_value >= potentially:
        return "POTENTIALLY_ACCEPTABLE"
    return "UNACCEPTABLE"


def _band_study():
    """A precomputed, band-evaluated ACCEPTABLE study (Ppk = 1.80 > 1.67)."""
    return assess_initial_process_study(
        precomputed_index_type="Ppk",
        precomputed_index_value=1.80,
        precomputed_sample_size=120,
        precomputed_subgroup_count=30,
    )


# ============================================================================
# 1 · valid workbook — 3 sheets in order, header, 18 data rows
# ============================================================================


def test_export_benchmark_is_a_loadable_xlsx_with_three_sheets():
    wb_bytes = export_ppap_workbook(benchmark_ppap_package())
    assert isinstance(wb_bytes, bytes)

    wb = _load(wb_bytes)
    assert wb.sheetnames == [_DEFAULT_TITLE, _COMPLETENESS, _CAPABILITY]

    checklist = wb[_DEFAULT_TITLE]
    assert [c.value for c in checklist[1]] == list(PPAP_EXPORT_COLUMNS)
    # Header row 1 + exactly 18 element data rows.
    assert checklist.max_row == len(PPAP_ELEMENT_IDS) + 1
    assert len(PPAP_ELEMENT_IDS) == 18

    completeness = wb[_COMPLETENESS]
    assert [c.value for c in completeness[1]] == ["Metric", "Value"]
    assert [completeness[f"A{r}"].value for r in range(2, 10)] == [
        "Submission Level",
        "Reason for Submission",
        "Package Verdict (Submission Readiness)",
        "Date Generated",
        "Total Elements",
        "Submitted",
        "Retained On File",
        "Missing",
    ]


def test_col_widths_cover_every_checklist_column():
    assert set(PPAP_COL_WIDTHS) == set(PPAP_EXPORT_COLUMNS)


def test_completeness_metadata_cells_are_engine_values():
    audit = audit_ppap_package(benchmark_ppap_package())
    completeness = _load(export_ppap_workbook(benchmark_ppap_package()))[_COMPLETENESS]
    assert completeness["B2"].value == audit.submission_level
    assert completeness["B3"].value == audit.reason_for_submission
    assert completeness["B4"].value == audit.package_verdict


# ============================================================================
# 2 · positive live-formula proof + exact formula strings
# ============================================================================


def test_completeness_and_gate_are_live_formulas_with_exact_strings():
    wb_bytes = export_ppap_workbook(
        benchmark_ppap_package(), process_study_result=_band_study()
    )

    for coord in ("B6", "B7", "B8", "B9"):
        # Does not raise: the cell carries a live <f> element.
        assert_cell_is_formula(wb_bytes, _COMPLETENESS, coord)
    assert_cell_is_formula(wb_bytes, _CAPABILITY, "B4")

    wb = _load(wb_bytes)
    completeness = wb[_COMPLETENESS]
    assert completeness["B6"].value == "=COUNTA('PPAP Checklist'!A2:A19)"
    assert completeness["B7"].value == '=COUNTIF(\'PPAP Checklist\'!E2:E19,"SUBMITTED")'
    assert (
        completeness["B8"].value == '=COUNTIF(\'PPAP Checklist\'!E2:E19,"RETAINED_ON_FILE")'
    )
    assert completeness["B9"].value == '=COUNTIF(\'PPAP Checklist\'!E2:E19,"MISSING")'

    gate = wb[_CAPABILITY]
    assert (
        gate["B4"].value
        == '=IF(B3>1.67,"ACCEPTABLE",IF(B3>=1.33,"POTENTIALLY_ACCEPTABLE","UNACCEPTABLE"))'
    )
    assert gate["B4"].value.count("=") >= 1  # sanity: it is a formula, not a literal


def test_threshold_constants_roundtrip_as_expected_literals():
    """The exporter interpolates the imported constants — verify (not assume) their str form."""
    assert f"{ACCEPTANCE_THRESHOLD_CAPABLE}" == "1.67"
    assert f"{ACCEPTANCE_THRESHOLD_POTENTIALLY_CAPABLE}" == "1.33"


# ============================================================================
# 3 · literal-build negative control (mandatory)
# ============================================================================


def test_literal_metric_build_fails_the_formula_verifier():
    """NEGATIVE CONTROL: hardcoded-value metric cells must FAIL the verifier.

    If a future edit swapped the ``Formula`` cells for plain ints/strings, this suite must
    catch it — so prove the verifier bites on a literal build of the same layout.
    """
    wb = openpyxl.Workbook()
    comp = wb.active
    comp.title = _COMPLETENESS
    comp.cell(row=6, column=2, value=18)  # literal ints, NOT Formulas
    comp.cell(row=7, column=2, value=11)
    comp.cell(row=8, column=2, value=2)
    comp.cell(row=9, column=2, value=3)
    gate = wb.create_sheet(_CAPABILITY)
    gate.cell(row=4, column=2, value="ACCEPTABLE")  # literal string, NOT a Formula
    saved = _saved(wb)

    for coord in ("B6", "B7", "B8", "B9"):
        with pytest.raises(AssertionError, match="literal, not a live formula"):
            assert_cell_is_formula(saved, _COMPLETENESS, coord)
    with pytest.raises(AssertionError, match="literal, not a live formula"):
        assert_cell_is_formula(saved, _CAPABILITY, "B4")


# ============================================================================
# 4 · injection negative control (mandatory)
# ============================================================================


def test_injection_freetext_is_inert_but_rollups_stay_live_and_correct():
    """NEGATIVE CONTROL: a document_reference starting with '=' renders inert, per-cell.

    The malicious §2.2.1 ``document_reference`` is apostrophe-escaped (never a live
    formula) on the Checklist sheet, while the Completeness roll-ups stay live AND still
    count §2.2.1 correctly under its ``SUBMITTED`` verdict — proving the ``Formula`` opt-in
    is per-cell and the roll-up counts a different column entirely.
    """
    payload = "=cmd|' /C calc'!A0"
    audit = audit_ppap_package(benchmark_ppap_package())
    # §2.2.1 audits SUBMITTED in the benchmark; poison only its document_reference.
    assert audit.elements["2.2.1"].verdict == "SUBMITTED"
    object.__setattr__(audit.elements["2.2.1"], "document_reference", payload)

    wb_bytes = export_ppap_workbook(audit_result=audit)
    wb = _load(wb_bytes)

    # Document_Reference is column G; §2.2.1 is row 2. Rendered inert (apostrophe-escaped).
    checklist = wb[_DEFAULT_TITLE]
    assert checklist[f"{_DOCREF_COL}2"].value == "'" + payload
    with pytest.raises(AssertionError, match="literal, not a live formula"):
        assert_cell_is_formula(wb_bytes, _DEFAULT_TITLE, f"{_DOCREF_COL}2")

    # The Completeness roll-ups are still live formulas...
    for coord in ("B6", "B7", "B8", "B9"):
        assert_cell_is_formula(wb_bytes, _COMPLETENESS, coord)

    # ...and B7 still counts §2.2.1's SUBMITTED verdict (poison lived in another column).
    completeness = wb[_COMPLETENESS]
    col, r1, r2 = _referenced_row_span(completeness["B7"].value)
    assert checklist[f"{col}2"].value == "SUBMITTED"
    submitted_engine = sum(1 for e in audit.elements.values() if e.verdict == "SUBMITTED")
    assert _countif(checklist, col, r1, r2, "SUBMITTED") == submitted_engine


# ============================================================================
# 5 · accuracy scorecard
# ============================================================================


def test_ppap_export_accuracy_scorecard():
    """The saved roll-up/gate FORMULAS must recompute to the engine's own values.

    # reproduce:
    #   cd packages/quality-core && uv run pytest \
    #     tests/test_ppap_export.py -k accuracy_scorecard -q

    Reads each formula's OWN referenced range / search-string / thresholds back out of the
    saved sheet and applies COUNTA / COUNTIF / the IF-IF band cascade in Python against the
    IMPORTED thresholds — never data_only=True, which returns None for a never-recalculated
    formula. These are the values a human opening the file in Excel would see after
    recalculation.
    """
    audit = audit_ppap_package(benchmark_ppap_package())
    study = _band_study()

    expected_total = len(PPAP_ELEMENT_IDS)
    expected = {
        v: sum(1 for e in audit.elements.values() if e.verdict == v)
        for _, v in _VERDICT_ROWS
    }
    # A genuinely partial, non-degenerate distribution — every rolled verdict is > 0.
    assert all(count > 0 for count in expected.values()), expected
    assert sum(expected.values()) < expected_total  # not a degenerate 18/18 either

    wb = _load(export_ppap_workbook(audit_result=audit, process_study_result=study))
    checklist = wb[_DEFAULT_TITLE]
    completeness = wb[_COMPLETENESS]
    gate = wb[_CAPABILITY]

    # B6 — COUNTA over the Element_ID column's own referenced range == 18.
    col, r1, r2 = _referenced_row_span(completeness["B6"].value)
    assert col == _ELEMENT_ID_COL
    assert _counta(checklist, col, r1, r2) == expected_total

    # B7/B8/B9 — COUNTIF(range,"<verdict>") over Audit_Verdict's own referenced range,
    # with the search string read back from the saved formula, not re-typed here.
    for coord, verdict in _VERDICT_ROWS:
        formula = completeness[coord].value
        criterion = _COUNTIF_CRITERION_RE.search(formula)
        assert criterion is not None, formula
        needle = criterion.group(1)
        assert needle == verdict
        col, r1, r2 = _referenced_row_span(formula)
        assert col == _AUDIT_VERDICT_COL
        assert _countif(checklist, col, r1, r2, needle) == expected[verdict]

    # Capability Gate B4 — read B3 (Index Value) back, recompute the IF/IF cascade in
    # Python over the IMPORTED thresholds, and assert it equals the engine's verdict.
    index_value = gate["B3"].value
    assert index_value == pytest.approx(study.index_value)
    recomputed = _eval_band_gate(
        index_value,
        ACCEPTANCE_THRESHOLD_CAPABLE,
        ACCEPTANCE_THRESHOLD_POTENTIALLY_CAPABLE,
    )
    assert recomputed == study.verdict == "ACCEPTABLE"

    # The two thresholds embedded in the SAVED formula equal the imported constants —
    # proving the exporter never drifted a copy of the threshold.
    embedded = [float(x) for x in _THRESHOLD_RE.findall(gate["B4"].value)]
    assert embedded == [
        ACCEPTANCE_THRESHOLD_CAPABLE,
        ACCEPTANCE_THRESHOLD_POTENTIALLY_CAPABLE,
    ]


# ============================================================================
# 6 · empty / None process_study_result branch
# ============================================================================


def test_no_study_renders_all_na_and_zero_gate_formulas():
    """process_study_result omitted → Capability Gate is 5 'N/A' literals, no live formula."""
    wb_bytes = export_ppap_workbook(benchmark_ppap_package())
    gate = _load(wb_bytes)[_CAPABILITY]
    assert [gate[f"B{r}"].value for r in range(2, 7)] == ["N/A"] * 5

    # B4 is correctly a literal here (expected-not-live), not the mandatory literal-build
    # negative control of test #3 — it proves the None branch writes no <f>.
    with pytest.raises(AssertionError, match="literal, not a live formula"):
        assert_cell_is_formula(wb_bytes, _CAPABILITY, "B4")


# ============================================================================
# 7 · non-band process_study_result branches (band is None, index populated or not)
# ============================================================================


def test_attribute_study_renders_na_index_and_engine_action():
    """is_attribute=True → band None AND index_type/value None → rows 2-4 'N/A'."""
    study = assess_initial_process_study(is_attribute=True)
    assert study.band is None and study.index_type is None and study.index_value is None

    wb_bytes = export_ppap_workbook(benchmark_ppap_package(), process_study_result=study)
    gate = _load(wb_bytes)[_CAPABILITY]
    assert gate["B2"].value == "N/A"  # Index Type
    assert gate["B3"].value == "N/A"  # Index Value
    assert gate["B4"].value == "N/A"  # gate literal, not a formula
    assert gate["B5"].value == "NOT_APPLICABLE_ATTRIBUTE_DATA"  # engine verdict
    assert gate["B6"].value == ACTION_ATTRIBUTE_DATA  # verbatim engine action

    with pytest.raises(AssertionError, match="literal, not a live formula"):
        assert_cell_is_formula(wb_bytes, _CAPABILITY, "B4")


def test_unstable_study_populates_index_but_gate_is_na():
    """Stability gate tripped → band None but index_type/value ARE populated (row 3 float)."""
    study = assess_initial_process_study(
        precomputed_index_type="Ppk",
        precomputed_index_value=1.80,
        precomputed_sample_size=120,
        precomputed_subgroup_count=30,
        violations=[{"rule": 1, "index": 4}],
    )
    assert study.band is None
    assert study.index_type == "Ppk"
    assert study.index_value == 1.80
    assert study.verdict == "INDETERMINATE"

    wb_bytes = export_ppap_workbook(benchmark_ppap_package(), process_study_result=study)
    gate = _load(wb_bytes)[_CAPABILITY]
    assert gate["B2"].value == "Ppk"
    assert gate["B3"].value == pytest.approx(1.80)
    assert gate["B3"].number_format == "0.0000"  # literal float, formatted
    assert gate["B4"].value == "N/A"
    assert gate["B5"].value == "INDETERMINATE"
    assert gate["B6"].value == ACTION_UNSTABLE

    with pytest.raises(AssertionError, match="literal, not a live formula"):
        assert_cell_is_formula(wb_bytes, _CAPABILITY, "B4")


# ============================================================================
# 8 · custom title propagation into the cross-sheet formulas
# ============================================================================


def test_custom_title_renames_checklist_and_threads_into_completeness_formulas():
    """A hardcoded checklist sheet name in the roll-ups would silently break here."""
    title = "PPAP Rev C 2026"
    wb_bytes = export_ppap_workbook(benchmark_ppap_package(), title=title)
    wb = _load(wb_bytes)
    assert wb.sheetnames == [title, _COMPLETENESS, _CAPABILITY]

    completeness = wb[_COMPLETENESS]
    assert completeness["B6"].value == f"=COUNTA('{title}'!A2:A19)"
    assert completeness["B7"].value == f'=COUNTIF(\'{title}\'!E2:E19,"SUBMITTED")'
    assert title in completeness["B6"].value

    # Still live and still resolving to the renamed sheet.
    for coord in ("B6", "B7", "B8", "B9"):
        assert_cell_is_formula(wb_bytes, _COMPLETENESS, coord)


# ============================================================================
# 9 · audit_result precedence over package
# ============================================================================


def test_audit_result_takes_precedence_over_package():
    """Passing both uses audit_result verbatim; package is not re-audited and discarded.

    Level 1 flips §2.2.1 from 'S' to 'R' in Table 4.1, so its audit verdict flips
    SUBMITTED -> RETAINED_ON_FILE. If the exporter re-audited the Level-3 package it would
    render SUBMITTED, so the two cases genuinely differ.
    """
    package = benchmark_ppap_package()
    level1 = audit_ppap_package(package, submission_level=1)
    assert level1.elements["2.2.1"].verdict == "RETAINED_ON_FILE"
    assert audit_ppap_package(package).elements["2.2.1"].verdict == "SUBMITTED"

    wb = _load(export_ppap_workbook(package, audit_result=level1))
    checklist = wb[_DEFAULT_TITLE]
    # §2.2.1 is row 2; Audit_Verdict is column E — reflects audit_result, not the package.
    assert checklist[f"{_AUDIT_VERDICT_COL}2"].value == "RETAINED_ON_FILE"
    assert wb[_COMPLETENESS]["B2"].value == 1  # submission level 1 from audit_result


# ============================================================================
# 10 · authority-invariant guard (mandatory)
# ============================================================================


def test_no_customer_disposition_token_anywhere_in_the_workbook():
    """No Section 5 disposition token (Approved / Interim Approval / Rejected) may appear.

    Positive regression guard, not a negative control: the exporter's inputs already cannot
    carry these tokens, so this test exists to catch a *future* accidental approval field.
    Exact, case-sensitive matches only — fuzzy matching would false-positive on legitimate
    AIAG element names like 'Customer Engineering Approval' (§2.2.3) and 'Appearance
    Approval Report (AAR)' (§2.2.13).
    """
    forbidden_exact = {"Approved", "Interim Approval", "Rejected"}
    wb = _load(export_ppap_workbook(benchmark_ppap_package(), process_study_result=_band_study()))

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    assert cell.value not in forbidden_exact, (
                        f"{sheet.title}!{cell.coordinate} = {cell.value!r}"
                    )

    # Substring 'Approval'/'Disposition' guard, restricted to STRUCTURE the exporter owns:
    # the Checklist header row and the metric labels in column A — never data cells, whose
    # legitimate element names contain 'Approval'.
    checklist = wb[_DEFAULT_TITLE]
    for cell in checklist[1]:
        assert "Approval" not in str(cell.value)
        assert "Disposition" not in str(cell.value)
    for sheet_name in (_COMPLETENESS, _CAPABILITY):
        sheet = wb[sheet_name]
        for r in range(1, sheet.max_row + 1):
            label = str(sheet[f"A{r}"].value)
            assert "Approval" not in label
            assert "Disposition" not in label


# ============================================================================
# 11 · re-export surface
# ============================================================================


def test_public_names_are_reexported_from_ppap_package():
    import quality_core.ppap as pkg
    from quality_core.ppap import export as mod

    assert pkg.build_ppap_workbook is mod.build_ppap_workbook
    assert pkg.export_ppap_workbook is mod.export_ppap_workbook
    assert pkg.benchmark_ppap_package is mod.benchmark_ppap_package


# ============================================================================
# 12 · fresh benchmark object per call
# ============================================================================


def test_benchmark_package_returns_a_fresh_object_each_call():
    a = benchmark_ppap_package()
    b = benchmark_ppap_package()
    assert a is not b
    assert a.elements is not b.elements


# ============================================================================
# 13 · ValueError when both package and audit_result are None
# ============================================================================


def test_raises_when_neither_package_nor_audit_result_supplied():
    with pytest.raises(ValueError, match="package or audit_result"):
        build_ppap_workbook()
    with pytest.raises(ValueError, match="package or audit_result"):
        export_ppap_workbook()


# ============================================================================
# 14 · canonical row order §2.2.1 → §2.2.18, not lexicographic
# ============================================================================


def test_row_order_is_canonical_not_sorted():
    """The Element_ID column must equal PPAP_ELEMENT_IDS exactly (2.2.10 follows 2.2.9)."""
    checklist = _load(export_ppap_workbook(benchmark_ppap_package()))[_DEFAULT_TITLE]
    ids = [
        checklist[f"{_ELEMENT_ID_COL}{r}"].value
        for r in range(2, len(PPAP_ELEMENT_IDS) + 2)
    ]
    assert ids == list(PPAP_ELEMENT_IDS)
    # Canonical, not lexicographic: "2.2.10" comes right after "2.2.9".
    assert ids[ids.index("2.2.9") + 1] == "2.2.10"


# ============================================================================
# 15 · nullable checklist fields round-trip as blank, not "None"
# ============================================================================


def test_absent_element_nullable_fields_roundtrip_as_blank():
    """A MISSING element (§2.2.12) has no evidence_status/document_reference → blank cells."""
    wb = _load(export_ppap_workbook(benchmark_ppap_package()))
    checklist = wb[_DEFAULT_TITLE]
    # §2.2.12 is row 13 (index 11 + header). document_reference is None for _absent items.
    row = PPAP_ELEMENT_IDS.index("2.2.12") + 2
    value = checklist[f"{_DOCREF_COL}{row}"].value
    assert value is None
    assert value != "None"
