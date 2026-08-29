"""AIAG Statistical Process Control (SPC) Reference Manual (4th Edition, 2005) Live-Formula Excel Exporter.

Provides deterministic, standards-compliant export of Shewhart variable (Xbar-R, Xbar-S, I-MR)
and attribute (p, c, u) control chart datasets and process capability studies to openpyxl
Workbooks and serialized Excel (.xlsx) bytes with live formulas.

Standards References:
- AIAG Statistical Process Control (SPC) Reference Manual, 4th Edition (2005):
  - Chapter II: Variables Control Charts (Xbar-R, Xbar-S, I-MR math & factor tables).
  - Chapter III: Attributes Control Charts (p, c, u math & variable sample size limits).
  - Chapter IV: Process Capability (Cp, Cpk, Pp, Ppk, within-subgroup and overall dispersion).
- Western Electric Statistical Quality Control Handbook (1956) & Nelson (1984) Run Rules.
"""

from __future__ import annotations

import io
import math
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from quality_core.io.export import Formula, now, sanitize_cell, write_formula_cell
from quality_core.spc.constants import (
    IMR_D2,
    IMR_D4,
    IMR_E2,
    XBAR_R_CONSTANTS,
    XBAR_S_CONSTANTS,
)
from quality_core.spc.control_charts import (
    compute_c,
    compute_imr,
    compute_p,
    compute_u,
    compute_xbar_r,
    compute_xbar_s,
)
from quality_core.spc.rule_detection import detect_violations

__all__ = ["export_spc_excel", "export_spc_to_workbook"]

_HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_BOLD_FONT = Font(bold=True, size=10)
_NORMAL_FONT = Font(size=10)
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)
_RIGHT_ALIGN = Alignment(horizontal="right", vertical="center", wrap_text=False)

_SUPPORTED_CHARTS: frozenset[str] = frozenset({"Xbar-R", "Xbar-S", "I-MR", "p", "c", "u"})
_VARIABLE_CHARTS: frozenset[str] = frozenset({"Xbar-R", "Xbar-S", "I-MR"})


def _set_cell(
    ws: Any,
    row: int,
    col: int,
    value: Any,
    *,
    font: Font = _NORMAL_FONT,
    fill: PatternFill | None = None,
    alignment: Alignment = _LEFT_ALIGN,
    number_format: str | None = None,
) -> Any:
    """Set a cell value with formatting, applying Formula bypass or sanitize_cell."""
    if isinstance(value, Formula):
        cell = write_formula_cell(
            ws, row, col, value.formula, number_format=value.number_format or number_format
        )
    else:
        sanitized = sanitize_cell(value)
        cell = ws.cell(row=row, column=col, value=sanitized)
        if number_format is not None:
            cell.number_format = number_format
    cell.font = font
    if fill is not None:
        cell.fill = fill
    cell.alignment = alignment
    return cell


def _auto_fit_columns(ws: Any, min_width: float = 12.0, max_width: float = 50.0) -> None:
    """Adjust worksheet column widths based on cell content lengths."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is not None:
                val_str = str(cell.value)
                if val_str.startswith("="):
                    val_str = "0.0000"
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = min(max_width, max(min_width, float(max_len + 3)))


def export_spc_to_workbook(
    chart_type: str,
    data: list[list[float]] | list[float],
    *,
    usl: float | None = None,
    lsl: float | None = None,
    sample_sizes: list[float] | None = None,
    rule_set: str = "Western Electric",
    title: str = "AIAG SPC Control Chart & Process Capability Analysis",
    part_name: str | None = None,
    part_number: str | None = None,
    characteristic: str | None = None,
) -> openpyxl.Workbook:
    """Export SPC control chart data and capability analysis to an openpyxl Workbook with live formulas.

    Parameters
    ----------
    chart_type : str
        Control chart type: "Xbar-R", "Xbar-S", "I-MR", "p", "c", or "u".
    data : list[list[float]] | list[float]
        Measurement data subgroups (2D list for Xbar-R, Xbar-S) or individual readings / counts (1D list).
    usl : float | None, optional
        Upper specification limit for capability analysis.
    lsl : float | None, optional
        Lower specification limit for capability analysis.
    sample_sizes : list[float] | None, optional
        Subgroup inspection unit / sample sizes (required for "p" and "u" charts).
    rule_set : str, optional
        Run-rule detection standard: "Western Electric" (default) or "Nelson".
    title : str, optional
        Analysis report title.
    part_name : str | None, optional
        Part name metadata.
    part_number : str | None, optional
        Part number metadata.
    characteristic : str | None, optional
        Inspection characteristic name metadata.

    Returns
    -------
    openpyxl.Workbook
        Constructed openpyxl Workbook containing "Control Chart Data", optional "Process Capability",
        and "Summary & Run Rules" worksheets with live OOXML formulas.
    """
    if chart_type not in _SUPPORTED_CHARTS:
        raise ValueError(f"Unsupported chart type: {chart_type!r}. Supported: {sorted(_SUPPORTED_CHARTS)}")

    if usl is not None and lsl is not None and usl < lsl:
        raise ValueError(f"USL ({usl}) cannot be less than LSL ({lsl}).")

    wb = openpyxl.Workbook()
    ws_data: Any = wb.active
    ws_data.title = "Control Chart Data"

    mean_coord = ""
    sigma_within_coord = ""
    data_range = ""
    violations: list[dict[str, Any]] = []

    # -------------------------------------------------------------------------
    # Sheet 1: Control Chart Data & Parameters
    # -------------------------------------------------------------------------
    if chart_type == "Xbar-R":
        if not isinstance(data, list) or not data or not isinstance(data[0], list):
            raise ValueError("Xbar-R chart requires data as a list of subgroups (list[list[float]]).")
        k = len(data)
        n = len(data[0])
        for sub in data:
            if not isinstance(sub, list) or len(sub) != n:
                raise ValueError("All subgroups in Xbar-R chart must have equal size.")
        if n not in XBAR_R_CONSTANTS:
            raise ValueError(f"Xbar-R requires subgroup size between 2 and 10, got {n}.")

        constants = XBAR_R_CONSTANTS[n]
        a2_val, d3_val, d4_val, d2_val = constants["A2"], constants["D3"], constants["D4"], constants["d2"]

        end_sample_col_letter = get_column_letter(1 + n)
        mean_col_letter = get_column_letter(2 + n)
        range_col_letter = get_column_letter(3 + n)
        param_col = 5 + n
        val_col = 6 + n
        v_letter = get_column_letter(val_col)

        # Headers
        headers = ["Subgroup"] + [f"X{i}" for i in range(1, n + 1)] + ["Subgroup Mean", "Subgroup Range"]
        for c_idx, h in enumerate(headers, start=1):
            _set_cell(ws_data, 1, c_idx, h, font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)
        ws_data.row_dimensions[1].height = 22

        # Data rows
        for r_idx, sub in enumerate(data, start=2):
            _set_cell(ws_data, r_idx, 1, r_idx - 1, font=_NORMAL_FONT, alignment=_CENTER_ALIGN)
            for j, val in enumerate(sub, start=2):
                _set_cell(ws_data, r_idx, j, float(val), font=_NORMAL_FONT, alignment=_RIGHT_ALIGN)
            _set_cell(
                ws_data,
                r_idx,
                2 + n,
                Formula(f"=AVERAGE(B{r_idx}:{end_sample_col_letter}{r_idx})", number_format="0.0000"),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )
            _set_cell(
                ws_data,
                r_idx,
                3 + n,
                Formula(
                    f"=MAX(B{r_idx}:{end_sample_col_letter}{r_idx})-MIN(B{r_idx}:{end_sample_col_letter}{r_idx})",
                    number_format="0.0000",
                ),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )

        # Parameter block
        _set_cell(ws_data, 1, param_col, "Parameter", font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)
        _set_cell(ws_data, 1, val_col, "Value", font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)

        _set_cell(ws_data, 2, param_col, "Grand Mean (Xbarbar)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            2,
            val_col,
            Formula(f"=AVERAGE({mean_col_letter}2:{mean_col_letter}{k+1})", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 3, param_col, "Range Mean (Rbar)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            3,
            val_col,
            Formula(f"=AVERAGE({range_col_letter}2:{range_col_letter}{k+1})", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 4, param_col, "Within Sigma (sigma_hat)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            4,
            val_col,
            Formula(f"=${v_letter}$3/{d2_val}", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 5, param_col, "UCL (Xbar)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            5,
            val_col,
            Formula(f"=${v_letter}$2+({a2_val}*${v_letter}$3)", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 6, param_col, "LCL (Xbar)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            6,
            val_col,
            Formula(f"=${v_letter}$2-({a2_val}*${v_letter}$3)", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 7, param_col, "UCL (R)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            7,
            val_col,
            Formula(f"={d4_val}*${v_letter}$3", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 8, param_col, "LCL (R)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            8,
            val_col,
            Formula(f"={d3_val}*${v_letter}$3", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        mean_coord = f"${v_letter}$2"
        sigma_within_coord = f"${v_letter}$4"
        data_range = f"B2:{end_sample_col_letter}{k+1}"

        # Recalculate violations for summary sheet
        subgroups_2d = [[float(x) for x in sub] for sub in data]
        xr = compute_xbar_r(subgroups_2d)
        sigma_pts = xr["sigma_hat"] / math.sqrt(n)
        violations = detect_violations(chart_type, xr["subgroup_means"], cl=xr["xbarbar"], sigma=sigma_pts, rule_set=rule_set)

    elif chart_type == "Xbar-S":
        if not isinstance(data, list) or not data or not isinstance(data[0], list):
            raise ValueError("Xbar-S chart requires data as a list of subgroups (list[list[float]]).")
        k = len(data)
        n = len(data[0])
        for sub in data:
            if not isinstance(sub, list) or len(sub) != n:
                raise ValueError("All subgroups in Xbar-S chart must have equal size.")
        if n not in XBAR_S_CONSTANTS:
            raise ValueError(f"Xbar-S requires subgroup size between 2 and 12, got {n}.")

        constants_s = XBAR_S_CONSTANTS[n]
        a3_val, b3_val, b4_val, c4_val = constants_s["A3"], constants_s["B3"], constants_s["B4"], constants_s["c4"]

        end_sample_col_letter = get_column_letter(1 + n)
        mean_col_letter = get_column_letter(2 + n)
        sd_col_letter = get_column_letter(3 + n)
        param_col = 5 + n
        val_col = 6 + n
        v_letter = get_column_letter(val_col)

        # Headers
        headers = ["Subgroup"] + [f"X{i}" for i in range(1, n + 1)] + ["Subgroup Mean", "Subgroup Std Dev"]
        for c_idx, h in enumerate(headers, start=1):
            _set_cell(ws_data, 1, c_idx, h, font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)
        ws_data.row_dimensions[1].height = 22

        # Data rows
        for r_idx, sub in enumerate(data, start=2):
            _set_cell(ws_data, r_idx, 1, r_idx - 1, font=_NORMAL_FONT, alignment=_CENTER_ALIGN)
            for j, val in enumerate(sub, start=2):
                _set_cell(ws_data, r_idx, j, float(val), font=_NORMAL_FONT, alignment=_RIGHT_ALIGN)
            _set_cell(
                ws_data,
                r_idx,
                2 + n,
                Formula(f"=AVERAGE(B{r_idx}:{end_sample_col_letter}{r_idx})", number_format="0.0000"),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )
            _set_cell(
                ws_data,
                r_idx,
                3 + n,
                Formula(f"=STDEV.S(B{r_idx}:{end_sample_col_letter}{r_idx})", number_format="0.0000"),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )

        # Parameter block
        _set_cell(ws_data, 1, param_col, "Parameter", font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)
        _set_cell(ws_data, 1, val_col, "Value", font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)

        _set_cell(ws_data, 2, param_col, "Grand Mean (Xbarbar)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            2,
            val_col,
            Formula(f"=AVERAGE({mean_col_letter}2:{mean_col_letter}{k+1})", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 3, param_col, "SD Mean (sbar)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            3,
            val_col,
            Formula(f"=AVERAGE({sd_col_letter}2:{sd_col_letter}{k+1})", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 4, param_col, "Within Sigma (sigma_hat)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            4,
            val_col,
            Formula(f"=${v_letter}$3/{c4_val}", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 5, param_col, "UCL (Xbar)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            5,
            val_col,
            Formula(f"=${v_letter}$2+({a3_val}*${v_letter}$3)", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 6, param_col, "LCL (Xbar)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            6,
            val_col,
            Formula(f"=${v_letter}$2-({a3_val}*${v_letter}$3)", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 7, param_col, "UCL (S)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            7,
            val_col,
            Formula(f"={b4_val}*${v_letter}$3", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 8, param_col, "LCL (S)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            8,
            val_col,
            Formula(f"={b3_val}*${v_letter}$3", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        mean_coord = f"${v_letter}$2"
        sigma_within_coord = f"${v_letter}$4"
        data_range = f"B2:{end_sample_col_letter}{k+1}"

        # Recalculate violations for summary sheet
        subgroups_2d_s = [[float(x) for x in sub] for sub in data]
        xs = compute_xbar_s(subgroups_2d_s)
        sigma_pts_s = xs["sigma_hat"] / math.sqrt(n)
        violations = detect_violations(chart_type, xs["subgroup_means"], cl=xs["xbarbar"], sigma=sigma_pts_s, rule_set=rule_set)

    elif chart_type == "I-MR":
        raw_vals = [float(x[0]) if isinstance(x, (list, tuple)) else float(x) for x in data]
        k = len(raw_vals)
        if k < 2:
            raise ValueError("I-MR chart requires at least two values.")

        param_col = 5
        val_col = 6
        v_letter = get_column_letter(val_col)

        # Headers
        headers = ["Observation", "Value (X)", "Moving Range (MR)"]
        for c_idx, h in enumerate(headers, start=1):
            _set_cell(ws_data, 1, c_idx, h, font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)
        ws_data.row_dimensions[1].height = 22

        # Data rows
        for r_idx, val in enumerate(raw_vals, start=2):
            _set_cell(ws_data, r_idx, 1, r_idx - 1, font=_NORMAL_FONT, alignment=_CENTER_ALIGN)
            _set_cell(ws_data, r_idx, 2, val, font=_NORMAL_FONT, alignment=_RIGHT_ALIGN)
            if r_idx == 2:
                _set_cell(ws_data, r_idx, 3, "", font=_NORMAL_FONT, alignment=_CENTER_ALIGN)
            else:
                _set_cell(
                    ws_data,
                    r_idx,
                    3,
                    Formula(f"=ABS(B{r_idx}-B{r_idx-1})", number_format="0.0000"),
                    font=_NORMAL_FONT,
                    alignment=_RIGHT_ALIGN,
                )

        # Parameter block
        _set_cell(ws_data, 1, param_col, "Parameter", font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)
        _set_cell(ws_data, 1, val_col, "Value", font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)

        _set_cell(ws_data, 2, param_col, "Process Mean (Xbar)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            2,
            val_col,
            Formula(f"=AVERAGE(B2:B{k+1})", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 3, param_col, "Moving Range Mean (MRbar)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            3,
            val_col,
            Formula(f"=AVERAGE(C3:C{k+1})", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 4, param_col, "Within Sigma (sigma_hat)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            4,
            val_col,
            Formula(f"=${v_letter}$3/{IMR_D2}", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 5, param_col, "UCL (X)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            5,
            val_col,
            Formula(f"=${v_letter}$2+({IMR_E2}*${v_letter}$3)", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 6, param_col, "LCL (X)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            6,
            val_col,
            Formula(f"=${v_letter}$2-({IMR_E2}*${v_letter}$3)", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 7, param_col, "UCL (MR)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            7,
            val_col,
            Formula(f"={IMR_D4}*${v_letter}$3", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 8, param_col, "LCL (MR)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            8,
            val_col,
            Formula("=0", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        mean_coord = f"${v_letter}$2"
        sigma_within_coord = f"${v_letter}$4"
        data_range = f"B2:B{k+1}"

        # Recalculate violations for summary sheet
        im = compute_imr(raw_vals)
        violations = detect_violations(chart_type, im["values"], cl=im["xbar"], sigma=im["sigma_hat"], rule_set=rule_set)

    elif chart_type == "p":
        if sample_sizes is None:
            raise ValueError("p chart requires sample_sizes.")
        counts = [float(x[0]) if isinstance(x, (list, tuple)) else float(x) for x in data]
        sizes = [float(s) for s in sample_sizes]
        k = len(counts)
        if k == 0 or len(sizes) != k:
            raise ValueError("counts and sample_sizes must have equal non-zero length.")
        if any(s <= 0 for s in sizes):
            raise ValueError("sample_sizes must be strictly positive.")

        param_col = 9
        val_col = 10
        v_letter = get_column_letter(val_col)
        pbar_cell = f"${v_letter}$2"

        # Headers
        headers = ["Subgroup", "Defectives", "Sample Size", "Proportion (p)", "UCL", "Centerline (pbar)", "LCL"]
        for c_idx, h in enumerate(headers, start=1):
            _set_cell(ws_data, 1, c_idx, h, font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)
        ws_data.row_dimensions[1].height = 22

        # Parameter block
        _set_cell(ws_data, 1, param_col, "Parameter", font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)
        _set_cell(ws_data, 1, val_col, "Value", font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)

        _set_cell(ws_data, 2, param_col, "Centerline (pbar)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            2,
            val_col,
            Formula(f"=SUM(B2:B{k+1})/SUM(C2:C{k+1})", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 3, param_col, "Total Defectives", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            3,
            val_col,
            Formula(f"=SUM(B2:B{k+1})", number_format="0"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 4, param_col, "Total Sample Size", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            4,
            val_col,
            Formula(f"=SUM(C2:C{k+1})", number_format="0"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        # Data rows
        for r_idx in range(2, k + 2):
            i = r_idx - 2
            _set_cell(ws_data, r_idx, 1, i + 1, font=_NORMAL_FONT, alignment=_CENTER_ALIGN)
            _set_cell(ws_data, r_idx, 2, counts[i], font=_NORMAL_FONT, alignment=_RIGHT_ALIGN)
            _set_cell(ws_data, r_idx, 3, sizes[i], font=_NORMAL_FONT, alignment=_RIGHT_ALIGN)
            _set_cell(
                ws_data,
                r_idx,
                4,
                Formula(f"=B{r_idx}/C{r_idx}", number_format="0.0000"),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )
            _set_cell(
                ws_data,
                r_idx,
                5,
                Formula(
                    f"={pbar_cell}+3*SQRT(({pbar_cell}*(1-{pbar_cell}))/C{r_idx})",
                    number_format="0.0000",
                ),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )
            _set_cell(
                ws_data,
                r_idx,
                6,
                Formula(f"={pbar_cell}", number_format="0.0000"),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )
            _set_cell(
                ws_data,
                r_idx,
                7,
                Formula(
                    f"=MAX(0,{pbar_cell}-3*SQRT(({pbar_cell}*(1-{pbar_cell}))/C{r_idx}))",
                    number_format="0.0000",
                ),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )

        p_res = compute_p(counts, sizes)
        mean_n = float(sum(sizes) / len(sizes)) if sizes else 1.0
        pbar = p_res["pbar"]
        sigma_p = math.sqrt(pbar * (1.0 - pbar) / mean_n) if (pbar * (1.0 - pbar) > 0 and mean_n > 0) else 0.0
        violations = detect_violations(chart_type, p_res["proportions"], cl=pbar, sigma=sigma_p, rule_set=rule_set)

    elif chart_type == "c":
        counts = [float(x[0]) if isinstance(x, (list, tuple)) else float(x) for x in data]
        k = len(counts)
        if k == 0:
            raise ValueError("c chart requires at least one data point.")

        param_col = 7
        val_col = 8
        v_letter = get_column_letter(val_col)
        cbar_cell = f"${v_letter}$2"

        # Headers
        headers = ["Sample", "Defects (c)", "UCL", "Centerline (cbar)", "LCL"]
        for c_idx, h in enumerate(headers, start=1):
            _set_cell(ws_data, 1, c_idx, h, font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)
        ws_data.row_dimensions[1].height = 22

        # Parameter block
        _set_cell(ws_data, 1, param_col, "Parameter", font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)
        _set_cell(ws_data, 1, val_col, "Value", font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)

        _set_cell(ws_data, 2, param_col, "Centerline (cbar)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            2,
            val_col,
            Formula(f"=AVERAGE(B2:B{k+1})", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 3, param_col, "UCL (c)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            3,
            val_col,
            Formula(f"={cbar_cell}+3*SQRT({cbar_cell})", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 4, param_col, "LCL (c)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            4,
            val_col,
            Formula(f"=MAX(0,{cbar_cell}-3*SQRT({cbar_cell}))", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        # Data rows
        for r_idx in range(2, k + 2):
            i = r_idx - 2
            _set_cell(ws_data, r_idx, 1, i + 1, font=_NORMAL_FONT, alignment=_CENTER_ALIGN)
            _set_cell(ws_data, r_idx, 2, counts[i], font=_NORMAL_FONT, alignment=_RIGHT_ALIGN)
            _set_cell(
                ws_data,
                r_idx,
                3,
                Formula(f"={cbar_cell}+3*SQRT({cbar_cell})", number_format="0.0000"),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )
            _set_cell(
                ws_data,
                r_idx,
                4,
                Formula(f"={cbar_cell}", number_format="0.0000"),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )
            _set_cell(
                ws_data,
                r_idx,
                5,
                Formula(f"=MAX(0,{cbar_cell}-3*SQRT({cbar_cell}))", number_format="0.0000"),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )

        c_res = compute_c(counts)
        cbar = c_res["cbar"]
        sigma_c = math.sqrt(cbar) if cbar > 0 else 0.0
        violations = detect_violations(chart_type, c_res["counts"], cl=cbar, sigma=sigma_c, rule_set=rule_set)

    else:  # chart_type == "u"
        if sample_sizes is None:
            raise ValueError("u chart requires sample_sizes.")
        counts = [float(x[0]) if isinstance(x, (list, tuple)) else float(x) for x in data]
        sizes = [float(s) for s in sample_sizes]
        k = len(counts)
        if k == 0 or len(sizes) != k:
            raise ValueError("counts and sample_sizes must have equal non-zero length.")
        if any(s <= 0 for s in sizes):
            raise ValueError("sample_sizes must be strictly positive.")

        param_col = 9
        val_col = 10
        v_letter = get_column_letter(val_col)
        ubar_cell = f"${v_letter}$2"

        # Headers
        headers = ["Sample", "Defects (c)", "Units (n)", "Defect Rate (u)", "UCL", "Centerline (ubar)", "LCL"]
        for c_idx, h in enumerate(headers, start=1):
            _set_cell(ws_data, 1, c_idx, h, font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)
        ws_data.row_dimensions[1].height = 22

        # Parameter block
        _set_cell(ws_data, 1, param_col, "Parameter", font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)
        _set_cell(ws_data, 1, val_col, "Value", font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)

        _set_cell(ws_data, 2, param_col, "Centerline (ubar)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            2,
            val_col,
            Formula(f"=SUM(B2:B{k+1})/SUM(C2:C{k+1})", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 3, param_col, "Total Defects", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            3,
            val_col,
            Formula(f"=SUM(B2:B{k+1})", number_format="0"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        _set_cell(ws_data, 4, param_col, "Total Units", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_data,
            4,
            val_col,
            Formula(f"=SUM(C2:C{k+1})", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        # Data rows
        for r_idx in range(2, k + 2):
            i = r_idx - 2
            _set_cell(ws_data, r_idx, 1, i + 1, font=_NORMAL_FONT, alignment=_CENTER_ALIGN)
            _set_cell(ws_data, r_idx, 2, counts[i], font=_NORMAL_FONT, alignment=_RIGHT_ALIGN)
            _set_cell(ws_data, r_idx, 3, sizes[i], font=_NORMAL_FONT, alignment=_RIGHT_ALIGN)
            _set_cell(
                ws_data,
                r_idx,
                4,
                Formula(f"=B{r_idx}/C{r_idx}", number_format="0.0000"),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )
            _set_cell(
                ws_data,
                r_idx,
                5,
                Formula(f"={ubar_cell}+3*SQRT({ubar_cell}/C{r_idx})", number_format="0.0000"),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )
            _set_cell(
                ws_data,
                r_idx,
                6,
                Formula(f"={ubar_cell}", number_format="0.0000"),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )
            _set_cell(
                ws_data,
                r_idx,
                7,
                Formula(f"=MAX(0,{ubar_cell}-3*SQRT({ubar_cell}/C{r_idx}))", number_format="0.0000"),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )

        u_res = compute_u(counts, sizes)
        mean_n = float(sum(sizes) / len(sizes)) if sizes else 1.0
        ubar = u_res["ubar"]
        sigma_u = math.sqrt(ubar / mean_n) if (ubar > 0 and mean_n > 0) else 0.0
        violations = detect_violations(chart_type, u_res["u_values"], cl=ubar, sigma=sigma_u, rule_set=rule_set)

    ws_data.freeze_panes = "A2"
    _auto_fit_columns(ws_data)

    # -------------------------------------------------------------------------
    # Sheet 2: Process Capability (Variable Charts Only)
    # -------------------------------------------------------------------------
    if chart_type in _VARIABLE_CHARTS:
        ws_cap = wb.create_sheet("Process Capability")

        _set_cell(ws_cap, 1, 1, "Capability Parameter", font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)
        _set_cell(ws_cap, 1, 2, "Value / Formula", font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)
        ws_cap.row_dimensions[1].height = 22

        # Row 2: USL
        _set_cell(ws_cap, 2, 1, "Upper Specification Limit (USL)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        if usl is not None:
            _set_cell(ws_cap, 2, 2, float(usl), font=_NORMAL_FONT, alignment=_RIGHT_ALIGN, number_format="0.0000")
        else:
            _set_cell(ws_cap, 2, 2, "N/A", font=_NORMAL_FONT, alignment=_CENTER_ALIGN)

        # Row 3: LSL
        _set_cell(ws_cap, 3, 1, "Lower Specification Limit (LSL)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        if lsl is not None:
            _set_cell(ws_cap, 3, 2, float(lsl), font=_NORMAL_FONT, alignment=_RIGHT_ALIGN, number_format="0.0000")
        else:
            _set_cell(ws_cap, 3, 2, "N/A", font=_NORMAL_FONT, alignment=_CENTER_ALIGN)

        # Row 4: Process Mean
        _set_cell(ws_cap, 4, 1, "Process Mean (mu)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_cap,
            4,
            2,
            Formula(f"='Control Chart Data'!{mean_coord}", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        # Row 5: Within Sigma
        _set_cell(ws_cap, 5, 1, "Within-Subgroup Sigma (sigma_hat)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_cap,
            5,
            2,
            Formula(f"='Control Chart Data'!{sigma_within_coord}", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        # Row 6: Overall Sigma
        _set_cell(ws_cap, 6, 1, "Overall Sigma (s)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_cap,
            6,
            2,
            Formula(f"=STDEV.S('Control Chart Data'!{data_range})", number_format="0.0000"),
            font=_NORMAL_FONT,
            alignment=_RIGHT_ALIGN,
        )

        # Row 7: Cp
        _set_cell(ws_cap, 7, 1, "Potential Capability (Cp)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        if usl is not None and lsl is not None:
            _set_cell(
                ws_cap,
                7,
                2,
                Formula("=(B2-B3)/(6*B5)", number_format="0.0000"),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )
        else:
            _set_cell(ws_cap, 7, 2, "N/A", font=_NORMAL_FONT, alignment=_CENTER_ALIGN)

        # Row 8: Cpk
        _set_cell(ws_cap, 8, 1, "Capability Index (Cpk)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        if usl is not None and lsl is not None:
            _set_cell(
                ws_cap,
                8,
                2,
                Formula("=MIN((B2-B4)/(3*B5),(B4-B3)/(3*B5))", number_format="0.0000"),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )
        elif usl is not None:
            _set_cell(
                ws_cap,
                8,
                2,
                Formula("=(B2-B4)/(3*B5)", number_format="0.0000"),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )
        elif lsl is not None:
            _set_cell(
                ws_cap,
                8,
                2,
                Formula("=(B4-B3)/(3*B5)", number_format="0.0000"),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )
        else:
            _set_cell(ws_cap, 8, 2, "N/A", font=_NORMAL_FONT, alignment=_CENTER_ALIGN)

        # Row 9: Pp
        _set_cell(ws_cap, 9, 1, "Potential Performance (Pp)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        if usl is not None and lsl is not None:
            _set_cell(
                ws_cap,
                9,
                2,
                Formula("=(B2-B3)/(6*B6)", number_format="0.0000"),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )
        else:
            _set_cell(ws_cap, 9, 2, "N/A", font=_NORMAL_FONT, alignment=_CENTER_ALIGN)

        # Row 10: Ppk
        _set_cell(ws_cap, 10, 1, "Performance Index (Ppk)", font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        if usl is not None and lsl is not None:
            _set_cell(
                ws_cap,
                10,
                2,
                Formula("=MIN((B2-B4)/(3*B6),(B4-B3)/(3*B6))", number_format="0.0000"),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )
        elif usl is not None:
            _set_cell(
                ws_cap,
                10,
                2,
                Formula("=(B2-B4)/(3*B6)", number_format="0.0000"),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )
        elif lsl is not None:
            _set_cell(
                ws_cap,
                10,
                2,
                Formula("=(B4-B3)/(3*B6)", number_format="0.0000"),
                font=_NORMAL_FONT,
                alignment=_RIGHT_ALIGN,
            )
        else:
            _set_cell(ws_cap, 10, 2, "N/A", font=_NORMAL_FONT, alignment=_CENTER_ALIGN)

        ws_cap.freeze_panes = "A2"
        _auto_fit_columns(ws_cap, min_width=25.0)

    # -------------------------------------------------------------------------
    # Sheet 3: Summary & Run Rules
    # -------------------------------------------------------------------------
    ws_summary = wb.create_sheet("Summary & Run Rules")

    _set_cell(ws_summary, 1, 1, "Metadata Field", font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)
    _set_cell(ws_summary, 1, 2, "Value", font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER_ALIGN)
    ws_summary.row_dimensions[1].height = 22

    meta_rows: list[tuple[str, Any]] = [
        ("Analysis Title", title),
        ("Part Name", part_name if part_name is not None else "N/A"),
        ("Part Number", part_number if part_number is not None else "N/A"),
        ("Characteristic", characteristic if characteristic is not None else "N/A"),
        ("Chart Type", chart_type),
        ("Standards Basis", "AIAG Statistical Process Control (SPC) Reference Manual (4th Edition, 2005)"),
        ("Run-Rule Standard", rule_set),
        ("Generated Timestamp", now()),
        (
            "Process Stability",
            "In Control (Common-Cause Variation Only)"
            if not violations
            else "Out of Control (Special Causes Detected)",
        ),
        ("Total Special Cause Violations", len(violations)),
    ]

    for r_idx, (label, val) in enumerate(meta_rows, start=2):
        _set_cell(ws_summary, r_idx, 1, label, font=_BOLD_FONT, alignment=_LEFT_ALIGN)
        _set_cell(ws_summary, r_idx, 2, val, font=_NORMAL_FONT, alignment=_LEFT_ALIGN)

    # Run-Rule Findings Table
    start_rule_row = len(meta_rows) + 3
    rule_headers = ["Subgroup / Observation", "Rule Violated", "Status"]
    for c_idx, h in enumerate(rule_headers, start=1):
        _set_cell(
            ws_summary,
            start_rule_row,
            c_idx,
            h,
            font=_HEADER_FONT,
            fill=_HEADER_FILL,
            alignment=_CENTER_ALIGN,
        )
    ws_summary.row_dimensions[start_rule_row].height = 22

    if violations:
        for v_idx, v in enumerate(violations, start=start_rule_row + 1):
            pt_idx = int(v.get("index", v.get("point_index", 0))) + 1
            rule_name = str(v.get("rule", "Unknown Rule"))
            _set_cell(ws_summary, v_idx, 1, pt_idx, font=_NORMAL_FONT, alignment=_CENTER_ALIGN)
            _set_cell(ws_summary, v_idx, 2, rule_name, font=_NORMAL_FONT, alignment=_LEFT_ALIGN)
            _set_cell(ws_summary, v_idx, 3, "Special Cause Detected", font=_NORMAL_FONT, alignment=_CENTER_ALIGN)
    else:
        _set_cell(ws_summary, start_rule_row + 1, 1, "-", font=_NORMAL_FONT, alignment=_CENTER_ALIGN)
        _set_cell(ws_summary, start_rule_row + 1, 2, "None", font=_NORMAL_FONT, alignment=_LEFT_ALIGN)
        _set_cell(
            ws_summary,
            start_rule_row + 1,
            3,
            "In Control (No Violations)",
            font=_NORMAL_FONT,
            alignment=_CENTER_ALIGN,
        )

    ws_summary.freeze_panes = "A2"
    _auto_fit_columns(ws_summary, min_width=25.0)

    return wb


def export_spc_excel(
    chart_type: str,
    data: list[list[float]] | list[float],
    *,
    usl: float | None = None,
    lsl: float | None = None,
    sample_sizes: list[float] | None = None,
    rule_set: str = "Western Electric",
    title: str = "AIAG SPC Control Chart & Process Capability Analysis",
    part_name: str | None = None,
    part_number: str | None = None,
    characteristic: str | None = None,
) -> bytes:
    """Export SPC control chart data and capability analysis to serialized .xlsx bytes.

    Parameters
    ----------
    chart_type : str
        Control chart type: "Xbar-R", "Xbar-S", "I-MR", "p", "c", or "u".
    data : list[list[float]] | list[float]
        Measurement data subgroups or individual readings / counts.
    usl : float | None, optional
        Upper specification limit.
    lsl : float | None, optional
        Lower specification limit.
    sample_sizes : list[float] | None, optional
        Subgroup sample sizes for "p" and "u" charts.
    rule_set : str, optional
        Run-rule standard: "Western Electric" (default) or "Nelson".
    title : str, optional
        Report title.
    part_name : str | None, optional
        Part name.
    part_number : str | None, optional
        Part number.
    characteristic : str | None, optional
        Characteristic name.

    Returns
    -------
    bytes
        Raw .xlsx workbook bytes.
    """
    wb = export_spc_to_workbook(
        chart_type=chart_type,
        data=data,
        usl=usl,
        lsl=lsl,
        sample_sizes=sample_sizes,
        rule_set=rule_set,
        title=title,
        part_name=part_name,
        part_number=part_number,
        characteristic=characteristic,
    )
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
