"""
quality_core/ncr/export.py
Nonconformance Reporting (NCR) live-formula / structured multi-sheet Excel exporter.

Generates multi-sheet .xlsx workbooks per ISO 9001:2015 Clause 8.7 ("Control of nonconforming outputs")
and IATF 16949:2016 Clause 8.7 (8.7.1.1 Customer concession, 8.7.1.4 Control of reworked product,
8.7.1.7 Nonconforming product disposition).

What ships in the workbook (three sheets, in this order):
  1. title (default "Nonconformance Records"): Styled table of records with sanitized text.
  2. "Dispositions & Containment": Live formulas for disposition counts (=COUNTIF),
     quantity sums (=SUMIF), %-of-quantity ratio (=IF($C$8=0, 0, C{r}/$C$8)), and grand totals (=SUM).
  3. "Summary & Metadata": Key-value summary with live =COUNTA and =SUM rollups, and MRB review counts.

Standards References:
- ISO 9001:2015 Clause 8.7 ("Control of nonconforming outputs"): Clause 8.7.1 & Clause 8.7.2.
- IATF 16949:2016 Clause 8.7 ("Control of nonconforming outputs"): Clause 8.7.1.1, 8.7.1.4 & 8.7.1.7.
"""

from __future__ import annotations

import io
from collections.abc import Sequence
from typing import Any

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from quality_core.io.export import (
    Formula,
    now,
    sanitize_cell,
    write_formula_cell,
    write_table_sheet,
)
from quality_core.ncr.schema import (
    DISPOSITION_VALUES,
    NCRDataset,
    NonconformanceRecord,
    validate_ncr,
)

__all__ = [
    "DISPOSITION_SUMMARY_COLUMNS",
    "NCR_COL_WIDTHS",
    "NCR_EXPORT_COLUMNS",
    "benchmark_ncr_dataset",
    "build_ncr_workbook",
    "export_ncr_excel",
    "export_ncr_workbook",
]

# ===========================================================================
# Layout Configuration & Column Descriptors
# ===========================================================================

NCR_EXPORT_COLUMNS: tuple[str, ...] = (
    "Record_ID",
    "Part_Lot_ID",
    "Defect_Description",
    "Requirement_Violated",
    "Quantity_Affected",
    "Detection_Point",
    "Severity",
    "Disposition",
    "Approval_Authority",
    "Rationale",
)

NCR_COL_WIDTHS: dict[str, float] = {
    "Record_ID": 16.0,
    "Part_Lot_ID": 18.0,
    "Defect_Description": 38.0,
    "Requirement_Violated": 34.0,
    "Quantity_Affected": 16.0,
    "Detection_Point": 24.0,
    "Severity": 14.0,
    "Disposition": 18.0,
    "Approval_Authority": 28.0,
    "Rationale": 40.0,
}

DISPOSITION_SUMMARY_COLUMNS: tuple[str, ...] = (
    "Disposition",
    "Record_Count",
    "Quantity_Affected",
    "Pct_of_Total_Quantity",
)

DISPOSITION_SUMMARY_COL_WIDTHS: dict[str, float] = {
    "Disposition": 22.0,
    "Record_Count": 16.0,
    "Quantity_Affected": 20.0,
    "Pct_of_Total_Quantity": 22.0,
}

_FIRST_DATA_ROW = 2
_DISP_SUMMARY_SHEET_TITLE = "Dispositions & Containment"
_SUMMARY_SHEET_TITLE = "Summary & Metadata"
_BOLD_FONT = Font(bold=True, size=10)
_NORMAL_FONT = Font(size=10)

_DISP_COL = get_column_letter(NCR_EXPORT_COLUMNS.index("Disposition") + 1)
_QTY_COL = get_column_letter(NCR_EXPORT_COLUMNS.index("Quantity_Affected") + 1)
_REC_ID_COL = get_column_letter(NCR_EXPORT_COLUMNS.index("Record_ID") + 1)


def _last_row(n_rows: int) -> int:
    """Return the last matrix sheet row the roll-up ranges should span.

    Floored at _FIRST_DATA_ROW (2) so empty datasets build valid 2:2 ranges.
    """
    return max(_FIRST_DATA_ROW + n_rows - 1, _FIRST_DATA_ROW)


def _row_record(record: NonconformanceRecord) -> dict[str, Any]:
    """Map one NonconformanceRecord to its matrix sheet dictionary."""
    return {
        "Record_ID": record.record_id,
        "Part_Lot_ID": record.part_lot_id,
        "Defect_Description": record.defect_description,
        "Requirement_Violated": record.requirement_violated,
        "Quantity_Affected": record.quantity_affected,
        "Detection_Point": record.detection_point,
        "Severity": record.severity,
        "Disposition": record.disposition,
        "Approval_Authority": record.approval_authority,
        "Rationale": record.rationale,
    }


def _write_dispositions_sheet(ws: Any, matrix_title: str, row_count: int) -> None:
    """Write the Dispositions & Containment summary sheet into ws."""
    last_row = _last_row(row_count)
    matrix = f"'{matrix_title}'!"
    disp_rows: list[dict[str, Any]] = []

    # Rows 2..6: Standard disposition categories
    for idx, disp in enumerate(DISPOSITION_VALUES):
        r = 2 + idx
        record_count = Formula(
            f'=COUNTIF({matrix}{_DISP_COL}{_FIRST_DATA_ROW}:{_DISP_COL}{last_row}, "{disp}")'
        )
        qty_affected = Formula(
            f'=SUMIF({matrix}{_DISP_COL}{_FIRST_DATA_ROW}:{_DISP_COL}{last_row}, "{disp}", {matrix}{_QTY_COL}{_FIRST_DATA_ROW}:{_QTY_COL}{last_row})'
        )
        pct_qty = Formula(f"=IF($C$8=0, 0, C{r}/$C$8)", number_format="0.0%")
        disp_rows.append(
            {
                "Disposition": disp,
                "Record_Count": record_count,
                "Quantity_Affected": qty_affected,
                "Pct_of_Total_Quantity": pct_qty,
            }
        )

    # Row 7: Unassigned disposition
    record_count_unassigned = Formula(
        f'=COUNTIF({matrix}{_DISP_COL}{_FIRST_DATA_ROW}:{_DISP_COL}{last_row}, "")'
    )
    qty_affected_unassigned = Formula(
        f'=SUMIF({matrix}{_DISP_COL}{_FIRST_DATA_ROW}:{_DISP_COL}{last_row}, "", {matrix}{_QTY_COL}{_FIRST_DATA_ROW}:{_QTY_COL}{last_row})'
    )
    pct_qty_unassigned = Formula("=IF($C$8=0, 0, C7/$C$8)", number_format="0.0%")
    disp_rows.append(
        {
            "Disposition": "Unassigned",
            "Record_Count": record_count_unassigned,
            "Quantity_Affected": qty_affected_unassigned,
            "Pct_of_Total_Quantity": pct_qty_unassigned,
        }
    )

    # Row 8: Grand total
    disp_rows.append(
        {
            "Disposition": "Total",
            "Record_Count": Formula("=SUM(B2:B7)"),
            "Quantity_Affected": Formula("=SUM(C2:C7)"),
            "Pct_of_Total_Quantity": Formula("=IF(C8=0, 0, SUM(D2:D7))", number_format="0.0%"),
        }
    )

    df_disp = pd.DataFrame(disp_rows, columns=list(DISPOSITION_SUMMARY_COLUMNS))
    write_table_sheet(
        ws,
        df_disp,
        title=_DISP_SUMMARY_SHEET_TITLE,
        columns=DISPOSITION_SUMMARY_COLUMNS,
        col_widths=DISPOSITION_SUMMARY_COL_WIDTHS,
    )


def _write_metadata_sheet(
    ws: Any,
    rows: Sequence[tuple[str, object]],
    *,
    title: str | None = None,
    key_width: float = 30.0,
    value_width: float = 48.0,
) -> None:
    """Write a two-column metadata sheet supporting live Formula instances."""
    if title is not None:
        ws.title = title
    for r_idx, (label, val) in enumerate(rows, start=1):
        ws.cell(r_idx, 1, sanitize_cell(label)).font = _BOLD_FONT
        if isinstance(val, Formula):
            cell = write_formula_cell(
                ws, r_idx, 2, val.formula, number_format=val.number_format
            )
            cell.font = _NORMAL_FONT
        else:
            cell = ws.cell(r_idx, 2, sanitize_cell(val))
            cell.font = _NORMAL_FONT
    ws.column_dimensions["A"].width = key_width
    ws.column_dimensions["B"].width = value_width


def _write_summary_metadata_sheet(
    ws: Any,
    matrix_title: str,
    row_count: int,
    mrb_count: int,
    report_title: str,
) -> None:
    """Write the Summary & Metadata sheet into ws."""
    last_row = _last_row(row_count)
    matrix = f"'{matrix_title}'!"
    rows: list[tuple[str, object]] = [
        ("Report Title", report_title),
        ("Date Generated", now()),
        ("Standards Basis", "ISO 9001:2015 §8.7 / IATF 16949:2016 §8.7"),
        ("Total Records", Formula(f"=COUNTA({matrix}{_REC_ID_COL}{_FIRST_DATA_ROW}:{_REC_ID_COL}{last_row})")),
        ("Total Quantity Affected", Formula(f"=SUM({matrix}{_QTY_COL}{_FIRST_DATA_ROW}:{_QTY_COL}{last_row})")),
        ("MRB Gate Reviews Required", mrb_count),
    ]
    _write_metadata_sheet(ws, rows, title=_SUMMARY_SHEET_TITLE, key_width=30.0, value_width=48.0)


# ===========================================================================
# Public Exporter API
# ===========================================================================


def build_ncr_workbook(
    dataset: NCRDataset | Sequence[NonconformanceRecord] | Sequence[dict[str, Any]] | pd.DataFrame | dict[str, Any],
    *,
    title: str = "Nonconformance Records",
) -> openpyxl.Workbook:
    """Build a 3-sheet openpyxl Workbook for NCR records with live summary roll-up formulas.

    Sheets:
      1. title (default: "Nonconformance Records"): Styled table of records with sanitized text.
      2. "Dispositions & Containment": Live formulas for disposition counts (=COUNTIF),
         quantity sums (=SUMIF), %-of-quantity ratio (=C{r}/$C${total_row}), and grand totals (=SUM).
      3. "Summary & Metadata": Key-value summary with live =COUNTA and =SUM rollups.
    """
    if isinstance(dataset, NCRDataset):
        records = dataset.records
    elif isinstance(dataset, (list, tuple)) and len(dataset) == 0:
        records = []
    elif isinstance(dataset, pd.DataFrame) and len(dataset) == 0:
        records = []
    elif isinstance(dataset, dict) and (dataset.get("records") == [] or dataset.get("rows") == []):
        records = []
    else:
        validated = validate_ncr(dataset)
        records = validated.records

    wb = Workbook()

    # Sheet 1: Matrix
    record_dicts = [_row_record(r) for r in records]
    df = pd.DataFrame(record_dicts, columns=list(NCR_EXPORT_COLUMNS))
    write_table_sheet(
        wb.active,
        df,
        title=title,
        columns=NCR_EXPORT_COLUMNS,
        col_widths=NCR_COL_WIDTHS,
    )

    # Sheet 2: Dispositions & Containment
    ws_disp = wb.create_sheet(_DISP_SUMMARY_SHEET_TITLE)
    _write_dispositions_sheet(ws_disp, title, len(records))

    # Sheet 3: Summary & Metadata
    mrb_count = sum(
        1
        for r in records
        if (
            r.disposition in ("UseAsIs", "Regrade")
            or (r.approval_authority is not None and "MRB" in r.approval_authority)
        )
    )
    ws_summary = wb.create_sheet(_SUMMARY_SHEET_TITLE)
    _write_summary_metadata_sheet(ws_summary, title, len(records), mrb_count, title)

    return wb


def export_ncr_workbook(
    dataset: NCRDataset | Sequence[NonconformanceRecord] | Sequence[dict[str, Any]] | pd.DataFrame | dict[str, Any],
    *,
    title: str = "Nonconformance Records",
) -> bytes:
    """Export NCR dataset to serialized .xlsx bytes."""
    wb = build_ncr_workbook(dataset, title=title)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_ncr_excel(
    dataset: NCRDataset | Sequence[NonconformanceRecord] | Sequence[dict[str, Any]] | pd.DataFrame | dict[str, Any],
    *,
    title: str = "Nonconformance Records",
) -> bytes:
    """Alias for export_ncr_workbook."""
    return export_ncr_workbook(dataset, title=title)


# ===========================================================================
# Benchmark Dataset
# ===========================================================================


def benchmark_ncr_dataset() -> NCRDataset:
    """Return a fresh 5-record benchmark NCRDataset covering all 5 disposition types."""
    records = [
        NonconformanceRecord(
            record_id="NCR-2026-001",
            part_lot_id="LOT-BRK-8821",
            defect_description="Cast porosity on brake caliper mounting flange exceeding max allowable void diameter.",
            requirement_violated="DWG-BRK-004 Rev D: Max surface pore diameter <= 0.50 mm; zero clustering permitted.",
            quantity_affected=45,
            detection_point="Receiving Inspection / CMM Cell 1",
            disposition="ReturnToVendor",
            severity="Major",
            rationale="Defect originated from external foundry supplier; nonconforming casting lot rejected and segregated for return per ISO 9001:2015 Clause 8.7.1(b).",
            approval_authority="Supplier Quality Assurance (SQA) / Purchasing",
        ),
        NonconformanceRecord(
            record_id="NCR-2026-002",
            part_lot_id="LOT-SHAFT-4410",
            defect_description="Drive shaft bearing journal outer diameter turned oversized at +0.035 mm above tolerance.",
            requirement_violated="SPEC-SFT-102: Bearing journal OD = 35.000 +0.005/-0.000 mm.",
            quantity_affected=120,
            detection_point="CNC Turning Station 3 / In-Process Post-Op Gauge",
            disposition="Rework",
            severity="Moderate",
            rationale="Product has excess stock and can be precision skim-ground to specification per ISO 9001:2015 Clause 8.7.1(a); risk analysis required per IATF 16949:2016 Clause 8.7.1.4.",
            approval_authority="Manufacturing Engineering & Quality Engineering",
        ),
        NonconformanceRecord(
            record_id="NCR-2026-003",
            part_lot_id="LOT-HSG-9904",
            defect_description="Inverter housing CNC internal bore undercut wall thickness below minimum structural limit.",
            requirement_violated="DWG-INV-012: Minimum wall thickness >= 3.20 mm across pressure envelope.",
            quantity_affected=18,
            detection_point="Final Machining CMM Inspection",
            disposition="Scrap",
            severity="Critical",
            rationale="Under-thickness structural wall cannot be restored to drawing requirements; must be defaced and rendered unusable per IATF 16949:2016 Clause 8.7.1.7.",
            approval_authority="Quality Manager / Scrap Authority",
        ),
        NonconformanceRecord(
            record_id="NCR-2026-004",
            part_lot_id="LOT-BKT-1102",
            defect_description="Zinc phosphate bracket e-coat surface finish minor gloss variation on non-cosmetic underbody bracket.",
            requirement_violated="SPEC-COAT-09: Gloss reading 60° target = 70 ± 5 GU; measured 58 GU on non-critical face.",
            quantity_affected=250,
            detection_point="E-Coat Line Unload Inspection",
            disposition="UseAsIs",
            severity="Minor",
            rationale="Deviation does not impair corrosion resistance, fit, form, function, or vehicle safety; customer concession permit obtained per ISO 9001:2015 Clause 8.7.1(d) and IATF 16949:2016 Clause 8.7.1.1.",
            approval_authority="Material Review Board (MRB) & Customer Concession Permit",
        ),
        NonconformanceRecord(
            record_id="NCR-2026-005",
            part_lot_id="LOT-ROD-3309",
            defect_description="Connecting rod tensile yield strength 780 MPa vs Grade A requirement 820 MPa; conforms to Grade B requirement (>= 750 MPa).",
            requirement_violated="MAT-SPEC-04 Grade A: Yield strength >= 820 MPa.",
            quantity_affected=300,
            detection_point="Metallurgical Lab Tensile Test Gate",
            disposition="Regrade",
            severity="Moderate",
            rationale="Lot meets Grade B application specification; authorized for secondary application transfer per IATF 16949:2016 Clause 8.7.1.7.",
            approval_authority="Material Review Board (MRB) & Customer Approval",
        ),
    ]
    return NCRDataset(records=records)
