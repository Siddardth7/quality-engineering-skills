"""
quality_core/ppap/export.py
PPAP submission-readiness Excel exporter — five live roll-up / gate formulas.

Consumer of the domain-agnostic ``quality_core.io.export`` primitives: that module owns
the cross-cutting machinery (formula-injection escaping, openpyxl table styling, the
``Formula`` opt-in marker) and stays free of domain knowledge; this module supplies the
PPAP *config* — column order, widths, and which cells are live formulas.

It lives in the ``ppap`` domain package rather than alongside ``io``'s primitives because
the dependency only points one way: ``ppap`` may import ``io`` (and ``ppap.schema``
already does, for the ingest boundary), but ``io`` sits structurally below every domain
engine and must never import one back. Putting a PPAP exporter inside ``io`` would close
that arrow into a cycle. For the same reason the benchmark package below is built here
from ``PPAPPackage``/``EvidenceItem`` directly and is **not** imported from
``quality_core.canvas``, which sits structurally *above* this package.

🔒 THE AUTHORITY INVARIANT (carried in from ``auditor.py`` / ``ASSUMPTIONS_LOG.md``
RULE 12): Section 5 dispositions — Approved, Interim Approval, Rejected — are assigned
exclusively by the customer's authorized representative. This workbook reports
supplier-side **submission readiness** only (``SUBMISSION_READY`` / ``NOT_READY`` /
``INDETERMINATE`` at package level; ``SUBMITTED`` / ``RETAINED_ON_FILE`` / ``MISSING`` /
``NOT_APPLICABLE`` / ``INDETERMINATE`` / ``EVIDENCE_INVALID`` per element). No column,
metric label, or literal in this module can carry an approval token, and none may be
added.

COMPOSITION DISCIPLINE: this module contains no standards data of its own. Every value
it writes is one the engine already computed — requirement codes come from
``ElementAuditResult.requirement_code`` (looked up by the auditor from AIAG Table 4.1),
the §2.2.11.3 acceptance thresholds are *imported* from
``quality_core.ppap.process_study`` and interpolated into the gate formula, and the
standard-mandated ``required_action`` text is passed through verbatim. Nothing is
recomputed, re-typed, or re-derived here.

What ships in the workbook (three sheets, in this order):

  1. the **checklist** sheet, named by ``title`` (default ``"PPAP Checklist"``). Rows
     2–19 are the 18 canonical elements in ``PPAP_ELEMENT_IDS`` order (§2.2.1 → §2.2.18),
     never re-sorted. Every cell on it is plain audit data written through
     ``write_table_sheet``'s ordinary ``sanitize_cell`` escaping. **No cell on this sheet
     is a live formula.**
  2. the **``"Completeness"``** sheet — package metadata plus the four ``COUNTA`` /
     ``COUNTIF`` roll-ups at fixed addresses ``B6``–``B9`` that never move, because the
     checklist is always exactly 18 data rows.
  3. the **``"Capability Gate"``** sheet — the §2.2.11 Initial Process Study result, whose
     ``B4`` band gate is a live ``IF``/``IF`` cascade over the ``B3`` index literal and the
     two imported thresholds. It is written **only** when the engine's own result reached
     the §2.2.11.3 band evaluation (``ProcessStudyResult.band is not None``); the
     attribute-data guard, insufficient-sample gate, and stability gate are qualitative
     decisions that stay in the engine and render ``"N/A"`` here.

Column letters are derived from ``PPAP_EXPORT_COLUMNS`` via ``get_column_letter``, never
hand-typed, so reordering the column tuple cannot silently desync a roll-up from the
column it is supposed to count.

No new standards constant lives here — see the "EXPORT BOUNDARY" entry in
``ppap/ASSUMPTIONS_LOG.md``.
"""

from __future__ import annotations

from collections.abc import Sequence
from io import BytesIO
from typing import Any, NamedTuple

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

# Imported from the sibling modules rather than the ``quality_core.ppap`` package, so
# this module is importable on its own and cannot see a partially-initialised package
# when it is loaded from that package's own ``__init__``.
from quality_core.io import (
    Formula,
    now,
    sanitize_cell,
    write_formula_cell,
    write_table_sheet,
)
from quality_core.ppap.auditor import (
    ElementAuditResult,
    ElementAuditVerdict,
    PPAPAuditResult,
    audit_ppap_package,
)
from quality_core.ppap.process_study import (
    ACCEPTANCE_THRESHOLD_CAPABLE,
    ACCEPTANCE_THRESHOLD_POTENTIALLY_CAPABLE,
    ProcessStudyResult,
)
from quality_core.ppap.schema import (
    PPAP_ELEMENT_IDS,
    EvidenceItem,
    PPAPElementId,
    PPAPPackage,
)

__all__ = [
    "PPAP_COL_WIDTHS",
    "PPAP_EXPORT_COLUMNS",
    "benchmark_ppap_package",
    "build_ppap_workbook",
    "export_ppap_workbook",
]


# ===========================================================================
# Checklist sheet layout
# ===========================================================================

#: Exported checklist columns, in sheet order (1-based column index = position + 1).
#: Readiness vocabulary only — there is deliberately no approval/disposition column.
PPAP_EXPORT_COLUMNS: tuple[str, ...] = (
    "Element_ID",
    "Element_Name",
    "Requirement_Code",
    "Applicability_Verdict",
    "Audit_Verdict",
    "Evidence_Status",
    "Document_Reference",
    "Rationale",
)

#: Column widths keyed by column name — wide for free text, narrow for codes and
#: verdicts. Presentation only; ``write_table_sheet`` falls back to ``default_width``
#: for anything missing, but every named column is given one here.
PPAP_COL_WIDTHS: dict[str, float] = {
    "Element_ID": 11,
    "Element_Name": 46,
    "Requirement_Code": 17,
    "Applicability_Verdict": 21,
    "Audit_Verdict": 18,
    "Evidence_Status": 16,
    "Document_Reference": 30,
    "Rationale": 70,
}

#: Sheet row of the first checklist data row: row 1 is the header written by
#: ``write_table_sheet``, whose body enumeration starts at 2.
_FIRST_DATA_ROW = 2

#: Sheet row of the last checklist data row. Fixed, not derived from the data: the
#: audit always covers exactly the 18 canonical elements, so — unlike the Control Plan
#: matrix — the roll-up ranges can never collapse to an empty or reversed span.
_LAST_DATA_ROW = _FIRST_DATA_ROW + len(PPAP_ELEMENT_IDS) - 1


def _column_letter(name: str) -> str:
    """Return the Excel column letter for ``name``'s position in the checklist layout."""
    return get_column_letter(PPAP_EXPORT_COLUMNS.index(name) + 1)


_ELEMENT_ID_COL = _column_letter("Element_ID")
_AUDIT_VERDICT_COL = _column_letter("Audit_Verdict")


def _element_record(element: ElementAuditResult) -> dict[str, object]:
    """Map one ``ElementAuditResult`` to its sheet record, keyed by the checklist columns.

    ``evidence_status`` and ``document_reference`` are nullable and pass through
    unchanged — ``write_table_sheet`` writes ``None`` as a blank cell, so no coercion
    (which would turn a blank into the literal text ``"None"``) happens here.
    """
    return {
        "Element_ID": element.element_id,
        "Element_Name": element.element_name,
        "Requirement_Code": element.requirement_code,
        "Applicability_Verdict": element.applicability_verdict,
        "Audit_Verdict": element.verdict,
        "Evidence_Status": element.evidence_status,
        "Document_Reference": element.document_reference,
        "Rationale": element.rationale,
    }


# ===========================================================================
# Shared two-column metric sheet
# ===========================================================================

_METRIC_HEADER: tuple[str, str] = ("Metric", "Value")

#: Sheet row of the first metric (row 1 is the Metric/Value header).
_FIRST_METRIC_ROW = 2
_METRIC_LABEL_COLUMN = 1
_METRIC_VALUE_COLUMN = 2
_METRIC_LABEL_WIDTH = 38.0

#: Rendered in a value cell whenever the engine has no value to report for it.
_NOT_AVAILABLE = "N/A"


class _MetricRow(NamedTuple):
    """One ``label`` / ``value`` row of a two-column metric sheet.

    ``value`` is written as a live formula if — and only if — it is a
    :class:`~quality_core.io.export.Formula`; anything else is routed through
    ``sanitize_cell`` and lands inert. ``number_format`` applies to literal values only
    (a ``Formula`` carries its own).
    """

    label: str
    value: object
    number_format: str | None = None


def _write_metric_sheet(ws: Any, rows: Sequence[_MetricRow], *, value_width: float) -> None:
    """Write a fixed two-column metric block into ``ws``.

    Column A holds plain sanitized labels; column B holds each row's value. Built with
    direct cell calls rather than ``write_table_sheet``/``write_keyvalue_sheet`` because
    this is a fixed metric block at addresses that must not move, and
    ``write_keyvalue_sheet`` has no ``Formula`` support.
    """
    for col_idx, label in enumerate(_METRIC_HEADER, start=1):
        ws.cell(row=1, column=col_idx, value=sanitize_cell(label))

    for offset, metric in enumerate(rows):
        sheet_row = _FIRST_METRIC_ROW + offset
        ws.cell(row=sheet_row, column=_METRIC_LABEL_COLUMN, value=sanitize_cell(metric.label))
        if isinstance(metric.value, Formula):
            # Explicit, type-based opt-in — the ONLY way a cell here becomes live.
            write_formula_cell(
                ws,
                sheet_row,
                _METRIC_VALUE_COLUMN,
                metric.value.formula,
                number_format=metric.value.number_format,
            )
        else:
            cell = ws.cell(
                row=sheet_row,
                column=_METRIC_VALUE_COLUMN,
                value=sanitize_cell(metric.value),
            )
            if metric.number_format is not None:
                cell.number_format = metric.number_format

    ws.column_dimensions[get_column_letter(_METRIC_LABEL_COLUMN)].width = _METRIC_LABEL_WIDTH
    ws.column_dimensions[get_column_letter(_METRIC_VALUE_COLUMN)].width = value_width


# ===========================================================================
# Completeness sheet — the COUNTA / COUNTIF roll-ups
# ===========================================================================

_COMPLETENESS_SHEET_TITLE = "Completeness"
_COMPLETENESS_VALUE_WIDTH = 30.0

#: The three element verdicts rolled up by ``COUNTIF``, paired with their sheet labels.
#: Typed against ``ElementAuditVerdict`` so a typo in a search string is a mypy error
#: rather than a formula that silently counts zero. Every member is readiness
#: vocabulary — the authority invariant forbids an approval token here.
_VERDICT_ROLLUPS: tuple[tuple[str, ElementAuditVerdict], ...] = (
    ("Submitted", "SUBMITTED"),
    ("Retained On File", "RETAINED_ON_FILE"),
    ("Missing", "MISSING"),
)


def _completeness_rows(audit_result: PPAPAuditResult, checklist_title: str) -> tuple[_MetricRow, ...]:
    """Return the Completeness sheet's metric rows, including its four live formulas.

    Exporter-authored: the ``Formula`` instances are built from fixed templates plus
    column letters this module derives itself. ``checklist_title`` is quoted
    unconditionally — the default title contains a space, and Excel requires quoting any
    sheet name that is not guaranteed free of spaces/specials.

    ``COUNTA`` targets ``Element_ID`` because that field is non-blank on all 18 rows, so
    it can neither under- nor over-count against ``len(PPAP_ELEMENT_IDS)``. ``COUNTIF``
    targets ``Audit_Verdict``, whose values are the auditor's own enumerated tokens, so
    each exact match is unambiguous.
    """
    checklist = f"'{checklist_title}'!"
    element_range = (
        f"{checklist}{_ELEMENT_ID_COL}{_FIRST_DATA_ROW}:{_ELEMENT_ID_COL}{_LAST_DATA_ROW}"
    )
    verdict_range = (
        f"{checklist}{_AUDIT_VERDICT_COL}{_FIRST_DATA_ROW}:{_AUDIT_VERDICT_COL}{_LAST_DATA_ROW}"
    )
    return (
        _MetricRow("Submission Level", audit_result.submission_level),
        _MetricRow("Reason for Submission", audit_result.reason_for_submission),
        # Readiness, never a customer disposition — see the module docstring.
        _MetricRow("Package Verdict (Submission Readiness)", audit_result.package_verdict),
        _MetricRow("Date Generated", now()),
        _MetricRow("Total Elements", Formula(f"=COUNTA({element_range})")),
        *(
            _MetricRow(label, Formula(f'=COUNTIF({verdict_range},"{verdict}")'))
            for label, verdict in _VERDICT_ROLLUPS
        ),
    )


# ===========================================================================
# Capability Gate sheet — the §2.2.11.3 acceptance-band formula
# ===========================================================================

_CAPABILITY_SHEET_TITLE = "Capability Gate"
_CAPABILITY_VALUE_WIDTH = 88.0

_INDEX_VALUE_NUMBER_FORMAT = "0.0000"

#: Sheet row offsets within the Capability Gate block (row 1 is the header).
_INDEX_VALUE_ROW = _FIRST_METRIC_ROW + 1

#: Address the band gate compares against: the Index Value literal on its own sheet.
#: Derived, never hand-typed, so the two cannot drift apart.
_INDEX_VALUE_CELL = f"{get_column_letter(_METRIC_VALUE_COLUMN)}{_INDEX_VALUE_ROW}"

_CAPABILITY_LABELS: tuple[str, str, str, str, str] = (
    "Index Type",
    "Index Value",
    "Acceptance Band Gate (§2.2.11.3)",
    "Engine Verdict",
    "Required Action",
)


def _band_gate_formula() -> Formula:
    """Return the live §2.2.11.3 acceptance-band gate over the Index Value cell.

    The two thresholds are **imported** from ``quality_core.ppap.process_study`` and
    interpolated, never re-typed here: a second copy of a standards value inside a
    presentation boundary could drift from the engine's. Scope is the numeric band
    comparison only (``Index > 1.67`` / ``1.33 <= Index <= 1.67`` / ``Index < 1.33``);
    the attribute-data guard, sample-size gate, and stability gate stay in the engine,
    which is why this formula is written only once ``band is not None`` proves the engine
    already cleared them.
    """
    return Formula(
        f"=IF({_INDEX_VALUE_CELL}>{ACCEPTANCE_THRESHOLD_CAPABLE},"
        f'"ACCEPTABLE",'
        f"IF({_INDEX_VALUE_CELL}>={ACCEPTANCE_THRESHOLD_POTENTIALLY_CAPABLE},"
        f'"POTENTIALLY_ACCEPTABLE","UNACCEPTABLE"))'
    )


def _or_na(value: object) -> object:
    """Return ``value``, or the ``"N/A"`` placeholder when the engine left it unset."""
    return _NOT_AVAILABLE if value is None else value


def _capability_rows(result: ProcessStudyResult | None) -> tuple[_MetricRow, ...]:
    """Return the Capability Gate sheet's metric rows for ``result``.

    With no study supplied every row is a plain ``"N/A"`` literal and the sheet carries
    zero live formulas. With a study supplied each field renders whatever the engine
    populated on the branch that produced it (``NOT_APPLICABLE_ATTRIBUTE_DATA`` leaves
    the index type/value unset, for instance), and the band gate goes live only when
    ``band is not None``.
    """
    if result is None:
        return tuple(_MetricRow(label, _NOT_AVAILABLE) for label in _CAPABILITY_LABELS)

    gate: object = _NOT_AVAILABLE if result.band is None else _band_gate_formula()
    return (
        _MetricRow(_CAPABILITY_LABELS[0], _or_na(result.index_type)),
        _MetricRow(
            _CAPABILITY_LABELS[1],
            _or_na(result.index_value),
            None if result.index_value is None else _INDEX_VALUE_NUMBER_FORMAT,
        ),
        _MetricRow(_CAPABILITY_LABELS[2], gate),
        _MetricRow(_CAPABILITY_LABELS[3], result.verdict),
        # Verbatim standard-mandated action text, engine-authored and passed through
        # the ordinary escaping unchanged — never re-typed here.
        _MetricRow(_CAPABILITY_LABELS[4], result.required_action),
    )


# ===========================================================================
# Public API
# ===========================================================================


def build_ppap_workbook(
    package: PPAPPackage | None = None,
    *,
    audit_result: PPAPAuditResult | None = None,
    process_study_result: ProcessStudyResult | None = None,
    title: str = "PPAP Checklist",
) -> openpyxl.Workbook:
    """Build a 3-sheet PPAP submission-readiness workbook with live formulas.

    Exactly one of ``package`` / ``audit_result`` must resolve the audit:

    - ``audit_result`` given: used directly, verbatim, with **no** re-audit. It takes
      precedence over ``package``, which is then unused — passing both never silently
      discards the caller's result.
    - ``audit_result`` is None and ``package`` given: ``audit_ppap_package(package)`` is
      called with the package's own ``submission_level`` / ``reason_for_submission`` and
      no override kwargs. A caller needing a different level/reason or explicit
      applicability overrides calls ``audit_ppap_package`` itself and passes
      ``audit_result=``.
    - both None: raises ``ValueError``.

    ``process_study_result`` is optional; when omitted the ``"Capability Gate"`` sheet
    renders five ``"N/A"`` literals and carries no live formula.

    ``title`` names the checklist worksheet (default ``"PPAP Checklist"``) and is the
    sheet name the ``Completeness`` formulas qualify their ranges with, so a custom
    title stays consistent across both sheets. It is passed to ``write_table_sheet``
    unvalidated, matching every other caller.

    The workbook reports submission readiness only; it never emits a Section 5 customer
    disposition (see the module docstring's authority invariant).
    """
    if audit_result is None:
        if package is None:
            raise ValueError("Either package or audit_result must be provided.")
        audit_result = audit_ppap_package(package)

    # Canonical §2.2.1 → §2.2.18 order, driven by PPAP_ELEMENT_IDS rather than by the
    # result dict's own iteration order, so the sheet's row order is guaranteed even if
    # a caller hands in a result assembled some other way. Never re-sorted.
    records = [_element_record(audit_result.elements[elem_id]) for elem_id in PPAP_ELEMENT_IDS]
    df = pd.DataFrame(records, columns=list(PPAP_EXPORT_COLUMNS))

    wb = openpyxl.Workbook()
    write_table_sheet(
        wb.active,
        df,
        title=title,
        columns=PPAP_EXPORT_COLUMNS,
        col_widths=PPAP_COL_WIDTHS,
    )
    _write_metric_sheet(
        wb.create_sheet(_COMPLETENESS_SHEET_TITLE),
        _completeness_rows(audit_result, title),
        value_width=_COMPLETENESS_VALUE_WIDTH,
    )
    _write_metric_sheet(
        wb.create_sheet(_CAPABILITY_SHEET_TITLE),
        _capability_rows(process_study_result),
        value_width=_CAPABILITY_VALUE_WIDTH,
    )
    return wb


def export_ppap_workbook(
    package: PPAPPackage | None = None,
    *,
    audit_result: PPAPAuditResult | None = None,
    process_study_result: ProcessStudyResult | None = None,
    title: str = "PPAP Checklist",
) -> bytes:
    """Serialize :func:`build_ppap_workbook` to .xlsx bytes.

    Returns the workbook ready to hand to a download button or write to disk; the
    arguments and their precedence are exactly :func:`build_ppap_workbook`'s.
    """
    wb = build_ppap_workbook(
        package,
        audit_result=audit_result,
        process_study_result=process_study_result,
        title=title,
    )
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===========================================================================
# Benchmark package
# ===========================================================================

# Self-contained PPAP evidence, built straight from ``EvidenceItem``/``PPAPPackage``.
# Deliberately NOT imported from ``quality_core.canvas`` (which sits structurally above
# this package, so depending on it here would invert the layering).


def _submitted(element_id: PPAPElementId, document_reference: str) -> EvidenceItem:
    """Evidence present and submitted to the customer, with a copy retained."""
    return EvidenceItem(
        element_id=element_id,
        present=True,
        status="submitted",
        submitted_to_customer=True,
        retained_at_organization=True,
        document_reference=document_reference,
    )


def _retained(element_id: PPAPElementId, document_reference: str) -> EvidenceItem:
    """Evidence present and retained at the organization, not submitted."""
    return EvidenceItem(
        element_id=element_id,
        present=True,
        status="retained",
        submitted_to_customer=False,
        retained_at_organization=True,
        document_reference=document_reference,
    )


def _absent(element_id: PPAPElementId) -> EvidenceItem:
    """Evidence confirmed absent — surveyed and not found, not merely un-surveyed."""
    return EvidenceItem(element_id=element_id, present=False, status="missing")


def benchmark_ppap_package() -> PPAPPackage:
    """Return a self-contained benchmark ``PPAPPackage`` for exports and demos.

    An automotive Level 3 initial submission carrying evidence for all 18 elements, in a
    genuinely partial distribution rather than a degenerate 0/18 or 18/18: 11 elements
    audit ``SUBMITTED``, 2 ``RETAINED_ON_FILE`` (§2.2.16/§2.2.17 are Table 4.1 coded
    ``R`` at Level 3), 3 ``MISSING``, and 2 ``INDETERMINATE``.

    The two ``INDETERMINATE`` elements are §2.2.3 (Customer Engineering Approval) and
    §2.2.15 (Master Sample): their applicability turns on
    ``customer_engineering_approval_required`` and ``master_sample_waived``, which are
    **not** ``PPAPPackage`` fields — ``assess_applicability`` reads them only from a raw
    dict or an explicit kwarg. Since the exporter audits a package with no override
    kwargs (by design — see :func:`build_ppap_workbook`), the engine correctly reports
    them as un-surveyed. Supplying evidence for them anyway makes it unambiguous that
    the verdict comes from applicability, not from absent evidence. A caller wanting
    those two decided passes their own ``audit_ppap_package(...)`` result via
    ``audit_result=``.

    A fresh ``PPAPPackage`` and a fresh set of ``EvidenceItem``s are built per call, so a
    caller mutating the returned package cannot affect the next one.
    """
    return PPAPPackage(
        part_name="Electric power steering pinion shaft",
        part_number="EPS-PIN-4417",
        organization="Northline Precision Machining",
        customer="Aurora Motors",
        submission_level=3,
        reason_for_submission="Initial Submission",
        application="C-segment EPS gear assembly",
        has_design_responsibility=True,
        appearance_item=True,
        has_checking_aid=True,
        elements=[
            _submitted("2.2.1", "DRW-EPS-PIN-4417 Rev D"),
            _submitted("2.2.2", "ECN-2026-0431"),
            _submitted("2.2.3", "CEA-AURORA-8812"),
            _submitted("2.2.4", "DFMEA-EPS-PIN-4417 Rev C"),
            _submitted("2.2.5", "PFD-EPS-PIN-4417 Rev B"),
            _submitted("2.2.6", "PFMEA-EPS-PIN-4417 Rev C"),
            _submitted("2.2.7", "CP-EPS-PIN-4417 Rev C"),
            _submitted("2.2.8", "MSA-GRR-PIN-BORE-2026-02"),
            _submitted("2.2.9", "DIM-REPORT-EPS-PIN-4417-01"),
            _submitted("2.2.10", "MAT-CERT-4140H-LOT-77120"),
            _submitted("2.2.11", "IPS-PPK-PIN-SPLINE-2026-02"),
            _absent("2.2.12"),
            _absent("2.2.13"),
            _absent("2.2.14"),
            _retained("2.2.15", "MASTER-SAMPLE-EPS-PIN-4417"),
            _retained("2.2.16", "CHK-AID-PIN-SPLINE-GO-NOGO"),
            _retained("2.2.17", "CSR-AURORA-QMS-2026 Rev 4"),
            _submitted("2.2.18", "PSW-EPS-PIN-4417-001"),
        ],
    )
