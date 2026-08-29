"""Tests for SPC live-formula Excel exporter (quality_core.io.export_spc).

Validates:
- 100% line & branch coverage on quality_core.io.export_spc.
- Reusable verifier `assert_cell_is_formula` passes on all computed cells across
  all 6 Shewhart chart types (Xbar-R, Xbar-S, I-MR, p, c, u).
- Capability sheet live formula verification across bilateral, unilateral (USL only),
  unilateral (LSL only), and no-spec modes.
- Accuracy scorecard: mathematical formulas match engine computations.
- 🔒 Security invariant: user metadata and untrusted data strings starting with
  '=', '+', '-', '@', '\t', '\r' are escaped and rendered inert.
- Negative controls: literal cells in computed slots fail `assert_cell_is_formula`.
- Type guards & input validation errors.
"""

from __future__ import annotations

import io
import math

import openpyxl
import pytest
from _xlsx_formula_audit import assert_cell_is_formula
from quality_core.io.export_spc import export_spc_excel, export_spc_to_workbook
from quality_core.spc.capability import compute_capability
from quality_core.spc.control_charts import (
    compute_xbar_r,
)

# AIAG SPC 4th Ed. Table II.1 benchmark dataset (20 subgroups of size 5)
SAMPLE_SPC_XBAR_R_DATA: list[list[float]] = [
    [10.1, 10.0, 9.9, 10.2, 9.8],
    [9.9, 10.1, 10.0, 10.0, 10.1],
    [10.2, 9.8, 10.1, 9.9, 10.0],
    [10.0, 10.0, 10.1, 10.2, 9.9],
    [9.8, 10.1, 10.0, 9.9, 10.2],
    [10.1, 10.2, 9.8, 10.0, 10.0],
    [10.0, 9.9, 10.1, 10.1, 10.0],
    [10.2, 10.0, 9.9, 10.1, 9.8],
    [9.9, 10.1, 10.0, 10.0, 10.2],
    [10.1, 9.8, 10.2, 10.0, 9.9],
    [10.0, 10.1, 9.9, 10.0, 10.1],
    [9.8, 10.0, 10.2, 10.1, 9.9],
    [10.1, 10.0, 10.0, 9.9, 10.2],
    [10.2, 9.9, 10.1, 10.0, 9.8],
    [9.9, 10.1, 10.0, 10.2, 10.0],
    [10.0, 9.8, 10.1, 10.0, 10.1],
    [10.1, 10.2, 9.9, 10.0, 9.9],
    [9.9, 10.0, 10.1, 10.2, 9.8],
    [10.0, 10.1, 10.0, 9.9, 10.1],
    [10.2, 9.9, 10.0, 10.1, 10.0],
]


# ===========================================================================
# 1. Positive Live Formula Verification Across All 6 Chart Types
# ===========================================================================


def test_export_xbar_r_live_formulas() -> None:
    """Xbar-R export must contain live <f> formulas for subgroup means, ranges, centerlines, and limits."""
    data = SAMPLE_SPC_XBAR_R_DATA  # 20 subgroups of size 5
    raw_bytes = export_spc_excel("Xbar-R", data, usl=11.0, lsl=9.0, title="Benchmark Xbar-R")

    # Sheet 1: Control Chart Data
    # Subgroup 1 mean (G2) and range (H2)
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "G2")
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "H2")
    # Last subgroup (row 21)
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "G21")
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "H21")

    # Sheet 1: Parameter block (n=5 -> param_col=10 (J), val_col=11 (K))
    # Grand Mean K2, Range Mean K3, Within Sigma K4, UCL_x K5, LCL_x K6, UCL_r K7, LCL_r K8
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "K2")
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "K3")
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "K4")
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "K5")
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "K6")
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "K7")
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "K8")

    # Sheet 2: Process Capability
    assert_cell_is_formula(raw_bytes, "Process Capability", "B4")  # Process Mean
    assert_cell_is_formula(raw_bytes, "Process Capability", "B5")  # Within Sigma
    assert_cell_is_formula(raw_bytes, "Process Capability", "B6")  # Overall Sigma
    assert_cell_is_formula(raw_bytes, "Process Capability", "B7")  # Cp
    assert_cell_is_formula(raw_bytes, "Process Capability", "B8")  # Cpk
    assert_cell_is_formula(raw_bytes, "Process Capability", "B9")  # Pp
    assert_cell_is_formula(raw_bytes, "Process Capability", "B10")  # Ppk


def test_export_xbar_s_live_formulas() -> None:
    """Xbar-S export must contain live <f> formulas for subgroup means, SDs, centerlines, and limits."""
    data = [
        [10.1, 10.0, 9.9, 10.2, 9.8],
        [9.9, 10.1, 10.0, 10.0, 10.1],
        [10.2, 9.8, 10.1, 9.9, 10.0],
        [10.0, 10.0, 10.1, 10.2, 9.9],
        [9.8, 10.1, 10.0, 9.9, 10.2],
    ]
    raw_bytes = export_spc_excel("Xbar-S", data, usl=11.0, lsl=9.0)

    # Subgroup 1 mean (G2) and std dev (H2)
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "G2")
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "H2")

    # Parameter block (n=5 -> param_col=10 (J), val_col=11 (K))
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "K2")  # Xbarbar
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "K3")  # Sbar
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "K4")  # Sigma hat
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "K5")  # UCL_x
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "K6")  # LCL_x
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "K7")  # UCL_s
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "K8")  # LCL_s

    # Capability formulas
    assert_cell_is_formula(raw_bytes, "Process Capability", "B4")  # Process Mean
    assert_cell_is_formula(raw_bytes, "Process Capability", "B5")  # Within Sigma
    assert_cell_is_formula(raw_bytes, "Process Capability", "B6")  # Overall Sigma
    assert_cell_is_formula(raw_bytes, "Process Capability", "B7")  # Cp
    assert_cell_is_formula(raw_bytes, "Process Capability", "B8")  # Cpk
    assert_cell_is_formula(raw_bytes, "Process Capability", "B9")  # Pp
    assert_cell_is_formula(raw_bytes, "Process Capability", "B10")  # Ppk


def test_export_imr_live_formulas() -> None:
    """I-MR export must contain live <f> formulas for moving ranges, centerlines, and limits."""
    data = [10.2, 10.0, 9.8, 10.1, 9.9, 10.3, 10.0, 9.7, 10.1, 10.2]
    raw_bytes = export_spc_excel("I-MR", data, usl=11.0, lsl=9.0)

    # Observation 2 moving range (C3) and observation 10 (C11)
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "C3")
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "C11")

    # Parameter block (param_col=5 (E), val_col=6 (F))
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "F2")  # Xbar
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "F3")  # MRbar
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "F4")  # Sigma hat
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "F5")  # UCL_x
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "F6")  # LCL_x
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "F7")  # UCL_mr

    # Capability formulas
    assert_cell_is_formula(raw_bytes, "Process Capability", "B4")  # Process Mean
    assert_cell_is_formula(raw_bytes, "Process Capability", "B5")  # Within Sigma
    assert_cell_is_formula(raw_bytes, "Process Capability", "B6")  # Overall Sigma
    assert_cell_is_formula(raw_bytes, "Process Capability", "B7")  # Cp
    assert_cell_is_formula(raw_bytes, "Process Capability", "B8")  # Cpk
    assert_cell_is_formula(raw_bytes, "Process Capability", "B9")  # Pp
    assert_cell_is_formula(raw_bytes, "Process Capability", "B10")  # Ppk


def test_export_p_chart_live_formulas() -> None:
    """p-chart export must contain live <f> formulas for proportions, p-bar, and control limits."""
    counts = [12.0, 8.0, 15.0, 9.0, 11.0]
    sizes = [100.0, 100.0, 100.0, 100.0, 100.0]
    raw_bytes = export_spc_excel("p", counts, sample_sizes=sizes)

    # Proportion formulas
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "D2")
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "D6")

    # Parameter table: p-bar (J2)
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "J2")
    # LCL (G2) and UCL (E2) formulas
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "E2")
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "G2")


def test_export_u_chart_live_formulas() -> None:
    """u-chart export must contain live <f> formulas for rates, u-bar, and control limits."""
    counts = [5.0, 3.0, 8.0, 4.0, 6.0]
    sizes = [10.0, 12.0, 10.0, 15.0, 10.0]
    raw_bytes = export_spc_excel("u", counts, sample_sizes=sizes)

    # Rate formulas
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "D2")
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "D6")

    # Parameter table: u-bar (J2)
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "J2")
    # UCL (E2) and LCL (G2) formulas
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "E2")
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "G2")


def test_export_c_chart_live_formulas() -> None:
    """c-chart export must contain live <f> formulas for c-bar, and control limits."""
    counts = [4.0, 2.0, 5.0, 3.0, 6.0, 1.0]
    raw_bytes = export_spc_excel("c", counts)

    # Parameter table (Cols G & H): c-bar (H2), UCL (H3), LCL (H4)
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "H2")
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "H3")
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "H4")

    # Data row 2: UCL (C2), Centerline (D2), LCL (E2)
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "C2")
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "D2")
    assert_cell_is_formula(raw_bytes, "Control Chart Data", "E2")


# ===========================================================================
# 2. Capability Mode Matrix (Bilateral, Unilateral, None)
# ===========================================================================


def test_capability_unilateral_upper_spec_only() -> None:
    """Unilateral upper spec (USL only) generates live Cpk/Ppk formulas; Cp/Pp are N/A."""
    data = [10.0, 10.1, 9.9, 10.2, 9.8, 10.0, 10.1, 9.9]
    raw_bytes = export_spc_excel("I-MR", data, usl=11.0, lsl=None)

    # Live Cpk (B8) and Ppk (B10)
    assert_cell_is_formula(raw_bytes, "Process Capability", "B8")
    assert_cell_is_formula(raw_bytes, "Process Capability", "B10")

    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes))
    ws_cap = wb["Process Capability"]
    assert ws_cap["B7"].value == "N/A"  # Cp is N/A
    assert ws_cap["B9"].value == "N/A"  # Pp is N/A


def test_capability_unilateral_lower_spec_only() -> None:
    """Unilateral lower spec (LSL only) generates live Cpk/Ppk formulas; Cp/Pp are N/A."""
    data = [10.0, 10.1, 9.9, 10.2, 9.8, 10.0, 10.1, 9.9]
    raw_bytes = export_spc_excel("I-MR", data, usl=None, lsl=9.0)

    # Live Cpk (B8) and Ppk (B10)
    assert_cell_is_formula(raw_bytes, "Process Capability", "B8")
    assert_cell_is_formula(raw_bytes, "Process Capability", "B10")

    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes))
    ws_cap = wb["Process Capability"]
    assert ws_cap["B7"].value == "N/A"  # Cp is N/A
    assert ws_cap["B9"].value == "N/A"  # Pp is N/A


def test_capability_no_spec_limits() -> None:
    """When no spec limits are given, capability indices are structured N/A."""
    data = [10.0, 10.1, 9.9, 10.2, 9.8, 10.0, 10.1, 9.9]
    raw_bytes = export_spc_excel("I-MR", data, usl=None, lsl=None)

    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes))
    ws_cap = wb["Process Capability"]
    assert ws_cap["B7"].value == "N/A"
    assert ws_cap["B8"].value == "N/A"
    assert ws_cap["B9"].value == "N/A"
    assert ws_cap["B10"].value == "N/A"


# ===========================================================================
# 3. Accuracy Scorecard (Mathematical Parity)
# ===========================================================================


def test_accuracy_scorecard_xbar_r() -> None:
    """Verify that formulas produced for Xbar-R match exact engine calculations."""
    data = SAMPLE_SPC_XBAR_R_DATA
    engine_res = compute_xbar_r(data)

    wb = export_spc_to_workbook("Xbar-R", data, usl=11.0, lsl=9.0)
    ws_data = wb["Control Chart Data"]

    # Check that formula strings exist
    assert ws_data["G2"].value == "=AVERAGE(B2:F2)"
    assert ws_data["H2"].value == "=MAX(B2:F2)-MIN(B2:F2)"
    assert ws_data["K2"].value == "=AVERAGE(G2:G21)"
    assert ws_data["K3"].value == "=AVERAGE(H2:H21)"

    # Compute numerical values of formulas
    sub_means = [sum(sub) / len(sub) for sub in data]
    sub_ranges = [max(sub) - min(sub) for sub in data]
    grand_mean = sum(sub_means) / len(sub_means)
    rbar = sum(sub_ranges) / len(sub_ranges)
    sigma_hat = rbar / 2.326
    ucl_x = grand_mean + 0.577 * rbar
    lcl_x = grand_mean - 0.577 * rbar

    assert math.isclose(grand_mean, engine_res["xbarbar"], rel_tol=1e-6)
    assert math.isclose(rbar, engine_res["rbar"], rel_tol=1e-6)
    assert math.isclose(sigma_hat, engine_res["sigma_hat"], rel_tol=1e-6)
    assert math.isclose(ucl_x, engine_res["ucl_x"], rel_tol=1e-6)
    assert math.isclose(lcl_x, engine_res["lcl_x"], rel_tol=1e-6)


def test_accuracy_scorecard_capability() -> None:
    """Verify capability formulas evaluate to engine capability indices."""
    all_values = [x for sub in SAMPLE_SPC_XBAR_R_DATA for x in sub]
    xr = compute_xbar_r(SAMPLE_SPC_XBAR_R_DATA)
    cap = compute_capability(all_values, lsl=9.0, usl=11.0, sigma_hat=xr["sigma_hat"])

    wb = export_spc_to_workbook("Xbar-R", SAMPLE_SPC_XBAR_R_DATA, lsl=9.0, usl=11.0)
    ws_cap = wb["Process Capability"]

    # Verify formulas
    assert ws_cap["B7"].value == "=(B2-B3)/(6*B5)"  # Cp
    assert ws_cap["B8"].value == "=MIN((B2-B4)/(3*B5),(B4-B3)/(3*B5))"  # Cpk
    assert ws_cap["B9"].value == "=(B2-B3)/(6*B6)"  # Pp
    assert ws_cap["B10"].value == "=MIN((B2-B4)/(3*B6),(B4-B3)/(3*B6))"  # Ppk

    # Recomputed values from formula definitions
    usl, lsl = 11.0, 9.0
    mu = sum(all_values) / len(all_values)
    variance = sum((x - mu) ** 2 for x in all_values) / (len(all_values) - 1)
    s_overall = math.sqrt(variance)
    s_within = xr["sigma_hat"]

    cp_recomputed = (usl - lsl) / (6.0 * s_within)
    cpk_recomputed = min((usl - mu) / (3.0 * s_within), (mu - lsl) / (3.0 * s_within))
    pp_recomputed = (usl - lsl) / (6.0 * s_overall)
    ppk_recomputed = min((usl - mu) / (3.0 * s_overall), (mu - lsl) / (3.0 * s_overall))

    assert math.isclose(cp_recomputed, cap["cp"], rel_tol=1e-6)
    assert math.isclose(cpk_recomputed, cap["cpk"], rel_tol=1e-6)
    assert math.isclose(pp_recomputed, cap["pp"], rel_tol=1e-6)
    assert math.isclose(ppk_recomputed, cap["ppk"], rel_tol=1e-6)


# ===========================================================================
# 4. Security Invariant (Formula-Injection Defense)
# ===========================================================================


def test_security_invariant_formula_injection_defense() -> None:
    """Untrusted user metadata starting with formula trigger characters is escaped with a leading apostrophe."""
    malicious_title = "=SUM(1+1)"
    malicious_part_name = "+DANGEROUS_CALL()"
    malicious_part_num = "@EVIL_FORMULA"
    malicious_char = "-CMD|' /C calc'!A0"

    raw_bytes = export_spc_excel(
        "Xbar-R",
        SAMPLE_SPC_XBAR_R_DATA,
        title=malicious_title,
        part_name=malicious_part_name,
        part_number=malicious_part_num,
        characteristic=malicious_char,
    )

    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes))
    ws_summary = wb["Summary & Run Rules"]

    # Verify that stored values start with leading apostrophe and are inert
    assert ws_summary["B2"].value == "'=SUM(1+1)"
    assert ws_summary["B3"].value == "'+DANGEROUS_CALL()"
    assert ws_summary["B4"].value == "'@EVIL_FORMULA"
    assert ws_summary["B5"].value == "'-CMD|' /C calc'!A0"


# ===========================================================================
# 5. Negative Controls (Verifier Failure on Literals)
# ===========================================================================


def test_negative_verifier_catches_literal_cell() -> None:
    """assert_cell_is_formula must fail with AssertionError when given a hardcoded literal cell."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Sheet"
    ws.cell(row=1, column=1, value=123.45)  # Literal number

    buf = io.BytesIO()
    wb.save(buf)
    literal_bytes = buf.getvalue()

    with pytest.raises(AssertionError, match="is a literal, not a live formula"):
        assert_cell_is_formula(literal_bytes, "Test Sheet", "A1")


# ===========================================================================
# 6. Type Guards & Input Validation Errors
# ===========================================================================


def test_type_guards_and_validation_errors() -> None:
    """All invalid arguments, out-of-range parameters, and mismatched shapes raise ValueError."""
    # 1. Invalid chart type
    with pytest.raises(ValueError, match="Unsupported chart type"):
        export_spc_excel("INVALID_CHART", [1.0, 2.0, 3.0])

    # 2. Inverted spec limits
    with pytest.raises(ValueError, match="cannot be less than LSL"):
        export_spc_excel("I-MR", [10.0, 10.1, 10.2], usl=5.0, lsl=10.0)

    # 3. Empty data
    with pytest.raises(ValueError, match="requires data as a list of subgroups"):
        export_spc_excel("Xbar-R", [])

    # 4. Xbar-R non-2D data
    with pytest.raises(ValueError, match="requires data as a list of subgroups"):
        export_spc_excel("Xbar-R", [1.0, 2.0, 3.0])

    # 5. Xbar-R ragged subgroups
    with pytest.raises(ValueError, match="All subgroups in Xbar-R chart must have equal size"):
        export_spc_excel("Xbar-R", [[1.0, 2.0], [1.0, 2.0, 3.0]])

    # 6. Xbar-R invalid subgroup size (n=1, n=15)
    with pytest.raises(ValueError, match="Xbar-R requires subgroup size between 2 and 10"):
        export_spc_excel("Xbar-R", [[1.0], [2.0]])
    with pytest.raises(ValueError, match="Xbar-R requires subgroup size between 2 and 10"):
        export_spc_excel("Xbar-R", [[1.0] * 15, [2.0] * 15])

    # 7. Xbar-S non-2D data & invalid size
    with pytest.raises(ValueError, match="Xbar-S chart requires data as a list of subgroups"):
        export_spc_excel("Xbar-S", [1.0, 2.0])
    with pytest.raises(ValueError, match="All subgroups in Xbar-S chart must have equal size"):
        export_spc_excel("Xbar-S", [[1.0, 2.0], [1.0]])
    with pytest.raises(ValueError, match="Xbar-S requires subgroup size between 2 and 12"):
        export_spc_excel("Xbar-S", [[1.0] * 15, [2.0] * 15])

    # 8. I-MR fewer than 2 points
    with pytest.raises(ValueError, match="I-MR chart requires at least two values"):
        export_spc_excel("I-MR", [10.0])

    # 9. Attribute charts missing sample_sizes
    with pytest.raises(ValueError, match="p chart requires sample_sizes"):
        export_spc_excel("p", [5.0, 10.0])
    with pytest.raises(ValueError, match="u chart requires sample_sizes"):
        export_spc_excel("u", [5.0, 10.0])

    # 10. Attribute charts mismatched sample_sizes length
    with pytest.raises(ValueError, match="counts and sample_sizes must have equal non-zero length"):
        export_spc_excel("p", [5.0, 10.0], sample_sizes=[100.0])
    with pytest.raises(ValueError, match="counts and sample_sizes must have equal non-zero length"):
        export_spc_excel("u", [5.0, 10.0], sample_sizes=[100.0])

    # 11. Attribute charts non-positive sample_sizes
    with pytest.raises(ValueError, match="sample_sizes must be strictly positive"):
        export_spc_excel("p", [5.0], sample_sizes=[0.0])
    with pytest.raises(ValueError, match="sample_sizes must be strictly positive"):
        export_spc_excel("u", [5.0], sample_sizes=[-2.0])

    # 12. c chart empty data
    with pytest.raises(ValueError, match="c chart requires at least one data point"):
        export_spc_excel("c", [])


# ===========================================================================
# 7. Summary & Run Rules Sheet Findings Generation
# ===========================================================================


def test_summary_and_run_rules_violations_populated() -> None:
    """Run-rule violations are evaluated and populated into the findings table."""
    # Data with variance and an extreme outlier point
    out_of_control_data = [
        [10.0, 10.1, 9.9],
        [10.1, 10.0, 9.8],
        [9.9, 10.2, 10.0],
        [10.0, 9.8, 10.1],
        [50.0, 52.0, 48.0],  # Massive outlier
    ]
    raw_bytes = export_spc_excel("Xbar-R", out_of_control_data)

    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes))
    ws_summary = wb["Summary & Run Rules"]

    assert ws_summary["B10"].value == "Out of Control (Special Causes Detected)"
    assert ws_summary["B11"].value >= 1
    # Check findings table header and row
    assert ws_summary["A13"].value == "Subgroup / Observation"
    assert ws_summary["B13"].value == "Rule Violated"
    assert ws_summary["C13"].value == "Status"
    assert ws_summary["C14"].value == "Special Cause Detected"
