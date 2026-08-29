"""
tests/test_rca_export.py
Comprehensive test suite for quality_core.rca.export (#146).

Covers:
1. Combined multi-sheet export (build_rca_workbook, export_rca_workbook) in canonical order.
2. Single-tool and partial combination exports (subsets, convenience exporters, title handling).
3. Data integrity & column mappings (5-Why, 6M Fishbone, Kepner-Tregoe Is-Is Not).
4. Formula injection safety & escaping with OWASP trigger characters.
5. Input polymorphism (Pydantic models, DataFrames, list of dicts, dicts) and error boundaries.
6. Benchmark data constructors and object isolation.
7. Re-exports and layout constant consistency.
"""

from __future__ import annotations

import io

import openpyxl
import pandas as pd
import pydantic
import pytest
import quality_core.rca.export as export_mod
from _xlsx_formula_audit import assert_cell_is_formula
from quality_core.rca import (
    FISHBONE_COLUMN_WIDTHS,
    FISHBONE_EXPORT_COLUMNS,
    FISHBONE_SHEET_TITLE,
    FIVE_WHY_COLUMN_WIDTHS,
    FIVE_WHY_EXPORT_COLUMNS,
    FIVE_WHY_SHEET_TITLE,
    IS_IS_NOT_COLUMN_WIDTHS,
    IS_IS_NOT_EXPORT_COLUMNS,
    IS_IS_NOT_SHEET_TITLE,
    AntiPatternFinding,
    FishboneCause,
    FishboneDataset,
    FiveWhyChain,
    FiveWhyStep,
    FiveWhyValidationResult,
    IsIsNotMatrix,
    IsIsNotRow,
    benchmark_fishbone_dataset,
    benchmark_five_why_chain,
    benchmark_is_is_not_matrix,
    benchmark_rca_datasets,
    build_rca_workbook,
    export_fishbone_workbook,
    export_five_why_workbook,
    export_is_is_not_workbook,
    export_rca_workbook,
    validate_five_why_chain,
)

# ==============================================================================
# Helper functions
# ==============================================================================


def _load_workbook_from_bytes(wb_bytes: bytes) -> openpyxl.Workbook:
    """Reload saved bytes into an openpyxl Workbook without data_only caching."""
    return openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)


# ==============================================================================
# 1. Combined Multi-Sheet Export Tests
# ==============================================================================


def test_export_rca_workbook_benchmark_creates_all_three_sheets_in_order() -> None:
    """Benchmark datasets create all 3 sheets in canonical order with matching headers."""
    five_why, fishbone, is_is_not = benchmark_rca_datasets()
    wb_bytes = export_rca_workbook(five_why=five_why, fishbone=fishbone, is_is_not=is_is_not)
    assert isinstance(wb_bytes, bytes)
    assert len(wb_bytes) > 0

    wb = _load_workbook_from_bytes(wb_bytes)
    expected_sheets = [FIVE_WHY_SHEET_TITLE, FISHBONE_SHEET_TITLE, IS_IS_NOT_SHEET_TITLE]
    assert wb.sheetnames == expected_sheets

    # Sheet 1: 5-Why
    ws_5w = wb[FIVE_WHY_SHEET_TITLE]
    headers_5w = [c.value for c in ws_5w[1]]
    assert headers_5w == list(FIVE_WHY_EXPORT_COLUMNS)
    assert ws_5w.max_row == len(five_why.steps) + 1

    # Sheet 2: Fishbone
    ws_fb = wb[FISHBONE_SHEET_TITLE]
    headers_fb = [c.value for c in ws_fb[1]]
    assert headers_fb == list(FISHBONE_EXPORT_COLUMNS)
    assert ws_fb.max_row == len(fishbone.causes) + 1

    # Sheet 3: Is-Is Not
    ws_in = wb[IS_IS_NOT_SHEET_TITLE]
    headers_in = [c.value for c in ws_in[1]]
    assert headers_in == list(IS_IS_NOT_EXPORT_COLUMNS)
    assert ws_in.max_row == len(is_is_not.rows) + 1


def test_build_rca_workbook_returns_openpyxl_workbook_object() -> None:
    """build_rca_workbook returns openpyxl.Workbook with correct active sheet."""
    five_why, fishbone, is_is_not = benchmark_rca_datasets()
    wb = build_rca_workbook(five_why=five_why, fishbone=fishbone, is_is_not=is_is_not)
    assert isinstance(wb, openpyxl.Workbook)
    assert wb.active.title == FIVE_WHY_SHEET_TITLE
    assert wb.sheetnames == [FIVE_WHY_SHEET_TITLE, FISHBONE_SHEET_TITLE, IS_IS_NOT_SHEET_TITLE]


# ==============================================================================
# 2. Single-Tool and Partial Combination Export Tests
# ==============================================================================


def test_export_rca_workbook_with_only_five_why() -> None:
    five_why = benchmark_five_why_chain()
    wb_bytes = export_rca_workbook(five_why=five_why)
    wb = _load_workbook_from_bytes(wb_bytes)
    assert wb.sheetnames == [FIVE_WHY_SHEET_TITLE]
    assert [c.value for c in wb[FIVE_WHY_SHEET_TITLE][1]] == list(FIVE_WHY_EXPORT_COLUMNS)


def test_export_rca_workbook_with_only_fishbone() -> None:
    fishbone = benchmark_fishbone_dataset()
    wb_bytes = export_rca_workbook(fishbone=fishbone)
    wb = _load_workbook_from_bytes(wb_bytes)
    assert wb.sheetnames == [FISHBONE_SHEET_TITLE]
    assert [c.value for c in wb[FISHBONE_SHEET_TITLE][1]] == list(FISHBONE_EXPORT_COLUMNS)


def test_export_rca_workbook_with_only_is_is_not() -> None:
    is_is_not = benchmark_is_is_not_matrix()
    wb_bytes = export_rca_workbook(is_is_not=is_is_not)
    wb = _load_workbook_from_bytes(wb_bytes)
    assert wb.sheetnames == [IS_IS_NOT_SHEET_TITLE]
    assert [c.value for c in wb[IS_IS_NOT_SHEET_TITLE][1]] == list(IS_IS_NOT_EXPORT_COLUMNS)


def test_export_rca_workbook_two_tool_combinations() -> None:
    five_why, fishbone, is_is_not = benchmark_rca_datasets()

    # 1. 5-Why + Fishbone
    wb1 = _load_workbook_from_bytes(export_rca_workbook(five_why=five_why, fishbone=fishbone))
    assert wb1.sheetnames == [FIVE_WHY_SHEET_TITLE, FISHBONE_SHEET_TITLE]

    # 2. 5-Why + Is-Is Not
    wb2 = _load_workbook_from_bytes(export_rca_workbook(five_why=five_why, is_is_not=is_is_not))
    assert wb2.sheetnames == [FIVE_WHY_SHEET_TITLE, IS_IS_NOT_SHEET_TITLE]

    # 3. Fishbone + Is-Is Not
    wb3 = _load_workbook_from_bytes(export_rca_workbook(fishbone=fishbone, is_is_not=is_is_not))
    assert wb3.sheetnames == [FISHBONE_SHEET_TITLE, IS_IS_NOT_SHEET_TITLE]


def test_export_single_tool_convenience_helpers() -> None:
    five_why, fishbone, is_is_not = benchmark_rca_datasets()

    # 5-Why helper
    wb_5w_bytes = export_five_why_workbook(five_why)
    wb_5w = _load_workbook_from_bytes(wb_5w_bytes)
    assert wb_5w.sheetnames == [FIVE_WHY_SHEET_TITLE]

    # Fishbone helper
    wb_fb_bytes = export_fishbone_workbook(fishbone)
    wb_fb = _load_workbook_from_bytes(wb_fb_bytes)
    assert wb_fb.sheetnames == [FISHBONE_SHEET_TITLE]

    # Is-Is Not helper
    wb_in_bytes = export_is_is_not_workbook(is_is_not)
    wb_in = _load_workbook_from_bytes(wb_in_bytes)
    assert wb_in.sheetnames == [IS_IS_NOT_SHEET_TITLE]


def test_single_tool_custom_title_overrides_sheet_name() -> None:
    five_why, fishbone, is_is_not = benchmark_rca_datasets()

    # Custom title on export_five_why_workbook
    wb1 = _load_workbook_from_bytes(export_five_why_workbook(five_why, title="Bearing 5-Why Study"))
    assert wb1.sheetnames == ["Bearing 5-Why Study"]

    # Custom title on export_fishbone_workbook
    wb2 = _load_workbook_from_bytes(export_fishbone_workbook(fishbone, title="Cylinder Ishikawa"))
    assert wb2.sheetnames == ["Cylinder Ishikawa"]

    # Custom title on export_is_is_not_workbook
    wb3 = _load_workbook_from_bytes(export_is_is_not_workbook(is_is_not, title="KT Scoping 2026"))
    assert wb3.sheetnames == ["KT Scoping 2026"]

    # Custom title via build_rca_workbook for single tool
    wb4 = build_rca_workbook(five_why=five_why, title="Custom Single 5W")
    assert wb4.sheetnames == ["Custom Single 5W"]

    wb5 = build_rca_workbook(fishbone=fishbone, title="Custom Single FB")
    assert wb5.sheetnames == ["Custom Single FB"]

    wb6 = build_rca_workbook(is_is_not=is_is_not, title="Custom Single IN")
    assert wb6.sheetnames == ["Custom Single IN"]


def test_multi_tool_custom_title_preserves_canonical_sheet_names() -> None:
    """When multiple tools are provided, sheet names stay canonical even if title is specified."""
    five_why, fishbone, is_is_not = benchmark_rca_datasets()
    wb = build_rca_workbook(five_why=five_why, fishbone=fishbone, is_is_not=is_is_not, title="Plant Investigation")
    assert wb.sheetnames == [FIVE_WHY_SHEET_TITLE, FISHBONE_SHEET_TITLE, IS_IS_NOT_SHEET_TITLE]


# ==============================================================================
# 3. Data Integrity & Detailed Column Mapping Tests
# ==============================================================================


def test_five_why_sheet_data_integrity() -> None:
    five_why = benchmark_five_why_chain()
    wb = _load_workbook_from_bytes(export_five_why_workbook(five_why))
    ws = wb[FIVE_WHY_SHEET_TITLE]

    # Verify rows
    for idx, step in enumerate(five_why.steps, start=2):
        assert ws[f"A{idx}"].value == step.step_number
        assert ws[f"B{idx}"].value == step.why
        assert ws[f"C{idx}"].value == step.because
        # Reverse therefore is formatted
        assert ws[f"D{idx}"].value.startswith("Because ")
        assert "therefore" in ws[f"D{idx}"].value
        assert ws[f"E{idx}"].value in ("Yes", "No")
        assert ws[f"F{idx}"].value in ("SYSTEMIC", "TECHNICAL_PROCESS", "HUMAN_INDIVIDUAL", "UNKNOWN")


def test_five_why_sheet_anti_patterns_and_non_reversible_rendering() -> None:
    """Verify anti-pattern finding codes and non-reversible flags are correctly rendered."""
    # Construct a chain with human blame anti-pattern (HUMAN_BLAME)
    chain = FiveWhyChain(
        problem_statement="Machine crashed",
        steps=[
            FiveWhyStep(step_number=1, why="Why did the machine crash?", because="The operator was careless"),
            FiveWhyStep(step_number=2, why="Why was the operator careless?", because="Lack of attention"),
        ],
        root_cause="Lack of attention",
    )
    wb = _load_workbook_from_bytes(export_five_why_workbook(chain))
    ws = wb[FIVE_WHY_SHEET_TITLE]

    # Check anti-pattern column (G) at terminal step (row 3)
    ap_val_2 = ws["G3"].value
    assert "BLAME_TERMINAL_OPERATOR_ERROR" in (ap_val_2 or "") or "PREMATURE_TERMINATION" in (ap_val_2 or "")
    # Check systemic classification (F)
    assert ws["F3"].value == "HUMAN_INDIVIDUAL"


def test_five_why_export_with_global_anti_pattern_without_step_number(monkeypatch: pytest.MonkeyPatch) -> None:
    """Anti-pattern finding with step_number=None is safely ignored in row mapping."""
    chain = benchmark_five_why_chain()
    real_val = validate_five_why_chain(chain)
    fake_ap = AntiPatternFinding(
        code="GLOBAL_FINDING",
        severity="warning",
        step_number=None,
        message="Chain-level anti-pattern",
        recommendation="Review entire chain",
    )
    patched_result = FiveWhyValidationResult(
        basis=real_val.basis,
        valid=real_val.valid,
        verdict=real_val.verdict,
        reversibility_score=real_val.reversibility_score,
        problem_statement=real_val.problem_statement,
        root_cause=real_val.root_cause,
        total_steps=real_val.total_steps,
        link_evaluations=real_val.link_evaluations,
        anti_patterns=[fake_ap],
        systemic_assessment=real_val.systemic_assessment,
        recommendations=real_val.recommendations,
    )
    monkeypatch.setattr(export_mod, "validate_five_why_chain", lambda c: patched_result)
    wb = _load_workbook_from_bytes(export_five_why_workbook(chain))
    ws = wb[FIVE_WHY_SHEET_TITLE]
    assert ws.max_row == len(chain.steps) + 1
    # Global anti-pattern with step_number=None is not attached to individual steps
    for r in range(2, ws.max_row + 1):
        assert ws[f"G{r}"].value in (None, "")


def test_fishbone_sheet_data_integrity() -> None:
    fishbone = benchmark_fishbone_dataset()
    wb = _load_workbook_from_bytes(export_fishbone_workbook(fishbone))
    ws = wb[FISHBONE_SHEET_TITLE]

    assert ws.max_row == len(fishbone.causes) + 1
    for idx, cause in enumerate(fishbone.causes, start=2):
        assert ws[f"A{idx}"].value == cause.category
        assert ws[f"B{idx}"].value == cause.cause
        assert ws[f"C{idx}"].value == (cause.sub_category or "")
        assert ws[f"D{idx}"].value in ("Active", "Bare Leg / Few Causes")
        assert ws[f"E{idx}"].value in ("Yes", "No")


def test_fishbone_duplicate_cause_and_bare_branch_handling() -> None:
    """Duplicate cause gets 'Yes' on duplicate occurrence, bare branch gets 'Bare Leg / Few Causes'."""
    causes = [
        FishboneCause(category="Man", cause="Operator fatigue", sub_category="Shift"),
        FishboneCause(category="Man", cause="operator fatigue  ", sub_category="Shift"),  # duplicate
        FishboneCause(category="Machine", cause="Spindle runout", sub_category=None),  # only 1 Machine cause -> bare branch
    ]
    ds = FishboneDataset(effect="Test Defect", causes=causes)
    wb = _load_workbook_from_bytes(export_fishbone_workbook(ds))
    ws = wb[FISHBONE_SHEET_TITLE]

    # Row 2: first "Operator fatigue" -> not duplicate, active branch (2 causes under Man)
    assert ws["E2"].value == "No"
    assert ws["D2"].value == "Active"

    # Row 3: second "operator fatigue  " -> duplicate
    assert ws["E3"].value == "Yes"
    assert ws["D3"].value == "Active"

    # Row 4: Machine (only 1 cause) -> bare branch status, sub_category blank
    assert ws["A4"].value == "Machine"
    assert ws["C4"].value in (None, "")
    assert ws["D4"].value == "Bare Leg / Few Causes"
    assert ws["E4"].value == "No"


def test_is_is_not_sheet_data_integrity() -> None:
    is_is_not = benchmark_is_is_not_matrix()
    wb = _load_workbook_from_bytes(export_is_is_not_workbook(is_is_not))
    ws = wb[IS_IS_NOT_SHEET_TITLE]

    assert ws.max_row == len(is_is_not.rows) + 1
    # Check canonical dimension order
    dims = [ws[f"A{r}"].value for r in range(2, ws.max_row + 1)]
    assert dims == ["WHAT", "WHERE", "WHEN", "EXTENT"]

    for idx in range(2, ws.max_row + 1):
        assert ws[f"B{idx}"].value is not None  # IS
        assert ws[f"C{idx}"].value is not None  # IS_NOT
        assert ws[f"D{idx}"].value is not None  # Distinction
        assert ws[f"E{idx}"].value is not None  # Change
        assert ws[f"F{idx}"].value is not None  # Candidate_Hypothesis
        assert ws[f"G{idx}"].value in ("Yes", "No")  # Hypothesis_Paired


def test_is_is_not_out_of_order_rows_sorted_to_canonical_kt_order() -> None:
    """Rows supplied in non-canonical order (e.g. EXTENT, WHEN, WHERE, WHAT) sort correctly."""
    rows = [
        IsIsNotRow(dimension="EXTENT", is_data="10 units", is_not_data="100 units"),
        IsIsNotRow(dimension="WHEN", is_data="Night shift", is_not_data="Day shift"),
        IsIsNotRow(dimension="WHERE", is_data="Line 2", is_not_data="Line 1"),
        IsIsNotRow(dimension="WHAT", is_data="Crack", is_not_data="Dent"),
    ]
    matrix = IsIsNotMatrix(rows=rows)
    wb = _load_workbook_from_bytes(export_is_is_not_workbook(matrix))
    ws = wb[IS_IS_NOT_SHEET_TITLE]

    dims = [ws[f"A{r}"].value for r in range(2, 6)]
    assert dims == ["WHAT", "WHERE", "WHEN", "EXTENT"]


def test_is_is_not_unpaired_and_blank_distinctions_changes() -> None:
    """Matrix rows with None distinctions/changes render as empty strings and unpaired hypothesis."""
    rows = [
        IsIsNotRow(dimension="WHAT", is_data="Crack", is_not_data="Dent", distinctions=None, changes=None),
        IsIsNotRow(dimension="WHERE", is_data="Line 2", is_not_data="Line 1", distinctions="Fixture A", changes=None),
    ]
    matrix = IsIsNotMatrix(rows=rows)
    wb = _load_workbook_from_bytes(export_is_is_not_workbook(matrix))
    ws = wb[IS_IS_NOT_SHEET_TITLE]

    # Row 2 (WHAT): distinctions and changes are empty, hypothesis is empty, paired is No
    assert ws["A2"].value == "WHAT"
    assert ws["D2"].value in (None, "")
    assert ws["E2"].value in (None, "")
    assert ws["G2"].value == "No"

    # Row 3 (WHERE): changes is empty, paired is No
    assert ws["A3"].value == "WHERE"
    assert ws["D3"].value == "Fixture A"
    assert ws["E3"].value in (None, "")
    assert ws["G3"].value == "No"


# ==============================================================================
# 4. Formula Injection Safety & Escaping (Negative Controls)
# ==============================================================================


@pytest.mark.parametrize(
    "payload",
    [
        "=cmd|' /C calc'!A0",
        "+SUM(A1:A10)",
        "-danger_calc()",
        "@HYPERLINK('http://evil.com')",
    ],
)
def test_rca_export_formula_injection_escaped_in_all_tools(payload: str) -> None:
    """NEGATIVE CONTROL: Untrusted strings with OWASP triggers must be escaped and stored as literals.

    Every cell containing injection payloads is apostrophe-escaped with "'" prefix,
    and assert_cell_is_formula fails (confirming literal, not live formula).
    """
    # 5-Why with injection in why and because
    chain = FiveWhyChain(
        problem_statement=f"{payload} problem",
        steps=[
            FiveWhyStep(step_number=1, why=f"{payload} why", because=f"{payload} because"),
        ],
        root_cause=f"{payload} root_cause",
    )

    # Fishbone with injection in cause and subcategory
    dataset = FishboneDataset(
        effect=f"{payload} effect",
        causes=[
            FishboneCause(category="Man", cause=f"{payload} cause", sub_category=f"{payload} sub"),
        ],
    )

    # Is-Is Not with injection in problem statement, IS, IS NOT, distinction, change
    matrix = IsIsNotMatrix(
        problem_statement=f"{payload} problem",
        rows=[
            IsIsNotRow(
                dimension="WHAT",
                is_data=f"{payload} is",
                is_not_data=f"{payload} is_not",
                distinctions=f"{payload} distinction",
                changes=f"{payload} change",
            )
        ],
    )

    wb_bytes = export_rca_workbook(five_why=chain, fishbone=dataset, is_is_not=matrix)
    wb = _load_workbook_from_bytes(wb_bytes)

    # Verify 5-Why cells are escaped
    ws_5w = wb[FIVE_WHY_SHEET_TITLE]
    val_why = ws_5w["B2"].value
    val_bec = ws_5w["C2"].value
    assert val_why == "'" + f"{payload} why"
    assert val_bec == "'" + f"{payload} because"

    with pytest.raises(AssertionError, match="literal, not a live formula"):
        assert_cell_is_formula(wb_bytes, FIVE_WHY_SHEET_TITLE, "B2")

    # Verify Fishbone cells are escaped
    ws_fb = wb[FISHBONE_SHEET_TITLE]
    val_cause = ws_fb["B2"].value
    val_sub = ws_fb["C2"].value
    assert val_cause == "'" + f"{payload} cause"
    assert val_sub == "'" + f"{payload} sub"

    with pytest.raises(AssertionError, match="literal, not a live formula"):
        assert_cell_is_formula(wb_bytes, FISHBONE_SHEET_TITLE, "B2")

    # Verify Is-Is Not cells are escaped
    ws_in = wb[IS_IS_NOT_SHEET_TITLE]
    assert ws_in["B2"].value == "'" + f"{payload} is"
    assert ws_in["C2"].value == "'" + f"{payload} is_not"
    assert ws_in["D2"].value == "'" + f"{payload} distinction"
    assert ws_in["E2"].value == "'" + f"{payload} change"

    with pytest.raises(AssertionError, match="literal, not a live formula"):
        assert_cell_is_formula(wb_bytes, IS_IS_NOT_SHEET_TITLE, "B2")


# ==============================================================================
# 5. Input Polymorphism & Validation Boundaries
# ==============================================================================


def test_build_rca_workbook_all_none_raises_value_error() -> None:
    with pytest.raises(ValueError, match="At least one RCA dataset .* must be provided"):
        build_rca_workbook(five_why=None, fishbone=None, is_is_not=None)


def test_export_rca_workbook_all_none_raises_value_error() -> None:
    with pytest.raises(ValueError, match="At least one RCA dataset .* must be provided"):
        export_rca_workbook(five_why=None, fishbone=None, is_is_not=None)


def test_export_rca_workbook_from_dataframes() -> None:
    df_5w = pd.DataFrame([
        {"step_number": 1, "why": "Why 1", "because": "Because 1"},
        {"step_number": 2, "why": "Why 2", "because": "Because 2"},
    ])
    df_fb = pd.DataFrame([
        {"category": "Man", "cause": "Fatigue", "sub_category": "Shift"},
        {"category": "Machine", "cause": "Wear", "sub_category": None},
    ])
    df_in = pd.DataFrame([
        {"dimension": "WHAT", "is_data": "Crack", "is_not_data": "Dent", "distinctions": None, "changes": None},
        {"dimension": "WHERE", "is_data": "Zone A", "is_not_data": "Zone B", "distinctions": None, "changes": None},
    ])

    wb_bytes = export_rca_workbook(five_why=df_5w, fishbone=df_fb, is_is_not=df_in)
    wb = _load_workbook_from_bytes(wb_bytes)
    assert wb.sheetnames == [FIVE_WHY_SHEET_TITLE, FISHBONE_SHEET_TITLE, IS_IS_NOT_SHEET_TITLE]
    assert wb[FIVE_WHY_SHEET_TITLE].max_row == 3
    assert wb[FISHBONE_SHEET_TITLE].max_row == 3
    assert wb[IS_IS_NOT_SHEET_TITLE].max_row == 3


def test_export_rca_workbook_from_list_of_dicts() -> None:
    records_5w = [
        {"step_number": 1, "why": "Why 1", "because": "Because 1"},
    ]
    records_fb = [
        {"category": "Method", "cause": "Missing step", "sub_category": "SOP"},
    ]
    records_in = [
        {"dimension": "WHEN", "is_data": "Morning", "is_not_data": "Night", "distinctions": "Temp", "changes": "Sun"},
    ]

    wb_bytes = export_rca_workbook(five_why=records_5w, fishbone=records_fb, is_is_not=records_in)
    wb = _load_workbook_from_bytes(wb_bytes)
    assert wb.sheetnames == [FIVE_WHY_SHEET_TITLE, FISHBONE_SHEET_TITLE, IS_IS_NOT_SHEET_TITLE]


def test_export_rca_workbook_from_dicts() -> None:
    dict_5w = {
        "problem_statement": "Issue",
        "steps": [{"step_number": 1, "why": "Why 1", "because": "Because 1"}],
    }
    dict_fb = {
        "effect": "Defect",
        "causes": [{"category": "Material", "cause": "Impurity"}],
    }
    dict_in = {
        "problem_statement": "Matrix",
        "rows": [{"dimension": "EXTENT", "is_data": "5%", "is_not_data": "50%"}],
    }

    wb_bytes = export_rca_workbook(five_why=dict_5w, fishbone=dict_fb, is_is_not=dict_in)
    wb = _load_workbook_from_bytes(wb_bytes)
    assert wb.sheetnames == [FIVE_WHY_SHEET_TITLE, FISHBONE_SHEET_TITLE, IS_IS_NOT_SHEET_TITLE]


def test_export_rca_workbook_rejects_unsupported_types() -> None:
    with pytest.raises(TypeError):
        build_rca_workbook(five_why=12345)  # type: ignore[arg-type]

    with pytest.raises(TypeError):
        build_rca_workbook(fishbone=True)  # type: ignore[arg-type]

    with pytest.raises(TypeError):
        build_rca_workbook(is_is_not=3.14159)  # type: ignore[arg-type]


def test_export_rca_workbook_rejects_invalid_schema_content() -> None:
    # 5-Why missing required field
    with pytest.raises(pydantic.ValidationError):
        build_rca_workbook(five_why=[{"step_number": 1, "why": "Why only"}])

    # Fishbone with invalid 6M category
    with pytest.raises(pydantic.ValidationError):
        build_rca_workbook(fishbone=[{"category": "InvalidCat", "cause": "Valid"}])

    # Is-Is Not with invalid dimension
    with pytest.raises(pydantic.ValidationError):
        build_rca_workbook(is_is_not=[{"dimension": "INVALID_DIM", "is_data": "A", "is_not_data": "B"}])


# ==============================================================================
# 6. Benchmark Data Constructors Tests
# ==============================================================================


def test_benchmark_five_why_chain_structure() -> None:
    chain = benchmark_five_why_chain()
    assert isinstance(chain, FiveWhyChain)
    assert len(chain.steps) == 5
    assert [s.step_number for s in chain.steps] == [1, 2, 3, 4, 5]
    assert chain.problem_statement == "Hole positions outside of tolerance on CNC drilling station"
    assert chain.root_cause == "The induction plan was not signed by Engineering"


def test_benchmark_fishbone_dataset_structure() -> None:
    ds = benchmark_fishbone_dataset()
    assert isinstance(ds, FishboneDataset)
    assert len(ds.causes) == 12
    categories_present = {c.category for c in ds.causes}
    assert categories_present == {"Man", "Machine", "Method", "Material", "Measurement", "Environment"}


def test_benchmark_is_is_not_matrix_structure() -> None:
    matrix = benchmark_is_is_not_matrix()
    assert isinstance(matrix, IsIsNotMatrix)
    assert len(matrix.rows) == 4
    dims_present = [r.dimension for r in matrix.rows]
    assert dims_present == ["WHAT", "WHERE", "WHEN", "EXTENT"]
    assert matrix.is_complete is True


def test_benchmark_rca_datasets_tuple() -> None:
    five_why, fishbone, is_is_not = benchmark_rca_datasets()
    assert isinstance(five_why, FiveWhyChain)
    assert isinstance(fishbone, FishboneDataset)
    assert isinstance(is_is_not, IsIsNotMatrix)


def test_benchmark_constructors_return_fresh_instances() -> None:
    """Benchmark functions return fresh object instances on each call to prevent mutation bugs."""
    fw1 = benchmark_five_why_chain()
    fw2 = benchmark_five_why_chain()
    assert fw1 is not fw2
    assert fw1.steps is not fw2.steps

    fb1 = benchmark_fishbone_dataset()
    fb2 = benchmark_fishbone_dataset()
    assert fb1 is not fb2
    assert fb1.causes is not fb2.causes

    in1 = benchmark_is_is_not_matrix()
    in2 = benchmark_is_is_not_matrix()
    assert in1 is not in2
    assert in1.rows is not in2.rows

    ds1 = benchmark_rca_datasets()
    ds2 = benchmark_rca_datasets()
    assert ds1[0] is not ds2[0]
    assert ds1[1] is not ds2[1]
    assert ds1[2] is not ds2[2]


# ==============================================================================
# 7. Package Surface & Re-exports Tests
# ==============================================================================


def test_rca_export_symbols_reexported_from_quality_core_rca() -> None:
    import quality_core.rca as rca_pkg

    # Builders & Exporters
    assert rca_pkg.build_rca_workbook is export_mod.build_rca_workbook
    assert rca_pkg.export_rca_workbook is export_mod.export_rca_workbook
    assert rca_pkg.export_five_why_workbook is export_mod.export_five_why_workbook
    assert rca_pkg.export_fishbone_workbook is export_mod.export_fishbone_workbook
    assert rca_pkg.export_is_is_not_workbook is export_mod.export_is_is_not_workbook

    # Benchmarks
    assert rca_pkg.benchmark_five_why_chain is export_mod.benchmark_five_why_chain
    assert rca_pkg.benchmark_fishbone_dataset is export_mod.benchmark_fishbone_dataset
    assert rca_pkg.benchmark_is_is_not_matrix is export_mod.benchmark_is_is_not_matrix
    assert rca_pkg.benchmark_rca_datasets is export_mod.benchmark_rca_datasets

    # Constants
    assert rca_pkg.FIVE_WHY_EXPORT_COLUMNS is export_mod.FIVE_WHY_EXPORT_COLUMNS
    assert rca_pkg.FIVE_WHY_COLUMN_WIDTHS is export_mod.FIVE_WHY_COLUMN_WIDTHS
    assert rca_pkg.FIVE_WHY_SHEET_TITLE is export_mod.FIVE_WHY_SHEET_TITLE

    assert rca_pkg.FISHBONE_EXPORT_COLUMNS is export_mod.FISHBONE_EXPORT_COLUMNS
    assert rca_pkg.FISHBONE_COLUMN_WIDTHS is export_mod.FISHBONE_COLUMN_WIDTHS
    assert rca_pkg.FISHBONE_SHEET_TITLE is export_mod.FISHBONE_SHEET_TITLE

    assert rca_pkg.IS_IS_NOT_EXPORT_COLUMNS is export_mod.IS_IS_NOT_EXPORT_COLUMNS
    assert rca_pkg.IS_IS_NOT_COLUMN_WIDTHS is export_mod.IS_IS_NOT_COLUMN_WIDTHS
    assert rca_pkg.IS_IS_NOT_SHEET_TITLE is export_mod.IS_IS_NOT_SHEET_TITLE


def test_column_widths_cover_all_export_columns() -> None:
    assert set(FIVE_WHY_COLUMN_WIDTHS.keys()) == set(FIVE_WHY_EXPORT_COLUMNS)
    assert set(FISHBONE_COLUMN_WIDTHS.keys()) == set(FISHBONE_EXPORT_COLUMNS)
    assert set(IS_IS_NOT_COLUMN_WIDTHS.keys()) == set(IS_IS_NOT_EXPORT_COLUMNS)
