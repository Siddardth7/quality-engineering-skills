"""
tests/test_export.py
Tests for quality_core/io/export.py — the shared export primitives (W04-1).

The formula-injection regression that used to live in the FMEA app now lives here,
against the core sanitizer, so every app that reuses `export_csv` inherits the
protection (and the guard travels with the code).
"""

import re

import numpy as np
import openpyxl
import pandas as pd
import pytest
from quality_core.io.export import (
    FORMULA_PREFIXES,
    add_image_page,
    export_csv,
    fmt,
    fmt_opt,
    generated_line,
    now,
    pdf_subheader,
    pdf_summary_cells,
    pdf_title,
    render_table,
    safe_text,
    sanitize_cell,
    sanitize_for_export,
    write_keyvalue_sheet,
    write_table_sheet,
)


class _RecordingPdf:
    """Minimal duck-typed stand-in for an fpdf2 PDF, so the core PDF helpers can be
    tested without depending on fpdf2 (a consumer-app dependency). Page geometry
    (``get_y``/``h``/``b_margin``) is configurable to drive render_table's
    page-break branch."""

    def __init__(self, *, y: float = 10.0, h: float = 297.0, b_margin: float = 12.0) -> None:
        self.calls: list[tuple] = []
        self.w, self.l_margin, self.r_margin = 210.0, 10.0, 10.0
        self.h, self.b_margin, self._y = h, b_margin, y

    def get_y(self) -> float:
        self.calls.append(("get_y", (), {}))
        return self._y

    def __getattr__(self, name):
        def record(*args, **kwargs):
            self.calls.append((name, args, kwargs))

        return record


def test_formula_prefixes():
    assert FORMULA_PREFIXES == ("=", "+", "-", "@", "\t", "\r")


def test_now():
    ts = now()
    assert isinstance(ts, str)
    assert re.match(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$", ts) is not None


def test_generated_line():
    line = generated_line("Test Detail")
    assert isinstance(line, str)
    assert re.match(r"^Generated: \d{4}-\d{2}-\d{2} \d{2}:\d{2}   \|   Test Detail$", line) is not None


def test_fmt():
    assert fmt(3.1415926) == "3.1416"
    assert fmt(3.1415926, precision=6) == "3.141593"


def test_fmt_opt():
    assert fmt_opt(None) == "N/A"
    assert fmt_opt(1.5) == "1.5000"
    assert fmt_opt(1.5, precision=6) == "1.500000"


def test_io_reexports():
    import quality_core.io as qio
    assert qio.now is now
    assert qio.generated_line is generated_line
    assert qio.fmt is fmt
    assert qio.fmt_opt is fmt_opt



# --- write_keyvalue_sheet title ---------------------------------------------


def test_write_keyvalue_sheet_sets_title_and_values():
    wb = openpyxl.Workbook()
    ws = wb.active
    write_keyvalue_sheet(ws, [("Key", "Val"), ("N", 3)], title="Meta")
    assert ws.title == "Meta"
    assert ws.cell(1, 1).value == "Key"
    assert ws.cell(2, 2).value == 3


def test_write_keyvalue_sheet_title_optional():
    wb = openpyxl.Workbook()
    ws = wb.active
    original = ws.title
    write_keyvalue_sheet(ws, [("a", "b")])
    assert ws.title == original  # unchanged when no title given


# --- PDF chrome helpers ------------------------------------------------------


def test_pdf_title_emits_sanitized_centered_bar():
    pdf = _RecordingPdf()
    pdf_title(pdf, "Report — Summary")  # em dash must be Latin-1 sanitized
    methods = [c[0] for c in pdf.calls]
    assert "set_fill_color" in methods
    cell_text = [c[1][2] for c in pdf.calls if c[0] == "cell"]
    assert cell_text == ["Report - Summary"]  # em dash → hyphen


def test_pdf_subheader_renders_one_centered_line():
    pdf = _RecordingPdf()
    pdf_subheader(pdf, "Generated: 2026")
    cells = [c for c in pdf.calls if c[0] == "cell"]
    assert len(cells) == 1
    assert cells[0][1][2] == "Generated: 2026"


def test_pdf_summary_cells_emits_label_and_value_per_metric():
    pdf = _RecordingPdf()
    pdf_summary_cells(pdf, [("Cp", "1.4"), ("Cpk", "1.2"), ("Pp", "1.3")])
    cells = [c for c in pdf.calls if c[0] == "cell"]
    assert len(cells) == 6  # 3 labels + 3 values
    expected_w = (210.0 - 20.0) / 3
    assert all(c[1][0] == pytest.approx(expected_w) for c in cells)
    assert cells[0][1][2] == "Cp" and cells[3][1][2] == "1.4"


def test_sanitize_cell_scalar_matches_dataframe_escaping():
    # The scalar helper escapes exactly what sanitize_for_export escapes per cell.
    assert sanitize_cell("=evil") == "'=evil"
    assert sanitize_cell("\t=cmd") == "'\t=cmd"  # control-prefixed formula
    assert sanitize_cell("  -1+2") == "'  -1+2"  # whitespace-bypass
    assert sanitize_cell("safe") == "safe"
    assert sanitize_cell(42) == 42  # non-strings pass through
    assert sanitize_cell(None) is None
    # Idempotent: an already-escaped value is not double-escaped.
    assert sanitize_cell(sanitize_cell("=x")) == "'=x"


def test_sanitize_escapes_formula_prefixes():
    df = pd.DataFrame(
        [{"a": "=evil", "b": "+bad", "c": "-exploit", "d": "@nope", "e": "safe"}]
    )
    out = sanitize_for_export(df)
    assert out.loc[0, "a"] == "'=evil"
    assert out.loc[0, "b"] == "'+bad"
    assert out.loc[0, "c"] == "'-exploit"
    assert out.loc[0, "d"] == "'@nope"
    assert out.loc[0, "e"] == "safe"  # untouched


def test_sanitize_escapes_whitespace_and_control_prefixed_formulas():
    """Leading Tab/CR/whitespace before a formula char must not bypass escaping.

    Spreadsheets strip leading whitespace before formula detection, so these are
    still evaluated if left unescaped (OWASP CSV-injection trigger set).
    """
    df = pd.DataFrame(
        [{
            "tab": "\t=cmd|'/C calc'!A1",
            "cr": "\r=evil",
            "space": " =1+1",
            "newline": "\n=x",
            "plainspace": "  hello",   # leading space, no formula → safe
        }]
    )
    out = sanitize_for_export(df)
    assert out.loc[0, "tab"] == "'\t=cmd|'/C calc'!A1"
    assert out.loc[0, "cr"] == "'\r=evil"
    assert out.loc[0, "space"] == "' =1+1"
    assert out.loc[0, "newline"] == "'\n=x"
    assert out.loc[0, "plainspace"] == "  hello"  # untouched


def test_sanitize_is_idempotent():
    """Re-sanitizing must not double-escape an already-escaped value."""
    df = pd.DataFrame([{"a": "=evil", "b": "\t=tabbed", "c": " =spaced"}])
    once = sanitize_for_export(df)
    twice = sanitize_for_export(once)
    assert twice.loc[0, "a"] == "'=evil"      # apostrophe-prefixed → no longer matches
    assert twice.loc[0, "b"] == "'\t=tabbed"  # not re-escaped
    assert twice.loc[0, "c"] == "' =spaced"


def test_sanitize_leaves_non_strings_and_copies():
    df = pd.DataFrame([{"n": 5, "f": 1.5, "b": True}])
    out = sanitize_for_export(df)
    assert out.loc[0, "n"] == 5 and out.loc[0, "f"] == 1.5 and bool(out.loc[0, "b"]) is True
    assert out is not df  # returns a copy


def test_export_csv_no_formula_injection():
    df = pd.DataFrame([{"x": "=SUM(1,2)", "y": "+bad", "z": "-exploit", "w": "@nope"}])
    csv_text = export_csv(df).decode("utf-8")
    for raw in ("=SUM(1,2)", "+bad", "-exploit", "@nope"):
        pattern = r"(?<!['])(?:(?<=,)|(?<=\n))" + re.escape(raw)
        assert not re.search(pattern, csv_text), f"unescaped {raw!r} found in CSV"


def test_safe_text_maps_unicode_to_latin1():
    assert safe_text("9–10 × ≥") == "9-10 x >="
    # Result must be encodable as Latin-1 (fpdf2 core-font safe).
    safe_text("greek Ω emoji 🎯").encode("latin-1")


# --- write_table_sheet (openpyxl) -------------------------------------------


def test_write_table_sheet_structure_widths_freeze_and_numpy_unwrap():
    wb = openpyxl.Workbook()
    ws = wb.active
    df = pd.DataFrame(
        [
            {"Name": "a", "Score": np.int64(5), "Tier": "Red"},
            {"Name": "b", "Score": np.int64(9), "Tier": "Green"},
        ]
    )
    tier_fill = {"Red": "FF0000", "Green": "00FF00"}
    write_table_sheet(
        ws, df,
        title="Data",
        columns=["Name", "Score", "Tier"],
        col_widths={"Name": 20, "Score": 10},
        row_fill_hex=lambda r: tier_fill.get(str(r["Tier"])),
    )
    assert ws.title == "Data"
    assert [c.value for c in ws[1]] == ["Name", "Score", "Tier"]  # header row
    assert ws.cell(1, 1).fill.fgColor.rgb.endswith("2C3E50")  # default header fill
    # numpy int64 unwrapped to a native Python int via .item()
    score = ws.cell(2, 2).value
    assert score == 5 and type(score) is int
    # per-row fills from the callback
    assert ws.cell(2, 1).fill.fgColor.rgb.endswith("FF0000")  # Red row
    assert ws.cell(3, 1).fill.fgColor.rgb.endswith("00FF00")  # Green row
    # explicit widths, then the default (14) for the unlisted column
    assert ws.column_dimensions["A"].width == 20
    assert ws.column_dimensions["B"].width == 10
    assert ws.column_dimensions["C"].width == 14
    assert ws.freeze_panes == "A2"  # default freeze


def test_write_table_sheet_writes_only_named_columns_in_order():
    wb = openpyxl.Workbook()
    ws = wb.active
    df = pd.DataFrame([{"A": 1, "B": 2, "Extra": 3}])
    write_table_sheet(
        ws, df,
        title="T",
        columns=["B", "Missing", "A"],  # Missing absent → skipped; Extra unlisted → omitted
        col_widths={},
        freeze=None,
    )
    assert [c.value for c in ws[1]] == ["B", "A"]  # named & present, in given order
    # data row is written in the SAME selected order (B then A), not df's order
    assert [ws.cell(2, 1).value, ws.cell(2, 2).value] == [2, 1]
    assert ws.freeze_panes is None


def test_write_table_sheet_no_fill_when_callback_returns_none():
    wb = openpyxl.Workbook()
    ws = wb.active
    df = pd.DataFrame([{"A": 1}])
    write_table_sheet(ws, df, title="T", columns=["A"], col_widths={}, row_fill_hex=lambda r: None)
    assert ws.cell(2, 1).fill.fill_type in (None, "none")


def test_write_table_sheet_caches_repeated_fill_colour(monkeypatch):
    # Two rows share a colour: the per-colour PatternFill cache must build that
    # fill exactly once and reuse it (proven by counting constructions, since
    # openpyxl never reports two fills as identical even when cached).
    import quality_core.io.export as export_mod

    real_pattern_fill = export_mod.PatternFill
    built: list[str | None] = []

    def counting_pattern_fill(*args, **kwargs):
        built.append(kwargs.get("start_color"))
        return real_pattern_fill(*args, **kwargs)

    monkeypatch.setattr(export_mod, "PatternFill", counting_pattern_fill)

    wb = openpyxl.Workbook()
    ws = wb.active
    df = pd.DataFrame([{"A": 1}, {"A": 2}])
    write_table_sheet(ws, df, title="T", columns=["A"], col_widths={}, row_fill_hex=lambda r: "123456")

    assert built.count("123456") == 1  # built once, reused for the second row
    assert ws.cell(2, 1).fill.fgColor.rgb.endswith("123456")
    assert ws.cell(3, 1).fill.fgColor.rgb.endswith("123456")


# --- render_table (fpdf2, faked) --------------------------------------------


def test_render_table_renders_header_then_rows():
    pdf = _RecordingPdf(y=10.0)  # plenty of vertical space → no page break
    df = pd.DataFrame([{"A": 1, "B": 2}, {"A": 3, "B": 4}])
    render_table(
        pdf, df,
        columns=[("A", 20), ("B", 20)],
        row_values=lambda r: [str(r["A"]), str(r["B"])],
        row_rgb=lambda r: (200, 200, 200),
    )
    cells = [c for c in pdf.calls if c[0] == "cell"]
    assert len(cells) == 6  # 2 header + 2 rows × 2 cells
    assert [cells[0][1][2], cells[1][1][2]] == ["A", "B"]  # header labels
    assert [cells[2][1][2], cells[3][1][2]] == ["1", "2"]  # first row values
    assert not any(c[0] == "add_page" for c in pdf.calls)
    # fills: the default header colour and the per-row colour from row_rgb are applied
    fills = [c[1] for c in pdf.calls if c[0] == "set_fill_color"]
    assert (44, 62, 80) in fills      # default header_fill_rgb
    assert (200, 200, 200) in fills   # row_rgb result


def test_render_table_repeats_header_on_page_break():
    # y near the bottom margin forces the page-break guard on the first row.
    pdf = _RecordingPdf(y=290.0, h=297.0, b_margin=12.0)
    df = pd.DataFrame([{"A": 1, "B": 2}])
    render_table(
        pdf, df,
        columns=[("A", 20), ("B", 20)],
        row_values=lambda r: [str(r["A"]), str(r["B"])],
        row_rgb=lambda r: (0, 0, 0),
    )
    assert any(c[0] == "add_page" for c in pdf.calls)
    # the header (an "A" cell) is rendered twice: once initially, once after the break
    assert sum(1 for c in pdf.calls if c[0] == "cell" and c[1][2] == "A") == 2


# --- add_image_page (fpdf2, faked) ------------------------------------------


def test_add_image_page_embeds_png_with_sanitized_title():
    pdf = _RecordingPdf()
    add_image_page(pdf, "/tmp/chart.png", "Pareto — RPN")
    assert any(c[0] == "add_page" for c in pdf.calls)
    titles = [c[1][2] for c in pdf.calls if c[0] == "cell"]
    assert titles == ["Pareto - RPN"]  # em dash sanitized
    images = [c for c in pdf.calls if c[0] == "image"]
    assert len(images) == 1
    assert images[0][1][0] == "/tmp/chart.png"
    assert images[0][2]["w"] == 277  # default landscape image width


# --- #198 · the export primitive owns the injection guarantee ----------------
#
# Every test below fails against the pre-#198 code. The fix was originally shipped
# with no test at all: the io gate still read 100% because existing tests crossed
# the new lines incidentally, so reverting the fix left the whole suite green.


def test_sanitize_for_export_escapes_column_labels():
    """The header row is a payload surface too — a label is written verbatim."""
    df = pd.DataFrame([{"=cmd|' /C calc'!A0": 1, "safe": 2}])
    out = sanitize_for_export(df)
    assert list(out.columns) == ["'=cmd|' /C calc'!A0", "safe"]


def test_sanitize_for_export_escapes_duplicate_column_labels():
    """Duplicate labels must both be escaped — label-keyed access would miss one."""
    df = pd.DataFrame([["=a", "=b"]], columns=["dup", "dup"])
    out = sanitize_for_export(df)
    assert list(out.columns) == ["dup", "dup"]
    assert out.iloc[0].tolist() == ["'=a", "'=b"]


def test_write_table_sheet_escapes_header_and_body():
    wb = openpyxl.Workbook()
    ws = wb.active
    df = pd.DataFrame([{"=header": "=body"}])
    write_table_sheet(
        ws, df,
        title="Data",
        columns=["=header"],
        col_widths={"=header": 20},
    )
    assert ws.cell(1, 1).value == "'=header"
    assert ws.cell(2, 1).value == "'=body"


def test_write_keyvalue_sheet_escapes_label_and_value():
    """The key/value sheet has four app callers; the primitive owns their safety."""
    wb = openpyxl.Workbook()
    ws = wb.active
    write_keyvalue_sheet(ws, [("=label", "=value")])
    assert ws.cell(1, 1).value == "'=label"
    assert ws.cell(1, 2).value == "'=value"


def test_numeric_text_is_not_escaped_anywhere():
    """Formatted metrics like "-3.0000" must survive intact.

    openpyxl stores a leading apostrophe as a literal character — Excel's typed-input
    convention does not apply — so escaping these would visibly corrupt every negative
    metric in the summary sheets.
    """
    for numeric in ("-3.0000", "+5", "-0.5", "1e-4", "-.5", "+1E+3"):
        assert sanitize_cell(numeric) == numeric, numeric

    wb = openpyxl.Workbook()
    ws = wb.active
    write_keyvalue_sheet(ws, [("Cpk lower bound", "-3.0000")])
    assert ws.cell(1, 2).value == "-3.0000"

    # ...but a payload that merely starts like a number is still escaped.
    assert sanitize_cell("-3.0000+cmd|' /C calc'!A0") == "'-3.0000+cmd|' /C calc'!A0"


def test_numeric_exemption_is_narrow():
    """The exemption is the one hole in the escaping, so pin its exact edges.

    Every string below is accepted by Python's ``float()`` but is NOT a plain decimal.
    A spreadsheet does not read them as the same number, and no formatter in this repo
    emits them — so widening the exemption back to ``float()`` would give up real
    coverage for nothing. This test fails if that happens.
    """
    for not_exempt in (
        "-inf", "-Infinity", "-nan",   # float() accepts; Excel renders #NAME? / text
        "-1_000",                      # PEP 515 separator; Excel reads it as text
        "-１２３", "-١٢٣",              # Unicode digits — `\d` would have matched these
    ):
        assert sanitize_cell(not_exempt) == f"'{not_exempt}", not_exempt

    # Anchoring is the other half: surrounding whitespace must never be exempt,
    # because a leading Tab/CR is itself a formula trigger.
    for spaced in ("\t-3.0", "\r-3.0", " -3.0", "-3.0 "):
        assert sanitize_cell(spaced) == f"'{spaced}", repr(spaced)
