"""
tests/test_controlplan_export.py
Tests for quality_core/controlplan/export.py — the Control Plan live-formula .xlsx
exporter (#145).

Adapted from the sibling FMEA suite (#142) to Control Plan's roll-up model: the only live
formulas are the three ``Coverage`` sheet cells B2 (``=COUNTA``), B3 (``=COUNTIF(...,"Yes")``)
and B4 (``=IF(B2=0,0,B3/B2)``). The matrix sheet — including the derived ``PFMEA_Linked``
and ``Sample_Plan_Placeholder`` display columns — carries no live formula.

The two mandatory negative controls: a literal-int Coverage build must FAIL
``assert_cell_is_formula`` (proving the check is load-bearing, not tautological), and a
free-text field starting with ``"="`` must render inert while the Coverage roll-ups stay
live (per-cell isolation). The accuracy scorecard reads each formula's OWN referenced range
back out of the *saved* sheet and recomputes COUNTA/COUNTIF/ratio in Python — it never
trusts ``data_only=True`` (openpyxl caches no value for a never-recalculated formula, so
that path reads ``None``).
"""

from __future__ import annotations

import io
import re

import openpyxl
import pytest
from _xlsx_formula_audit import assert_cell_is_formula
from openpyxl.utils import get_column_letter
from quality_core.controlplan import (
    CONTROLPLAN_COL_WIDTHS,
    CONTROLPLAN_EXPORT_COLUMNS,
    benchmark_controlplan_dataset,
    export_controlplan_workbook,
)
from quality_core.controlplan.schema import ControlPlanDataset, ControlPlanRow

# Column letters derived the same way the exporter does, so a reorder of
# CONTROLPLAN_EXPORT_COLUMNS moves the test's expectations with the implementation's.
_CHAR_COL = get_column_letter(CONTROLPLAN_EXPORT_COLUMNS.index("Characteristic") + 1)
_LINKED_COL = get_column_letter(CONTROLPLAN_EXPORT_COLUMNS.index("PFMEA_Linked") + 1)
_PLACEHOLDER_COL = get_column_letter(
    CONTROLPLAN_EXPORT_COLUMNS.index("Sample_Plan_Placeholder") + 1
)
_SRCID_COL = get_column_letter(CONTROLPLAN_EXPORT_COLUMNS.index("Source_Cause_ID") + 1)
_LSL_COL = get_column_letter(CONTROLPLAN_EXPORT_COLUMNS.index("LSL") + 1)
_CHART_COL = get_column_letter(CONTROLPLAN_EXPORT_COLUMNS.index("Recommended_Chart") + 1)

_DEFAULT_TITLE = "Control Plan"

# Full-field row template so mypy sees explicit values for the nullable float fields —
# ``Annotated[float | None, Field(default=None)]`` is not treated as a default by mypy,
# unlike ``recommended_chart``/``source_cause_id``/``sample_plan_is_placeholder``.
_VALID_ROW: dict[str, object] = {
    "characteristic": "Bore diameter",
    "lsl": None,
    "target": None,
    "usl": None,
    "measurement_method": "Air gauge",
    "sample_size": 5,
    "frequency": "Every 30 minutes",
    "recommended_chart": None,
    "reaction_plan": "Quarantine and re-gauge",
    "source_cause_id": None,
}


def _make_row(**overrides: object) -> ControlPlanRow:
    return ControlPlanRow(**{**_VALID_ROW, **overrides})  # type: ignore[arg-type]


def _saved(wb: openpyxl.Workbook) -> bytes:
    """Serialize a workbook to .xlsx bytes (tests never touch disk)."""
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _load(wb_bytes: bytes) -> openpyxl.Workbook:
    """Reload saved bytes with formulas intact (never data_only — no cache exists)."""
    return openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)


def _last_row(n_rows: int) -> int:
    """Last matrix row a roll-up range spans, floored at 2 (the empty-dataset guard)."""
    return max(n_rows + 1, 2)


def _expected_total_formula(title: str, n_rows: int) -> str:
    return f"=COUNTA('{title}'!{_CHAR_COL}2:{_CHAR_COL}{_last_row(n_rows)})"


def _expected_linked_formula(title: str, n_rows: int) -> str:
    return f'=COUNTIF(\'{title}\'!{_LINKED_COL}2:{_LINKED_COL}{_last_row(n_rows)},"Yes")'


_RANGE_RE = re.compile(r"([A-Z]+)(\d+):([A-Z]+)(\d+)")


def _referenced_row_span(formula: str) -> tuple[str, int, int]:
    """Parse ``col``, ``first_row``, ``last_row`` out of a saved COUNTA/COUNTIF formula.

    Reads the formula's OWN referenced range so the recomputation below counts exactly the
    cells Excel would, rather than a range the test re-derives independently.
    """
    m = _RANGE_RE.search(formula)
    assert m is not None, f"no A1 range found in formula {formula!r}"
    col, r1, _col2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
    assert col == m.group(3), f"range spans two columns in {formula!r}"
    return col, r1, r2


def _counta(ws: openpyxl.worksheet.worksheet.Worksheet, col: str, r1: int, r2: int) -> int:
    """Python equivalent of COUNTA over ``col`` rows ``r1..r2`` (count non-empty)."""
    return sum(
        1
        for r in range(r1, r2 + 1)
        if ws[f"{col}{r}"].value not in (None, "")
    )


def _countif_yes(
    ws: openpyxl.worksheet.worksheet.Worksheet, col: str, r1: int, r2: int
) -> int:
    """Python equivalent of COUNTIF(range, "Yes") over ``col`` rows ``r1..r2``."""
    return sum(1 for r in range(r1, r2 + 1) if ws[f"{col}{r}"].value == "Yes")


# --- 1 · valid workbook -------------------------------------------------------


def test_export_benchmark_is_a_loadable_xlsx_with_both_sheets():
    dataset = benchmark_controlplan_dataset()
    wb_bytes = export_controlplan_workbook(dataset)
    assert isinstance(wb_bytes, bytes)

    wb = _load(wb_bytes)
    assert wb.sheetnames == [_DEFAULT_TITLE, "Coverage"]

    matrix = wb[_DEFAULT_TITLE]
    assert [c.value for c in matrix[1]] == list(CONTROLPLAN_EXPORT_COLUMNS)
    assert matrix.max_row == len(dataset.rows) + 1

    coverage = wb["Coverage"]
    assert [c.value for c in coverage[1]] == ["Metric", "Value"]
    assert [coverage[f"A{r}"].value for r in (2, 3, 4)] == [
        "Total Characteristics",
        "PFMEA-Linked Characteristics",
        "Linkage Coverage %",
    ]


def test_col_widths_cover_every_matrix_column():
    assert set(CONTROLPLAN_COL_WIDTHS) == set(CONTROLPLAN_EXPORT_COLUMNS)


# --- 2 · custom title threads into cross-sheet formulas -----------------------


def test_custom_title_renames_matrix_and_appears_in_coverage_formulas():
    """Edge case 7: a hardcoded matrix sheet name would silently break here."""
    title = "CP-Rev-C 2026"
    dataset = benchmark_controlplan_dataset()
    wb_bytes = export_controlplan_workbook(dataset, title=title)
    wb = _load(wb_bytes)
    assert wb.sheetnames == [title, "Coverage"]

    n = len(dataset.rows)
    coverage = wb["Coverage"]
    assert coverage["B2"].value == _expected_total_formula(title, n)
    assert coverage["B3"].value == _expected_linked_formula(title, n)
    assert title in coverage["B2"].value
    assert title in coverage["B3"].value


# --- 3 · positive live-formula proof + exact formula strings ------------------


def test_coverage_rollups_are_live_formulas_with_exact_strings():
    dataset = benchmark_controlplan_dataset()
    wb_bytes = export_controlplan_workbook(dataset)
    n = len(dataset.rows)

    for coord in ("B2", "B3", "B4"):
        # Does not raise: the cell carries a live <f> element.
        assert_cell_is_formula(wb_bytes, "Coverage", coord)

    coverage = _load(wb_bytes)["Coverage"]
    assert coverage["B2"].value == _expected_total_formula(_DEFAULT_TITLE, n)
    assert coverage["B3"].value == _expected_linked_formula(_DEFAULT_TITLE, n)
    assert coverage["B4"].value == "=IF(B2=0,0,B3/B2)"
    assert coverage["B4"].number_format == "0.0%"


# --- 4 · literal-build negative control ---------------------------------------


def test_literal_coverage_build_fails_the_formula_verifier():
    """NEGATIVE CONTROL: hardcoded-value Coverage cells must FAIL the verifier.

    If a future edit swapped the three ``Formula`` cells for plain ints/floats, this suite
    must catch it — so prove the verifier bites on a literal build of the same layout.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Coverage"
    ws.cell(row=2, column=2, value=6)  # literal int, NOT a Formula
    ws.cell(row=3, column=2, value=4)
    ws.cell(row=4, column=2, value=0.667)
    saved = _saved(wb)

    for coord in ("B2", "B3", "B4"):
        with pytest.raises(AssertionError, match="literal, not a live formula"):
            assert_cell_is_formula(saved, "Coverage", coord)


# --- 5 · accuracy scorecard ---------------------------------------------------


def test_controlplan_export_accuracy_scorecard():
    """The saved roll-up FORMULAS must recompute to the engine/dataset's own counts.

    # reproduce:
    #   cd packages/quality-core && uv run pytest \
    #     tests/test_controlplan_export.py -k accuracy_scorecard -q

    Recomputes total/linked/coverage from the dataset, then reads each formula's OWN
    referenced range back out of the saved sheet and applies COUNTA/COUNTIF/ratio in
    Python — never data_only=True, which returns None for a never-recalculated formula.
    These are the values a human opening the file in Excel would see after recalculation.
    """
    dataset = benchmark_controlplan_dataset()
    expected_total = len(dataset.rows)
    expected_linked = sum(1 for r in dataset.rows if r.source_cause_id is not None)
    expected_coverage = expected_linked / expected_total if expected_total else 0

    # The benchmark must exercise a genuine partial-coverage case, not a degenerate edge.
    assert 0 < expected_coverage < 1

    wb = _load(export_controlplan_workbook(dataset))
    matrix = wb[_DEFAULT_TITLE]
    coverage = wb["Coverage"]

    # B2 — COUNTA over the Characteristic column's own referenced range.
    b2 = coverage["B2"].value
    col, r1, r2 = _referenced_row_span(b2)
    assert col == _CHAR_COL
    assert _counta(matrix, col, r1, r2) == expected_total

    # B3 — COUNTIF(range,"Yes") over the PFMEA_Linked column's own referenced range.
    b3 = coverage["B3"].value
    col, r1, r2 = _referenced_row_span(b3)
    assert col == _LINKED_COL
    assert _countif_yes(matrix, col, r1, r2) == expected_linked

    # B4 — the ratio of the two, as Excel would compute it.
    assert coverage["B4"].value == "=IF(B2=0,0,B3/B2)"
    computed_ratio = expected_linked / expected_total if expected_total else 0
    assert computed_ratio == pytest.approx(expected_coverage)


# --- 6 · derived display columns are literals, not formulas -------------------


def test_derived_display_columns_are_not_live_formulas():
    dataset = benchmark_controlplan_dataset()
    wb_bytes = export_controlplan_workbook(dataset)

    for r in range(2, len(dataset.rows) + 2):
        for col in (_LINKED_COL, _PLACEHOLDER_COL):
            with pytest.raises(AssertionError, match="literal, not a live formula"):
                assert_cell_is_formula(wb_bytes, _DEFAULT_TITLE, f"{col}{r}")

    matrix = _load(wb_bytes)[_DEFAULT_TITLE]
    # Sanity: the values really are the Yes/No sentinels the roll-up counts against.
    assert set(matrix[f"{_LINKED_COL}{r}"].value for r in range(2, len(dataset.rows) + 2)) <= {
        "Yes",
        "No",
    }


# --- 7 · injection safety negative control ------------------------------------


def test_injection_freetext_is_inert_but_rollups_stay_live_and_correct():
    """NEGATIVE CONTROL: a data field starting with '=' renders inert, per-cell.

    The malicious ``reaction_plan`` is apostrophe-escaped (never a live formula), while the
    Coverage roll-ups stay live AND still count the injected row correctly — proving the
    ``Formula`` opt-in is per-cell and the roll-up counts a different column entirely.
    """
    payload = "=cmd|' /C calc'!A0"
    dataset = ControlPlanDataset(
        rows=[
            _make_row(
                characteristic="Injected row",
                reaction_plan=payload,
                source_cause_id="F9::FM9::C9",
            ),
            _make_row(characteristic="Plain row", source_cause_id=None),
        ]
    )
    wb_bytes = export_controlplan_workbook(dataset)

    reaction_col = get_column_letter(
        CONTROLPLAN_EXPORT_COLUMNS.index("Reaction_Plan") + 1
    )
    matrix = _load(wb_bytes)[_DEFAULT_TITLE]
    assert matrix[f"{reaction_col}2"].value == "'" + payload
    with pytest.raises(AssertionError, match="literal, not a live formula"):
        assert_cell_is_formula(wb_bytes, _DEFAULT_TITLE, f"{reaction_col}2")

    # The Coverage roll-ups are still live formulas...
    for coord in ("B2", "B3", "B4"):
        assert_cell_is_formula(wb_bytes, "Coverage", coord)

    # ...and still count the injected row: 2 total, 1 linked.
    coverage = _load(wb_bytes)["Coverage"]
    col, r1, r2 = _referenced_row_span(coverage["B2"].value)
    assert _counta(matrix, col, r1, r2) == 2
    col, r1, r2 = _referenced_row_span(coverage["B3"].value)
    assert _countif_yes(matrix, col, r1, r2) == 1


# --- 8 · empty dataset branch -------------------------------------------------


def test_empty_dataset_yields_header_only_matrix_with_live_zero_rollups():
    """Empty-rows path: matrix header-only, roll-up ranges floor to A2:A2, B4 guard string.

    Exercises ``_last_matrix_row``'s floor branch (row_count == 0) and the unconditional
    ``IF(B2=0,0,...)`` divide-by-zero guard.
    """
    title = "CP Custom"
    wb_bytes = export_controlplan_workbook(ControlPlanDataset(rows=[]), title=title)
    wb = _load(wb_bytes)

    matrix = wb[title]
    assert matrix.max_row == 1
    assert [c.value for c in matrix[1]] == list(CONTROLPLAN_EXPORT_COLUMNS)

    coverage = wb["Coverage"]
    # B2/B3 are still live formulas (not omitted), ranges floored to A2:A2 / L2:L2.
    for coord in ("B2", "B3", "B4"):
        assert_cell_is_formula(wb_bytes, "Coverage", coord)
    assert coverage["B2"].value == f"=COUNTA('{title}'!{_CHAR_COL}2:{_CHAR_COL}2)"
    assert coverage["B3"].value == f'=COUNTIF(\'{title}\'!{_LINKED_COL}2:{_LINKED_COL}2,"Yes")'
    assert coverage["B4"].value == "=IF(B2=0,0,B3/B2)"

    # The referenced range is empty (nothing written past the header) → COUNTA/COUNTIF 0.
    col, r1, r2 = _referenced_row_span(coverage["B2"].value)
    assert _counta(matrix, col, r1, r2) == 0
    col, r1, r2 = _referenced_row_span(coverage["B3"].value)
    assert _countif_yes(matrix, col, r1, r2) == 0


# --- 9 · all-linked / all-unlinked edges --------------------------------------


def test_all_linked_dataset_counts_equal():
    dataset = ControlPlanDataset(
        rows=[
            _make_row(characteristic="A", source_cause_id="F1::FM1::C1"),
            _make_row(characteristic="B", source_cause_id="F1::FM1::C2"),
        ]
    )
    wb = _load(export_controlplan_workbook(dataset))
    matrix, coverage = wb[_DEFAULT_TITLE], wb["Coverage"]

    col, r1, r2 = _referenced_row_span(coverage["B2"].value)
    total = _counta(matrix, col, r1, r2)
    col, r1, r2 = _referenced_row_span(coverage["B3"].value)
    linked = _countif_yes(matrix, col, r1, r2)
    assert total == linked == 2  # coverage would be 100%, not an error


def test_all_unlinked_dataset_has_zero_linked_but_no_div_zero():
    dataset = ControlPlanDataset(
        rows=[
            _make_row(characteristic="A", source_cause_id=None),
            _make_row(characteristic="B", source_cause_id=None),
        ]
    )
    wb = _load(export_controlplan_workbook(dataset))
    matrix, coverage = wb[_DEFAULT_TITLE], wb["Coverage"]

    col, r1, r2 = _referenced_row_span(coverage["B2"].value)
    total = _counta(matrix, col, r1, r2)
    col, r1, r2 = _referenced_row_span(coverage["B3"].value)
    linked = _countif_yes(matrix, col, r1, r2)
    assert total == 2
    assert linked == 0  # 0/2 is a valid division; only B2==0 trips the IF guard
    assert coverage["B4"].value == "=IF(B2=0,0,B3/B2)"


# --- 10 · row order preserved -------------------------------------------------


def test_row_order_is_preserved_not_sorted():
    dataset = ControlPlanDataset(
        rows=[
            _make_row(characteristic="Zeta"),
            _make_row(characteristic="Alpha"),
            _make_row(characteristic="Mike"),
        ]
    )
    matrix = _load(export_controlplan_workbook(dataset))[_DEFAULT_TITLE]
    assert [matrix[f"{_CHAR_COL}{r}"].value for r in (2, 3, 4)] == ["Zeta", "Alpha", "Mike"]


# --- 11 · nullable fields round-trip as blank, not "None" ---------------------


def test_none_valued_nullable_fields_roundtrip_as_blank():
    dataset = ControlPlanDataset(
        rows=[
            _make_row(
                characteristic="All-None row",
                lsl=None,
                target=None,
                usl=None,
                recommended_chart=None,
                source_cause_id=None,
            )
        ]
    )
    matrix = _load(export_controlplan_workbook(dataset))[_DEFAULT_TITLE]
    for col in (_LSL_COL, _CHART_COL, _SRCID_COL):
        value = matrix[f"{col}2"].value
        assert value is None, f"{col}2 should be blank, got {value!r}"
        assert value != "None"


# --- 12 · re-export surface ---------------------------------------------------


def test_public_names_are_reexported_from_controlplan_package():
    import quality_core.controlplan as pkg
    from quality_core.controlplan import export as mod

    assert pkg.export_controlplan_workbook is mod.export_controlplan_workbook
    assert pkg.benchmark_controlplan_dataset is mod.benchmark_controlplan_dataset
    assert pkg.CONTROLPLAN_EXPORT_COLUMNS is mod.CONTROLPLAN_EXPORT_COLUMNS
    assert pkg.CONTROLPLAN_COL_WIDTHS is mod.CONTROLPLAN_COL_WIDTHS


# --- 13 · fresh benchmark object per call -------------------------------------


def test_benchmark_dataset_returns_a_fresh_object_each_call():
    a = benchmark_controlplan_dataset()
    b = benchmark_controlplan_dataset()
    assert a is not b
    assert a.rows is not b.rows


def test_benchmark_placeholder_flag_spans_both_yes_and_no():
    """Both branches of the Yes/No placeholder display must be exercised."""
    dataset = benchmark_controlplan_dataset()
    matrix = _load(export_controlplan_workbook(dataset))[_DEFAULT_TITLE]
    values = {
        matrix[f"{_PLACEHOLDER_COL}{r}"].value for r in range(2, len(dataset.rows) + 2)
    }
    assert values == {"Yes", "No"}
