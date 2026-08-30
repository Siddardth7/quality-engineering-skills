"""
tests/test_copq_export.py
Tests for quality_core/copq/export.py — Cost of Poor Quality (COPQ) live-formula .xlsx exporter (#147).

Validates:
- 100% line & branch coverage on quality_core.copq.export.
- Positive live OOXML <f> formula verification via assert_cell_is_formula on:
  - COPQ Ledger sheet: per-line =(C{r}*D{r})+E{r} formulas in Column F.
  - PAF Category Summary sheet: SUMIF subtotals (C2:C5), Total CoQ SUM (C6), Pct ratios (D2:D6).
  - Executive Summary & Metadata sheet: cross-sheet formulas for CoQ (B4), COPQ (B5), CoGQ (B6),
    Failure-to-Conformance Ratio (B7), COPQ % of Revenue (B9), Line Item Count (B10).
- Accuracy scorecards: recomputed line totals and category sums equal engine computed values
  from estimate_copq() and COPQDataset properties.
- 🔒 Security invariant: untrusted data strings starting with '=', '+', '-', '@', '\\t', '\\r' render inert.
- Reusable verifier negative controls: literal cells in computed slots fail assert_cell_is_formula.
- Type guards & error paths: empty datasets, invalid revenue_base (<= 0, nan, inf, bool, type),
  polymorphism (COPQDataset, list[CostItem], list[dict], DataFrame, dict), custom titles.
- Re-exports from quality_core.copq package.
"""

from __future__ import annotations

import io

import openpyxl
import pytest
from _xlsx_formula_audit import assert_cell_is_formula
from openpyxl.utils import get_column_letter
from quality_core.copq import (
    COPQ_COL_WIDTHS,
    COPQ_LEDGER_COLUMNS,
    PAF_CATEGORY_VALUES,
    PAF_SUMMARY_COL_WIDTHS,
    PAF_SUMMARY_COLUMNS,
    COPQDataset,
    CostItem,
    benchmark_copq_dataset,
    build_copq_workbook,
    estimate_copq,
    export_copq_excel,
    export_copq_workbook,
)
from quality_core.copq.export import (
    _EXECUTIVE_SUMMARY_SHEET_TITLE,
    _PAF_SUMMARY_SHEET_TITLE,
    _item_drivers,
    _last_row,
    _write_metadata_sheet,
)

_DEFAULT_TITLE = "COPQ Ledger"
_CAT_COL = get_column_letter(COPQ_LEDGER_COLUMNS.index("Category") + 1)
_QTY_COL = get_column_letter(COPQ_LEDGER_COLUMNS.index("Quantity_or_Hours") + 1)
_RATE_COL = get_column_letter(COPQ_LEDGER_COLUMNS.index("Unit_Rate") + 1)
_DIRECT_COL = get_column_letter(COPQ_LEDGER_COLUMNS.index("Direct_Expense") + 1)
_TOTAL_COL = get_column_letter(COPQ_LEDGER_COLUMNS.index("Line_Total") + 1)


def _saved(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _load(wb_bytes: bytes) -> openpyxl.Workbook:
    return openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)


# ===========================================================================
# 1. Structure & Layout Tests
# ===========================================================================


def test_export_benchmark_is_a_loadable_3sheet_workbook() -> None:
    dataset = benchmark_copq_dataset()
    wb_bytes = export_copq_workbook(dataset)
    assert isinstance(wb_bytes, bytes)

    alias_bytes = export_copq_excel(dataset)
    assert isinstance(alias_bytes, bytes)

    wb = _load(wb_bytes)
    assert wb.sheetnames == [_DEFAULT_TITLE, _PAF_SUMMARY_SHEET_TITLE, _EXECUTIVE_SUMMARY_SHEET_TITLE]

    ledger = wb[_DEFAULT_TITLE]
    assert [c.value for c in ledger[1]] == list(COPQ_LEDGER_COLUMNS)
    assert ledger.max_row == len(dataset.items) + 1

    paf_sheet = wb[_PAF_SUMMARY_SHEET_TITLE]
    assert [c.value for c in paf_sheet[1]] == list(PAF_SUMMARY_COLUMNS)
    assert paf_sheet.max_row == 6
    paf_categories = [paf_sheet[f"A{r}"].value for r in range(2, 6)]
    assert paf_categories == list(PAF_CATEGORY_VALUES)
    assert paf_sheet["A6"].value == "Total Cost of Quality"

    exec_sheet = wb[_EXECUTIVE_SUMMARY_SHEET_TITLE]
    assert [exec_sheet[f"A{r}"].value for r in range(1, 11)] == [
        "Report Title",
        "Date Generated",
        "Standards Basis",
        "Total Cost of Quality (CoQ)",
        "Cost of Poor Quality (COPQ)",
        "Cost of Good Quality (CoGQ)",
        "COPQ / CoGQ Failure-to-Conformance Ratio",
        "Revenue Base",
        "COPQ % of Revenue",
        "Total Cost Line Items",
    ]
    assert exec_sheet["B1"].value == _DEFAULT_TITLE
    assert exec_sheet["B3"].value == "ASQ CSSGB BoK / PAF Model (Feigenbaum & Juran)"
    assert exec_sheet["B8"].value == 500000.0


def test_col_widths_and_constants() -> None:
    assert set(COPQ_COL_WIDTHS) == set(COPQ_LEDGER_COLUMNS)
    assert set(PAF_SUMMARY_COL_WIDTHS) == set(PAF_SUMMARY_COLUMNS)
    for col, width in COPQ_COL_WIDTHS.items():
        assert width > 0
    for col, width in PAF_SUMMARY_COL_WIDTHS.items():
        assert width > 0


# ===========================================================================
# 2. Positive Live Formula Verification
# ===========================================================================


def test_positive_live_formula_audit_copq() -> None:
    dataset = benchmark_copq_dataset()
    wb_bytes = export_copq_workbook(dataset)
    n = len(dataset.items)
    last_row = _last_row(n)
    ledger_prefix = f"'{_DEFAULT_TITLE}'!"
    paf_prefix = f"'{_PAF_SUMMARY_SHEET_TITLE}'!"

    # 1. Sheet 1: COPQ Ledger (Column F live formulas)
    for r in range(2, n + 2):
        assert_cell_is_formula(wb_bytes, _DEFAULT_TITLE, f"F{r}")

    # 2. Sheet 2: PAF Category Summary (Columns C & D live formulas)
    for r in range(2, 7):
        assert_cell_is_formula(wb_bytes, _PAF_SUMMARY_SHEET_TITLE, f"C{r}")
        assert_cell_is_formula(wb_bytes, _PAF_SUMMARY_SHEET_TITLE, f"D{r}")

    # 3. Sheet 3: Executive Summary (B4..B7, B9, B10 live formulas)
    for coord in ("B4", "B5", "B6", "B7", "B9", "B10"):
        assert_cell_is_formula(wb_bytes, _EXECUTIVE_SUMMARY_SHEET_TITLE, coord)

    # Verify exact formula strings via openpyxl reload
    wb = _load(wb_bytes)

    # Sheet 1 formulas
    ws_ledger = wb[_DEFAULT_TITLE]
    for r in range(2, n + 2):
        assert ws_ledger[f"F{r}"].value == f"=({_QTY_COL}{r}*{_RATE_COL}{r})+{_DIRECT_COL}{r}"
        assert ws_ledger[f"F{r}"].number_format == "0.00"

    # Sheet 2 formulas
    ws_paf = wb[_PAF_SUMMARY_SHEET_TITLE]
    paf_configs = [
        ("Prevention", 2),
        ("Appraisal", 3),
        ("InternalFailure", 4),
        ("ExternalFailure", 5),
    ]
    for cat_name, r in paf_configs:
        assert (
            ws_paf[f"C{r}"].value
            == f'=SUMIF({ledger_prefix}{_CAT_COL}2:{_CAT_COL}{last_row}, "{cat_name}", {ledger_prefix}{_TOTAL_COL}2:{_TOTAL_COL}{last_row})'
        )
        assert ws_paf[f"C{r}"].number_format == "0.00"
        assert ws_paf[f"D{r}"].value == f"=IF($C$6=0, 0, C{r}/$C$6)"
        assert ws_paf[f"D{r}"].number_format == "0.00%"

    assert ws_paf["C6"].value == "=SUM(C2:C5)"
    assert ws_paf["C6"].number_format == "0.00"
    assert ws_paf["D6"].value == "=IF(C6=0, 0, SUM(D2:D5))"
    assert ws_paf["D6"].number_format == "0.00%"

    # Sheet 3 formulas
    ws_exec = wb[_EXECUTIVE_SUMMARY_SHEET_TITLE]
    assert ws_exec["B4"].value == f"={paf_prefix}C6"
    assert ws_exec["B5"].value == f"={paf_prefix}C4+{paf_prefix}C5"
    assert ws_exec["B6"].value == f"={paf_prefix}C2+{paf_prefix}C3"
    assert ws_exec["B7"].value == "=IF(B6=0, 0, B5/B6)"
    assert ws_exec["B9"].value == "=(B5/B8)*100"
    assert ws_exec["B10"].value == f"=COUNTA({ledger_prefix}{_CAT_COL}2:{_CAT_COL}{last_row})"


# ===========================================================================
# 3. Accuracy Scorecard vs estimate_copq & COPQDataset
# ===========================================================================


def test_accuracy_scorecard_copq() -> None:
    dataset = benchmark_copq_dataset(revenue_base=500000.0)

    # 1. Dataset property validation
    assert dataset.prevention_cost == 7300.0
    assert dataset.appraisal_cost == 11300.0
    assert dataset.internal_failure_cost == 9925.0
    assert dataset.external_failure_cost == 13800.0
    assert dataset.copq == 23725.0
    assert dataset.cogq == 18600.0
    assert dataset.total_cost == 42325.0
    assert dataset.copq_pct_revenue == pytest.approx(4.745)

    # estimate_copq with equivalent category inputs
    engine_result = estimate_copq(
        prevention_cost=dataset.prevention_cost,
        appraisal_cost=dataset.appraisal_cost,
        scrap_qty=45,
        unit_cost=120.0,
        rework_hours=35.0,
        labor_rate=65.0,
        added_material_cost=450.0,
        warranty_units=12,
        warranty_unit_cost=850.0,
        recall_cost=3600.0 + 1800.0,  # remaining direct external + containment
        revenue_base=dataset.revenue_base,
    )
    assert engine_result.total_coq == dataset.total_cost
    assert engine_result.total_copq == dataset.copq
    assert engine_result.cogq_total == dataset.cogq
    assert engine_result.copq_percentage_of_revenue == pytest.approx(dataset.copq_pct_revenue)

    # 2. Sheet recomputation check
    wb_bytes = export_copq_workbook(dataset)
    wb = _load(wb_bytes)
    ledger = wb[_DEFAULT_TITLE]

    # Recompute line totals from sheet
    recomputed_totals: list[float] = []
    category_totals: dict[str, float] = {
        "Prevention": 0.0,
        "Appraisal": 0.0,
        "InternalFailure": 0.0,
        "ExternalFailure": 0.0,
    }
    for r in range(2, len(dataset.items) + 2):
        cat = ledger[f"{_CAT_COL}{r}"].value
        qty = float(ledger[f"{_QTY_COL}{r}"].value)
        rate = float(ledger[f"{_RATE_COL}{r}"].value)
        direct = float(ledger[f"{_DIRECT_COL}{r}"].value)
        line_total = (qty * rate) + direct
        recomputed_totals.append(line_total)
        category_totals[cat] += line_total

    for idx, item in enumerate(dataset.items):
        assert recomputed_totals[idx] == pytest.approx(item.total_cost)

    assert category_totals["Prevention"] == pytest.approx(dataset.prevention_cost)
    assert category_totals["Appraisal"] == pytest.approx(dataset.appraisal_cost)
    assert category_totals["InternalFailure"] == pytest.approx(dataset.internal_failure_cost)
    assert category_totals["ExternalFailure"] == pytest.approx(dataset.external_failure_cost)

    total_coq = sum(category_totals.values())
    assert total_coq == pytest.approx(dataset.total_cost)

    copq_total = category_totals["InternalFailure"] + category_totals["ExternalFailure"]
    assert copq_total == pytest.approx(dataset.copq)

    cogq_total = category_totals["Prevention"] + category_totals["Appraisal"]
    assert cogq_total == pytest.approx(dataset.cogq)


# ===========================================================================
# 4. Revenue Base Variations & Error Guards
# ===========================================================================


def test_revenue_base_none_omits_revenue_formula() -> None:
    dataset = benchmark_copq_dataset(revenue_base=None)
    wb_bytes = export_copq_workbook(dataset, revenue_base=None)
    wb = _load(wb_bytes)
    ws_exec = wb[_EXECUTIVE_SUMMARY_SHEET_TITLE]

    assert ws_exec["B8"].value == "N/A"
    assert ws_exec["B9"].value == "N/A"

    # B9 is literal 'N/A', so assert_cell_is_formula fails on B9
    with pytest.raises(AssertionError, match="literal, not a live formula"):
        assert_cell_is_formula(wb_bytes, _EXECUTIVE_SUMMARY_SHEET_TITLE, "B9")

    # Other computed cells are still live formulas
    for coord in ("B4", "B5", "B6", "B7", "B10"):
        assert_cell_is_formula(wb_bytes, _EXECUTIVE_SUMMARY_SHEET_TITLE, coord)


def test_revenue_base_explicit_override_and_dataset_propagation() -> None:
    # 1. Propagate from dataset
    dataset = benchmark_copq_dataset(revenue_base=750000.0)
    wb = _load(export_copq_workbook(dataset))
    ws_exec = wb[_EXECUTIVE_SUMMARY_SHEET_TITLE]
    assert ws_exec["B8"].value == 750000.0
    assert ws_exec["B9"].value == "=(B5/B8)*100"

    # 2. Explicit parameter override
    wb_override = _load(export_copq_workbook(dataset, revenue_base=1200000.0))
    ws_exec_override = wb_override[_EXECUTIVE_SUMMARY_SHEET_TITLE]
    assert ws_exec_override["B8"].value == 1200000.0


def test_invalid_revenue_base_raises_appropriate_errors() -> None:
    dataset = benchmark_copq_dataset()

    # Boolean revenue base
    with pytest.raises(TypeError, match="cannot be a boolean"):
        build_copq_workbook(dataset, revenue_base=True)  # type: ignore[arg-type]

    # Non-number type
    with pytest.raises(TypeError, match="must be a number or None"):
        build_copq_workbook(dataset, revenue_base="500000")  # type: ignore[arg-type]

    # Non-positive values
    with pytest.raises(ValueError, match="must be a positive finite number"):
        build_copq_workbook(dataset, revenue_base=0.0)

    with pytest.raises(ValueError, match="must be a positive finite number"):
        build_copq_workbook(dataset, revenue_base=-500.0)

    # Non-finite values (NaN / Inf)
    with pytest.raises(ValueError, match="must be a positive finite number"):
        build_copq_workbook(dataset, revenue_base=float("nan"))

    with pytest.raises(ValueError, match="must be a positive finite number"):
        build_copq_workbook(dataset, revenue_base=float("inf"))


# ===========================================================================
# 5. Custom Title Handling
# ===========================================================================


def test_custom_title_threads_into_all_cross_sheet_formulas_copq() -> None:
    custom_title = "FY26 Q3 Cost of Quality Report"
    dataset = benchmark_copq_dataset()
    wb_bytes = export_copq_workbook(dataset, title=custom_title)
    wb = _load(wb_bytes)

    assert wb.sheetnames == [custom_title, _PAF_SUMMARY_SHEET_TITLE, _EXECUTIVE_SUMMARY_SHEET_TITLE]

    ws_paf = wb[_PAF_SUMMARY_SHEET_TITLE]
    prefix = f"'{custom_title}'!"
    assert prefix in ws_paf["C2"].value

    ws_exec = wb[_EXECUTIVE_SUMMARY_SHEET_TITLE]
    assert ws_exec["B1"].value == custom_title
    assert prefix in ws_exec["B10"].value

    # Check live formula elements
    assert_cell_is_formula(wb_bytes, custom_title, "F2")
    assert_cell_is_formula(wb_bytes, _PAF_SUMMARY_SHEET_TITLE, "C2")
    assert_cell_is_formula(wb_bytes, _PAF_SUMMARY_SHEET_TITLE, "D2")
    assert_cell_is_formula(wb_bytes, _EXECUTIVE_SUMMARY_SHEET_TITLE, "B4")
    assert_cell_is_formula(wb_bytes, _EXECUTIVE_SUMMARY_SHEET_TITLE, "B10")


# ===========================================================================
# 6. Negative Controls (Literal & Injection)
# ===========================================================================


def test_literal_negative_control_fails_formula_verifier_copq() -> None:
    wb = openpyxl.Workbook()
    ws_ledger = wb.active
    ws_ledger.title = _DEFAULT_TITLE
    ws_ledger.cell(row=2, column=6, value=5000.0)  # Literal float in F2

    ws_paf = wb.create_sheet(_PAF_SUMMARY_SHEET_TITLE)
    ws_paf.cell(row=2, column=3, value=7300.0)  # Literal float in C2
    ws_paf.cell(row=2, column=4, value=0.15)    # Literal float in D2
    ws_paf.cell(row=6, column=3, value=48000.0) # Literal float in C6
    ws_paf.cell(row=6, column=4, value=1.0)     # Literal float in D6

    ws_exec = wb.create_sheet(_EXECUTIVE_SUMMARY_SHEET_TITLE)
    ws_exec.cell(row=4, column=2, value=48000.0)
    ws_exec.cell(row=5, column=2, value=30000.0)
    ws_exec.cell(row=6, column=2, value=18000.0)
    ws_exec.cell(row=7, column=2, value=1.67)
    ws_exec.cell(row=9, column=2, value=6.0)
    ws_exec.cell(row=10, column=2, value=9)

    saved = _saved(wb)

    with pytest.raises(AssertionError, match="literal, not a live formula"):
        assert_cell_is_formula(saved, _DEFAULT_TITLE, "F2")

    for coord in ("C2", "D2", "C6", "D6"):
        with pytest.raises(AssertionError, match="literal, not a live formula"):
            assert_cell_is_formula(saved, _PAF_SUMMARY_SHEET_TITLE, coord)

    for coord in ("B4", "B5", "B6", "B7", "B9", "B10"):
        with pytest.raises(AssertionError, match="literal, not a live formula"):
            assert_cell_is_formula(saved, _EXECUTIVE_SUMMARY_SHEET_TITLE, coord)


def test_security_invariant_formula_injection_defense_copq() -> None:
    malicious_inputs = [
        ("=SUM(A1:A10)", "formula starting with ="),
        ("+1+1", "formula starting with +"),
        ("-1-1", "negative number prefix"),
        ("@SUM(B1:B5)", "formula starting with @"),
        ("\t=CALC()", "tab prefixed formula"),
        ("\r+2+2", "CR prefixed formula"),
    ]

    items = [
        CostItem(
            category="InternalFailure",
            description=payload,
            direct_cost=100.0,
        )
        for payload, _ in malicious_inputs
    ]

    dataset = COPQDataset(items=items, revenue_base=100000.0)
    malicious_title = "=DANGEROUS_COPQ_TITLE"
    wb_bytes = export_copq_workbook(dataset, title=malicious_title)
    wb = _load(wb_bytes)

    # Sheet tab title
    assert malicious_title in wb.sheetnames

    matrix = wb[malicious_title]
    for r in range(2, len(items) + 2):
        desc_cell = matrix[f"B{r}"].value
        assert str(desc_cell).startswith("'"), f"Description was not escaped: {desc_cell!r}"
        with pytest.raises(AssertionError, match="literal, not a live formula"):
            assert_cell_is_formula(wb_bytes, malicious_title, f"B{r}")

    # Computed line totals are STILL live formulas
    for r in range(2, len(items) + 2):
        assert_cell_is_formula(wb_bytes, malicious_title, f"F{r}")

    # Summary sheet title value must be escaped
    ws_exec = wb[_EXECUTIVE_SUMMARY_SHEET_TITLE]
    assert ws_exec["B1"].value == "'" + malicious_title

    # Summary formulas remain live
    for r in range(2, 7):
        assert_cell_is_formula(wb_bytes, _PAF_SUMMARY_SHEET_TITLE, f"C{r}")
        assert_cell_is_formula(wb_bytes, _PAF_SUMMARY_SHEET_TITLE, f"D{r}")

    for coord in ("B4", "B5", "B6", "B7", "B9", "B10"):
        assert_cell_is_formula(wb_bytes, _EXECUTIVE_SUMMARY_SHEET_TITLE, coord)


# ===========================================================================
# 7. Empty Dataset & Boundary Conditions
# ===========================================================================


def test_empty_dataset_exports_cleanly_without_division_errors_copq() -> None:
    """Empty dataset produces valid 3-sheet workbook with floored 2:2 ranges and zero ratios."""
    wb_bytes = export_copq_workbook([], revenue_base=100000.0)
    wb = _load(wb_bytes)

    matrix = wb[_DEFAULT_TITLE]
    assert matrix.max_row == 1

    ws_paf = wb[_PAF_SUMMARY_SHEET_TITLE]
    assert ws_paf["C2"].value == f'=SUMIF(\'{_DEFAULT_TITLE}\'!{_CAT_COL}2:{_CAT_COL}2, "Prevention", \'{_DEFAULT_TITLE}\'!{_TOTAL_COL}2:{_TOTAL_COL}2)'
    assert ws_paf["D2"].value == "=IF($C$6=0, 0, C2/$C$6)"

    ws_exec = wb[_EXECUTIVE_SUMMARY_SHEET_TITLE]
    assert ws_exec["B7"].value == "=IF(B6=0, 0, B5/B6)"
    assert ws_exec["B10"].value == f"=COUNTA('{_DEFAULT_TITLE}'!{_CAT_COL}2:{_CAT_COL}2)"

    # Live formulas pass verifier
    for r in range(2, 7):
        assert_cell_is_formula(wb_bytes, _PAF_SUMMARY_SHEET_TITLE, f"C{r}")
        assert_cell_is_formula(wb_bytes, _PAF_SUMMARY_SHEET_TITLE, f"D{r}")
    for coord in ("B4", "B5", "B6", "B7", "B9", "B10"):
        assert_cell_is_formula(wb_bytes, _EXECUTIVE_SUMMARY_SHEET_TITLE, coord)


# ===========================================================================
# 8. CostItem Drivers & Nullable Handling
# ===========================================================================


def test_cost_item_drivers_branches() -> None:
    # 1. scrap_qty & unit_cost
    item1 = CostItem(category="InternalFailure", description="Scrap", scrap_qty=20, unit_cost=50.0)
    assert _item_drivers(item1) == (20.0, 50.0, 0.0)

    # 2. rework_hours & labor_rate
    item2 = CostItem(category="InternalFailure", description="Rework", rework_hours=15.0, labor_rate=60.0)
    assert _item_drivers(item2) == (15.0, 60.0, 0.0)

    # 3. containment_hours & labor_rate
    item3 = CostItem(category="InternalFailure", description="Containment", containment_hours=10.0, labor_rate=40.0)
    assert _item_drivers(item3) == (10.0, 40.0, 0.0)

    # 4. warranty_units & warranty_unit_cost
    item4 = CostItem(category="ExternalFailure", description="Warranty", warranty_units=5, warranty_unit_cost=300.0)
    assert _item_drivers(item4) == (5.0, 300.0, 0.0)

    # 5. direct_cost only
    item5 = CostItem(category="Prevention", description="Audit", direct_cost=2500.0)
    assert _item_drivers(item5) == (0.0, 0.0, 2500.0)

    # 6. None direct cost
    item6 = CostItem(category="Prevention", description="Audit", direct_cost=None)
    assert _item_drivers(item6) == (0.0, 0.0, 0.0)


def test_write_metadata_sheet_without_title_copq() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    _write_metadata_sheet(ws, [("Metric", "Value")], title=None)
    assert ws["A1"].value == "Metric"
    assert ws["B1"].value == "Value"


# ===========================================================================
# 9. Input Polymorphism & Validation
# ===========================================================================


def test_input_polymorphism_copq() -> None:
    import pandas as pd

    item_dicts = [
        {"category": "Prevention", "description": "Training", "direct_cost": 1500.0},
        {"category": "Appraisal", "description": "Testing", "direct_cost": 2200.0},
    ]
    items = [CostItem(**d) for d in item_dicts]
    df = pd.DataFrame(item_dicts)
    dict_payload = {"items": [item_dicts[0]], "revenue_base": 400000.0}
    rows_dict_payload = {"rows": [item_dicts[0]]}

    # 1. Sequence of CostItem
    wb1 = _load(export_copq_workbook(items))
    assert wb1[_DEFAULT_TITLE].max_row == 3

    # 2. Sequence of dict
    wb2 = _load(export_copq_workbook(item_dicts))
    assert wb2[_DEFAULT_TITLE].max_row == 3

    # 3. pandas DataFrame
    wb3 = _load(export_copq_workbook(df))
    assert wb3[_DEFAULT_TITLE].max_row == 3

    # 4. dict with 'items'
    wb4 = _load(export_copq_workbook(dict_payload))
    assert wb4[_DEFAULT_TITLE].max_row == 2
    assert wb4[_EXECUTIVE_SUMMARY_SHEET_TITLE]["B8"].value == 400000.0

    # 5. dict with 'rows'
    wb5 = _load(export_copq_workbook(rows_dict_payload))
    assert wb5[_DEFAULT_TITLE].max_row == 2

    # 6. Empty sequence forms
    wb_empty_list = _load(export_copq_workbook([]))
    assert wb_empty_list[_DEFAULT_TITLE].max_row == 1

    wb_empty_df = _load(export_copq_workbook(pd.DataFrame()))
    assert wb_empty_df[_DEFAULT_TITLE].max_row == 1

    wb_empty_dict1 = _load(export_copq_workbook({"items": [], "revenue_base": 300000.0}))
    assert wb_empty_dict1[_DEFAULT_TITLE].max_row == 1
    assert wb_empty_dict1[_EXECUTIVE_SUMMARY_SHEET_TITLE]["B8"].value == 300000.0

    wb_empty_dict2 = _load(export_copq_workbook({"rows": []}))
    assert wb_empty_dict2[_DEFAULT_TITLE].max_row == 1

    wb_empty_dict3 = _load(export_copq_workbook({"items": [], "revenue_base": "invalid_str"}))
    assert wb_empty_dict3[_DEFAULT_TITLE].max_row == 1
    assert wb_empty_dict3[_EXECUTIVE_SUMMARY_SHEET_TITLE]["B8"].value == "N/A"


# ===========================================================================
# 10. Re-Export Integrity & Fresh Benchmarks
# ===========================================================================


def test_reexport_integrity_copq() -> None:
    import quality_core.copq as pkg
    import quality_core.copq.export as mod

    assert pkg.build_copq_workbook is mod.build_copq_workbook
    assert pkg.export_copq_workbook is mod.export_copq_workbook
    assert pkg.export_copq_excel is mod.export_copq_excel
    assert pkg.benchmark_copq_dataset is mod.benchmark_copq_dataset
    assert pkg.COPQ_LEDGER_COLUMNS is mod.COPQ_LEDGER_COLUMNS
    assert pkg.COPQ_COL_WIDTHS is mod.COPQ_COL_WIDTHS
    assert pkg.PAF_SUMMARY_COLUMNS is mod.PAF_SUMMARY_COLUMNS
    assert pkg.PAF_SUMMARY_COL_WIDTHS is mod.PAF_SUMMARY_COL_WIDTHS


def test_benchmark_dataset_returns_fresh_instances_copq() -> None:
    a = benchmark_copq_dataset()
    b = benchmark_copq_dataset()
    assert a is not b
    assert a.items is not b.items
    assert len(a.items) == 9

    categories_present = {item.category for item in a.items}
    assert categories_present == set(PAF_CATEGORY_VALUES)
