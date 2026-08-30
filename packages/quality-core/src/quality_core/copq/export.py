"""
quality_core/copq/export.py
Cost of Poor Quality (COPQ) live-formula Excel exporter.

Generates multi-sheet .xlsx workbooks for Cost of Poor Quality financial analysis
structured on the Prevention-Appraisal-Failure (PAF) model (Feigenbaum 1956, Juran 1951)
per ASQ Certified Six Sigma Green Belt (CSSGB) Body of Knowledge and CSSC Lean Six Sigma Manual (2018).

What ships in the workbook (three sheets, in this order):
  1. title (default "COPQ Ledger"): Table of cost items with live per-line formula
     =(C{r}*D{r})+E{r} for Line_Total.
  2. "PAF Category Summary": Prevention, Appraisal, Internal Failure, External Failure
     subtotals via =SUMIF, Total CoQ via =SUM(C2:C5), and %-of-CoQ via =IF($C$6=0, 0, C{r}/$C$6).
  3. "Executive Summary & Metadata": Executive KPI metrics with live cross-sheet
     formulas for CoQ, COPQ (=C4+C5), CoGQ (=C2+C3), Failure-to-Conformance ratio (=IF(B6=0, 0, B5/B6)),
     revenue base, COPQ % of revenue (=IF(B8=0, 0, (B5/B8)*100)), and total cost line items (=COUNTA).

Standards References:
- ASQ Certified Six Sigma Green Belt (CSSGB) Body of Knowledge.
- PAF (Prevention-Appraisal-Failure) Model: Feigenbaum (1956), Juran (1951).
- CSSC Lean Six Sigma Manual (2018): Cost of Quality financial categorization and COPQ % of sales metric.
"""

from __future__ import annotations

import io
import math
from collections.abc import Sequence
from typing import Any

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from quality_core.copq.schema import (
    COPQDataset,
    CostItem,
    validate_copq,
)
from quality_core.io.export import (
    Formula,
    now,
    sanitize_cell,
    write_formula_cell,
    write_table_sheet,
)

__all__ = [
    "COPQ_COL_WIDTHS",
    "COPQ_LEDGER_COLUMNS",
    "PAF_SUMMARY_COLUMNS",
    "PAF_SUMMARY_COL_WIDTHS",
    "benchmark_copq_dataset",
    "build_copq_workbook",
    "export_copq_excel",
    "export_copq_workbook",
]

# ===========================================================================
# Layout Configuration & Column Descriptors
# ===========================================================================

COPQ_LEDGER_COLUMNS: tuple[str, ...] = (
    "Category",
    "Description",
    "Quantity_or_Hours",
    "Unit_Rate",
    "Direct_Expense",
    "Line_Total",
)

COPQ_COL_WIDTHS: dict[str, float] = {
    "Category": 18.0,
    "Description": 38.0,
    "Quantity_or_Hours": 18.0,
    "Unit_Rate": 16.0,
    "Direct_Expense": 16.0,
    "Line_Total": 18.0,
}

PAF_SUMMARY_COLUMNS: tuple[str, ...] = (
    "PAF_Category",
    "Classification",
    "Subtotal_Cost",
    "Pct_of_Total_CoQ",
)

PAF_SUMMARY_COL_WIDTHS: dict[str, float] = {
    "PAF_Category": 22.0,
    "Classification": 18.0,
    "Subtotal_Cost": 18.0,
    "Pct_of_Total_CoQ": 18.0,
}

_FIRST_DATA_ROW = 2
_PAF_SUMMARY_SHEET_TITLE = "PAF Category Summary"
_EXECUTIVE_SUMMARY_SHEET_TITLE = "Executive Summary & Metadata"
_BOLD_FONT = Font(bold=True, size=10)
_NORMAL_FONT = Font(size=10)

_CAT_COL = get_column_letter(COPQ_LEDGER_COLUMNS.index("Category") + 1)
_DESC_COL = get_column_letter(COPQ_LEDGER_COLUMNS.index("Description") + 1)
_QTY_COL = get_column_letter(COPQ_LEDGER_COLUMNS.index("Quantity_or_Hours") + 1)
_RATE_COL = get_column_letter(COPQ_LEDGER_COLUMNS.index("Unit_Rate") + 1)
_DIRECT_COL = get_column_letter(COPQ_LEDGER_COLUMNS.index("Direct_Expense") + 1)
_TOTAL_COL = get_column_letter(COPQ_LEDGER_COLUMNS.index("Line_Total") + 1)


def _last_row(n_rows: int) -> int:
    """Return the last ledger sheet row the roll-up ranges should span.

    Floored at _FIRST_DATA_ROW (2) so empty datasets build valid 2:2 ranges.
    """
    return max(_FIRST_DATA_ROW + n_rows - 1, _FIRST_DATA_ROW)


def _item_drivers(item: CostItem) -> tuple[float, float, float]:
    """Extract (quantity_or_hours, unit_rate, direct_expense) from a CostItem."""
    if item.scrap_qty is not None and item.unit_cost is not None:
        qty = float(item.scrap_qty)
        rate = float(item.unit_cost)
    elif item.rework_hours is not None and item.labor_rate is not None:
        qty = float(item.rework_hours)
        rate = float(item.labor_rate)
    elif item.containment_hours is not None and item.labor_rate is not None:
        qty = float(item.containment_hours)
        rate = float(item.labor_rate)
    elif item.warranty_units is not None and item.warranty_unit_cost is not None:
        qty = float(item.warranty_units)
        rate = float(item.warranty_unit_cost)
    else:
        qty = 0.0
        rate = 0.0
    direct = float(item.direct_cost) if item.direct_cost is not None else 0.0
    return qty, rate, direct


def _row_record(item: CostItem, row_idx: int) -> dict[str, Any]:
    """Map one CostItem to its ledger dictionary with live formula for Line_Total."""
    qty, rate, direct = _item_drivers(item)
    formula = Formula(
        f"=({_QTY_COL}{row_idx}*{_RATE_COL}{row_idx})+{_DIRECT_COL}{row_idx}",
        number_format="0.00",
    )
    return {
        "Category": item.category,
        "Description": item.description,
        "Quantity_or_Hours": qty,
        "Unit_Rate": rate,
        "Direct_Expense": direct,
        "Line_Total": formula,
    }


def _write_paf_summary_sheet(ws: Any, ledger_title: str, row_count: int) -> None:
    """Write the PAF Category Summary sheet into ws."""
    last_row = _last_row(row_count)
    ledger = f"'{ledger_title}'!"

    paf_configs: list[tuple[str, str, int]] = [
        ("Prevention", "CoGQ", 2),
        ("Appraisal", "CoGQ", 3),
        ("InternalFailure", "COPQ", 4),
        ("ExternalFailure", "COPQ", 5),
    ]

    paf_rows: list[dict[str, Any]] = []
    for cat_name, classification, r_idx in paf_configs:
        subtotal = Formula(
            f'=SUMIF({ledger}{_CAT_COL}{_FIRST_DATA_ROW}:{_CAT_COL}{last_row}, "{cat_name}", {ledger}{_TOTAL_COL}{_FIRST_DATA_ROW}:{_TOTAL_COL}{last_row})',
            number_format="0.00",
        )
        pct_coq = Formula(f"=IF($C$6=0, 0, C{r_idx}/$C$6)", number_format="0.00%")
        paf_rows.append(
            {
                "PAF_Category": cat_name,
                "Classification": classification,
                "Subtotal_Cost": subtotal,
                "Pct_of_Total_CoQ": pct_coq,
            }
        )

    # Row 6: Total Cost of Quality
    paf_rows.append(
        {
            "PAF_Category": "Total Cost of Quality",
            "Classification": "Total",
            "Subtotal_Cost": Formula("=SUM(C2:C5)", number_format="0.00"),
            "Pct_of_Total_CoQ": Formula("=IF(C6=0, 0, SUM(D2:D5))", number_format="0.00%"),
        }
    )

    df_paf = pd.DataFrame(paf_rows, columns=list(PAF_SUMMARY_COLUMNS))
    write_table_sheet(
        ws,
        df_paf,
        title=_PAF_SUMMARY_SHEET_TITLE,
        columns=PAF_SUMMARY_COLUMNS,
        col_widths=PAF_SUMMARY_COL_WIDTHS,
    )


def _write_metadata_sheet(
    ws: Any,
    rows: Sequence[tuple[str, object]],
    *,
    title: str | None = None,
    key_width: float = 38.0,
    value_width: float = 48.0,
) -> None:
    """Write a two-column metadata sheet supporting live Formula instances."""
    if title is not None:
        ws.title = title
    for r_idx, (label, val) in enumerate(rows, start=1):
        ws.cell(r_idx, 1, sanitize_cell(label)).font = _BOLD_FONT
        if isinstance(val, Formula):
            cell = write_formula_cell(
                ws, r_idx, 2, val.formula, number_format=val.number_format
            )
            cell.font = _NORMAL_FONT
        else:
            cell = ws.cell(r_idx, 2, sanitize_cell(val))
            cell.font = _NORMAL_FONT
    ws.column_dimensions["A"].width = key_width
    ws.column_dimensions["B"].width = value_width


def _write_executive_summary_sheet(
    ws: Any,
    ledger_title: str,
    row_count: int,
    revenue_base: float | None,
    report_title: str,
) -> None:
    """Write the Executive Summary & Metadata sheet into ws."""
    last_row = _last_row(row_count)
    ledger = f"'{ledger_title}'!"
    paf_sheet = f"'{_PAF_SUMMARY_SHEET_TITLE}'!"

    coq_formula = Formula(f"={paf_sheet}C6", number_format="0.00")
    copq_formula = Formula(f"={paf_sheet}C4+{paf_sheet}C5", number_format="0.00")
    cogq_formula = Formula(f"={paf_sheet}C2+{paf_sheet}C3", number_format="0.00")
    ratio_formula = Formula("=IF(B6=0, 0, B5/B6)", number_format="0.00")
    items_formula = Formula(f"=COUNTA({ledger}{_CAT_COL}{_FIRST_DATA_ROW}:{_CAT_COL}{last_row})")

    revenue_val: object = revenue_base if revenue_base is not None else "N/A"
    copq_pct_rev_formula: object = (
        Formula("=(B5/B8)*100", number_format="0.00%") if revenue_base is not None else "N/A"
    )

    rows: list[tuple[str, object]] = [
        ("Report Title", report_title),
        ("Date Generated", now()),
        ("Standards Basis", "ASQ CSSGB BoK / PAF Model (Feigenbaum & Juran)"),
        ("Total Cost of Quality (CoQ)", coq_formula),
        ("Cost of Poor Quality (COPQ)", copq_formula),
        ("Cost of Good Quality (CoGQ)", cogq_formula),
        ("COPQ / CoGQ Failure-to-Conformance Ratio", ratio_formula),
        ("Revenue Base", revenue_val),
        ("COPQ % of Revenue", copq_pct_rev_formula),
        ("Total Cost Line Items", items_formula),
    ]

    _write_metadata_sheet(
        ws,
        rows,
        title=_EXECUTIVE_SUMMARY_SHEET_TITLE,
        key_width=38.0,
        value_width=48.0,
    )


# ===========================================================================
# Public Exporter API
# ===========================================================================


def build_copq_workbook(
    dataset: COPQDataset | Sequence[CostItem] | Sequence[dict[str, Any]] | pd.DataFrame | dict[str, Any],
    *,
    revenue_base: float | None = None,
    title: str = "COPQ Ledger",
) -> openpyxl.Workbook:
    """Build a 3-sheet openpyxl Workbook for COPQ analysis with live formulas.

    Sheets:
      1. title (default: "COPQ Ledger"): Table of cost items with live per-line
         formula =(C{r}*D{r})+E{r} for Line_Total.
      2. "PAF Category Summary": Prevention, Appraisal, Internal Failure, External Failure
         subtotals via =SUMIF, Total CoQ via =SUM(C2:C5), and %-of-CoQ via =IF($C$6=0, 0, C{r}/$C$6).
      3. "Executive Summary & Metadata": Executive KPI metrics with live cross-sheet
         formulas for CoQ, COPQ (=C4+C5), CoGQ (=C2+C3), COPQ/CoGQ ratio, and % of revenue.
    """
    # Validate revenue_base if supplied
    effective_revenue_base: float | None = revenue_base
    if effective_revenue_base is not None:
        if isinstance(effective_revenue_base, bool):
            raise TypeError("revenue_base cannot be a boolean")
        if not isinstance(effective_revenue_base, (int, float)):
            raise TypeError(f"revenue_base must be a number or None, got {type(effective_revenue_base).__name__}")
        f_rev = float(effective_revenue_base)
        if math.isnan(f_rev) or math.isinf(f_rev) or f_rev <= 0.0:
            raise ValueError(f"revenue_base must be a positive finite number, got {f_rev}")
        effective_revenue_base = f_rev

    if isinstance(dataset, COPQDataset):
        items = dataset.items
        if effective_revenue_base is None and dataset.revenue_base is not None:
            effective_revenue_base = dataset.revenue_base
    elif isinstance(dataset, (list, tuple)) and len(dataset) == 0:
        items = []
    elif isinstance(dataset, pd.DataFrame) and len(dataset) == 0:
        items = []
    elif isinstance(dataset, dict) and (dataset.get("items") == [] or dataset.get("rows") == []):
        items = []
        if effective_revenue_base is None and dataset.get("revenue_base") is not None:
            raw_rev = dataset.get("revenue_base")
            if isinstance(raw_rev, (int, float)) and not isinstance(raw_rev, bool) and raw_rev > 0:
                effective_revenue_base = float(raw_rev)
    else:
        validated = validate_copq(dataset, revenue_base=effective_revenue_base)
        items = validated.items
        if effective_revenue_base is None and validated.revenue_base is not None:
            effective_revenue_base = validated.revenue_base

    wb = Workbook()

    # Sheet 1: Ledger
    item_dicts = [_row_record(item, row_idx=idx + _FIRST_DATA_ROW) for idx, item in enumerate(items)]
    df = pd.DataFrame(item_dicts, columns=list(COPQ_LEDGER_COLUMNS))
    write_table_sheet(
        wb.active,
        df,
        title=title,
        columns=COPQ_LEDGER_COLUMNS,
        col_widths=COPQ_COL_WIDTHS,
    )

    # Sheet 2: PAF Category Summary
    ws_paf = wb.create_sheet(_PAF_SUMMARY_SHEET_TITLE)
    _write_paf_summary_sheet(ws_paf, title, len(items))

    # Sheet 3: Executive Summary & Metadata
    ws_exec = wb.create_sheet(_EXECUTIVE_SUMMARY_SHEET_TITLE)
    _write_executive_summary_sheet(ws_exec, title, len(items), effective_revenue_base, title)

    return wb


def export_copq_workbook(
    dataset: COPQDataset | Sequence[CostItem] | Sequence[dict[str, Any]] | pd.DataFrame | dict[str, Any],
    *,
    revenue_base: float | None = None,
    title: str = "COPQ Ledger",
) -> bytes:
    """Export COPQ dataset to serialized .xlsx bytes."""
    wb = build_copq_workbook(dataset, revenue_base=revenue_base, title=title)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_copq_excel(
    dataset: COPQDataset | Sequence[CostItem] | Sequence[dict[str, Any]] | pd.DataFrame | dict[str, Any],
    *,
    revenue_base: float | None = None,
    title: str = "COPQ Ledger",
) -> bytes:
    """Alias for export_copq_workbook."""
    return export_copq_workbook(dataset, revenue_base=revenue_base, title=title)


# ===========================================================================
# Benchmark Dataset
# ===========================================================================


def benchmark_copq_dataset(revenue_base: float | None = 500000.0) -> COPQDataset:
    """Return a fresh 9-item benchmark COPQDataset with revenue base."""
    items = [
        CostItem(
            category="Prevention",
            description="APQP Quality Planning & DFM Design Review",
            direct_cost=4500.0,
        ),
        CostItem(
            category="Prevention",
            description="Operator Error-Proofing / Poka-Yoke Assembly Training",
            direct_cost=2800.0,
        ),
        CostItem(
            category="Appraisal",
            description="Receiving CMM Dimensional Verification & Metallurgical Testing",
            direct_cost=6200.0,
        ),
        CostItem(
            category="Appraisal",
            description="In-Process Automated Optical Inspection (AOI) Station Audits",
            direct_cost=5100.0,
        ),
        CostItem(
            category="InternalFailure",
            description="Machined Bore Casting Porosity Scrap (45 pcs)",
            scrap_qty=45,
            unit_cost=120.0,
        ),
        CostItem(
            category="InternalFailure",
            description="Connecting Rod Undersized Journal Rework (35 hrs labor)",
            rework_hours=35.0,
            labor_rate=65.0,
            direct_cost=450.0,
        ),
        CostItem(
            category="InternalFailure",
            description="Plant Containment Sorting & Gauge Re-inspection (40 hrs)",
            containment_hours=40.0,
            labor_rate=45.0,
        ),
        CostItem(
            category="ExternalFailure",
            description="Customer Field Warranty Claims & Replacement Assembly (12 units)",
            warranty_units=12,
            warranty_unit_cost=850.0,
        ),
        CostItem(
            category="ExternalFailure",
            description="Customer Returned Defective Batch Logistics & Restocking Loss",
            direct_cost=3600.0,
        ),
    ]
    return COPQDataset(items=items, revenue_base=revenue_base)
