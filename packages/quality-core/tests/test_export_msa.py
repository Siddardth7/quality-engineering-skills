"""Tests for quality_core/msa/export.py — Live-formula Excel exporter for MSA Gage R&R.

Covers:
- Live formula assertions (OOXML <f> tags) on Gage R&R Summary and ANOVA Table sheets.
- Negative controls:
  - Mutating a computed cell to a literal value fails assert_cell_is_formula.
  - Untrusted strings (part/appraiser names, study titles) are safely sanitized
    with leading apostrophe and do NOT become live formulas.
- Accuracy scorecard: formula evaluation equivalence with compute_gage_rr outputs.
- Method dispatch: Average-and-Range vs ANOVA (with interaction vs pooled).
- Tolerance conditioning: tolerance=None vs tolerance=float.
- Input polymorphism: pd.DataFrame, list[dict], data=None with results dict.
- Error paths: both data and results None, invalid method, invalid/non-positive/nan/inf tolerance.
- Re-exports from quality_core.msa and quality_core.msa.export.
"""

from __future__ import annotations

import io
import math
import zipfile
from typing import Any

import openpyxl
import pandas as pd
import pytest
from _xlsx_formula_audit import assert_cell_is_formula
from quality_core.canvas.msa import SAMPLE_MSA_STUDY_DATA
from quality_core.msa import (
    METHOD,
    METHOD_ANOVA,
    build_msa_workbook,
    compute_gage_rr,
    export_msa_workbook,
)
from quality_core.msa.export import (
    build_msa_workbook as export_build,
)
from quality_core.msa.export import (
    export_msa_workbook as export_export,
)


def _saved(wb: openpyxl.Workbook) -> bytes:
    """Serialize an openpyxl Workbook to in-memory bytes."""
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def sample_study_df() -> pd.DataFrame:
    """Standard AIAG MSA 4th Edition benchmark dataset (10 parts x 3 appraisers x 3 trials)."""
    return pd.DataFrame(SAMPLE_MSA_STUDY_DATA)


@pytest.fixture
def sample_study_dicts() -> list[dict[str, Any]]:
    """List of dicts representation of the benchmark dataset."""
    return list(SAMPLE_MSA_STUDY_DATA)


@pytest.fixture
def interaction_anova_study_df() -> pd.DataFrame:
    """Dataset with genuine part x appraiser interaction (significant at alpha=0.05)."""
    return pd.DataFrame([
        {
            "part": f"P{p}",
            "appraiser": a,
            "trial": t,
            "measurement": (
                float(p)
                + (0.01 if t == 2 else 0.0)
                + (0.02 if t == 3 else 0.0)
                + (1.2 if (p == 3 and a == "C") else 0.0)
            ),
        }
        for p in range(1, 7)
        for a in ["A", "B", "C"]
        for t in [1, 2, 3]
    ])


@pytest.fixture
def pooled_anova_study_df() -> pd.DataFrame:
    """Synthetic dataset where part x appraiser interaction is not significant (pooled)."""
    rows = []
    for part in [f"P{i:02d}" for i in range(1, 6)]:
        part_base = float(int(part[1:])) * 5.0
        for app in ["A", "B", "C"]:
            app_bias = {"A": 0.0, "B": 0.05, "C": -0.05}[app]
            for trial in [1, 2, 3]:
                # Pure additive model with no interaction
                val = part_base + app_bias + (trial * 0.001)
                rows.append({
                    "part": part,
                    "appraiser": app,
                    "trial": trial,
                    "measurement": val,
                })
    return pd.DataFrame(rows)


# =============================================================================
# 1. Live Formula Verification (assert_cell_is_formula)
# =============================================================================


def test_summary_sheet_live_formulas_average_and_range(sample_study_df: pd.DataFrame) -> None:
    """Verify all computed variance component cells in Average-and-Range hold live formulas."""
    raw_bytes = export_msa_workbook(
        sample_study_df,
        tolerance=4.42,
        method=METHOD,
    )

    sheet = "Gage R&R Summary"
    # Row 10: ndc live formula
    assert_cell_is_formula(raw_bytes, sheet, "B10")

    # In Average-and-Range:
    # Row 13: EV (6*SD, %SV, %Tol)
    assert_cell_is_formula(raw_bytes, sheet, "C13")
    assert_cell_is_formula(raw_bytes, sheet, "D13")
    assert_cell_is_formula(raw_bytes, sheet, "E13")

    # Row 14: AV (6*SD, %SV, %Tol)
    assert_cell_is_formula(raw_bytes, sheet, "C14")
    assert_cell_is_formula(raw_bytes, sheet, "D14")
    assert_cell_is_formula(raw_bytes, sheet, "E14")

    # Row 15: GRR (SD, 6*SD, %SV, %Tol)
    assert_cell_is_formula(raw_bytes, sheet, "B15")
    assert_cell_is_formula(raw_bytes, sheet, "C15")
    assert_cell_is_formula(raw_bytes, sheet, "D15")
    assert_cell_is_formula(raw_bytes, sheet, "E15")

    # Row 16: PV (6*SD, %SV, %Tol)
    assert_cell_is_formula(raw_bytes, sheet, "C16")
    assert_cell_is_formula(raw_bytes, sheet, "D16")
    assert_cell_is_formula(raw_bytes, sheet, "E16")

    # Row 17: TV (SD, 6*SD, %SV, %Tol)
    assert_cell_is_formula(raw_bytes, sheet, "B17")
    assert_cell_is_formula(raw_bytes, sheet, "C17")
    assert_cell_is_formula(raw_bytes, sheet, "D17")
    assert_cell_is_formula(raw_bytes, sheet, "E17")


def test_summary_sheet_live_formulas_anova(interaction_anova_study_df: pd.DataFrame) -> None:
    """Verify all computed variance component cells in ANOVA hold live formulas."""
    raw_bytes = export_msa_workbook(
        interaction_anova_study_df,
        tolerance=4.42,
        method=METHOD_ANOVA,
    )

    sheet = "Gage R&R Summary"
    # Row 10: ndc live formula
    assert_cell_is_formula(raw_bytes, sheet, "B10")

    # In ANOVA:
    # Row 13: EV
    assert_cell_is_formula(raw_bytes, sheet, "C13")
    assert_cell_is_formula(raw_bytes, sheet, "D13")
    assert_cell_is_formula(raw_bytes, sheet, "E13")

    # Row 14: AV
    assert_cell_is_formula(raw_bytes, sheet, "C14")
    assert_cell_is_formula(raw_bytes, sheet, "D14")
    assert_cell_is_formula(raw_bytes, sheet, "E14")

    # Row 15: Interaction (INT)
    assert_cell_is_formula(raw_bytes, sheet, "C15")
    assert_cell_is_formula(raw_bytes, sheet, "D15")
    assert_cell_is_formula(raw_bytes, sheet, "E15")

    # Row 16: GRR (SD, 6*SD, %SV, %Tol)
    assert_cell_is_formula(raw_bytes, sheet, "B16")
    assert_cell_is_formula(raw_bytes, sheet, "C16")
    assert_cell_is_formula(raw_bytes, sheet, "D16")
    assert_cell_is_formula(raw_bytes, sheet, "E16")

    # Row 17: PV (6*SD, %SV, %Tol)
    assert_cell_is_formula(raw_bytes, sheet, "C17")
    assert_cell_is_formula(raw_bytes, sheet, "D17")
    assert_cell_is_formula(raw_bytes, sheet, "E17")

    # Row 18: TV (SD, 6*SD, %SV, %Tol)
    assert_cell_is_formula(raw_bytes, sheet, "B18")
    assert_cell_is_formula(raw_bytes, sheet, "C18")
    assert_cell_is_formula(raw_bytes, sheet, "D18")
    assert_cell_is_formula(raw_bytes, sheet, "E18")


def test_anova_sheet_live_formulas(sample_study_df: pd.DataFrame) -> None:
    """Verify live formulas on the ANOVA Table sheet."""
    raw_bytes = export_msa_workbook(
        sample_study_df,
        tolerance=4.42,
        method=METHOD_ANOVA,
    )

    sheet = "ANOVA Table"
    # Row 2 (Part): MS formula
    assert_cell_is_formula(raw_bytes, sheet, "D2")
    # Row 3 (Appraiser): MS formula
    assert_cell_is_formula(raw_bytes, sheet, "D3")
    # Row 4 (Interaction): MS formula and F-Statistic formula
    assert_cell_is_formula(raw_bytes, sheet, "D4")
    assert_cell_is_formula(raw_bytes, sheet, "E4")
    # Row 5 (Equipment / Error): MS formula
    assert_cell_is_formula(raw_bytes, sheet, "D5")
    # Row 6 (Total): DF Total SUM and SS Total SUM formulas
    assert_cell_is_formula(raw_bytes, sheet, "B6")
    assert_cell_is_formula(raw_bytes, sheet, "C6")


# =============================================================================
# 2. Negative Controls (Load-Bearing Proof)
# =============================================================================


def test_negative_control_hardcoded_literal_fails_formula_audit(sample_study_df: pd.DataFrame) -> None:
    """NEGATIVE CONTROL: mutating a computed cell to a literal constant must fail assert_cell_is_formula."""
    wb = build_msa_workbook(sample_study_df, tolerance=4.42, method=METHOD)
    ws = wb["Gage R&R Summary"]

    # Deliberately replace the live formula with a hardcoded float literal
    ws.cell(row=10, column=2, value=4.0)
    ws.cell(row=13, column=4, value=14.28)

    saved = _saved(wb)

    with pytest.raises(AssertionError, match="is a literal, not a live formula"):
        assert_cell_is_formula(saved, "Gage R&R Summary", "B10")

    with pytest.raises(AssertionError, match="is a literal, not a live formula"):
        assert_cell_is_formula(saved, "Gage R&R Summary", "D13")


def test_negative_control_untrusted_input_sanitization_is_inert() -> None:
    """NEGATIVE CONTROL: formula injection strings in titles, part IDs, and appraiser names are sanitized.

    Proves they are escaped with a leading apostrophe and do NOT become live OOXML <f> elements.
    """
    malicious_data = [
        {"part": "=cmd|' /C calc'!A0", "appraiser": "+evil_user", "trial": 1, "measurement": 10.1},
        {"part": "=cmd|' /C calc'!A0", "appraiser": "+evil_user", "trial": 2, "measurement": 10.2},
        {"part": "P02", "appraiser": "-danger_appraiser", "trial": 1, "measurement": 12.1},
        {"part": "P02", "appraiser": "-danger_appraiser", "trial": 2, "measurement": 12.3},
    ]

    wb = build_msa_workbook(
        malicious_data,
        tolerance=2.0,
        method=METHOD,
        title="=SUM(A1:A10)",
    )

    # 1. Summary sheet metadata escaping
    ws_summary = wb["Gage R&R Summary"]
    assert ws_summary.cell(1, 2).value == "'=SUM(A1:A10)"

    # 2. Measurements sheet cell escaping
    ws_meas = wb["Measurements"]
    assert ws_meas.cell(2, 1).value == "'=cmd|' /C calc'!A0"
    assert ws_meas.cell(2, 2).value == "'+evil_user"
    assert ws_meas.cell(4, 2).value == "'-danger_appraiser"

    # 3. Assert assert_cell_is_formula raises on sanitized text cells
    saved = _saved(wb)
    with pytest.raises(AssertionError, match="is a literal, not a live formula"):
        assert_cell_is_formula(saved, "Gage R&R Summary", "B1")

    with pytest.raises(AssertionError, match="is a literal, not a live formula"):
        assert_cell_is_formula(saved, "Measurements", "A2")

    with pytest.raises(AssertionError, match="is a literal, not a live formula"):
        assert_cell_is_formula(saved, "Measurements", "B2")

    with pytest.raises(AssertionError, match="is a literal, not a live formula"):
        assert_cell_is_formula(saved, "Measurements", "B4")


# =============================================================================
# 3. Accuracy Scorecard (Formula Evaluation vs compute_gage_rr)
# =============================================================================


def test_accuracy_scorecard_average_and_range(sample_study_df: pd.DataFrame) -> None:
    """Evaluate live formula expressions against compute_gage_rr reference outputs for Average-and-Range."""
    tolerance = 4.42
    results = compute_gage_rr(sample_study_df, tolerance=tolerance, method=METHOD)
    wb = build_msa_workbook(sample_study_df, tolerance=tolerance, method=METHOD)
    ws = wb["Gage R&R Summary"]

    ev_sd = float(ws.cell(13, 2).value)
    av_sd = float(ws.cell(14, 2).value)
    pv_sd = float(ws.cell(16, 2).value)

    # Formula for GRR SD in B15: SQRT(B13^2 + B14^2)
    assert ws.cell(15, 2).value == "=SQRT(B13^2 + B14^2)"
    eval_grr = math.sqrt(ev_sd**2 + av_sd**2)
    assert eval_grr == pytest.approx(results["grr"], rel=1e-5)

    # Formula for TV SD in B17: SQRT(B15^2 + B16^2)
    assert ws.cell(17, 2).value == "=SQRT(B15^2 + B16^2)"
    eval_tv = math.sqrt(eval_grr**2 + pv_sd**2)
    assert eval_tv == pytest.approx(results["tv"], rel=1e-5)

    # Formula for ndc in B10: MAX(1, INT(1.41 * (B16 / B15)))
    assert ws.cell(10, 2).value == "=MAX(1, INT(1.41 * (B16 / B15)))"
    eval_ndc = max(1, int(1.41 * (pv_sd / eval_grr)))
    assert eval_ndc == results["ndc"]

    # 6*SD formulas
    assert ws.cell(13, 3).value == "=B13*6"
    assert ws.cell(14, 3).value == "=B14*6"
    assert ws.cell(15, 3).value == "=B15*6"
    assert ws.cell(16, 3).value == "=B16*6"
    assert ws.cell(17, 3).value == "=B17*6"

    # %SV formulas: (B{row}/$B$17)*100
    assert ws.cell(13, 4).value == "=(B13/$B$17)*100"
    assert ws.cell(14, 4).value == "=(B14/$B$17)*100"
    assert ws.cell(15, 4).value == "=(B15/$B$17)*100"
    assert ws.cell(16, 4).value == "=(B16/$B$17)*100"
    assert ws.cell(17, 4).value == "=(B17/$B$17)*100"

    assert (ev_sd / eval_tv) * 100 == pytest.approx(results["pev_study"], rel=1e-5)
    assert (av_sd / eval_tv) * 100 == pytest.approx(results["pav_study"], rel=1e-5)
    assert (eval_grr / eval_tv) * 100 == pytest.approx(results["pgrr_study"], rel=1e-5)
    assert (pv_sd / eval_tv) * 100 == pytest.approx(results["ppv_study"], rel=1e-5)

    # %Tol formulas: (C{row}/$B$8)*100
    assert ws.cell(13, 5).value == "=(C13/$B$8)*100"
    assert ws.cell(14, 5).value == "=(C14/$B$8)*100"
    assert ws.cell(15, 5).value == "=(C15/$B$8)*100"
    assert ws.cell(16, 5).value == "=(C16/$B$8)*100"
    assert ws.cell(17, 5).value == "=(C17/$B$8)*100"

    assert (ev_sd * 6 / tolerance) * 100 == pytest.approx(results["pev_tolerance"], rel=1e-5)
    assert (av_sd * 6 / tolerance) * 100 == pytest.approx(results["pav_tolerance"], rel=1e-5)
    assert (eval_grr * 6 / tolerance) * 100 == pytest.approx(results["pgrr_tolerance"], rel=1e-5)
    assert (pv_sd * 6 / tolerance) * 100 == pytest.approx(results["ppv_tolerance"], rel=1e-5)


def test_accuracy_scorecard_anova(interaction_anova_study_df: pd.DataFrame) -> None:
    """Evaluate live formula expressions against compute_gage_rr reference outputs for ANOVA."""
    tolerance = 4.42
    results = compute_gage_rr(interaction_anova_study_df, tolerance=tolerance, method=METHOD_ANOVA)
    wb = build_msa_workbook(interaction_anova_study_df, tolerance=tolerance, method=METHOD_ANOVA)
    ws = wb["Gage R&R Summary"]

    ev_sd = float(ws.cell(13, 2).value)
    av_sd = float(ws.cell(14, 2).value)
    int_sd = float(ws.cell(15, 2).value)
    pv_sd = float(ws.cell(17, 2).value)

    # GRR SD formula in B16: SQRT(B13^2 + B14^2 + B15^2)
    assert ws.cell(16, 2).value == "=SQRT(B13^2 + B14^2 + B15^2)"
    eval_grr = math.sqrt(ev_sd**2 + av_sd**2 + int_sd**2)
    assert eval_grr == pytest.approx(results["grr"], rel=1e-5)

    # TV SD formula in B18: SQRT(B16^2 + B17^2)
    assert ws.cell(18, 2).value == "=SQRT(B16^2 + B17^2)"
    eval_tv = math.sqrt(eval_grr**2 + pv_sd**2)
    assert eval_tv == pytest.approx(results["tv"], rel=1e-5)

    # ndc formula in B10: MAX(1, INT(1.41 * (B17 / B16)))
    assert ws.cell(10, 2).value == "=MAX(1, INT(1.41 * (B17 / B16)))"
    eval_ndc = max(1, int(1.41 * (pv_sd / eval_grr)))
    assert eval_ndc == results["ndc"]

    # %SV formulas: (B{row}/$B$18)*100
    assert ws.cell(13, 4).value == "=(B13/$B$18)*100"
    assert ws.cell(14, 4).value == "=(B14/$B$18)*100"
    assert ws.cell(15, 4).value == "=(B15/$B$18)*100"
    assert ws.cell(16, 4).value == "=(B16/$B$18)*100"
    assert ws.cell(17, 4).value == "=(B17/$B$18)*100"
    assert ws.cell(18, 4).value == "=(B18/$B$18)*100"

    assert (ev_sd / eval_tv) * 100 == pytest.approx(results["pev_study"], rel=1e-5)
    assert (av_sd / eval_tv) * 100 == pytest.approx(results["pav_study"], rel=1e-5)
    assert (eval_grr / eval_tv) * 100 == pytest.approx(results["pgrr_study"], rel=1e-5)
    assert (pv_sd / eval_tv) * 100 == pytest.approx(results["ppv_study"], rel=1e-5)

    # %Tol formulas: (C{row}/$B$8)*100
    assert ws.cell(13, 5).value == "=(C13/$B$8)*100"
    assert ws.cell(14, 5).value == "=(C14/$B$8)*100"
    assert ws.cell(15, 5).value == "=(C15/$B$8)*100"
    assert ws.cell(16, 5).value == "=(C16/$B$8)*100"
    assert ws.cell(17, 5).value == "=(C17/$B$8)*100"
    assert ws.cell(18, 5).value == "=(C18/$B$8)*100"

    assert (ev_sd * 6 / tolerance) * 100 == pytest.approx(results["pev_tolerance"], rel=1e-5)
    assert (av_sd * 6 / tolerance) * 100 == pytest.approx(results["pav_tolerance"], rel=1e-5)
    assert (eval_grr * 6 / tolerance) * 100 == pytest.approx(results["pgrr_tolerance"], rel=1e-5)
    assert (pv_sd * 6 / tolerance) * 100 == pytest.approx(results["ppv_tolerance"], rel=1e-5)

    # Check ANOVA table formulas
    ws_anova = wb["ANOVA Table"]
    assert ws_anova.cell(2, 4).value == "=C2/B2"
    assert ws_anova.cell(3, 4).value == "=C3/B3"
    assert ws_anova.cell(4, 4).value == "=C4/B4"
    assert ws_anova.cell(4, 5).value == "=D4/D5"
    assert ws_anova.cell(5, 4).value == "=C5/B5"
    assert ws_anova.cell(6, 2).value == "=SUM(B2:B5)"
    assert ws_anova.cell(6, 3).value == "=SUM(C2:C5)"


# =============================================================================
# 4. Method Dispatch & Sheet Layout
# =============================================================================


def test_method_dispatch_average_and_range(sample_study_df: pd.DataFrame) -> None:
    """Average-and-Range produces a 2-sheet workbook with expected metadata and layout."""
    wb = build_msa_workbook(sample_study_df, method=METHOD)
    assert wb.sheetnames == ["Gage R&R Summary", "Measurements"]

    ws = wb["Gage R&R Summary"]
    assert ws.cell(4, 2).value == "Average and Range"
    assert ws.cell(13, 1).value == "Repeatability (EV)"
    assert ws.cell(14, 1).value == "Reproducibility (AV)"
    assert ws.cell(15, 1).value == "Total Gage R&R (GRR)"
    assert ws.cell(16, 1).value == "Part Variation (PV)"
    assert ws.cell(17, 1).value == "Total Variation (TV)"


def test_method_dispatch_anova_significant_interaction(interaction_anova_study_df: pd.DataFrame) -> None:
    """ANOVA with significant interaction produces a 3-sheet workbook with 'Yes' significance."""
    wb = build_msa_workbook(interaction_anova_study_df, method=METHOD_ANOVA)
    assert wb.sheetnames == ["Gage R&R Summary", "Measurements", "ANOVA Table"]

    ws_sum = wb["Gage R&R Summary"]
    assert ws_sum.cell(4, 2).value == "ANOVA (Crossed Two-Factor with Replication)"
    assert ws_sum.cell(15, 1).value == "Part × Appraiser Interaction (INT)"
    assert ws_sum.cell(16, 1).value == "Total Gage R&R (GRR)"

    ws_anova = wb["ANOVA Table"]
    # Interaction row is row 4, column 6 is 'Significant (alpha=0.05)'
    assert ws_anova.cell(4, 6).value == "Yes"


def test_method_dispatch_anova_pooled_interaction(pooled_anova_study_df: pd.DataFrame) -> None:
    """ANOVA with non-significant interaction displays 'No (Pooled)' and zero interaction."""
    wb = build_msa_workbook(pooled_anova_study_df, method=METHOD_ANOVA)
    assert wb.sheetnames == ["Gage R&R Summary", "Measurements", "ANOVA Table"]

    ws_sum = wb["Gage R&R Summary"]
    # Interaction value reported as 0.0
    assert ws_sum.cell(15, 2).value == 0.0

    ws_anova = wb["ANOVA Table"]
    assert ws_anova.cell(4, 6).value == "No (Pooled)"


def test_method_dispatch_anova_sig_text_empty_when_interaction_sig_none() -> None:
    """When interaction_significant is None or missing from results, sig_text defaults to empty string."""
    results_mock = {
        "method": METHOD_ANOVA,
        "n_parts": 5,
        "n_appraisers": 2,
        "n_trials": 2,
        "ev": 0.1,
        "av": 0.1,
        "interaction": 0.0,
        "grr": 0.1414,
        "pv": 0.5,
        "tv": 0.5196,
        "interaction_significant": None,
        "verdict": "Accept",
    }
    dummy_data = pd.DataFrame([
        {"part": "P1", "appraiser": "A", "trial": 1, "measurement": 1.0},
        {"part": "P1", "appraiser": "A", "trial": 2, "measurement": 1.1},
        {"part": "P1", "appraiser": "B", "trial": 1, "measurement": 1.0},
        {"part": "P1", "appraiser": "B", "trial": 2, "measurement": 1.1},
        {"part": "P2", "appraiser": "A", "trial": 1, "measurement": 2.0},
        {"part": "P2", "appraiser": "A", "trial": 2, "measurement": 2.1},
        {"part": "P2", "appraiser": "B", "trial": 1, "measurement": 2.0},
        {"part": "P2", "appraiser": "B", "trial": 2, "measurement": 2.1},
    ])
    wb = build_msa_workbook(dummy_data, results=results_mock, method=METHOD_ANOVA)
    ws_anova = wb["ANOVA Table"]
    assert ws_anova.cell(4, 6).value == ""


# =============================================================================
# 5. Tolerance Conditioning
# =============================================================================


def test_tolerance_none_omits_tol_column(sample_study_df: pd.DataFrame) -> None:
    """When tolerance=None, % Tolerance column is omitted and metadata shows 'N/A'."""
    wb = build_msa_workbook(sample_study_df, tolerance=None, method=METHOD)
    ws = wb["Gage R&R Summary"]

    # Metadata row 8
    assert ws.cell(8, 1).value == "Tolerance"
    assert ws.cell(8, 2).value == "N/A"

    # Headers in row 12 have 4 columns
    headers = [ws.cell(12, c).value for c in range(1, 6) if ws.cell(12, c).value is not None]
    assert headers == [
        "Source",
        "Standard Deviation (SD)",
        "Study Variation (6×SD)",
        "% Study Variation (%SV)",
    ]

    # Column E cells are empty/None
    assert ws.cell(13, 5).value is None
    assert ws.cell(17, 5).value is None


def test_tolerance_provided_includes_tol_column_and_formulas(sample_study_df: pd.DataFrame) -> None:
    """When tolerance is provided, % Tolerance column is present with live formulas."""
    tolerance = 5.0
    wb = build_msa_workbook(sample_study_df, tolerance=tolerance, method=METHOD)
    ws = wb["Gage R&R Summary"]

    # Metadata row 8
    assert ws.cell(8, 1).value == "Tolerance"
    assert ws.cell(8, 2).value == 5.0
    assert ws.cell(8, 2).number_format == "0.0000"

    # Header in row 12 includes % Tolerance
    assert ws.cell(12, 5).value == "% Tolerance (%Tol)"

    # Formulas reference tolerance cell $B$8
    assert ws.cell(13, 5).value == "=(C13/$B$8)*100"
    assert ws.cell(17, 5).value == "=(C17/$B$8)*100"


# =============================================================================
# 6. Input Polymorphism & Error Handling
# =============================================================================


def test_input_as_sequence_of_dicts(sample_study_dicts: list[dict[str, Any]]) -> None:
    """build_msa_workbook accepts Sequence[Mapping[str, Any]]."""
    wb = build_msa_workbook(sample_study_dicts, method=METHOD)
    assert wb.sheetnames == ["Gage R&R Summary", "Measurements"]
    ws_meas = wb["Measurements"]
    assert ws_meas.cell(1, 1).value == "Part"
    assert ws_meas.cell(2, 1).value == "P01"


def test_results_only_without_data() -> None:
    """When results is supplied but data=None, generates only the 1-sheet summary workbook."""
    results = {
        "method": METHOD,
        "n_parts": 10,
        "n_appraisers": 3,
        "n_trials": 3,
        "ev": 0.05,
        "av": 0.03,
        "grr": 0.0583,
        "pv": 0.25,
        "tv": 0.2567,
        "pev_study": 19.47,
        "pav_study": 11.68,
        "pgrr_study": 22.71,
        "ppv_study": 97.39,
        "ndc": 6,
        "verdict": "Accept",
    }
    wb = build_msa_workbook(data=None, results=results, tolerance=2.5, method=METHOD)
    assert wb.sheetnames == ["Gage R&R Summary"]
    ws = wb["Gage R&R Summary"]
    assert ws.cell(1, 2).value == "AIAG MSA Gage R&R Study"
    assert ws.cell(5, 2).value == 10
    assert ws.cell(6, 2).value == 3
    assert ws.cell(7, 2).value == 3
    assert ws.cell(9, 2).value == "Accept"


def test_results_and_data_both_provided(sample_study_df: pd.DataFrame) -> None:
    """When results and data are both provided, uses results and writes data to Measurements."""
    results = {
        "method": METHOD,
        "n_parts": 10,
        "n_appraisers": 3,
        "n_trials": 3,
        "ev": 0.05,
        "av": 0.03,
        "grr": 0.0583,
        "pv": 0.25,
        "tv": 0.2567,
        "pev_study": 19.47,
        "pav_study": 11.68,
        "pgrr_study": 22.71,
        "ppv_study": 97.39,
        "ndc": 6,
        "verdict": "Accept",
    }
    wb = build_msa_workbook(sample_study_df, results=results, method=METHOD)
    assert wb.sheetnames == ["Gage R&R Summary", "Measurements"]
    ws_meas = wb["Measurements"]
    assert ws_meas.cell(1, 1).value == "Part"
    assert ws_meas.cell(2, 1).value == "P01"


def test_error_both_data_and_results_none() -> None:
    """Raises ValueError when both data and results are None."""
    with pytest.raises(ValueError, match="Either data or results must be provided"):
        build_msa_workbook(data=None, results=None)


def test_error_invalid_method(sample_study_df: pd.DataFrame) -> None:
    """Raises ValueError for unknown method name."""
    with pytest.raises(ValueError, match="Unknown method: 'invalid_method'"):
        build_msa_workbook(sample_study_df, method="invalid_method")


@pytest.mark.parametrize("bad_tolerance", [0, -1.0, float("nan"), float("inf"), float("-inf")])
def test_error_invalid_tolerance(sample_study_df: pd.DataFrame, bad_tolerance: float) -> None:
    """Raises ValueError for non-positive or non-finite tolerance values."""
    with pytest.raises(ValueError, match="must be a positive finite number"):
        build_msa_workbook(sample_study_df, tolerance=bad_tolerance)


def test_custom_study_title(sample_study_df: pd.DataFrame) -> None:
    """Custom title is reflected in the summary metadata."""
    wb = build_msa_workbook(sample_study_df, title="Custom Stamping Gage R&R")
    ws = wb["Gage R&R Summary"]
    assert ws.cell(1, 2).value == "Custom Stamping Gage R&R"


# =============================================================================
# 7. Serialization & Re-exports
# =============================================================================


def test_export_msa_workbook_returns_valid_xlsx_bytes(sample_study_df: pd.DataFrame) -> None:
    """export_msa_workbook produces valid zip/OOXML bytes."""
    raw_bytes = export_msa_workbook(sample_study_df, tolerance=4.42)
    assert isinstance(raw_bytes, bytes)
    assert raw_bytes.startswith(b"PK\x03\x04")

    # Verify zip integrity and presence of expected OOXML components
    with zipfile.ZipFile(io.BytesIO(raw_bytes)) as zf:
        names = zf.namelist()
        assert "[Content_Types].xml" in names
        assert "xl/workbook.xml" in names


def test_reexports_quality_core_msa(sample_study_df: pd.DataFrame) -> None:
    """Verify build_msa_workbook and export_msa_workbook are re-exported from quality_core.msa and msa.export."""
    import quality_core.msa as qcmsa

    assert hasattr(qcmsa, "build_msa_workbook")
    assert hasattr(qcmsa, "export_msa_workbook")
    assert qcmsa.build_msa_workbook is export_build
    assert qcmsa.export_msa_workbook is export_export

    wb = qcmsa.build_msa_workbook(sample_study_df)
    assert isinstance(wb, openpyxl.Workbook)
    raw = qcmsa.export_msa_workbook(sample_study_df)
    assert isinstance(raw, bytes)
