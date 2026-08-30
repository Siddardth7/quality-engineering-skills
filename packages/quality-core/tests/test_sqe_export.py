"""
tests/test_sqe_export.py
Tests for quality_core/sqe/export.py — SQE vendor-rating live-formula .xlsx exporter (#149).

Validates:
- 100% line & branch coverage on quality_core.sqe.export.
- Accuracy scorecard: for each RATED row, the sheet's PPM / OTIF / composite formulas,
  recomputed independently in Python, equal the engine's own computed values
  (source_evidence["ppm"], source_evidence["otif_pct"] with the 0-100 scale correction,
  and the raw ScorecardResult.composite_score).
- Positive live OOXML <f> formula verification via assert_cell_is_formula on I/M/U for
  every data row including the INDETERMINATE row, plus exact formula-string checks.
- Negative controls: literal cells in computed slots fail assert_cell_is_formula;
  formula-injection payloads in supplier_name render inert ('-prefixed) while I/M/U stay live.
- Edge cases: dimension absent (weight 0.0 -> NOT_SCORED/0.0), full 3-way SUMPRODUCT row,
  fully INDETERMINATE row (blank data cells + live N/A formulas), empty rows, dict-row form,
  Band None -> blank, custom title threading, metadata omitted-dimension fallbacks.
- Row validation guards and re-export integrity.

Reproduce (coverage):
  uv run pytest packages/quality-core/tests/test_sqe_export.py \
    --cov=quality_core.sqe.export --cov-report=term-missing --cov-fail-under=100
"""

from __future__ import annotations

import datetime
import io

import openpyxl
import pytest
from _xlsx_formula_audit import assert_cell_is_formula
from openpyxl.utils import get_column_letter
from quality_core.sqe.escalation import EscalationResult, evaluate_escalation
from quality_core.sqe.export import (
    _METADATA_SHEET_TITLE,
    _STANDARDS_BASIS,
    VENDOR_SCORECARD_COL_WIDTHS,
    VENDOR_SCORECARD_COLUMNS,
    SQEVendorRow,
    benchmark_sqe_vendor_rows,
    build_sqe_workbook,
    export_sqe_excel,
    export_sqe_workbook,
)
from quality_core.sqe.schema import DeliveryRecord, ReceiptLot, SupplierPeriod
from quality_core.sqe.scorecard import (
    LinearScoringCurve,
    ScorecardConfig,
    ScorecardResult,
    calculate_vendor_scorecard,
)

_DEFAULT_TITLE = "SQE Vendor Scorecard"


def _col(name: str) -> str:
    return get_column_letter(VENDOR_SCORECARD_COLUMNS.index(name) + 1)


_DEFECTS = _col("Defects")
_TOTAL_RECEIVED = _col("Total_Received")
_PPM = _col("PPM")
_OTIF_COUNT = _col("On_Time_In_Full")
_DELIVERY_COUNT = _col("Total_Deliveries")
_OTIF = _col("OTIF")
_QUALITY_SCORE = _col("Quality_Score")
_DELIVERY_SCORE = _col("Delivery_Score")
_COST_SCORE = _col("Cost_Score")
_QUALITY_WEIGHT = _col("Quality_Weight_(HEURISTIC)")
_DELIVERY_WEIGHT = _col("Delivery_Weight_(HEURISTIC)")
_COST_WEIGHT = _col("Cost_Weight_(HEURISTIC)")
_SCORECARD_VERDICT = _col("Scorecard_Verdict")
_COMPOSITE = _col("Composite_Score")
_BAND = _col("Band")


def _saved(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _load(wb_bytes: bytes) -> openpyxl.Workbook:
    return openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)


def _period(supplier_id: str) -> SupplierPeriod:
    return SupplierPeriod(
        supplier_id=supplier_id,
        period_start=datetime.date(2026, 1, 1),
        period_end=datetime.date(2026, 1, 31),
        period_label="January 2026",
    )


# ===========================================================================
# 1. Structure & Layout
# ===========================================================================


def test_export_benchmark_is_a_loadable_2sheet_workbook() -> None:
    rows = benchmark_sqe_vendor_rows()
    wb_bytes = export_sqe_workbook(rows)
    assert isinstance(wb_bytes, bytes)

    alias_bytes = export_sqe_excel(rows)
    assert isinstance(alias_bytes, bytes)

    wb = _load(wb_bytes)
    assert wb.sheetnames == [_DEFAULT_TITLE, _METADATA_SHEET_TITLE]

    ws = wb[_DEFAULT_TITLE]
    assert [c.value for c in ws[1]] == list(VENDOR_SCORECARD_COLUMNS)
    assert ws.max_row == len(rows) + 1


def test_col_widths_and_constants() -> None:
    assert set(VENDOR_SCORECARD_COL_WIDTHS) == set(VENDOR_SCORECARD_COLUMNS)
    assert len(VENDOR_SCORECARD_COLUMNS) == 24
    for width in VENDOR_SCORECARD_COL_WIDTHS.values():
        assert width > 0


# ===========================================================================
# 2. Accuracy Scorecard (headline requirement)
# ===========================================================================


def test_accuracy_scorecard_sqe() -> None:
    rows = benchmark_sqe_vendor_rows()
    wb_bytes = export_sqe_workbook(rows)
    wb = _load(wb_bytes)
    ws = wb[_DEFAULT_TITLE]

    rated_seen = 0
    unrated_seen = 0
    for idx, row in enumerate(rows):
        r = idx + 2
        dims = {d.name: d for d in row.scorecard.dimensions}
        if ws[f"{_SCORECARD_VERDICT}{r}"].value == "RATED":
            rated_seen += 1

            # PPM: defects / total_received * 1_000_000
            defects = float(ws[f"{_DEFECTS}{r}"].value)
            total_received = float(ws[f"{_TOTAL_RECEIVED}{r}"].value)
            recomputed_ppm = defects / total_received * 1_000_000
            assert recomputed_ppm == pytest.approx(
                dims["quality"].source_evidence["ppm"]
            )

            # OTIF: on_time_in_full / total_deliveries, engine otif_pct is 0-100
            on_time = float(ws[f"{_OTIF_COUNT}{r}"].value)
            total_deliveries = float(ws[f"{_DELIVERY_COUNT}{r}"].value)
            recomputed_otif_fraction = on_time / total_deliveries
            assert recomputed_otif_fraction * 100 == pytest.approx(
                dims["delivery"].source_evidence["otif_pct"]
            )

            # Composite: elementwise N*Q + O*R + P*S (SUMPRODUCT semantics)
            n = float(ws[f"{_QUALITY_SCORE}{r}"].value)
            o = float(ws[f"{_DELIVERY_SCORE}{r}"].value)
            p = float(ws[f"{_COST_SCORE}{r}"].value)
            q = float(ws[f"{_QUALITY_WEIGHT}{r}"].value)
            rr = float(ws[f"{_DELIVERY_WEIGHT}{r}"].value)
            s = float(ws[f"{_COST_WEIGHT}{r}"].value)
            recomputed_composite = n * q + o * rr + p * s
            assert recomputed_composite == pytest.approx(row.scorecard.composite_score)
        else:
            unrated_seen += 1
            # The guarded IF true-branch is wired even though no engine evaluates it here.
            assert '"N/A"' in ws[f"{_PPM}{r}"].value
            assert '"N/A"' in ws[f"{_OTIF}{r}"].value
            assert '"N/A"' in ws[f"{_COMPOSITE}{r}"].value

    assert rated_seen >= 2  # SUP-A (2-way) and SUP-B (3-way SUMPRODUCT)
    assert unrated_seen >= 1  # SUP-C fully INDETERMINATE


def test_full_three_way_sumproduct_row_present() -> None:
    """The benchmark must include a row with all three weights > 0."""
    rows = benchmark_sqe_vendor_rows()
    three_way = [
        row
        for row in rows
        if {d.name for d in row.scorecard.dimensions} >= {"quality", "delivery", "cost"}
        and all(d.weight > 0.0 for d in row.scorecard.dimensions)
    ]
    assert len(three_way) >= 1
    # And its cost sub-score is a genuine (non-zero) third term.
    cost = {d.name: d for d in three_way[0].scorecard.dimensions}["cost"]
    assert cost.sub_score not in (None, 0.0)
    assert cost.weight > 0.0


def test_absent_dimension_renders_not_scored_and_zero() -> None:
    """SUP-A carries no cost dimension (cost_weight 0.0) -> 0.0 score/weight cells."""
    rows = benchmark_sqe_vendor_rows()
    wb = _load(export_sqe_workbook(rows))
    ws = wb[_DEFAULT_TITLE]

    # Find the row whose cost dimension is omitted.
    for idx, row in enumerate(rows):
        names = {d.name for d in row.scorecard.dimensions}
        if "cost" not in names and row.scorecard.verdict == "RATED":
            r = idx + 2
            assert ws[f"{_COST_SCORE}{r}"].value == 0.0
            assert ws[f"{_COST_WEIGHT}{r}"].value == 0.0
            return
    pytest.fail("benchmark lacks a RATED row with an omitted cost dimension")


def test_indeterminate_row_blank_counts_and_blank_band() -> None:
    rows = benchmark_sqe_vendor_rows()
    wb = _load(export_sqe_workbook(rows))
    ws = wb[_DEFAULT_TITLE]

    for idx, row in enumerate(rows):
        if row.scorecard.verdict != "RATED":
            r = idx + 2
            # numerator/otif_count are None on a blocked dimension -> blank cells.
            assert ws[f"{_DEFECTS}{r}"].value is None
            assert ws[f"{_OTIF_COUNT}{r}"].value is None
            # Band is None on an INDETERMINATE row -> blank, not the string "None".
            assert ws[f"{_BAND}{r}"].value is None
            return
    pytest.fail("benchmark lacks an INDETERMINATE row")


# ===========================================================================
# 3. Positive Live Formula Verification
# ===========================================================================


def test_positive_live_formula_audit_sqe() -> None:
    rows = benchmark_sqe_vendor_rows()
    wb_bytes = export_sqe_workbook(rows)

    # I/M/U are live <f> on EVERY data row, including the INDETERMINATE one.
    for r in range(2, len(rows) + 2):
        assert_cell_is_formula(wb_bytes, _DEFAULT_TITLE, f"{_PPM}{r}")
        assert_cell_is_formula(wb_bytes, _DEFAULT_TITLE, f"{_OTIF}{r}")
        assert_cell_is_formula(wb_bytes, _DEFAULT_TITLE, f"{_COMPOSITE}{r}")

    wb = _load(wb_bytes)
    ws = wb[_DEFAULT_TITLE]
    for r in range(2, len(rows) + 2):
        assert ws[f"{_PPM}{r}"].value == (
            f'=IF({_col("PPM_Verdict")}{r}<>"MEASURED","N/A",'
            f"{_DEFECTS}{r}/{_TOTAL_RECEIVED}{r}*1000000)"
        )
        assert ws[f"{_OTIF}{r}"].value == (
            f'=IF({_col("OTIF_Verdict")}{r}<>"MEASURED","N/A",'
            f"{_OTIF_COUNT}{r}/{_DELIVERY_COUNT}{r})"
        )
        assert ws[f"{_OTIF}{r}"].number_format == "0.0%"
        assert ws[f"{_COMPOSITE}{r}"].value == (
            f'=IF({_SCORECARD_VERDICT}{r}<>"RATED","N/A",'
            f"SUMPRODUCT({_QUALITY_WEIGHT}{r}:{_COST_WEIGHT}{r},"
            f"{_QUALITY_SCORE}{r}:{_COST_SCORE}{r}))"
        )


# ===========================================================================
# 4. Negative Controls
# ===========================================================================


def test_literal_negative_control_fails_formula_verifier_sqe() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _DEFAULT_TITLE
    ws.cell(row=2, column=VENDOR_SCORECARD_COLUMNS.index("PPM") + 1, value=2000.0)
    ws.cell(row=2, column=VENDOR_SCORECARD_COLUMNS.index("OTIF") + 1, value=0.75)
    ws.cell(row=2, column=VENDOR_SCORECARD_COLUMNS.index("Composite_Score") + 1, value=78.0)
    saved = _saved(wb)

    for coord in (f"{_PPM}2", f"{_OTIF}2", f"{_COMPOSITE}2"):
        with pytest.raises(AssertionError, match="literal, not a live formula"):
            assert_cell_is_formula(saved, _DEFAULT_TITLE, coord)


def test_security_invariant_formula_injection_defense_sqe() -> None:
    malicious_inputs = [
        "=SUM(A1:A10)",
        "+1+1",
        "@EVIL()",
        "\t=CALC()",
        "\r+2+2",
    ]
    base = benchmark_sqe_vendor_rows()[0]
    rows = [
        SQEVendorRow(
            supplier_id=f"INJ-{i}",
            scorecard=base.scorecard,
            escalation=base.escalation,
            supplier_name=payload,
        )
        for i, payload in enumerate(malicious_inputs)
    ]
    wb_bytes = export_sqe_workbook(rows)
    wb = _load(wb_bytes)
    ws = wb[_DEFAULT_TITLE]
    name_col = _col("Supplier_Name")

    for i in range(len(rows)):
        r = i + 2
        name_cell = ws[f"{name_col}{r}"].value
        assert str(name_cell).startswith("'"), f"name not escaped: {name_cell!r}"
        # The escaped freetext cell is inert (not a live formula) ...
        with pytest.raises(AssertionError, match="literal, not a live formula"):
            assert_cell_is_formula(wb_bytes, _DEFAULT_TITLE, f"{name_col}{r}")
        # ... while the computed cells on the same row remain live formulas.
        assert_cell_is_formula(wb_bytes, _DEFAULT_TITLE, f"{_PPM}{r}")
        assert_cell_is_formula(wb_bytes, _DEFAULT_TITLE, f"{_OTIF}{r}")
        assert_cell_is_formula(wb_bytes, _DEFAULT_TITLE, f"{_COMPOSITE}{r}")


# ===========================================================================
# 5. Empty, Dict-Row, Custom Title, Metadata
# ===========================================================================


def test_empty_rows_produces_header_only_sheet_and_metadata_notice() -> None:
    wb_bytes = export_sqe_workbook([])
    wb = _load(wb_bytes)

    ws = wb[_DEFAULT_TITLE]
    assert ws.max_row == 1
    assert [c.value for c in ws[1]] == list(VENDOR_SCORECARD_COLUMNS)

    meta = wb[_METADATA_SHEET_TITLE]
    assert meta["A1"].value == "Heuristic Configuration"
    assert "unavailable" in meta["B1"].value


def test_dict_row_input_form() -> None:
    row = benchmark_sqe_vendor_rows()[0]
    dict_rows = [
        {
            "supplier_id": row.supplier_id,
            "scorecard": row.scorecard,
            "escalation": row.escalation,
            "supplier_name": row.supplier_name,
        }
    ]
    wb = _load(export_sqe_workbook(dict_rows))
    ws = wb[_DEFAULT_TITLE]
    assert ws.max_row == 2
    assert ws[f"{_col('Supplier_ID')}2"].value == row.scorecard.supplier_id


def test_custom_title_threads_into_sheet_name() -> None:
    custom = "FY26 Q1 Supplier Rating"
    rows = benchmark_sqe_vendor_rows()
    wb = _load(export_sqe_workbook(rows, title=custom))
    assert wb.sheetnames == [custom, _METADATA_SHEET_TITLE]
    meta = wb[_METADATA_SHEET_TITLE]
    # Report Title metadata row echoes the custom title.
    assert meta["B1"].value == custom


def test_metadata_rows_populated_from_first_supplier() -> None:
    rows = benchmark_sqe_vendor_rows()
    wb = _load(export_sqe_workbook(rows))
    meta = wb[_METADATA_SHEET_TITLE]
    kv = {meta[f"A{r}"].value: meta[f"B{r}"].value for r in range(1, meta.max_row + 1)}

    # Metadata is built from rows[0]; assert that row exercises the *populated*
    # (non-NOT_SCORED) branch so the value assertions below are load-bearing.
    first = rows[0]
    dims = {d.name: d for d in first.scorecard.dimensions}
    assert "quality" in dims and "delivery" in dims

    # Every value is copied verbatim from an engine payload — pin each to its source,
    # not just the label's presence. A mutation of the source key (e.g. "minimum" ->
    # "basis") must fail here.
    sample_adequacy = dims["quality"].source_evidence["sample_adequacy"]
    assert kv["Report Title"] == _DEFAULT_TITLE
    assert kv["Date Generated"]  # now() -> non-empty timestamp string
    assert kv["Standards Basis (Vendor Rating)"] == _STANDARDS_BASIS
    assert kv["PPM Sample-Adequacy Minimum (HEURISTIC)"] == sample_adequacy["minimum"]
    assert kv["PPM Sample-Adequacy Basis"] == sample_adequacy["basis"]
    assert kv["OTIF Heuristic Configuration (HEURISTIC)"] == str(
        dims["delivery"].source_evidence["heuristic_configuration"]
    )
    assert kv["Scorecard Weights / Curves / Bands (HEURISTIC)"] == str(
        first.scorecard.heuristic_configuration
    )
    assert kv["Escalation Thresholds (HEURISTIC)"] == str(
        first.escalation.heuristic_configuration
    )


def test_metadata_falls_back_when_first_row_omits_quality_and_delivery() -> None:
    """rows[0] with quality & delivery omitted hits both NOT_SCORED metadata arms."""
    period = _period("SUP-COST")
    lots = [
        ReceiptLot(
            supplier_id="SUP-COST",
            lot_id="LOT-1",
            quantity_received=1500,
            receipt_date=datetime.date(2026, 1, 10),
            defect_count=2,
        )
    ]
    deliveries = [
        DeliveryRecord(
            supplier_id="SUP-COST",
            order_id="PO-1",
            quantity_ordered=100,
            quantity_delivered=100,
            promised_date=datetime.date(2026, 1, 11),
            actual_delivery_date=datetime.date(2026, 1, 11),
        )
    ]
    config = ScorecardConfig(
        quality_weight=0.0,
        delivery_weight=0.0,
        cost_weight=1.0,
        cost_curve=LinearScoringCurve(best_value=0.0, worst_value=10.0),
    )
    scorecard = calculate_vendor_scorecard(
        period,
        lots,
        deliveries,
        copq_items=[
            {
                "category": "InternalFailure",
                "description": "scrap",
                "direct_cost": 5000.0,
            }
        ],
        revenue_base=500_000.0,
        config=config,
    )
    names = {d.name for d in scorecard.dimensions}
    assert "quality" not in names and "delivery" not in names

    rows = [
        SQEVendorRow(
            supplier_id="SUP-COST",
            scorecard=scorecard,
            escalation=evaluate_escalation(scorecard),
            supplier_name="Cost Only",
        )
    ]
    wb = _load(export_sqe_workbook(rows))
    meta = wb[_METADATA_SHEET_TITLE]
    kv = {meta[f"A{r}"].value: meta[f"B{r}"].value for r in range(1, meta.max_row + 1)}
    assert kv["PPM Sample-Adequacy Minimum (HEURISTIC)"] == "NOT_SCORED"
    assert kv["PPM Sample-Adequacy Basis"] == "NOT_SCORED"
    assert kv["OTIF Heuristic Configuration (HEURISTIC)"] == "NOT_SCORED"


# ===========================================================================
# 6. Row Validation & Coercion Guards
# ===========================================================================


def _valid_kwargs() -> dict[str, object]:
    row = benchmark_sqe_vendor_rows()[0]
    return {
        "supplier_id": "S1",
        "scorecard": row.scorecard,
        "escalation": row.escalation,
    }


def test_sqevendorrow_rejects_bad_supplier_id() -> None:
    kwargs = _valid_kwargs()
    with pytest.raises(TypeError, match="supplier_id must be a non-empty string"):
        SQEVendorRow(**{**kwargs, "supplier_id": "   "})  # type: ignore[arg-type]
    with pytest.raises(TypeError, match="supplier_id must be a non-empty string"):
        SQEVendorRow(**{**kwargs, "supplier_id": 123})  # type: ignore[arg-type]


def test_sqevendorrow_rejects_bad_scorecard_and_escalation() -> None:
    kwargs = _valid_kwargs()
    with pytest.raises(TypeError, match="scorecard must be a ScorecardResult"):
        SQEVendorRow(**{**kwargs, "scorecard": object()})  # type: ignore[arg-type]
    with pytest.raises(TypeError, match="escalation must be an EscalationResult"):
        SQEVendorRow(**{**kwargs, "escalation": object()})  # type: ignore[arg-type]


def test_sqevendorrow_rejects_bad_supplier_name_type() -> None:
    kwargs = _valid_kwargs()
    with pytest.raises(TypeError, match="supplier_name must be a string or None"):
        SQEVendorRow(**{**kwargs, "supplier_name": 42})  # type: ignore[arg-type]


def test_sqevendorrow_blank_supplier_name_becomes_none() -> None:
    kwargs = _valid_kwargs()
    row = SQEVendorRow(**{**kwargs, "supplier_name": "   "})  # type: ignore[arg-type]
    assert row.supplier_name is None
    # And a supplied name is stripped/kept.
    named = SQEVendorRow(**{**kwargs, "supplier_name": "  Acme  "})  # type: ignore[arg-type]
    assert named.supplier_name == "Acme"


def test_sqevendorrow_none_supplier_name_renders_blank_cell() -> None:
    kwargs = _valid_kwargs()
    row = SQEVendorRow(**kwargs)  # type: ignore[arg-type]
    assert row.supplier_name is None
    wb = _load(export_sqe_workbook([row]))
    ws = wb[_DEFAULT_TITLE]
    assert ws[f"{_col('Supplier_Name')}2"].value is None


def test_coerce_rows_rejects_unknown_type() -> None:
    with pytest.raises(TypeError, match="Expected SQEVendorRow or dict"):
        build_sqe_workbook([object()])  # type: ignore[list-item]


# ===========================================================================
# 7. Re-export Integrity & Fresh Benchmarks
# ===========================================================================


def test_reexport_integrity_sqe() -> None:
    import quality_core.sqe as pkg
    import quality_core.sqe.export as mod

    assert pkg.build_sqe_workbook is mod.build_sqe_workbook
    assert pkg.export_sqe_workbook is mod.export_sqe_workbook
    assert pkg.export_sqe_excel is mod.export_sqe_excel
    assert pkg.benchmark_sqe_vendor_rows is mod.benchmark_sqe_vendor_rows
    assert pkg.SQEVendorRow is mod.SQEVendorRow
    assert pkg.VENDOR_SCORECARD_COLUMNS is mod.VENDOR_SCORECARD_COLUMNS
    assert pkg.VENDOR_SCORECARD_COL_WIDTHS is mod.VENDOR_SCORECARD_COL_WIDTHS


def test_benchmark_returns_fresh_instances_sqe() -> None:
    a = benchmark_sqe_vendor_rows()
    b = benchmark_sqe_vendor_rows()
    assert a is not b
    assert len(a) == 3
    assert all(isinstance(row, SQEVendorRow) for row in a)
    assert all(isinstance(row.scorecard, ScorecardResult) for row in a)
    assert all(isinstance(row.escalation, EscalationResult) for row in a)
    # Fresh nested objects, not shared references.
    assert a[0].scorecard is not b[0].scorecard
