"""
quality_core/msa/export.py
Live-formula Excel exporter for Measurement Systems Analysis (MSA) Gage R&R studies.

Generates multi-sheet .xlsx workbooks per AIAG MSA (4th Edition) for crossed Gage R&R
studies (Average-and-Range and ANOVA methods). Computed cells (%EV, %AV, %GRR, %PV,
%TV vs study variation; %EV, %AV, %GRR, %PV, %TV vs tolerance; 6*SD; GRR SD; TV SD;
ndc; and ANOVA MS/F/Total sums) are live openpyxl Formula elements that dynamically
recalculate in spreadsheet viewers upon editing underlying cells. Qualitative verdicts
remain structured strings ("Accept", "Marginal", "Reject"), and untrusted data cells
route through sanitize_cell to prevent CSV/formula injection.
"""

from __future__ import annotations

import io
from collections.abc import Mapping, Sequence
from typing import Any

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from quality_core.io import (
    Formula,
    now,
    sanitize_cell,
    write_formula_cell,
    write_table_sheet,
)
from quality_core.msa.gage_rr import (
    METHOD,
    METHOD_ANOVA,
    compute_gage_rr,
)

__all__ = [
    "METHOD",
    "METHOD_ANOVA",
    "build_msa_workbook",
    "export_msa_workbook",
]

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_BOLD_FONT = Font(bold=True, size=10)
_NORMAL_FONT = Font(size=10)
_HEADER_FILL_HEX = "2C3E50"


def build_msa_workbook(
    data: pd.DataFrame | Sequence[Mapping[str, Any]] | None = None,
    *,
    results: Mapping[str, Any] | None = None,
    tolerance: float | None = None,
    method: str = METHOD,
    title: str = "AIAG MSA Gage R&R Study",
) -> openpyxl.Workbook:
    """Build an openpyxl Workbook for a Gage R&R study with live formulas.

    Args:
        data: Optional measurement DataFrame or list of dicts with keys:
            part, appraiser, trial, measurement.
        results: Optional precomputed result mapping from `compute_gage_rr`.
            If omitted, computed automatically from `data`.
        tolerance: USL - LSL (study tolerance). If provided, tolerance basis
            percentages and formulas are included.
        method: Study method — ``METHOD`` ("average_and_range") or
            ``METHOD_ANOVA`` ("anova").
        title: Study title for the metadata header.

    Returns:
        openpyxl.Workbook with "Gage R&R Summary", optional "Measurements",
        and optional "ANOVA Table" sheets.

    Raises:
        ValueError: If both data and results are None, if method is unknown,
            if tolerance is non-positive or non-finite, or if data is invalid.
    """
    if method not in (METHOD, METHOD_ANOVA):
        raise ValueError(
            f"Unknown method: {method!r}. Supported: {METHOD!r}, {METHOD_ANOVA!r}."
        )

    if tolerance is not None:
        if tolerance <= 0 or not np.isfinite(tolerance):
            raise ValueError("Tolerance (USL - LSL) must be a positive finite number.")

    if data is None and results is None:
        raise ValueError("Either data or results must be provided.")

    # Convert data to DataFrame if provided
    df_data: pd.DataFrame | None = None
    if data is not None:
        if isinstance(data, pd.DataFrame):
            df_data = data.copy()
        else:
            df_data = pd.DataFrame(list(data))

    # Compute results if not supplied
    res: dict[str, Any]
    if results is None:
        assert df_data is not None
        res = compute_gage_rr(df_data, tolerance=tolerance, method=method)
    else:
        res = dict(results)

    method_used = res.get("method", method)
    is_anova = method_used == METHOD_ANOVA

    wb = openpyxl.Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1: Gage R&R Summary
    # -----------------------------------------------------------------------
    ws_summary: Any = wb.active
    ws_summary.title = "Gage R&R Summary"

    method_display = (
        "ANOVA (Crossed Two-Factor with Replication)"
        if is_anova
        else "Average and Range"
    )

    metadata_rows: list[tuple[str, Any]] = [
        ("Study Title", title),
        ("Date Generated", now()),
        ("Basis", "AIAG MSA 4th Edition (Crossed)"),
        ("Method", method_display),
        ("Number of Parts", res.get("n_parts", "N/A")),
        ("Number of Appraisers", res.get("n_appraisers", "N/A")),
        ("Number of Trials", res.get("n_trials", "N/A")),
        ("Tolerance", tolerance if tolerance is not None else "N/A"),
        ("AIAG Verdict", res.get("verdict", "N/A")),
    ]

    for r_idx, (label, val) in enumerate(metadata_rows, start=1):
        cell_lbl = ws_summary.cell(row=r_idx, column=1, value=sanitize_cell(label))
        cell_lbl.font = _BOLD_FONT
        cell_lbl.alignment = Alignment(vertical="center")

        if r_idx == 8 and tolerance is not None:
            cell_val = ws_summary.cell(row=r_idx, column=2, value=float(tolerance))
            cell_val.number_format = "0.0000"
        else:
            cell_val = ws_summary.cell(row=r_idx, column=2, value=sanitize_cell(val))
        cell_val.font = _NORMAL_FONT
        cell_val.alignment = Alignment(vertical="center")

    # Determine row layout for Variance Components table
    # Metadata takes rows 1-10; row 11 is blank; row 12 is table header.
    if is_anova:
        ev_row = 13
        av_row = 14
        int_row = 15
        grr_row = 16
        pv_row = 17
        tv_row = 18
    else:
        ev_row = 13
        av_row = 14
        grr_row = 15
        pv_row = 16
        tv_row = 17

    # Row 10: ndc live formula
    cell_ndc_lbl = ws_summary.cell(
        row=10, column=1, value=sanitize_cell("Number of Distinct Categories (ndc)")
    )
    cell_ndc_lbl.font = _BOLD_FONT
    cell_ndc_lbl.alignment = Alignment(vertical="center")

    cell_ndc_val = write_formula_cell(
        ws_summary,
        row=10,
        column=2,
        formula=f"=MAX(1, INT(1.41 * (B{pv_row} / B{grr_row})))",
        number_format="0",
    )
    cell_ndc_val.font = _NORMAL_FONT
    cell_ndc_val.alignment = Alignment(vertical="center")

    # Header Row for Variance Components (Row 12)
    header_fill = PatternFill(
        start_color=_HEADER_FILL_HEX,
        end_color=_HEADER_FILL_HEX,
        fill_type="solid",
    )
    table_headers = [
        "Source",
        "Standard Deviation (SD)",
        "Study Variation (6×SD)",
        "% Study Variation (%SV)",
    ]
    if tolerance is not None:
        table_headers.append("% Tolerance (%Tol)")

    ws_summary.row_dimensions[12].height = 22
    for col_idx, h_text in enumerate(table_headers, start=1):
        h_cell = ws_summary.cell(row=12, column=col_idx, value=sanitize_cell(h_text))
        h_cell.fill = header_fill
        h_cell.font = _HEADER_FONT
        h_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    # Helper to populate a variance component row
    def _write_vc_row(
        row: int,
        source_name: str,
        sd_value: float | None,
        sd_formula: str | None = None,
    ) -> None:
        c_src = ws_summary.cell(row=row, column=1, value=sanitize_cell(source_name))
        c_src.font = _NORMAL_FONT
        c_src.alignment = Alignment(vertical="center")

        if sd_formula is not None:
            c_sd = write_formula_cell(
                ws_summary,
                row=row,
                column=2,
                formula=sd_formula,
                number_format="0.0000",
            )
        else:
            assert sd_value is not None
            c_sd = ws_summary.cell(row=row, column=2, value=float(sd_value))
            c_sd.number_format = "0.0000"
        c_sd.font = _NORMAL_FONT
        c_sd.alignment = Alignment(vertical="center")

        # 6×SD: =B{row}*6
        c_6sd = write_formula_cell(
            ws_summary,
            row=row,
            column=3,
            formula=f"=B{row}*6",
            number_format="0.0000",
        )
        c_6sd.font = _NORMAL_FONT
        c_6sd.alignment = Alignment(vertical="center")

        # %SV: =(B{row}/$B${tv_row})*100
        c_psv = write_formula_cell(
            ws_summary,
            row=row,
            column=4,
            formula=f"=(B{row}/$B${tv_row})*100",
            number_format="0.0000",
        )
        c_psv.font = _NORMAL_FONT
        c_psv.alignment = Alignment(vertical="center")

        if tolerance is not None:
            # %Tol: =(C{row}/$B$8)*100
            c_ptol = write_formula_cell(
                ws_summary,
                row=row,
                column=5,
                formula=f"=(C{row}/$B$8)*100",
                number_format="0.0000",
            )
            c_ptol.font = _NORMAL_FONT
            c_ptol.alignment = Alignment(vertical="center")

    # Populate rows for Repeatability and Reproducibility
    _write_vc_row(ev_row, "Repeatability (EV)", res.get("ev", 0.0))
    _write_vc_row(av_row, "Reproducibility (AV)", res.get("av", 0.0))

    if is_anova:
        interaction_val = res.get("interaction", 0.0)
        _write_vc_row(
            int_row,
            "Part × Appraiser Interaction (INT)",
            interaction_val if interaction_val is not None else 0.0,
        )
        grr_formula = f"=SQRT(B{ev_row}^2 + B{av_row}^2 + B{int_row}^2)"
    else:
        grr_formula = f"=SQRT(B{ev_row}^2 + B{av_row}^2)"

    _write_vc_row(grr_row, "Total Gage R&R (GRR)", None, sd_formula=grr_formula)
    _write_vc_row(pv_row, "Part Variation (PV)", res.get("pv", 0.0))
    _write_vc_row(
        tv_row,
        "Total Variation (TV)",
        None,
        sd_formula=f"=SQRT(B{grr_row}^2 + B{pv_row}^2)",
    )

    ws_summary.column_dimensions["A"].width = 38
    ws_summary.column_dimensions["B"].width = 24
    ws_summary.column_dimensions["C"].width = 24
    ws_summary.column_dimensions["D"].width = 24
    if tolerance is not None:
        ws_summary.column_dimensions["E"].width = 24

    # -----------------------------------------------------------------------
    # Sheet 2: Measurements (if data provided)
    # -----------------------------------------------------------------------
    if df_data is not None:
        ws_meas = wb.create_sheet(title="Measurements")

        # Normalize column names for display
        col_rename = {
            "part": "Part",
            "appraiser": "Appraiser",
            "trial": "Trial",
            "measurement": "Measurement",
        }
        meas_display = df_data.rename(
            columns=lambda c: col_rename.get(str(c).lower(), str(c))
        )
        write_table_sheet(
            ws_meas,
            meas_display,
            title="Measurements",
            columns=["Part", "Appraiser", "Trial", "Measurement"],
            col_widths={
                "Part": 16,
                "Appraiser": 16,
                "Trial": 12,
                "Measurement": 18,
            },
        )

        # -------------------------------------------------------------------
        # Sheet 3: ANOVA Table (if method == "anova" and data provided)
        # -------------------------------------------------------------------
        if is_anova:
            ws_anova = wb.create_sheet(title="ANOVA Table")

            # Calculate ANOVA SS and DF values
            df_calc = df_data.copy()
            df_calc.columns = [str(c).lower() for c in df_calc.columns]
            df_calc["measurement"] = pd.to_numeric(
                df_calc["measurement"], errors="raise"
            )

            n_parts = df_calc["part"].nunique()
            n_appraisers = df_calc["appraiser"].nunique()
            cell_means = df_calc.groupby(["part", "appraiser"])["measurement"].mean()
            n_trials = int(df_calc.groupby(["part", "appraiser"]).size().iloc[0])

            grand_mean = float(df_calc["measurement"].mean())
            ss_total = float(((df_calc["measurement"] - grand_mean) ** 2).sum())
            part_means = df_calc.groupby("part")["measurement"].mean()
            ss_parts = float(
                n_appraisers * n_trials * ((part_means - grand_mean) ** 2).sum()
            )
            appraiser_means = df_calc.groupby("appraiser")["measurement"].mean()
            ss_appraiser = float(
                n_parts * n_trials * ((appraiser_means - grand_mean) ** 2).sum()
            )
            ss_cells = float(n_trials * ((cell_means - grand_mean) ** 2).sum())
            ss_interaction = ss_cells - ss_parts - ss_appraiser
            ss_equipment = ss_total - ss_cells

            df_parts = n_parts - 1
            df_appraiser = n_appraisers - 1
            df_interaction = df_parts * df_appraiser
            df_equipment = n_parts * n_appraisers * (n_trials - 1)

            interaction_sig = res.get("interaction_significant")
            sig_text = (
                "Yes"
                if interaction_sig is True
                else "No (Pooled)"
                if interaction_sig is False
                else ""
            )

            anova_rows = [
                {
                    "Source": "Part",
                    "Degrees of Freedom (DF)": df_parts,
                    "Sum of Squares (SS)": ss_parts,
                    "Mean Square (MS)": Formula("=C2/B2", number_format="0.0000"),
                    "F-Statistic": "",
                    "Significant (α=0.05)": "",
                },
                {
                    "Source": "Appraiser",
                    "Degrees of Freedom (DF)": df_appraiser,
                    "Sum of Squares (SS)": ss_appraiser,
                    "Mean Square (MS)": Formula("=C3/B3", number_format="0.0000"),
                    "F-Statistic": "",
                    "Significant (α=0.05)": "",
                },
                {
                    "Source": "Part × Appraiser Interaction",
                    "Degrees of Freedom (DF)": df_interaction,
                    "Sum of Squares (SS)": ss_interaction,
                    "Mean Square (MS)": Formula("=C4/B4", number_format="0.0000"),
                    "F-Statistic": Formula("=D4/D5", number_format="0.0000"),
                    "Significant (α=0.05)": sig_text,
                },
                {
                    "Source": "Equipment (Error)",
                    "Degrees of Freedom (DF)": df_equipment,
                    "Sum of Squares (SS)": ss_equipment,
                    "Mean Square (MS)": Formula("=C5/B5", number_format="0.0000"),
                    "F-Statistic": "",
                    "Significant (α=0.05)": "",
                },
                {
                    "Source": "Total",
                    "Degrees of Freedom (DF)": Formula(
                        "=SUM(B2:B5)", number_format="0"
                    ),
                    "Sum of Squares (SS)": Formula(
                        "=SUM(C2:C5)", number_format="0.0000"
                    ),
                    "Mean Square (MS)": "",
                    "F-Statistic": "",
                    "Significant (α=0.05)": "",
                },
            ]

            anova_df = pd.DataFrame(anova_rows)
            write_table_sheet(
                ws_anova,
                anova_df,
                title="ANOVA Table",
                columns=[
                    "Source",
                    "Degrees of Freedom (DF)",
                    "Sum of Squares (SS)",
                    "Mean Square (MS)",
                    "F-Statistic",
                    "Significant (α=0.05)",
                ],
                col_widths={
                    "Source": 30,
                    "Degrees of Freedom (DF)": 24,
                    "Sum of Squares (SS)": 22,
                    "Mean Square (MS)": 20,
                    "F-Statistic": 16,
                    "Significant (α=0.05)": 22,
                },
            )

    return wb


def export_msa_workbook(
    data: pd.DataFrame | Sequence[Mapping[str, Any]] | None = None,
    *,
    results: Mapping[str, Any] | None = None,
    tolerance: float | None = None,
    method: str = METHOD,
    title: str = "AIAG MSA Gage R&R Study",
) -> bytes:
    """Export a Gage R&R study as serialized .xlsx bytes with live formulas.

    Args:
        data: Optional measurement DataFrame or sequence of row dicts.
        results: Optional precomputed result mapping from `compute_gage_rr`.
        tolerance: USL - LSL (study tolerance). If provided, tolerance basis
            percentages and formulas are included.
        method: Study method — ``METHOD`` ("average_and_range") or
            ``METHOD_ANOVA`` ("anova").
        title: Study title for the metadata header.

    Returns:
        Serialized .xlsx file bytes.
    """
    wb = build_msa_workbook(
        data=data,
        results=results,
        tolerance=tolerance,
        method=method,
        title=title,
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
