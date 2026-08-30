"""
tests/test_ncr_export.py
Tests for quality_core/ncr/export.py — Nonconformance Reporting (NCR) live-formula .xlsx exporter (#147).

Validates:
- 100% line & branch coverage on quality_core.ncr.export.
- Positive live OOXML <f> formula verification via assert_cell_is_formula on:
  - Dispositions & Containment sheet: COUNTIF (B2:B7), SUMIF (C2:C7), Pct ratio (D2:D7), SUM totals (B8:D8).
  - Summary & Metadata sheet: COUNTA (B4), SUM (B5).
- Accuracy scorecards: recomputed disposition counts and affected quantities match dataset.
- 🔒 Security invariant: untrusted data strings starting with '=', '+', '-', '@', '\\t', '\\r' render inert.
- Reusable verifier negative controls: literal cells in computed slots fail assert_cell_is_formula.
- Type guards & edge cases: empty datasets, polymorphism (NCRDataset, list[Record], list[dict], DataFrame, dict), custom titles.
- Re-exports from quality_core.ncr package.
"""

from __future__ import annotations

import io
from typing import Any

import openpyxl
import pytest
from _xlsx_formula_audit import assert_cell_is_formula
from openpyxl.utils import get_column_letter
from quality_core.ncr import (
    DISPOSITION_SUMMARY_COLUMNS,
    DISPOSITION_VALUES,
    NCR_COL_WIDTHS,
    NCR_EXPORT_COLUMNS,
    NCRDataset,
    NonconformanceRecord,
    benchmark_ncr_dataset,
    export_ncr_excel,
    export_ncr_workbook,
)
from quality_core.ncr.export import (
    _DISP_SUMMARY_SHEET_TITLE,
    _SUMMARY_SHEET_TITLE,
    DISPOSITION_SUMMARY_COL_WIDTHS,
    _last_row,
    _write_metadata_sheet,
)

_DEFAULT_TITLE = "Nonconformance Records"
_DISP_COL = get_column_letter(NCR_EXPORT_COLUMNS.index("Disposition") + 1)
_QTY_COL = get_column_letter(NCR_EXPORT_COLUMNS.index("Quantity_Affected") + 1)
_REC_ID_COL = get_column_letter(NCR_EXPORT_COLUMNS.index("Record_ID") + 1)

_SAMPLE_RECORD_DATA: dict[str, Any] = {
    "record_id": "NCR-2026-001",
    "part_lot_id": "LOT-BRK-8821",
    "defect_description": "Cast porosity on brake caliper mounting flange exceeding max allowable void diameter.",
    "requirement_violated": "DWG-BRK-004 Rev D: Max surface pore diameter <= 0.50 mm.",
    "quantity_affected": 45,
    "detection_point": "Receiving Inspection",
    "severity": "Major",
    "disposition": "ReturnToVendor",
    "approval_authority": "Supplier Quality Assurance",
    "rationale": "Nonconforming casting lot rejected and segregated for return.",
}


def _make_record(**overrides: Any) -> NonconformanceRecord:
    return NonconformanceRecord(**{**_SAMPLE_RECORD_DATA, **overrides})


def _saved(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _load(wb_bytes: bytes) -> openpyxl.Workbook:
    return openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)


# ===========================================================================
# 1. Structure & Layout Tests
# ===========================================================================


def test_export_benchmark_is_a_loadable_3sheet_workbook() -> None:
    dataset = benchmark_ncr_dataset()
    wb_bytes = export_ncr_workbook(dataset)
    assert isinstance(wb_bytes, bytes)

    alias_bytes = export_ncr_excel(dataset)
    assert isinstance(alias_bytes, bytes)

    wb = _load(wb_bytes)
    assert wb.sheetnames == [_DEFAULT_TITLE, _DISP_SUMMARY_SHEET_TITLE, _SUMMARY_SHEET_TITLE]

    matrix = wb[_DEFAULT_TITLE]
    assert [c.value for c in matrix[1]] == list(NCR_EXPORT_COLUMNS)
    assert matrix.max_row == len(dataset.records) + 1

    ws_disp = wb[_DISP_SUMMARY_SHEET_TITLE]
    assert [c.value for c in ws_disp[1]] == list(DISPOSITION_SUMMARY_COLUMNS)
    assert ws_disp.max_row == 8
    disposition_labels = [ws_disp[f"A{r}"].value for r in range(2, 9)]
    assert disposition_labels == list(DISPOSITION_VALUES) + ["Unassigned", "Total"]

    ws_summary = wb[_SUMMARY_SHEET_TITLE]
    assert [ws_summary[f"A{r}"].value for r in range(1, 7)] == [
        "Report Title",
        "Date Generated",
        "Standards Basis",
        "Total Records",
        "Total Quantity Affected",
        "MRB Gate Reviews Required",
    ]
    assert ws_summary["B1"].value == _DEFAULT_TITLE
    assert ws_summary["B3"].value == "ISO 9001:2015 §8.7 / IATF 16949:2016 §8.7"


def test_col_widths_and_constants() -> None:
    assert set(NCR_COL_WIDTHS) == set(NCR_EXPORT_COLUMNS)
    assert set(DISPOSITION_SUMMARY_COL_WIDTHS) == set(DISPOSITION_SUMMARY_COLUMNS)
    for col, width in NCR_COL_WIDTHS.items():
        assert width > 0
    for col, width in DISPOSITION_SUMMARY_COL_WIDTHS.items():
        assert width > 0


# ===========================================================================
# 2. Positive Live Formula Verification
# ===========================================================================


def test_positive_live_formula_audit_ncr() -> None:
    dataset = benchmark_ncr_dataset()
    wb_bytes = export_ncr_workbook(dataset)
    n = len(dataset.records)
    last_row = _last_row(n)
    matrix_prefix = f"'{_DEFAULT_TITLE}'!"

    for r in range(2, 9):
        assert_cell_is_formula(wb_bytes, _DISP_SUMMARY_SHEET_TITLE, f"B{r}")
        assert_cell_is_formula(wb_bytes, _DISP_SUMMARY_SHEET_TITLE, f"C{r}")
        assert_cell_is_formula(wb_bytes, _DISP_SUMMARY_SHEET_TITLE, f"D{r}")

    assert_cell_is_formula(wb_bytes, _SUMMARY_SHEET_TITLE, "B4")
    assert_cell_is_formula(wb_bytes, _SUMMARY_SHEET_TITLE, "B5")

    wb = _load(wb_bytes)
    ws_disp = wb[_DISP_SUMMARY_SHEET_TITLE]

    for idx, disp in enumerate(DISPOSITION_VALUES):
        r = 2 + idx
        assert ws_disp[f"B{r}"].value == f'=COUNTIF({matrix_prefix}{_DISP_COL}2:{_DISP_COL}{last_row}, "{disp}")'
        assert ws_disp[f"C{r}"].value == f'=SUMIF({matrix_prefix}{_DISP_COL}2:{_DISP_COL}{last_row}, "{disp}", {matrix_prefix}{_QTY_COL}2:{_QTY_COL}{last_row})'
        assert ws_disp[f"D{r}"].value == f"=IF($C$8=0, 0, C{r}/$C$8)"
        assert ws_disp[f"D{r}"].number_format == "0.0%"

    assert ws_disp["B7"].value == f'=COUNTIF({matrix_prefix}{_DISP_COL}2:{_DISP_COL}{last_row}, "")'
    assert ws_disp["C7"].value == f'=SUMIF({matrix_prefix}{_DISP_COL}2:{_DISP_COL}{last_row}, "", {matrix_prefix}{_QTY_COL}2:{_QTY_COL}{last_row})'
    assert ws_disp["D7"].value == "=IF($C$8=0, 0, C7/$C$8)"
    assert ws_disp["D7"].number_format == "0.0%"

    assert ws_disp["B8"].value == "=SUM(B2:B7)"
    assert ws_disp["C8"].value == "=SUM(C2:C7)"
    assert ws_disp["D8"].value == "=IF(C8=0, 0, SUM(D2:D7))"
    assert ws_disp["D8"].number_format == "0.0%"

    ws_summary = wb[_SUMMARY_SHEET_TITLE]
    assert ws_summary["B4"].value == f"=COUNTA({matrix_prefix}{_REC_ID_COL}2:{_REC_ID_COL}{last_row})"
    assert ws_summary["B5"].value == f"=SUM({matrix_prefix}{_QTY_COL}2:{_QTY_COL}{last_row})"


# ===========================================================================
# 3. Accuracy Scorecard
# ===========================================================================


def test_accuracy_scorecard_ncr() -> None:
    dataset = benchmark_ncr_dataset()
    wb_bytes = export_ncr_workbook(dataset)
    wb = _load(wb_bytes)
    matrix = wb[_DEFAULT_TITLE]

    total_records = len(dataset.records)
    total_qty = sum(r.quantity_affected for r in dataset.records)

    sheet_records = matrix.max_row - 1
    assert sheet_records == total_records

    sheet_qty = sum(matrix[f"{_QTY_COL}{r}"].value for r in range(2, matrix.max_row + 1))
    assert sheet_qty == total_qty

    for disp in DISPOSITION_VALUES:
        expected_count = sum(1 for r in dataset.records if r.disposition == disp)
        expected_disp_qty = sum(r.quantity_affected for r in dataset.records if r.disposition == disp)
        sheet_count = sum(1 for r in range(2, matrix.max_row + 1) if matrix[f"{_DISP_COL}{r}"].value == disp)
        sheet_disp_qty = sum(
            matrix[f"{_QTY_COL}{r}"].value
            for r in range(2, matrix.max_row + 1)
            if matrix[f"{_DISP_COL}{r}"].value == disp
        )
        assert sheet_count == expected_count
        assert sheet_disp_qty == expected_disp_qty

    ws_summary = wb[_SUMMARY_SHEET_TITLE]
    expected_mrb = sum(
        1
        for r in dataset.records
        if (
            r.disposition in ("UseAsIs", "Regrade")
            or (r.approval_authority is not None and "MRB" in r.approval_authority)
        )
    )
    assert ws_summary["B6"].value == expected_mrb


# ===========================================================================
# 4. Custom Title Handling
# ===========================================================================


def test_custom_title_threads_into_all_cross_sheet_formulas() -> None:
    custom_title = "Plant 4 Nonconformance Log"
    dataset = benchmark_ncr_dataset()
    wb_bytes = export_ncr_workbook(dataset, title=custom_title)
    wb = _load(wb_bytes)

    assert wb.sheetnames == [custom_title, _DISP_SUMMARY_SHEET_TITLE, _SUMMARY_SHEET_TITLE]

    ws_disp = wb[_DISP_SUMMARY_SHEET_TITLE]
    prefix = f"'{custom_title}'!"
    assert prefix in ws_disp["B2"].value
    assert prefix in ws_disp["C2"].value

    ws_summary = wb[_SUMMARY_SHEET_TITLE]
    assert ws_summary["B1"].value == custom_title
    assert prefix in ws_summary["B4"].value
    assert prefix in ws_summary["B5"].value

    for r in range(2, 9):
        assert_cell_is_formula(wb_bytes, _DISP_SUMMARY_SHEET_TITLE, f"B{r}")
        assert_cell_is_formula(wb_bytes, _DISP_SUMMARY_SHEET_TITLE, f"C{r}")
        assert_cell_is_formula(wb_bytes, _DISP_SUMMARY_SHEET_TITLE, f"D{r}")
    assert_cell_is_formula(wb_bytes, _SUMMARY_SHEET_TITLE, "B4")
    assert_cell_is_formula(wb_bytes, _SUMMARY_SHEET_TITLE, "B5")


# ===========================================================================
# 5. Negative Controls (Literal & Injection)
# ===========================================================================


def test_literal_negative_control_fails_formula_verifier() -> None:
    wb = openpyxl.Workbook()
    ws_disp = wb.active
    ws_disp.title = _DISP_SUMMARY_SHEET_TITLE
    ws_disp.cell(row=2, column=2, value=10)
    ws_disp.cell(row=2, column=3, value=450)
    ws_disp.cell(row=2, column=4, value=0.5)

    ws_summary = wb.create_sheet(_SUMMARY_SHEET_TITLE)
    ws_summary.cell(row=4, column=2, value=5)
    ws_summary.cell(row=5, column=2, value=1200)

    saved = _saved(wb)

    for coord in ("B2", "C2", "D2"):
        with pytest.raises(AssertionError, match="literal, not a live formula"):
            assert_cell_is_formula(saved, _DISP_SUMMARY_SHEET_TITLE, coord)

    for coord in ("B4", "B5"):
        with pytest.raises(AssertionError, match="literal, not a live formula"):
            assert_cell_is_formula(saved, _SUMMARY_SHEET_TITLE, coord)


def test_security_invariant_formula_injection_defense() -> None:
    malicious_inputs = [
        ("=SUM(A1:A10)", "formula starting with ="),
        ("+1+1", "formula starting with +"),
        ("-1-1", "negative formula prefix"),
        ("@SUM(B1:B5)", "formula starting with @"),
        ("\t=CALC()", "tab prefixed formula"),
        ("\r+2+2", "CR prefixed formula"),
    ]

    records = [
        _make_record(
            record_id=f"NCR-INJ-{i}",
            part_lot_id=f"=LOT-{i}",
            defect_description=payload,
            requirement_violated=f"+REQ-{i}",
            rationale=f"-RATIONALE-{i}",
            approval_authority=f"@AUTHORITY-{i}",
            quantity_affected=10,
            disposition="Scrap",
        )
        for i, (payload, _) in enumerate(malicious_inputs)
    ]

    dataset = NCRDataset(records=records)
    malicious_title = "=DANGEROUS_TITLE"
    wb_bytes = export_ncr_workbook(dataset, title=malicious_title)
    wb = _load(wb_bytes)

    # The title on sheet 1 tab is set
    assert malicious_title in wb.sheetnames

    matrix = wb[malicious_title]
    # Matrix cells must have leading apostrophe
    for r in range(2, len(records) + 2):
        desc_cell = matrix[f"C{r}"].value
        assert str(desc_cell).startswith("'"), f"Defect description was not escaped: {desc_cell!r}"
        # Matrix cells must NOT be formulas
        with pytest.raises(AssertionError, match="literal, not a live formula"):
            assert_cell_is_formula(wb_bytes, malicious_title, f"C{r}")

    # Summary sheet title value must be escaped
    ws_summary = wb[_SUMMARY_SHEET_TITLE]
    assert ws_summary["B1"].value == "'" + malicious_title

    # But computed rollups must STILL be live formulas
    for r in range(2, 9):
        assert_cell_is_formula(wb_bytes, _DISP_SUMMARY_SHEET_TITLE, f"B{r}")
        assert_cell_is_formula(wb_bytes, _DISP_SUMMARY_SHEET_TITLE, f"C{r}")
        assert_cell_is_formula(wb_bytes, _DISP_SUMMARY_SHEET_TITLE, f"D{r}")
    assert_cell_is_formula(wb_bytes, _SUMMARY_SHEET_TITLE, "B4")
    assert_cell_is_formula(wb_bytes, _SUMMARY_SHEET_TITLE, "B5")


# ===========================================================================
# 6. Empty Dataset & Boundary Conditions
# ===========================================================================


def test_empty_dataset_exports_cleanly_without_division_errors() -> None:
    """Empty dataset (e.g. empty sequence) produces valid 3-sheet workbook with floored 2:2 ranges and zero ratios."""
    wb_bytes = export_ncr_workbook([])
    wb = _load(wb_bytes)

    matrix = wb[_DEFAULT_TITLE]
    assert matrix.max_row == 1

    ws_disp = wb[_DISP_SUMMARY_SHEET_TITLE]
    assert ws_disp["B2"].value == f'=COUNTIF(\'{_DEFAULT_TITLE}\'!{_DISP_COL}2:{_DISP_COL}2, "Scrap")'
    assert ws_disp["C2"].value == f'=SUMIF(\'{_DEFAULT_TITLE}\'!{_DISP_COL}2:{_DISP_COL}2, "Scrap", \'{_DEFAULT_TITLE}\'!{_QTY_COL}2:{_QTY_COL}2)'
    assert ws_disp["D2"].value == "=IF($C$8=0, 0, C2/$C$8)"

    ws_summary = wb[_SUMMARY_SHEET_TITLE]
    assert ws_summary["B4"].value == f"=COUNTA('{_DEFAULT_TITLE}'!{_REC_ID_COL}2:{_REC_ID_COL}2)"
    assert ws_summary["B5"].value == f"=SUM('{_DEFAULT_TITLE}'!{_QTY_COL}2:{_QTY_COL}2)"
    assert ws_summary["B6"].value == 0

    for r in range(2, 9):
        assert_cell_is_formula(wb_bytes, _DISP_SUMMARY_SHEET_TITLE, f"B{r}")
        assert_cell_is_formula(wb_bytes, _DISP_SUMMARY_SHEET_TITLE, f"C{r}")
        assert_cell_is_formula(wb_bytes, _DISP_SUMMARY_SHEET_TITLE, f"D{r}")
    assert_cell_is_formula(wb_bytes, _SUMMARY_SHEET_TITLE, "B4")
    assert_cell_is_formula(wb_bytes, _SUMMARY_SHEET_TITLE, "B5")


# ===========================================================================
# 7. Input Polymorphism & Validation
# ===========================================================================


def test_input_polymorphism_and_unassigned_disposition() -> None:
    import pandas as pd

    record_dicts = [_SAMPLE_RECORD_DATA.copy()]
    records = [_make_record()]
    df = pd.DataFrame(record_dicts)
    dict_payload = {"records": record_dicts}
    rows_dict_payload = {"rows": record_dicts}

    wb1 = _load(export_ncr_workbook(records))
    assert wb1[_DEFAULT_TITLE].max_row == 2

    wb2 = _load(export_ncr_workbook(record_dicts))
    assert wb2[_DEFAULT_TITLE].max_row == 2

    wb3 = _load(export_ncr_workbook(df))
    assert wb3[_DEFAULT_TITLE].max_row == 2

    wb4 = _load(export_ncr_workbook(dict_payload))
    assert wb4[_DEFAULT_TITLE].max_row == 2

    wb5 = _load(export_ncr_workbook(rows_dict_payload))
    assert wb5[_DEFAULT_TITLE].max_row == 2

    wb_empty_list = _load(export_ncr_workbook([]))
    assert wb_empty_list[_DEFAULT_TITLE].max_row == 1

    wb_empty_df = _load(export_ncr_workbook(pd.DataFrame()))
    assert wb_empty_df[_DEFAULT_TITLE].max_row == 1

    wb_empty_dict1 = _load(export_ncr_workbook({"records": []}))
    assert wb_empty_dict1[_DEFAULT_TITLE].max_row == 1

    wb_empty_dict2 = _load(export_ncr_workbook({"rows": []}))
    assert wb_empty_dict2[_DEFAULT_TITLE].max_row == 1

    unassigned_record = _make_record(disposition=None, approval_authority="Plant QA")
    wb_unassigned = _load(export_ncr_workbook([unassigned_record]))
    assert wb_unassigned[_DEFAULT_TITLE]["H2"].value is None


# ===========================================================================
# 8. MRB Gate Reviews Logic & Nullable Pass-Through
# ===========================================================================


def test_mrb_gate_reviews_logic_coverage() -> None:
    records = [
        _make_record(record_id="NCR-1", disposition="UseAsIs", approval_authority="Plant Manager"),
        _make_record(record_id="NCR-2", disposition="Regrade", approval_authority="Engineering"),
        _make_record(record_id="NCR-3", disposition="Scrap", approval_authority="MRB Chairperson"),
        _make_record(record_id="NCR-4", disposition="Rework", approval_authority="Production Lead"),
        _make_record(record_id="NCR-5", disposition="Scrap", approval_authority=None),
    ]
    wb = _load(export_ncr_workbook(records))
    ws_summary = wb[_SUMMARY_SHEET_TITLE]
    assert ws_summary["B6"].value == 3


def test_none_valued_nullable_fields_pass_through_without_string_coercion() -> None:
    record = _make_record(
        severity=None,
        disposition=None,
        approval_authority=None,
        rationale=None,
    )
    wb = _load(export_ncr_workbook([record]))
    matrix = wb[_DEFAULT_TITLE]

    for col_name in ("Severity", "Disposition", "Approval_Authority", "Rationale"):
        col_letter = get_column_letter(NCR_EXPORT_COLUMNS.index(col_name) + 1)
        val = matrix[f"{col_letter}2"].value
        assert val is None, f"{col_name} should be None, got {val!r}"
        assert val != "None"


def test_write_metadata_sheet_without_title() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    _write_metadata_sheet(ws, [("Key", "Value")], title=None)
    assert ws["A1"].value == "Key"
    assert ws["B1"].value == "Value"


# ===========================================================================
# 9. Re-Export Integrity & Fresh Benchmarks
# ===========================================================================


def test_reexport_integrity_ncr() -> None:
    import quality_core.ncr as pkg
    import quality_core.ncr.export as mod

    assert pkg.build_ncr_workbook is mod.build_ncr_workbook
    assert pkg.export_ncr_workbook is mod.export_ncr_workbook
    assert pkg.export_ncr_excel is mod.export_ncr_excel
    assert pkg.benchmark_ncr_dataset is mod.benchmark_ncr_dataset
    assert pkg.NCR_EXPORT_COLUMNS is mod.NCR_EXPORT_COLUMNS
    assert pkg.NCR_COL_WIDTHS is mod.NCR_COL_WIDTHS
    assert pkg.DISPOSITION_SUMMARY_COLUMNS is mod.DISPOSITION_SUMMARY_COLUMNS


def test_benchmark_dataset_returns_fresh_instances() -> None:
    a = benchmark_ncr_dataset()
    b = benchmark_ncr_dataset()
    assert a is not b
    assert a.records is not b.records
    assert len(a.records) == 5

    dispositions_present = {r.disposition for r in a.records}
    assert dispositions_present == set(DISPOSITION_VALUES)
