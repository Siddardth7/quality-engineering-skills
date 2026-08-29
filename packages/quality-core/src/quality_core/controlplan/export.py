"""
quality_core/controlplan/export.py
Control Plan Excel exporter — three live linkage-coverage roll-up formulas.

Consumer of the domain-agnostic ``quality_core.io.export`` primitives: that module owns
the cross-cutting machinery (formula-injection escaping, openpyxl table styling, the
``Formula`` opt-in marker) and stays free of domain knowledge; this module supplies the
Control Plan *config* — column order, widths, and which cells are live formulas.

It lives in the ``controlplan`` domain package rather than alongside ``io``'s primitives
because the dependency only points one way: ``controlplan`` may import ``io`` (and
``controlplan.schema`` already does, for the ingest boundary), but ``io`` sits structurally
below every domain engine and must never import one back. Putting a Control Plan exporter
inside ``io`` would close that arrow into a cycle.

What ships in the workbook (two sheets, in this order):

  1. the **matrix** sheet, named by ``title`` (default ``"Control Plan"``). Every cell on
     it is plain data from :class:`~quality_core.controlplan.ControlPlanRow`, written
     through ``write_table_sheet``'s ordinary ``sanitize_cell`` escaping. **No cell on
     this sheet is a live formula.** Two of its columns are exporter-authored *structured
     display strings*, not formulas: ``Sample_Plan_Placeholder`` and ``PFMEA_Linked``,
     each a plain ``"Yes"``/``"No"`` rendering of one row field. They go through the same
     escaper as every other data cell — there is no special case that skips it.
  2. the **``"Coverage"``** sheet, holding the only three
     :class:`~quality_core.io.export.Formula` cells in the workbook, at fixed addresses
     that never move with ``len(dataset.rows)``:

     ``B2`` ``=COUNTA(...)`` over the matrix ``Characteristic`` column — total
     characteristics; ``B3`` ``=COUNTIF(...,"Yes")`` over the matrix ``PFMEA_Linked``
     column — linked characteristics; ``B4`` ``=IF(B2=0,0,B3/B2)`` — linkage coverage as
     a ratio, rendered ``0.0%``. The ``IF`` guard is unconditional, so an empty dataset
     shows ``0`` rather than ``#DIV/0!``.

**"Linked" here means *declared* linkage, not *validated* linkage.** The ``PFMEA_Linked``
column and the ``B3``/``B4`` roll-ups count rows where ``source_cause_id is not None`` —
rows that *declare* an FMEA source. That is a strictly weaker claim than
:func:`quality_core.controlplan.connector.validate_pfmea_linkage`, which additionally requires the
ID to resolve against a specific ``RelationalFMEA``; that check needs both objects and is
out of reach of a dataset-only exporter call, so it is deliberately not made here. The
exported "Linkage Coverage %" must never be read as a validated-linkage number.

Row order is the caller's: rows are exported in ``dataset.rows`` list order, **never
re-sorted**. Column letters are derived from ``CONTROLPLAN_EXPORT_COLUMNS`` via
``get_column_letter``, never hand-typed, so reordering the column tuple cannot silently
desync a roll-up formula from the column it is supposed to count.

No new standards constant lives here — see ``controlplan/ASSUMPTIONS_LOG.md`` (RULE 4).
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Imported from the sibling module rather than the ``quality_core.controlplan`` package,
# so this module is importable on its own and cannot see a partially-initialised package
# when it is loaded from that package's own ``__init__``.
from quality_core.controlplan.schema import ControlPlanDataset, ControlPlanRow
from quality_core.io import (
    Formula,
    sanitize_cell,
    write_formula_cell,
    write_table_sheet,
)

# ===========================================================================
# Matrix sheet layout
# ===========================================================================

#: Exported matrix columns, in sheet order (1-based column index = position + 1).
#: Every ``ControlPlanRow`` field is represented — nothing is silently dropped — plus
#: the two derived ``"Yes"``/``"No"`` display columns.
CONTROLPLAN_EXPORT_COLUMNS: tuple[str, ...] = (
    "Characteristic",
    "LSL",
    "Target",
    "USL",
    "Measurement_Method",
    "Sample_Size",
    "Frequency",
    "Recommended_Chart",
    "Reaction_Plan",
    "Source_Cause_ID",
    "Sample_Plan_Placeholder",
    "PFMEA_Linked",
)

#: Column widths keyed by column name — wide for free text, narrow for numbers and
#: the Yes/No flags. Presentation only; ``write_table_sheet`` falls back to
#: ``default_width`` for anything missing, but every named column is given one here.
CONTROLPLAN_COL_WIDTHS: dict[str, float] = {
    "Characteristic": 34,
    "LSL": 9,
    "Target": 9,
    "USL": 9,
    "Measurement_Method": 28,
    "Sample_Size": 11,
    "Frequency": 18,
    "Recommended_Chart": 17,
    "Reaction_Plan": 34,
    "Source_Cause_ID": 18,
    "Sample_Plan_Placeholder": 21,
    "PFMEA_Linked": 13,
}

#: Sheet row of the first matrix data row: row 1 is the header written by
#: ``write_table_sheet``, whose body enumeration starts at 2.
_FIRST_DATA_ROW = 2


def _column_letter(name: str) -> str:
    """Return the Excel column letter for ``name``'s position in the matrix layout."""
    return get_column_letter(CONTROLPLAN_EXPORT_COLUMNS.index(name) + 1)


_CHARACTERISTIC_COL = _column_letter("Characteristic")
_PFMEA_LINKED_COL = _column_letter("PFMEA_Linked")


def _yes_no(flag: bool) -> str:
    """Render a boolean as the sheet's ``"Yes"``/``"No"`` display string."""
    return "Yes" if flag else "No"


def _row_record(row: ControlPlanRow) -> dict[str, object]:
    """Map one ``ControlPlanRow`` to its sheet record, keyed by the matrix columns.

    ``lsl``/``target``/``usl``/``recommended_chart``/``source_cause_id`` are nullable and
    pass through unchanged — ``write_table_sheet`` writes ``None`` as a blank cell, so no
    coercion (which would turn a blank into the literal text ``"None"``) happens here.
    """
    return {
        "Characteristic": row.characteristic,
        "LSL": row.lsl,
        "Target": row.target,
        "USL": row.usl,
        "Measurement_Method": row.measurement_method,
        "Sample_Size": row.sample_size,
        "Frequency": row.frequency,
        "Recommended_Chart": row.recommended_chart,
        "Reaction_Plan": row.reaction_plan,
        "Source_Cause_ID": row.source_cause_id,
        "Sample_Plan_Placeholder": _yes_no(row.sample_plan_is_placeholder),
        # Declared linkage, not validated linkage — see the module docstring.
        "PFMEA_Linked": _yes_no(row.source_cause_id is not None),
    }


# ===========================================================================
# Coverage sheet — the workbook's only live formulas
# ===========================================================================

#: Name of the second sheet. Its metric rows are at fixed addresses (B2/B3/B4)
#: precisely because they must not move with the matrix's row count.
_COVERAGE_SHEET_TITLE = "Coverage"

_COVERAGE_HEADER: tuple[str, str] = ("Metric", "Value")
_COVERAGE_LABELS: tuple[str, str, str] = (
    "Total Characteristics",
    "PFMEA-Linked Characteristics",
    "Linkage Coverage %",
)

#: Sheet row of the first Coverage metric (row 1 is the Metric/Value header).
_COVERAGE_FIRST_METRIC_ROW = 2
_COVERAGE_LABEL_COLUMN = 1
_COVERAGE_VALUE_COLUMN = 2
_COVERAGE_LABEL_WIDTH = 30.0
_COVERAGE_VALUE_WIDTH = 14.0

#: Ratio of B3 to B2, guarded so an empty dataset yields ``0`` rather than
#: ``#DIV/0!``. Same-sheet references, so no sheet qualification is needed; the
#: string is a constant — it never varies with the data, only its result does.
_COVERAGE_RATIO_FORMULA = "=IF(B2=0,0,B3/B2)"
_COVERAGE_RATIO_NUMBER_FORMAT = "0.0%"


def _last_matrix_row(row_count: int) -> int:
    """Return the last matrix sheet row the roll-up ranges should span.

    ``_FIRST_DATA_ROW + row_count - 1`` is the last row ``write_table_sheet`` actually
    wrote. The floor at ``_FIRST_DATA_ROW`` is the **empty-dataset guard**: with zero
    rows that expression is 1, which would build the invalid reversed range ``2:1``.
    Flooring makes it ``2:2`` — a single row nothing was ever written to, which
    ``COUNTA``/``COUNTIF`` correctly evaluate as empty. The floor only ever applies at
    ``row_count == 0``; for one row or more the value is unchanged.
    """
    return max(_FIRST_DATA_ROW + row_count - 1, _FIRST_DATA_ROW)


def _coverage_formulas(matrix_title: str, row_count: int) -> tuple[Formula, Formula, Formula]:
    """Return the (total, linked, coverage) live formulas for the Coverage sheet.

    Exporter-authored: the only ``Formula`` instances this module ever builds, from the
    fixed templates below plus column letters it derives itself. ``matrix_title`` is
    quoted unconditionally — the default title contains a space, and Excel requires
    quoting any sheet name that is not guaranteed free of spaces/specials.

    ``COUNTA`` targets ``Characteristic`` because that field is required and non-blank on
    every row, so it can neither under- nor over-count against ``len(dataset.rows)``.
    ``COUNTIF`` targets the ``PFMEA_Linked`` sentinel rather than counting non-blank
    ``Source_Cause_ID`` cells, because a ``None`` source ID may serialize as either a
    blank cell or an empty string; the ``"Yes"``/``"No"`` column is always a real
    non-blank string, so the exact match is unambiguous.
    """
    last_row = _last_matrix_row(row_count)
    matrix = f"'{matrix_title}'!"
    total = Formula(
        f"=COUNTA({matrix}{_CHARACTERISTIC_COL}{_FIRST_DATA_ROW}"
        f":{_CHARACTERISTIC_COL}{last_row})"
    )
    linked = Formula(
        f"=COUNTIF({matrix}{_PFMEA_LINKED_COL}{_FIRST_DATA_ROW}"
        f':{_PFMEA_LINKED_COL}{last_row},"Yes")'
    )
    coverage = Formula(_COVERAGE_RATIO_FORMULA, _COVERAGE_RATIO_NUMBER_FORMAT)
    return total, linked, coverage


def _write_coverage_sheet(ws: Any, matrix_title: str, row_count: int) -> None:
    """Write the fixed three-metric Coverage sheet into ``ws``.

    Column A holds plain sanitized labels (written the way ``write_keyvalue_sheet``
    writes its keys); column B holds the three live formulas. Built with direct cell
    calls rather than ``write_table_sheet``/``write_keyvalue_sheet`` because this is a
    fixed three-metric block, and ``write_keyvalue_sheet`` has no ``Formula`` support.
    """
    for col_idx, label in enumerate(_COVERAGE_HEADER, start=1):
        ws.cell(row=1, column=col_idx, value=sanitize_cell(label))

    formulas = _coverage_formulas(matrix_title, row_count)
    for offset, (label, formula) in enumerate(zip(_COVERAGE_LABELS, formulas, strict=True)):
        sheet_row = _COVERAGE_FIRST_METRIC_ROW + offset
        ws.cell(row=sheet_row, column=_COVERAGE_LABEL_COLUMN, value=sanitize_cell(label))
        write_formula_cell(
            ws,
            sheet_row,
            _COVERAGE_VALUE_COLUMN,
            formula.formula,
            number_format=formula.number_format,
        )

    ws.column_dimensions["A"].width = _COVERAGE_LABEL_WIDTH
    ws.column_dimensions["B"].width = _COVERAGE_VALUE_WIDTH


# ===========================================================================
# Public API
# ===========================================================================


def export_controlplan_workbook(
    dataset: ControlPlanDataset, *, title: str = "Control Plan"
) -> bytes:
    """Export ``dataset`` to a two-sheet .xlsx workbook with live coverage roll-ups.

    Returns the serialized workbook as bytes, ready to hand to a download button or write
    to disk. ``title`` names the matrix worksheet (default ``"Control Plan"``) and is the
    sheet name the ``Coverage`` formulas qualify their ranges with, so a custom title
    stays consistent across both sheets. It is passed to ``write_table_sheet``
    unvalidated, matching every other caller.

    Rows are written in ``dataset.rows`` order. An empty dataset yields a valid workbook
    whose matrix carries the header row only and whose ``Coverage`` formulas evaluate
    to ``0``.
    """
    records = [_row_record(row) for row in dataset.rows]
    # Explicit ``columns=`` so an empty dataset still carries the full header: a
    # DataFrame built from an empty record list would otherwise have no columns at
    # all, and ``write_table_sheet`` would filter every column out.
    df = pd.DataFrame(records, columns=list(CONTROLPLAN_EXPORT_COLUMNS))

    wb = Workbook()
    write_table_sheet(
        wb.active,
        df,
        title=title,
        columns=CONTROLPLAN_EXPORT_COLUMNS,
        col_widths=CONTROLPLAN_COL_WIDTHS,
    )
    _write_coverage_sheet(wb.create_sheet(_COVERAGE_SHEET_TITLE), title, len(dataset.rows))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===========================================================================
# Benchmark dataset
# ===========================================================================

# Self-contained automotive Control Plan rows, built straight from ``ControlPlanRow``.
# Deliberately NOT built via ``connector.build_control_plan`` (which needs a whole
# ``RelationalFMEA`` to derive rows from) and NOT imported from ``quality_core.canvas``
# (which sits structurally above this package, so depending on it here would invert the
# layering). Four of the six rows carry a ``source_cause_id`` in the
# connector's own ``f"{function.id}::{failure_mode.id}::{cause.id}"`` join-key shape and
# two carry ``None``, so the exported coverage is a genuine partial fraction (4/6) rather
# than a degenerate 0% or 100%. Tolerance fields, ``recommended_chart``, and the
# placeholder flag each span both the set and unset case.

_BENCHMARK_ROWS: tuple[ControlPlanRow, ...] = (
    ControlPlanRow(
        characteristic="Brake caliper bore diameter",
        lsl=42.90,
        target=43.00,
        usl=43.10,
        measurement_method="Three-point air gauge",
        sample_size=5,
        frequency="Every 30 minutes",
        recommended_chart="Xbar-R",
        reaction_plan="Quarantine the interval's parts, re-gauge, notify process engineering",
        source_cause_id="F1::FM1::C1",
    ),
    ControlPlanRow(
        characteristic="Steering column weld penetration depth",
        lsl=1.20,
        target=1.50,
        usl=1.80,
        measurement_method="Destructive macro-section",
        sample_size=1,
        frequency="Once per shift",
        recommended_chart="I-MR",
        reaction_plan="Hold the shift's production and escalate to weld engineering",
        source_cause_id="F2::FM1::C2",
        sample_plan_is_placeholder=True,
    ),
    ControlPlanRow(
        characteristic="Stator winding dielectric withstand",
        lsl=1500.0,
        target=None,
        usl=None,
        measurement_method="In-line high-potential tester",
        sample_size=1,
        frequency="Every part",
        recommended_chart="p",
        reaction_plan="Scrap the failing stator and audit the impregnation batch",
        source_cause_id="F3::FM2::C1",
    ),
    ControlPlanRow(
        characteristic="Gearbox output seal press depth",
        lsl=None,
        target=None,
        usl=None,
        measurement_method="Press-force displacement monitor",
        sample_size=1,
        frequency="Every part",
        reaction_plan="Reject the assembly and verify press setup before restarting",
    ),
    ControlPlanRow(
        characteristic="Inverter gate-drive rise time",
        lsl=0.80,
        target=1.00,
        usl=1.20,
        measurement_method="Oscilloscope capture at end-of-line bench",
        sample_size=3,
        frequency="Hourly",
        recommended_chart="Xbar-S",
        reaction_plan="Contain the hour's inverters and re-run the gate-drive sweep",
        source_cause_id="F4::FM1::C3",
    ),
    ControlPlanRow(
        characteristic="Door trim panel gap",
        lsl=None,
        target=None,
        usl=2.50,
        measurement_method="Feeler gauge appearance check",
        sample_size=2,
        frequency="First-off each shift",
        reaction_plan="Adjust clip retention fixture and re-check the first-off part",
    ),
)


def benchmark_controlplan_dataset() -> ControlPlanDataset:
    """Return a self-contained benchmark ``ControlPlanDataset`` for exports and demos.

    Six automotive Control Plan characteristics, four of which declare a
    ``source_cause_id`` and two of which do not — so the exported linkage coverage is a
    genuine partial fraction (4 of 6) rather than 0% or 100%. A fresh
    ``ControlPlanDataset`` is built per call, so a caller mutating the returned dataset
    cannot affect the next one.
    """
    return ControlPlanDataset(rows=list(_BENCHMARK_ROWS))
