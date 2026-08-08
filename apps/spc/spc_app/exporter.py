"""
exporter.py
Manufacturing SPC Dashboard — Export Layer

SPC-specific export *config* (which fields, metadata, layout) composed over the
shared, app-agnostic primitives in ``quality_core.io.export`` (CSV/formula-injection
escaping, openpyxl styling, fpdf2 table rendering). The cross-cutting machinery is
written once in core and reused here, exactly as the FMEA exporter does.

Two report kinds, each rendered to Excel (.xlsx) and PDF:

    Control chart report  — points, control limits, rule violations, chart metrics
    Capability report     — Cp/Cpk/Pp/Ppk, distribution summary, normality, stability

Builders take a frozen report dataclass (the values a page has already computed) and
return raw bytes suitable for ``st.download_button``.
"""

from __future__ import annotations

import io
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from typing import Any

import openpyxl
import pandas as pd
from quality_core.io.export import (
    fmt,
    fmt_opt,
    generated_line,
    now,
    pdf_subheader,
    pdf_summary_cells,
    pdf_title,
    render_table,
    safe_text,
    sanitize_cell,
    sanitize_for_export,
    write_keyvalue_sheet,
    write_table_sheet,
)

from spc_app import __version__

_TOOL_VERSION = __version__
_ENGINEERING_REF = "AIAG SPC Reference Manual, 4th Ed. (2005)"

# Cpk capability bands — mirror the Capability page's interpretation table, which
# cites docs/ASSUMPTIONS_LOG.md (1.33 = common minimum target; 1.00 = not capable).
_CPK_CAPABLE = 1.33
_CPK_MARGINAL = 1.00

_VIOLATION_FILL_HEX = "F8D7DA"  # light red for out-of-control rows
_VIOLATION_RGB = (248, 215, 218)
_WHITE_RGB = (255, 255, 255)


# ===========================================================================
# Report inputs
# ===========================================================================


@dataclass(frozen=True)
class ControlChartReport:
    """Everything the Control Charts page already computed for one chart."""

    chart_label: str
    stream: str
    rule_set: str
    points: Sequence[float]
    cl: float
    # p- and u-charts have per-point (vector) control limits; the others are scalar.
    ucl: float | Sequence[float]
    lcl: float | Sequence[float]
    violations: Sequence[Mapping[str, Any]]  # [{"index": int, "rule": str}, ...]
    metrics: Sequence[tuple[str, str]]  # summarize_metrics() output
    # Optional second per-point series (e.g. CUSUM's C- lower arm, stored as the
    # positive accumulator) rendered as its own column, (label, values). `None`
    # (the default) keeps every existing caller/report byte-for-byte unchanged.
    secondary_points: tuple[str, Sequence[float]] | None = None


@dataclass(frozen=True)
class CapabilityReport:
    """Everything the Process Capability page already computed for one stream."""

    stream_label: str
    values: Sequence[float]
    capability: Mapping[str, Any]  # cp/cpk/pp/ppk/mean/sigma_hat/sigma_overall
    lsl: float | None
    usl: float | None
    normality: Mapping[str, Any]  # is_normal/p_value/w_stat
    oos_signal_count: int


# ===========================================================================
# Formatting helpers
# ===========================================================================


def _fmt_ci(ci: tuple[float, float] | None) -> str:
    return "N/A" if ci is None else f"[{ci[0]:.4f}, {ci[1]:.4f}]"


def _limit_at(limit: float | Sequence[float], index: int) -> float:
    """Resolve a control limit at a point: a scalar applies to every point, a
    vector (p/u charts) gives the per-point limit."""
    if isinstance(limit, (int, float)):
        return float(limit)
    return float(limit[index])


def _fmt_limit(limit: float | Sequence[float]) -> str:
    """A scalar limit prints as a number; a vector limit varies per subgroup."""
    if isinstance(limit, (int, float)):
        return fmt(float(limit))
    return "varies (per subgroup)"


def _cpk_rating(cpk: float | None) -> str:
    if cpk is None:
        return "N/A"
    if cpk >= _CPK_CAPABLE:
        return "Capable"
    if cpk >= _CPK_MARGINAL:
        return "Marginal"
    return "Not capable"


def _violations_by_index(report: ControlChartReport) -> dict[int, list[str]]:
    by_index: dict[int, list[str]] = {}
    for violation in report.violations:
        by_index.setdefault(int(violation["index"]), []).append(str(violation["rule"]))
    return by_index


_POINT_COLUMNS = ["Point", "Value", "UCL", "LCL", "Status"]


def _point_columns(report: ControlChartReport) -> list[str]:
    """Column order for the per-point table — `Value` plus, when present, the
    secondary series (e.g. CUSUM's C-) right next to it, then UCL/LCL/Status."""
    if report.secondary_points is None:
        return _POINT_COLUMNS
    label, _ = report.secondary_points
    return ["Point", "Value", label, "UCL", "LCL", "Status"]


def _points_frame(report: ControlChartReport) -> pd.DataFrame:
    """Per-point table: 1-based Point, Value, the UCL/LCL it was tested against
    (constant for most charts, per-point for p/u), and OK / rule-violation Status.
    Optionally a secondary per-point series (e.g. CUSUM's C- lower arm)."""
    by_index = _violations_by_index(report)
    secondary_label: str | None = None
    secondary_values: Sequence[float] | None = None
    if report.secondary_points is not None:
        secondary_label, secondary_values = report.secondary_points

    rows = []
    for index, value in enumerate(report.points):
        row: dict[str, Any] = {"Point": index + 1, "Value": round(float(value), 6)}
        if secondary_label is not None and secondary_values is not None:
            row[secondary_label] = round(float(secondary_values[index]), 6)
        row["UCL"] = round(_limit_at(report.ucl, index), 6)
        row["LCL"] = round(_limit_at(report.lcl, index), 6)
        row["Status"] = "; ".join(by_index[index]) if index in by_index else "OK"
        rows.append(row)
    return pd.DataFrame(rows, columns=_point_columns(report))


def _values_frame(values: Sequence[float]) -> pd.DataFrame:
    rows = [{"Point": i + 1, "Value": round(float(v), 6)} for i, v in enumerate(values)]
    return pd.DataFrame(rows, columns=["Point", "Value"])


# ===========================================================================
# Control chart report
# ===========================================================================


def _control_chart_summary_rows(report: ControlChartReport) -> list[tuple[str, object]]:
    rows: list[tuple[str, object]] = [
        ("Generated", now()),
        ("Tool Version", _TOOL_VERSION),
        ("Engineering Ref", _ENGINEERING_REF),
        ("", ""),
        ("Chart Type", sanitize_cell(report.chart_label)),
        ("Process Stream", sanitize_cell(report.stream)),
        ("Rule Set", sanitize_cell(report.rule_set)),
        ("Center Line (CL)", fmt(report.cl)),
        ("UCL", _fmt_limit(report.ucl)),
        ("LCL", _fmt_limit(report.lcl)),
        ("Data Points", len(report.points)),
        ("Rule Violations", len(report.violations)),
        ("", ""),
    ]
    # Metrics are app-formatted numeric strings (e.g. "10.0000", "-3.0000"). Since #198
    # write_keyvalue_sheet sanitizes everything it writes, so these do pass through
    # sanitize_cell — they survive intact because a plain decimal is exempt, not because
    # this call site is trusted.
    rows += list(report.metrics)
    if report.violations:
        rows.append(("", ""))
        for violation in report.violations:
            point = int(violation["index"]) + 1
            rows.append((f"Violation @ Point {point}", sanitize_cell(str(violation["rule"]))))
    return rows


def _control_chart_row_fill(row: pd.Series) -> str | None:
    return _VIOLATION_FILL_HEX if str(row.get("Status")) != "OK" else None


def build_control_chart_report_excel(report: ControlChartReport) -> bytes:
    """Excel workbook: a coloured per-point sheet + a summary/metadata sheet."""
    columns = _point_columns(report)
    points = sanitize_for_export(_points_frame(report))
    col_widths = {"Point": 8, "Value": 14, "UCL": 14, "LCL": 14, "Status": 40}
    if report.secondary_points is not None:
        col_widths[report.secondary_points[0]] = 14

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None  # a freshly created workbook always has an active sheet
    write_table_sheet(
        ws,
        points,
        title="Control Chart",
        columns=columns,
        col_widths=col_widths,
        row_fill_hex=_control_chart_row_fill,
    )
    write_keyvalue_sheet(wb.create_sheet("Summary"), _control_chart_summary_rows(report))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_control_chart_report_pdf(report: ControlChartReport) -> bytes:
    """PDF report: title, summary metric strip, and a per-point violations table."""
    from fpdf import FPDF

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.set_margins(10, 10, 10)
    pdf.add_page()

    pdf_title(pdf, "SPC Control Chart Report")
    pdf_subheader(pdf, generated_line(f"{report.chart_label}  |  {_ENGINEERING_REF}"))
    pdf_summary_cells(
        pdf,
        [
            ("CL", fmt(report.cl)),
            ("UCL", _fmt_limit(report.ucl)),
            ("LCL", _fmt_limit(report.lcl)),
            ("Points", str(len(report.points))),
            ("Violations", str(len(report.violations))),
        ],
    )

    if report.secondary_points is None:
        pdf_columns = [("Point", 18), ("Value", 30), ("UCL", 30), ("LCL", 30), ("Status", 82)]
    else:
        label = report.secondary_points[0]
        pdf_columns = [
            ("Point", 18), ("Value", 26), (label, 26), ("UCL", 26), ("LCL", 26), ("Status", 68),
        ]

    render_table(
        pdf,
        _points_frame(report),
        columns=pdf_columns,
        row_values=lambda r: [safe_text(str(r[name])) for name, _ in pdf_columns],
        row_rgb=lambda r: _VIOLATION_RGB if str(r["Status"]) != "OK" else _WHITE_RGB,
    )
    return bytes(pdf.output())


# ===========================================================================
# Capability report
# ===========================================================================


def _capability_detail_rows(report: CapabilityReport) -> list[tuple[str, object]]:
    cap = report.capability
    norm = report.normality
    stability = (
        "In statistical control"
        if report.oos_signal_count == 0
        else f"{report.oos_signal_count} WE signal(s) — Cpk indicative only"
    )
    return [
        ("Process Stream", sanitize_cell(report.stream_label)),
        ("Data Points", len(report.values)),
        ("LSL", fmt_opt(report.lsl)),
        ("USL", fmt_opt(report.usl)),
        ("Method", str(cap.get("method", "normal"))),
        ("Box-Cox lambda", fmt_opt(cap.get("lambda_used"))),
        ("Cp", fmt_opt(cap["cp"])),
        ("Cpk", fmt_opt(cap["cpk"])),
        # Populated on the percentile method only (bootstrap CIs); the parametric
        # χ²/Bissell CIs live on Pp/Ppk below (#193).
        ("Cp 95% CI", _fmt_ci(cap.get("cp_ci"))),
        ("Cpk 95% CI", _fmt_ci(cap.get("cpk_ci"))),
        ("Cpk lower bound", fmt_opt(cap.get("cpk_lower"))),
        ("Pp", fmt_opt(cap["pp"])),
        ("Pp 95% CI", _fmt_ci(cap.get("pp_ci"))),
        ("Ppk", fmt_opt(cap["ppk"])),
        ("Ppk 95% CI", _fmt_ci(cap.get("ppk_ci"))),
        ("Ppk lower bound", fmt_opt(cap.get("ppk_lower"))),
        ("CI basis", f"{cap.get('ci_estimator')} (df={cap.get('ci_df')})"),
        ("Cpk Rating", _cpk_rating(cap["cpk"])),
        ("Fitted distribution", sanitize_cell(str(cap["fitted_dist"])) if cap.get("fitted_dist") else "N/A"),
        ("Mean", fmt(cap["mean"])),
        ("Sigma Hat (within)", fmt(cap["sigma_hat"])),
        ("Sigma Overall", fmt(cap["sigma_overall"])),
        ("Normality (Shapiro-Wilk p)", fmt(norm["p_value"])),
        ("Approximately Normal?", "Yes" if norm["is_normal"] else "No"),
        ("Stability", stability),
    ]


def _capability_summary_rows(report: CapabilityReport) -> list[tuple[str, object]]:
    return [
        ("Generated", now()),
        ("Tool Version", _TOOL_VERSION),
        ("Engineering Ref", _ENGINEERING_REF),
        ("", ""),
        *_capability_detail_rows(report),
    ]


def build_capability_report_excel(report: CapabilityReport) -> bytes:
    """Excel workbook: a capability/metadata summary sheet + the raw data sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    write_keyvalue_sheet(ws, _capability_summary_rows(report), title="Capability")
    write_table_sheet(
        wb.create_sheet("Data"),
        sanitize_for_export(_values_frame(report.values)),
        title="Data",
        columns=["Point", "Value"],
        col_widths={"Point": 8, "Value": 16},
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_capability_report_pdf(report: CapabilityReport) -> bytes:
    """PDF report: title, Cp/Cpk/Pp/Ppk strip, and a capability detail table."""
    from fpdf import FPDF

    cap = report.capability
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.set_margins(10, 10, 10)
    pdf.add_page()

    pdf_title(pdf, "SPC Process Capability Report")
    pdf_subheader(pdf, generated_line(f"{report.stream_label}  |  {_ENGINEERING_REF}"))
    pdf_summary_cells(
        pdf,
        [
            ("Cp", fmt_opt(cap["cp"])),
            ("Cpk", fmt_opt(cap["cpk"])),
            ("Pp", fmt_opt(cap["pp"])),
            ("Ppk", fmt_opt(cap["ppk"])),
        ],
    )

    details = pd.DataFrame(_capability_detail_rows(report), columns=["Metric", "Value"])
    render_table(
        pdf,
        details,
        columns=[("Metric", 90), ("Value", 100)],
        row_values=lambda r: [safe_text(str(r["Metric"])), safe_text(str(r["Value"]))],
        row_rgb=lambda r: _WHITE_RGB,
    )
    return bytes(pdf.output())
