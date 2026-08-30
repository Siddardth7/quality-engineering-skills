"""
quality_core/sqe/export.py
Supplier Quality Engineering (SQE) vendor-rating live-formula Excel exporter.

Consumer of the domain-agnostic ``quality_core.io.export`` primitives: that module owns
the cross-cutting machinery (formula-injection escaping, openpyxl table styling, the
``Formula`` opt-in marker) and stays free of domain knowledge; this module supplies the
SQE *config* — column order, widths, and which cells are live formulas.

It lives in the ``sqe`` domain package rather than alongside ``io``'s primitives because
the dependency only points one way: ``sqe`` may import ``io`` (and ``sqe.schema``
already does, for the ingest boundary), but ``io`` sits structurally below every domain
engine and must never import one back. Putting an SQE exporter inside ``io`` would close
that arrow into a cycle. For the same reason :class:`SQEVendorRow` is defined here rather
than imported from ``quality_core.canvas.sqe`` (whose ``SQECanvasRow`` has the same
shape): ``canvas`` sits structurally *above* this package.

COMPOSITION DISCIPLINE: this module computes nothing and introduces no numeric constant.
PPM (``=defects/total*1000000``), OTIF (``=on_time_in_full/total_deliveries``), and the
weighted composite (``=SUMPRODUCT(weights, metrics)``) are the same arithmetic
``ppm.py``/``otif.py``/``scorecard.py`` already perform, re-expressed as live Excel
formulas over cells populated from each engine's own result payload. Every heuristic value
the workbook renders — the PPM sample-adequacy minimum, the OTIF window/tolerance
configuration, the scorecard weights/curves/bands, and the escalation thresholds — is
copied verbatim out of that payload's own ``is_heuristic``/``basis`` disclosure and is
never re-typed or re-derived here. PPM and OTIF numerators/denominators are read from the
scorecard dimension's ``source_evidence`` (which *is* ``PPMResult.to_dict()`` /
``OTIFResult.to_dict()``), so the workbook has exactly one source of truth for them.

🔒 THE NO-STANDARD-IMPLIED INVARIANT (carried in from ``ASSUMPTIONS_LOG.md``, RULE-SQE-018
and the "Note on the No-Standard-Implied Invariant" section): ISO 9001:2015 §8.4/§10.2 and
IATF 16949:2016 §8.4 require supplier evaluation against organization-determined criteria;
they define none of the numeric criteria this workbook renders. Every weight and threshold
column/row is therefore visibly labelled ``(HEURISTIC)``, and no column, label, or literal
here may present one as a standards requirement.

What ships in the workbook (two sheets, in this order):

  1. the vendor-scorecard sheet, named by ``title`` (default ``"SQE Vendor Scorecard"``),
     one row per supplier. Three cells per row are live formulas — ``PPM`` (column I),
     ``OTIF`` (column M, storing the raw 0–1 fraction and formatted ``0.0%``), and
     ``Composite_Score`` (column U) — each wrapped in an ``IF`` guard on that row's own
     verdict cell so an INDETERMINATE/NOT_SCORED row resolves to ``"N/A"`` instead of
     ``#DIV/0!``. Every other cell is plain data written through ``write_table_sheet``'s
     ordinary ``sanitize_cell`` escaping.
  2. the ``"Heuristic Config & Metadata"`` sheet — a key/value disclosure of the
     heuristics behind sheet 1. Nothing on it is a live formula: every value is an
     already-computed configuration, not a recomputation. (The name is abbreviated because
     ECMA-376 caps a sheet name at 31 characters; the unabbreviated
     "Heuristic Configuration & Metadata" is 34 and makes Excel report the workbook as
     needing repair.)

Column letters are derived from ``VENDOR_SCORECARD_COLUMNS`` via ``get_column_letter``,
never hand-typed, so reordering the column tuple cannot silently desync a formula from the
column it references.
"""

from __future__ import annotations

import datetime
import io
from collections.abc import Sequence
from dataclasses import dataclass
from typing import Any

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from quality_core.io.export import (
    Formula,
    now,
    write_keyvalue_sheet,
    write_table_sheet,
)
from quality_core.sqe.escalation import EscalationResult, evaluate_escalation
from quality_core.sqe.schema import DeliveryRecord, ReceiptLot, SupplierPeriod
from quality_core.sqe.scorecard import (
    LinearScoringCurve,
    ScorecardConfig,
    ScorecardDimensionResult,
    ScorecardResult,
    calculate_vendor_scorecard,
)

__all__ = [
    "VENDOR_SCORECARD_COLUMNS",
    "VENDOR_SCORECARD_COL_WIDTHS",
    "SQEVendorRow",
    "benchmark_sqe_vendor_rows",
    "build_sqe_workbook",
    "export_sqe_excel",
    "export_sqe_workbook",
]

# ===========================================================================
# Layout Configuration & Column Descriptors
# ===========================================================================

VENDOR_SCORECARD_COLUMNS: tuple[str, ...] = (
    "Supplier_ID",
    "Supplier_Name",
    "Period_Label",
    "Period_Start",
    "Period_End",
    "PPM_Verdict",
    "Defects",
    "Total_Received",
    "PPM",
    "OTIF_Verdict",
    "On_Time_In_Full",
    "Total_Deliveries",
    "OTIF",
    "Quality_Score",
    "Delivery_Score",
    "Cost_Score",
    "Quality_Weight_(HEURISTIC)",
    "Delivery_Weight_(HEURISTIC)",
    "Cost_Weight_(HEURISTIC)",
    "Scorecard_Verdict",
    "Composite_Score",
    "Band",
    "Escalation_Tier",
    "Reason",
)

VENDOR_SCORECARD_COL_WIDTHS: dict[str, float] = {
    "Supplier_ID": 16.0,
    "Supplier_Name": 26.0,
    "Period_Label": 18.0,
    "Period_Start": 14.0,
    "Period_End": 14.0,
    "PPM_Verdict": 16.0,
    "Defects": 12.0,
    "Total_Received": 16.0,
    "PPM": 14.0,
    "OTIF_Verdict": 16.0,
    "On_Time_In_Full": 18.0,
    "Total_Deliveries": 18.0,
    "OTIF": 12.0,
    "Quality_Score": 14.0,
    "Delivery_Score": 14.0,
    "Cost_Score": 14.0,
    "Quality_Weight_(HEURISTIC)": 22.0,
    "Delivery_Weight_(HEURISTIC)": 22.0,
    "Cost_Weight_(HEURISTIC)": 22.0,
    "Scorecard_Verdict": 20.0,
    "Composite_Score": 18.0,
    "Band": 10.0,
    "Escalation_Tier": 22.0,
    "Reason": 48.0,
}

_FIRST_DATA_ROW = 2
#: Abbreviated to stay inside ECMA-376's 31-character sheet-name limit — Excel reports a
#: workbook carrying a longer name as needing repair.
_METADATA_SHEET_TITLE = "Heuristic Config & Metadata"

#: Rendered in the verdict columns when a dimension carries weight 0.0 and was therefore
#: never scored (``ScorecardResult.omitted_dimensions``) — distinct from INDETERMINATE,
#: which means the dimension was weighted but its evidence was undecided.
_NOT_SCORED = "NOT_SCORED"

#: Paraphrase already on record at the top of ``sqe/ASSUMPTIONS_LOG.md`` (and flagged there
#: as a tracked PROCUREMENT-GAP): the ISO/IATF excerpts are not on-machine, so this is a
#: good-faith statement of clause intent, never a verbatim quotation.
_STANDARDS_BASIS = (
    "ISO 9001:2015 §8.4/§10.2 and IATF 16949:2016 §8.4 require supplier evaluation against "
    "organization-determined criteria; those clauses define none of the numeric criteria "
    "below. Every value on this sheet is a caller-configurable engineering heuristic, not a "
    "standard."
)

_PPM_VERDICT_COL = get_column_letter(VENDOR_SCORECARD_COLUMNS.index("PPM_Verdict") + 1)
_DEFECTS_COL = get_column_letter(VENDOR_SCORECARD_COLUMNS.index("Defects") + 1)
_TOTAL_RECEIVED_COL = get_column_letter(VENDOR_SCORECARD_COLUMNS.index("Total_Received") + 1)
_OTIF_VERDICT_COL = get_column_letter(VENDOR_SCORECARD_COLUMNS.index("OTIF_Verdict") + 1)
_OTIF_COUNT_COL = get_column_letter(VENDOR_SCORECARD_COLUMNS.index("On_Time_In_Full") + 1)
_DELIVERY_COUNT_COL = get_column_letter(VENDOR_SCORECARD_COLUMNS.index("Total_Deliveries") + 1)
_QUALITY_SCORE_COL = get_column_letter(VENDOR_SCORECARD_COLUMNS.index("Quality_Score") + 1)
_COST_SCORE_COL = get_column_letter(VENDOR_SCORECARD_COLUMNS.index("Cost_Score") + 1)
_QUALITY_WEIGHT_COL = get_column_letter(
    VENDOR_SCORECARD_COLUMNS.index("Quality_Weight_(HEURISTIC)") + 1
)
_COST_WEIGHT_COL = get_column_letter(
    VENDOR_SCORECARD_COLUMNS.index("Cost_Weight_(HEURISTIC)") + 1
)
_SCORECARD_VERDICT_COL = get_column_letter(
    VENDOR_SCORECARD_COLUMNS.index("Scorecard_Verdict") + 1
)


@dataclass
class SQEVendorRow:
    """One already-evaluated supplier scorecard + escalation result for one period.

    Deliberately independent of ``quality_core.canvas.sqe.SQECanvasRow``: ``canvas`` sits
    structurally above the domain packages, so this module may not import it (see the
    module docstring).
    """

    supplier_id: str
    scorecard: ScorecardResult
    escalation: EscalationResult
    supplier_name: str | None = None

    def __post_init__(self) -> None:
        """Validate identity and result types, mirroring ``SQECanvasRow``'s ingest rules."""
        if not isinstance(self.supplier_id, str) or not self.supplier_id.strip():
            raise TypeError("supplier_id must be a non-empty string")
        self.supplier_id = self.supplier_id.strip()
        if not isinstance(self.scorecard, ScorecardResult):
            raise TypeError("scorecard must be a ScorecardResult")
        if not isinstance(self.escalation, EscalationResult):
            raise TypeError("escalation must be an EscalationResult")
        if self.supplier_name is not None:
            if not isinstance(self.supplier_name, str):
                raise TypeError("supplier_name must be a string or None")
            self.supplier_name = self.supplier_name.strip() or None


def _coerce_rows(
    rows: Sequence[SQEVendorRow] | Sequence[dict[str, Any]],
) -> list[SQEVendorRow]:
    """Normalize the polymorphic ``rows`` argument to a list of :class:`SQEVendorRow`."""
    coerced: list[SQEVendorRow] = []
    for row in rows:
        if isinstance(row, SQEVendorRow):
            coerced.append(row)
        elif isinstance(row, dict):
            coerced.append(SQEVendorRow(**row))
        else:
            raise TypeError(f"Expected SQEVendorRow or dict, got {type(row).__name__}")
    return coerced


def _dimensions(scorecard: ScorecardResult) -> dict[str, ScorecardDimensionResult]:
    """Index one scorecard's scored dimensions by name.

    A dimension is absent exactly when its configured weight is ``0.0``
    (``scorecard.py`` omits it and records the omission in ``omitted_dimensions``), so an
    absent dimension's score and weight are provably ``0.0`` — writing those literals
    fabricates nothing and leaves ``SUMPRODUCT`` unchanged.
    """
    return {dimension.name: dimension for dimension in scorecard.dimensions}


def _sub_score(dimension: ScorecardDimensionResult | None) -> float | None:
    """Return a dimension's sub-score, or ``0.0`` for a dimension that carries no weight."""
    return 0.0 if dimension is None else dimension.sub_score


def _weight(dimension: ScorecardDimensionResult | None) -> float:
    """Return a dimension's weight, or ``0.0`` for a dimension that carries no weight."""
    return 0.0 if dimension is None else dimension.weight


def _evidence(dimension: ScorecardDimensionResult | None) -> dict[str, Any]:
    """Return a dimension's source-engine payload, or an empty mapping when unscored."""
    return {} if dimension is None else dimension.source_evidence


def _row_record(row: SQEVendorRow, row_idx: int) -> dict[str, Any]:
    """Map one evaluated supplier to its scorecard-sheet row, with three live formulas.

    PPM/OTIF counts come from the scorecard dimension's ``source_evidence`` — the source
    engines' own ``to_dict()`` payloads — never from a second call to ``ppm``/``otif``.
    """
    scorecard = row.scorecard
    dimensions = _dimensions(scorecard)
    quality = dimensions.get("quality")
    delivery = dimensions.get("delivery")
    cost = dimensions.get("cost")
    quality_evidence = _evidence(quality)
    delivery_evidence = _evidence(delivery)

    # Each IF guards its own row's verdict cell. Excel does not evaluate the untaken
    # branch, so the division is never reached on an INDETERMINATE/NOT_SCORED row whose
    # numerator/denominator cells are blank.
    ppm_formula = Formula(
        f'=IF({_PPM_VERDICT_COL}{row_idx}<>"MEASURED","N/A",'
        f"{_DEFECTS_COL}{row_idx}/{_TOTAL_RECEIVED_COL}{row_idx}*1000000)"
    )
    # Stores the raw 0–1 fraction (the engine's otif_pct is on a 0–100 scale); the
    # percentage is a display concern, handled by the number format, not by the arithmetic.
    otif_formula = Formula(
        f'=IF({_OTIF_VERDICT_COL}{row_idx}<>"MEASURED","N/A",'
        f"{_OTIF_COUNT_COL}{row_idx}/{_DELIVERY_COUNT_COL}{row_idx})",
        number_format="0.0%",
    )
    composite_formula = Formula(
        f'=IF({_SCORECARD_VERDICT_COL}{row_idx}<>"RATED","N/A",'
        f"SUMPRODUCT({_QUALITY_WEIGHT_COL}{row_idx}:{_COST_WEIGHT_COL}{row_idx},"
        f"{_QUALITY_SCORE_COL}{row_idx}:{_COST_SCORE_COL}{row_idx}))"
    )

    return {
        "Supplier_ID": scorecard.supplier_id,
        "Supplier_Name": row.supplier_name,
        "Period_Label": scorecard.period_label,
        "Period_Start": scorecard.period_start.isoformat(),
        "Period_End": scorecard.period_end.isoformat(),
        "PPM_Verdict": quality_evidence.get("verdict", _NOT_SCORED),
        "Defects": quality_evidence.get("numerator"),
        "Total_Received": quality_evidence.get("denominator"),
        "PPM": ppm_formula,
        "OTIF_Verdict": delivery_evidence.get("verdict", _NOT_SCORED),
        "On_Time_In_Full": delivery_evidence.get("otif_count"),
        "Total_Deliveries": delivery_evidence.get("delivery_count"),
        "OTIF": otif_formula,
        "Quality_Score": _sub_score(quality),
        "Delivery_Score": _sub_score(delivery),
        "Cost_Score": _sub_score(cost),
        "Quality_Weight_(HEURISTIC)": _weight(quality),
        "Delivery_Weight_(HEURISTIC)": _weight(delivery),
        "Cost_Weight_(HEURISTIC)": _weight(cost),
        "Scorecard_Verdict": scorecard.verdict,
        "Composite_Score": composite_formula,
        "Band": scorecard.band,
        "Escalation_Tier": row.escalation.tier,
        "Reason": scorecard.reason or row.escalation.reason,
    }


def _metadata_rows(rows: Sequence[SQEVendorRow], title: str) -> list[tuple[str, object]]:
    """Build the heuristic-disclosure key/value rows from the first supplier's payloads.

    All suppliers in one export are expected to share one configuration; where they do not,
    this sheet documents the first row's configuration — an accepted limitation, stated
    here rather than papered over. Every value is copied verbatim from an engine payload.
    """
    if not rows:
        return [
            (
                "Heuristic Configuration",
                "No supplier rows supplied; heuristic configuration unavailable.",
            )
        ]

    first = rows[0]
    dimensions = _dimensions(first.scorecard)
    quality = dimensions.get("quality")
    delivery = dimensions.get("delivery")

    sample_adequacy_minimum: object = _NOT_SCORED
    sample_adequacy_basis: object = _NOT_SCORED
    if quality is not None:
        # PPMResult.to_dict()["sample_adequacy"] — the engine's own labelled disclosure.
        sample_adequacy = quality.source_evidence["sample_adequacy"]
        sample_adequacy_minimum = sample_adequacy["minimum"]
        sample_adequacy_basis = sample_adequacy["basis"]

    otif_configuration: object = _NOT_SCORED
    if delivery is not None:
        otif_configuration = str(delivery.source_evidence["heuristic_configuration"])

    return [
        ("Report Title", title),
        ("Date Generated", now()),
        ("Standards Basis (Vendor Rating)", _STANDARDS_BASIS),
        ("PPM Sample-Adequacy Minimum (HEURISTIC)", sample_adequacy_minimum),
        ("PPM Sample-Adequacy Basis", sample_adequacy_basis),
        ("OTIF Heuristic Configuration (HEURISTIC)", otif_configuration),
        (
            "Scorecard Weights / Curves / Bands (HEURISTIC)",
            str(first.scorecard.heuristic_configuration),
        ),
        ("Escalation Thresholds (HEURISTIC)", str(first.escalation.heuristic_configuration)),
    ]


# ===========================================================================
# Public Exporter API
# ===========================================================================


def build_sqe_workbook(
    rows: Sequence[SQEVendorRow] | Sequence[dict[str, Any]],
    *,
    title: str = "SQE Vendor Scorecard",
) -> openpyxl.Workbook:
    """Build the 2-sheet vendor-rating openpyxl Workbook with live formulas.

    Sheets:
      1. ``title`` (default ``"SQE Vendor Scorecard"``): one row per supplier, with live
         formulas for ``PPM`` (``=IF(F{r}<>"MEASURED","N/A",G{r}/H{r}*1000000)``), ``OTIF``
         (``=IF(J{r}<>"MEASURED","N/A",K{r}/L{r})``) and ``Composite_Score``
         (``=IF(T{r}<>"RATED","N/A",SUMPRODUCT(Q{r}:S{r},N{r}:P{r}))``).
      2. ``"Heuristic Config & Metadata"``: the heuristic disclosure behind sheet 1.

    An empty ``rows`` sequence yields a valid, loadable header-only scorecard sheet and a
    metadata sheet stating that the heuristic configuration is unavailable.
    """
    resolved_rows = _coerce_rows(rows)

    wb = Workbook()

    # Sheet 1: Vendor Scorecard. dtype=object keeps an unmeasured count as None (a blank
    # cell) instead of letting pandas widen the column to float and write NaN.
    records = [
        _row_record(row, row_idx=idx + _FIRST_DATA_ROW)
        for idx, row in enumerate(resolved_rows)
    ]
    df = pd.DataFrame(records, columns=list(VENDOR_SCORECARD_COLUMNS), dtype=object)
    write_table_sheet(
        wb.active,
        df,
        title=title,
        columns=VENDOR_SCORECARD_COLUMNS,
        col_widths=VENDOR_SCORECARD_COL_WIDTHS,
    )

    # Sheet 2: Heuristic Configuration & Metadata (no live formulas).
    ws_meta = wb.create_sheet(_METADATA_SHEET_TITLE)
    write_keyvalue_sheet(
        ws_meta,
        _metadata_rows(resolved_rows, title),
        title=_METADATA_SHEET_TITLE,
        key_width=44.0,
        value_width=96.0,
    )

    return wb


def export_sqe_workbook(
    rows: Sequence[SQEVendorRow] | Sequence[dict[str, Any]],
    *,
    title: str = "SQE Vendor Scorecard",
) -> bytes:
    """Export evaluated supplier rows to serialized .xlsx bytes."""
    wb = build_sqe_workbook(rows, title=title)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_sqe_excel(
    rows: Sequence[SQEVendorRow] | Sequence[dict[str, Any]],
    *,
    title: str = "SQE Vendor Scorecard",
) -> bytes:
    """Alias for :func:`export_sqe_workbook`."""
    return export_sqe_workbook(rows, title=title)


# ===========================================================================
# Benchmark Dataset
# ===========================================================================


def benchmark_sqe_vendor_rows() -> list[SQEVendorRow]:
    """Return a fresh 3-supplier benchmark, evaluated through the SQE engines.

    A fresh list of fresh ``ScorecardResult``/``EscalationResult`` objects on every call.
    The three suppliers exercise the workbook's three shapes:

      * ``SUP-A`` — RATED on the default config, so the zero-weight cost dimension is
        omitted from the scorecard and renders as ``0.0`` score / ``0.0`` weight.
      * ``SUP-B`` — RATED with all three dimensions positively weighted, so the
        ``SUMPRODUCT`` is exercised as a real 3-term product.
      * ``SUP-C`` — fully INDETERMINATE (no receipt lots, no deliveries), so every live
        formula renders its guarded ``"N/A"`` branch over blank count cells.
    """
    period_start = datetime.date(2026, 1, 1)
    period_end = datetime.date(2026, 1, 31)

    # --- SUP-A: default weights (cost_weight 0.0 => cost dimension omitted) -------------
    period_a = SupplierPeriod(
        supplier_id="SUP-A",
        period_start=period_start,
        period_end=period_end,
        period_label="January 2026",
    )
    lots_a = [
        ReceiptLot(
            supplier_id="SUP-A",
            lot_id="LOT-A-001",
            quantity_received=1200,
            receipt_date=datetime.date(2026, 1, 8),
            defect_count=3,
        ),
        ReceiptLot(
            supplier_id="SUP-A",
            lot_id="LOT-A-002",
            quantity_received=800,
            receipt_date=datetime.date(2026, 1, 22),
            defect_count=1,
        ),
    ]
    deliveries_a = [
        DeliveryRecord(
            supplier_id="SUP-A",
            order_id="PO-A-001",
            quantity_ordered=400,
            quantity_delivered=400,
            promised_date=datetime.date(2026, 1, 9),
            actual_delivery_date=datetime.date(2026, 1, 9),
        ),
        DeliveryRecord(
            supplier_id="SUP-A",
            order_id="PO-A-002",
            quantity_ordered=300,
            quantity_delivered=300,
            promised_date=datetime.date(2026, 1, 16),
            actual_delivery_date=datetime.date(2026, 1, 17),
        ),
        DeliveryRecord(
            supplier_id="SUP-A",
            order_id="PO-A-003",
            quantity_ordered=500,
            quantity_delivered=460,
            promised_date=datetime.date(2026, 1, 23),
            actual_delivery_date=datetime.date(2026, 1, 23),
        ),
        DeliveryRecord(
            supplier_id="SUP-A",
            order_id="PO-A-004",
            quantity_ordered=250,
            quantity_delivered=250,
            promised_date=datetime.date(2026, 1, 29),
            actual_delivery_date=datetime.date(2026, 1, 30),
        ),
    ]
    scorecard_a = calculate_vendor_scorecard(period_a, lots_a, deliveries_a)

    # --- SUP-B: all three dimensions positively weighted --------------------------------
    period_b = SupplierPeriod(
        supplier_id="SUP-B",
        period_start=period_start,
        period_end=period_end,
        period_label="January 2026",
    )
    lots_b = [
        ReceiptLot(
            supplier_id="SUP-B",
            lot_id="LOT-B-001",
            quantity_received=2500,
            receipt_date=datetime.date(2026, 1, 12),
            defect_count=10,
        ),
    ]
    deliveries_b = [
        DeliveryRecord(
            supplier_id="SUP-B",
            order_id="PO-B-001",
            quantity_ordered=900,
            quantity_delivered=900,
            promised_date=datetime.date(2026, 1, 13),
            actual_delivery_date=datetime.date(2026, 1, 13),
        ),
        DeliveryRecord(
            supplier_id="SUP-B",
            order_id="PO-B-002",
            quantity_ordered=800,
            quantity_delivered=800,
            promised_date=datetime.date(2026, 1, 26),
            actual_delivery_date=datetime.date(2026, 1, 28),
        ),
    ]
    copq_items_b: list[dict[str, Any]] = [
        {
            "category": "InternalFailure",
            "description": "Incoming bore-porosity scrap at receiving inspection",
            "direct_cost": 8200.0,
        },
        {
            "category": "ExternalFailure",
            "description": "Customer line-stop containment sort and premium freight",
            "direct_cost": 4300.0,
        },
    ]
    config_b = ScorecardConfig(
        quality_weight=0.50,
        delivery_weight=0.30,
        cost_weight=0.20,
        cost_curve=LinearScoringCurve(best_value=0.0, worst_value=10.0),
    )
    scorecard_b = calculate_vendor_scorecard(
        period_b,
        lots_b,
        deliveries_b,
        copq_items=copq_items_b,
        revenue_base=750_000.0,
        config=config_b,
    )

    # --- SUP-C: no evidence at all => INDETERMINATE on both weighted dimensions ---------
    period_c = SupplierPeriod(
        supplier_id="SUP-C",
        period_start=period_start,
        period_end=period_end,
        period_label="January 2026",
    )
    scorecard_c = calculate_vendor_scorecard(period_c)

    return [
        SQEVendorRow(
            supplier_id="SUP-A",
            scorecard=scorecard_a,
            escalation=evaluate_escalation(scorecard_a),
            supplier_name="Benchmark Precision Machining",
        ),
        SQEVendorRow(
            supplier_id="SUP-B",
            scorecard=scorecard_b,
            escalation=evaluate_escalation(scorecard_b),
            supplier_name="Benchmark Castings & Forgings",
        ),
        SQEVendorRow(
            supplier_id="SUP-C",
            scorecard=scorecard_c,
            escalation=evaluate_escalation(scorecard_c),
            supplier_name="Benchmark Surface Treatment",
        ),
    ]
