"""
tests/test_export_fmea.py
Tests for quality_core/io/export_fmea.py — the FMEA live-formula .xlsx exporter (#142).

The point of this suite is the two mandatory negative controls: a literal-int RPN build
must FAIL ``assert_cell_is_formula`` (proving the live-formula check is load-bearing, not
tautological), and a free-text field starting with ``"="`` must render inert while its
row's RPN cell stays a live formula (proving the ``Formula`` opt-in is per-cell). The
accuracy scorecard reads S/O/D back out of the *saved* sheet and multiplies them — it does
not trust ``data_only=True`` (openpyxl caches no value for a never-recalculated formula,
so that path reads ``None``).
"""

from __future__ import annotations

import io

import openpyxl
import pandas as pd
import pytest
from _xlsx_formula_audit import assert_cell_is_formula
from openpyxl.utils import get_column_letter
from quality_core import scoring
from quality_core.io.export import write_table_sheet
from quality_core.io.export_fmea import (
    FMEA_COL_WIDTHS,
    FMEA_EXPORT_COLUMNS,
    benchmark_fmea_dataset,
    export_fmea_workbook,
)
from quality_core.schema.fmea import FMEADataset, FMEARow

# Column letters derived the same way the exporter does, so a reorder of
# FMEA_EXPORT_COLUMNS moves the test's expectations with the implementation's.
_S_COL = get_column_letter(FMEA_EXPORT_COLUMNS.index("Severity") + 1)
_O_COL = get_column_letter(FMEA_EXPORT_COLUMNS.index("Occurrence") + 1)
_D_COL = get_column_letter(FMEA_EXPORT_COLUMNS.index("Detection") + 1)
_ID_COL = get_column_letter(FMEA_EXPORT_COLUMNS.index("ID") + 1)
_RPN_COL = get_column_letter(FMEA_EXPORT_COLUMNS.index("RPN") + 1)
_AP_COL = get_column_letter(FMEA_EXPORT_COLUMNS.index("Action_Priority") + 1)

_VALID_ROW: dict[str, object] = {
    "ID": 1,
    "Process_Step": "Mix",
    "Component": "Resin",
    "Function": "Bond layers",
    "Failure_Mode": "Incomplete cure",
    "Effect": "Delamination",
    "Severity": 8,
    "Cause": "Low temperature",
    "Occurrence": 4,
    "Current_Control": "Oven thermocouple",
    "Detection": 5,
}


def _make_row(**overrides: object) -> FMEARow:
    return FMEARow(**{**_VALID_ROW, **overrides})  # type: ignore[arg-type]


def _saved(wb: openpyxl.Workbook) -> bytes:
    """Serialize a workbook to .xlsx bytes (tests never touch disk)."""
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _load(wb_bytes: bytes) -> openpyxl.Workbook:
    """Reload saved bytes with formulas intact (never data_only — no cache exists)."""
    return openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)


# --- 1 · valid workbook -------------------------------------------------------


def test_export_benchmark_is_a_loadable_xlsx_with_header_and_row_count():
    dataset = benchmark_fmea_dataset()
    wb_bytes = export_fmea_workbook(dataset)
    assert isinstance(wb_bytes, bytes)

    wb = _load(wb_bytes)
    ws = wb["FMEA"]
    assert [c.value for c in ws[1]] == list(FMEA_EXPORT_COLUMNS)
    assert ws.max_row == len(dataset.rows) + 1


def test_export_uses_custom_title_as_sheet_name():
    wb_bytes = export_fmea_workbook(benchmark_fmea_dataset(), title="PFMEA-2026")
    wb = _load(wb_bytes)
    assert wb.sheetnames == ["PFMEA-2026"]


# --- 2 · positive live-formula proof -----------------------------------------


def test_rpn_cells_are_live_formulas_first_and_last_row():
    dataset = benchmark_fmea_dataset()
    wb_bytes = export_fmea_workbook(dataset)
    n = len(dataset.rows)

    first_r = 2
    last_r = n + 1
    for r in (first_r, last_r):
        # Does not raise: the cell carries a live <f> element.
        assert_cell_is_formula(wb_bytes, "FMEA", f"{_RPN_COL}{r}")

    ws = _load(wb_bytes)["FMEA"]
    for r in (first_r, last_r):
        assert ws[f"{_RPN_COL}{r}"].value == f"={_S_COL}{r}*{_O_COL}{r}*{_D_COL}{r}"


# --- 3 · literal-build negative control ---------------------------------------


def test_literal_rpn_build_fails_the_formula_verifier():
    """NEGATIVE CONTROL: a hardcoded-int RPN column must FAIL ``assert_cell_is_formula``.

    If a future edit swapped ``Formula(...)`` for a plain ``s*o*d`` int in the exporter,
    this suite must catch it — so prove the verifier bites on a literal build of the same
    column layout.
    """
    row = _make_row(Severity=8, Occurrence=4, Detection=5)
    record = {name: getattr(row, name) for name in FMEA_EXPORT_COLUMNS[:-2]}
    record["RPN"] = row.Severity * row.Occurrence * row.Detection  # literal int, NOT Formula
    record["Action_Priority"] = scoring.action_priority(
        row.Severity, row.Occurrence, row.Detection
    )
    df = pd.DataFrame([record], columns=list(FMEA_EXPORT_COLUMNS))

    wb = openpyxl.Workbook()
    write_table_sheet(
        wb.active,
        df,
        title="FMEA",
        columns=FMEA_EXPORT_COLUMNS,
        col_widths=FMEA_COL_WIDTHS,
    )
    saved = _saved(wb)
    with pytest.raises(AssertionError, match="literal, not a live formula"):
        assert_cell_is_formula(saved, "FMEA", f"{_RPN_COL}2")


# --- 4 · accuracy scorecard ---------------------------------------------------


def test_fmea_export_accuracy_scorecard_rpn_and_ap():
    """The saved FORMULA itself must recompute to the engine's own RPN, per row.

    # reproduce:
    #   uv run pytest packages/quality-core -k fmea_export_accuracy -q
    #   (from repo root; or `cd packages/quality-core && uv run pytest \
    #    tests/test_export_fmea.py -k fmea_export_accuracy -q`)

    Reads S/O/D back out of the saved sheet and multiplies — never data_only=True, which
    returns None for a never-recalculated formula.
    """
    dataset = benchmark_fmea_dataset()
    wb_bytes = export_fmea_workbook(dataset)
    ws = _load(wb_bytes)["FMEA"]

    for i, row in enumerate(dataset.rows):
        r = i + 2

        # The stored formula matches this exporter's own template exactly.
        assert ws[f"{_RPN_COL}{r}"].value == f"={_S_COL}{r}*{_O_COL}{r}*{_D_COL}{r}"

        # Read the formula's OWN referenced S/O/D data cells from the saved sheet.
        s = ws[f"{_S_COL}{r}"].value
        o = ws[f"{_O_COL}{r}"].value
        d = ws[f"{_D_COL}{r}"].value
        product = s * o * d

        assert product == row.RPN
        assert product == scoring.rpn(s, o, d)

        # AP is a structured string, equal to the engine's lookup for this row.
        assert ws[f"{_AP_COL}{r}"].value == scoring.action_priority(
            row.Severity, row.Occurrence, row.Detection
        )


def test_benchmark_exercises_all_three_ap_bands():
    """The benchmark must span High/Medium/Low so the AP scorecard is not one-note."""
    dataset = benchmark_fmea_dataset()
    bands = {
        scoring.action_priority(row.Severity, row.Occurrence, row.Detection)
        for row in dataset.rows
    }
    assert bands == {scoring.HIGH, scoring.MEDIUM, scoring.LOW}


# --- 5 · AP is a structured value, not a formula ------------------------------


def test_action_priority_cell_is_not_a_live_formula():
    dataset = benchmark_fmea_dataset()
    wb_bytes = export_fmea_workbook(dataset)

    for i, row in enumerate(dataset.rows):
        r = i + 2
        with pytest.raises(AssertionError, match="literal, not a live formula"):
            assert_cell_is_formula(wb_bytes, "FMEA", f"{_AP_COL}{r}")


# --- 6 · injection safety negative control ------------------------------------


def test_injection_freetext_is_inert_but_row_rpn_stays_live():
    """NEGATIVE CONTROL: a data field starting with '=' must render inert, per-cell.

    The malicious ``Current_Control`` must be apostrophe-escaped (never a live formula),
    while the SAME row's RPN cell stays live — proving the ``Formula`` opt-in is per-cell,
    not suppressed by an untrusted neighbor.
    """
    payload = "=cmd|' /C calc'!A0"
    dataset = FMEADataset(rows=[_make_row(Current_Control=payload)])
    wb_bytes = export_fmea_workbook(dataset)

    control_col = get_column_letter(FMEA_EXPORT_COLUMNS.index("Current_Control") + 1)
    ws = _load(wb_bytes)["FMEA"]
    assert ws[f"{control_col}2"].value == "'" + payload

    with pytest.raises(AssertionError, match="literal, not a live formula"):
        assert_cell_is_formula(wb_bytes, "FMEA", f"{control_col}2")

    # ...but the adjacent RPN cell in the same row is still live.
    assert_cell_is_formula(wb_bytes, "FMEA", f"{_RPN_COL}2")


# --- 7 · empty dataset branch -------------------------------------------------


def test_empty_dataset_yields_header_only_workbook():
    wb_bytes = export_fmea_workbook(FMEADataset(rows=[]))
    ws = _load(wb_bytes)["FMEA"]
    assert ws.max_row == 1
    assert [c.value for c in ws[1]] == list(FMEA_EXPORT_COLUMNS)


# --- extra · row order preserved & re-export surface --------------------------


def test_row_order_is_preserved_not_sorted():
    dataset = FMEADataset(
        rows=[_make_row(ID=3), _make_row(ID=1), _make_row(ID=2)]
    )
    wb_bytes = export_fmea_workbook(dataset)
    ws = _load(wb_bytes)["FMEA"]
    assert [ws[f"{_ID_COL}{r}"].value for r in (2, 3, 4)] == [3, 1, 2]


def test_export_fmea_workbook_is_reexported_from_quality_core_io():
    import quality_core.io as qio
    from quality_core.io import export_fmea as mod

    assert qio.export_fmea_workbook is mod.export_fmea_workbook
    assert qio.benchmark_fmea_dataset is mod.benchmark_fmea_dataset


def test_benchmark_dataset_returns_a_fresh_object_each_call():
    a = benchmark_fmea_dataset()
    b = benchmark_fmea_dataset()
    assert a is not b
    assert a.rows is not b.rows
